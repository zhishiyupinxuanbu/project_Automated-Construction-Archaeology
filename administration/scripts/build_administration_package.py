#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable


HEADERS = [
    "项目",
    "卷",
    "文号",
    "编制单位（承担单位）",
    "题名",
    "编制时间",
    "批准单位",
    "批准时间",
    "张数",
    "备注",
    "扫描时间",
    "归还时间",
]

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}
PDF_EXTS = {".pdf"}
SKIP_NAMES = {".DS_Store"}
INVALID_CHARS_RE = re.compile(r'[\\/:*?"<>|]+')
NUM_RE = re.compile(r"(\d+)")

DOCNO_PATTERNS = [
    re.compile(r"[\u4e00-\u9fffA-Za-z]{1,16}(?:发|字|函|报|文|政|办|办法|政办发|政字|文旅发|文物发)\s*[〔【\[\(（]\s*\d{4}\s*[〕】\]\)）]\s*\d+\s*号?"),
    re.compile(r"[\u4e00-\u9fffA-Za-z]{1,16}(?:发|字|函|报|文|政|办|办法|政办发|政字|文旅发|文物发)\s+\d{4}\s+\d+\s*号?"),
]
FULL_DATE_RE = re.compile(r"((?:19|20)\d{2})年(\d{1,2})月(\d{1,2})日")
ISO_DATE_RE = re.compile(r"((?:19|20)\d{2})[-./](\d{1,2})[-./](\d{1,2})")
ISSUER_RE = re.compile(
    r"([\u4e00-\u9fff·]{2,40}(?:人民政府办公室|人民政府|文化和旅游局|文物局|文物保护管理所|文物保护和旅游事业发展中心|财政局|公安局|委员会|办公室|文化厅|文物局办公室))"
)


@dataclass
class SourceItem:
    source_path: str
    source_kind: str
    source_name: str
    page_sources: list[str] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)
    parent_note: str = ""


@dataclass
class BuiltRecord:
    seq: int
    input_order: int
    source_path: str
    source_kind: str
    original_name: str
    title: str
    folder_title: str
    docno: str
    issuer: str
    date: str
    date_source: str
    date_candidates: list[str]
    page_count: int
    output_folder: str
    notes: list[str]
    first_page_text: str = ""
    last_page_text: str = ""
    blank_pages_removed: int = 0


def run(cmd: list[str], *, timeout: int = 600, cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=timeout,
        check=False,
    )


def command_exists(name: str) -> bool:
    return shutil.which(name) is not None


def natural_key(value: str) -> list[object]:
    return [int(part) if part.isdigit() else part.lower() for part in NUM_RE.split(value)]


def image_files(path: Path) -> list[Path]:
    return sorted(
        (
            child
            for child in path.iterdir()
            if child.is_file()
            and child.name not in SKIP_NAMES
            and not child.name.startswith("._")
            and child.suffix.lower() in IMAGE_EXTS
        ),
        key=lambda p: natural_key(p.name),
    )


def safe_name(value: str, *, limit: int = 120) -> str:
    value = INVALID_CHARS_RE.sub("-", value)
    value = re.sub(r"\s+", " ", value).strip(" .-_")
    return (value or "题名待核")[:limit]


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def discover_inputs(paths: Iterable[Path]) -> tuple[list[SourceItem], list[dict[str, str]]]:
    items: list[SourceItem] = []
    unsupported: list[dict[str, str]] = []

    def add_directory_item(path: Path, note: str = "") -> None:
        imgs = image_files(path)
        if not imgs:
            return
        issues = page_sequence_issues(imgs)
        if note:
            issues.append(note)
        items.append(
            SourceItem(
                source_path=str(path.resolve()),
                source_kind="image_folder",
                source_name=path.name,
                page_sources=[str(p.resolve()) for p in imgs],
                issues=issues,
                parent_note=note,
            )
        )

    def walk(path: Path, top_level: bool = False) -> None:
        if not path.exists():
            unsupported.append({"path": str(path), "reason": "路径不存在"})
            return
        if path.is_file():
            if path.name in SKIP_NAMES or path.name.startswith("._"):
                return
            if path.suffix.lower() in PDF_EXTS:
                items.append(SourceItem(str(path.resolve()), "pdf", path.stem))
            elif path.suffix.lower() in IMAGE_EXTS:
                items.append(SourceItem(str(path.resolve()), "single_image", path.stem, [str(path.resolve())]))
            else:
                unsupported.append({"path": str(path.resolve()), "reason": f"暂不支持扩展名 {path.suffix}"})
            return

        direct_imgs = image_files(path)
        children = sorted(
            [child for child in path.iterdir() if child.name not in SKIP_NAMES and not child.name.startswith("._")],
            key=lambda p: natural_key(p.name),
        )
        has_child_records = any(child.is_dir() or child.suffix.lower() in PDF_EXTS for child in children if child.is_file() or child.is_dir())
        note = "同一目录同时存在根目录页图和子资料，根目录页图按独立资料处理" if direct_imgs and has_child_records else ""
        add_directory_item(path, note)

        for child in children:
            if child.is_dir():
                walk(child)
            elif child.is_file() and child.suffix.lower() in PDF_EXTS:
                items.append(SourceItem(str(child.resolve()), "pdf", child.stem))
            elif child.is_file() and child.suffix.lower() not in IMAGE_EXTS:
                unsupported.append({"path": str(child.resolve()), "reason": f"暂不支持扩展名 {child.suffix}"})

    for raw in paths:
        walk(raw, top_level=True)

    dedup: dict[tuple[str, str], SourceItem] = {}
    for item in items:
        key = (item.source_kind, item.source_path)
        dedup.setdefault(key, item)
    return list(dedup.values()), unsupported


def page_sequence_issues(files: list[Path]) -> list[str]:
    numbers: list[int] = []
    missing_number = False
    for path in files:
        matches = NUM_RE.findall(path.stem)
        if matches:
            numbers.append(int(matches[-1]))
        else:
            missing_number = True
    issues: list[str] = []
    if missing_number:
        issues.append("部分页图文件名未含数字页码，已按自然文件名排序")
    if numbers:
        seen: set[int] = set()
        duplicates = sorted({n for n in numbers if n in seen or seen.add(n)})
        if duplicates:
            issues.append(f"页码数字重复: {duplicates}")
        if len(numbers) == len(files):
            ordered = sorted(numbers)
            gaps = [n for n in range(ordered[0], ordered[-1] + 1) if n not in set(ordered)]
            if gaps:
                issues.append(f"页码数字不连续，缺: {gaps[:20]}")
    return issues


def pdf_page_count(pdf: Path) -> int:
    if not command_exists("pdfinfo"):
        return 0
    cp = run(["pdfinfo", str(pdf)], timeout=120)
    if cp.returncode != 0:
        return 0
    match = re.search(r"^Pages:\s+(\d+)", cp.stdout, re.M)
    return int(match.group(1)) if match else 0


def pdf_page_text(pdf: Path, page: int) -> str:
    if not command_exists("pdftotext"):
        return ""
    cp = run(["pdftotext", "-f", str(page), "-l", str(page), str(pdf), "-"], timeout=180)
    if cp.returncode != 0:
        return ""
    return normalize_spaces(cp.stdout)


def pdf_first_page_text(pdf: Path) -> str:
    return pdf_page_text(pdf, 1)


def is_blank_image(image_path: Path, *, dark_ratio_threshold: float, stddev_threshold: float) -> bool:
    from PIL import Image, ImageStat

    with Image.open(image_path) as image:
        gray = image.convert("L")
        gray.thumbnail((600, 600))
        stat = ImageStat.Stat(gray)
        stddev = stat.stddev[0] if stat.stddev else 0
        hist = gray.histogram()
        dark_pixels = sum(hist[:245])
        total = sum(hist) or 1
        dark_ratio = dark_pixels / total
    return dark_ratio <= dark_ratio_threshold and stddev <= stddev_threshold


def render_pdf(
    pdf: Path,
    out_dir: Path,
    *,
    dpi: int,
    remove_blank_pages: bool,
    dark_ratio_threshold: float,
    stddev_threshold: float,
) -> tuple[list[Path], list[str], int, list[int]]:
    issues: list[str] = []
    blank_removed = 0
    if not command_exists("pdftoppm"):
        raise RuntimeError("pdftoppm not found; cannot render PDFs")
    with tempfile.TemporaryDirectory() as tmp:
        prefix = Path(tmp) / "page"
        cp = run(["pdftoppm", "-r", str(dpi), "-png", str(pdf), str(prefix)], timeout=1800)
        if cp.returncode != 0:
            raise RuntimeError((cp.stderr or cp.stdout or "pdftoppm failed")[-2000:])
        rendered = sorted(Path(tmp).glob("page-*.png"), key=lambda p: natural_key(p.name))
        if not rendered:
            raise RuntimeError("pdftoppm produced no pages")
        out_dir.mkdir(parents=True, exist_ok=True)
        pages: list[Path] = []
        expected = pdf_page_count(pdf)
        if expected and expected != len(rendered):
            issues.append(f"PDF页数 {expected} 与渲染页图数 {len(rendered)} 不一致")
        kept: list[Path] = []
        kept_original_pages: list[int] = []
        for original_idx, page in enumerate(rendered, 1):
            if remove_blank_pages and is_blank_image(
                page,
                dark_ratio_threshold=dark_ratio_threshold,
                stddev_threshold=stddev_threshold,
            ):
                blank_removed += 1
                issues.append(f"删除空白页: 原PDF第{original_idx}页")
                continue
            kept.append(page)
            kept_original_pages.append(original_idx)
        for idx, page in enumerate(kept, 1):
            dest = out_dir / f"page_{idx:03d}.png"
            shutil.copy2(page, dest)
            pages.append(dest)
        if rendered and not kept:
            issues.append("全部页被判定为空白页，输出页数为0，需人工复核")
        return pages, issues, blank_removed, kept_original_pages


def copy_images_as_png(
    files: list[Path],
    out_dir: Path,
    *,
    remove_blank_pages: bool,
    dark_ratio_threshold: float,
    stddev_threshold: float,
) -> tuple[list[Path], list[str], int, list[str]]:
    from PIL import Image

    out_dir.mkdir(parents=True, exist_ok=True)
    pages: list[Path] = []
    issues = page_sequence_issues(files)
    blank_removed = 0
    kept: list[Path] = []
    kept_source_names: list[str] = []
    for original_idx, src in enumerate(files, 1):
        try:
            if remove_blank_pages and is_blank_image(
                src,
                dark_ratio_threshold=dark_ratio_threshold,
                stddev_threshold=stddev_threshold,
            ):
                blank_removed += 1
                issues.append(f"删除空白页: 原图{original_idx} {src.name}")
                continue
        except Exception as exc:  # noqa: BLE001
            issues.append(f"空白页检测失败 {src.name}: {exc}")
        kept.append(src)
        kept_source_names.append(src.name)
    for idx, src in enumerate(kept, 1):
        dest = out_dir / f"page_{idx:03d}.png"
        try:
            with Image.open(src) as image:
                if image.mode not in {"RGB", "RGBA"}:
                    image = image.convert("RGB")
                image.save(dest)
        except Exception as exc:  # noqa: BLE001
            issues.append(f"页图转换失败 {src.name}: {exc}")
            continue
        pages.append(dest)
    if files and not kept:
        issues.append("全部页被判定为空白页，输出页数为0，需人工复核")
    return pages, issues, blank_removed, kept_source_names


def ocr_image(image: Path, *, ocr_dir: Path) -> tuple[str, str]:
    script = ocr_dir / "scripts" / "ocr_image.sh"
    if not script.exists():
        return "", f"OCR脚本不存在: {script}"
    with tempfile.TemporaryDirectory() as tmp:
        out_json = Path(tmp) / "ocr.json"
        cp = run([str(script), str(image), "--output", str(out_json)], timeout=900, cwd=ocr_dir)
        if cp.returncode != 0:
            return "", compact_error(cp.stderr or cp.stdout or "OCR failed")
        try:
            data = json.loads(out_json.read_text(encoding="utf-8"))
        except Exception as exc:  # noqa: BLE001
            return "", f"OCR结果读取失败: {exc}"
    if isinstance(data.get("text"), str):
        return normalize_spaces(data["text"]), ""
    lines = data.get("lines") or []
    text = " ".join(str(line.get("text", "")) for line in lines if isinstance(line, dict))
    return normalize_spaces(text), ""


def compact_error(value: str, limit: int = 220) -> str:
    value = normalize_spaces(value)
    if "Connection refused" in value or "Failed to establish a new connection" in value:
        return "OCR服务连接失败，请确认本地Paddle OCR服务已启动"
    return value[-limit:]


def extract_docno(text: str) -> str:
    for pattern in DOCNO_PATTERNS:
        match = pattern.search(text)
        if match:
            return normalize_spaces(match.group(0))
    return "待核"


def is_reference_context(text: str, start: int, end: int) -> bool:
    ctx = context_for(text, start, end, radius=80)
    return bool(
        re.search(r"(按照|依据|根据|参照|贯彻|落实|引用|《[^》]{0,40}$)", ctx)
        or re.search(r"(指导意见|实施意见|条例|办法|法律|国务院|办公厅关于)", ctx)
    )


def extract_docno_current(text: str, *, source_name: str = "") -> tuple[str, str]:
    # File/folder names usually describe the current item; full OCR text may include cited policy numbers.
    for pattern in DOCNO_PATTERNS:
        match = pattern.search(source_name)
        if match:
            return normalize_spaces(match.group(0)), "文件名"

    search_area = text[:500]
    for pattern in DOCNO_PATTERNS:
        for match in pattern.finditer(search_area):
            if not is_reference_context(search_area, match.start(), match.end()):
                return normalize_spaces(match.group(0)), "首页/标题区"
    return "待核", ""


def extract_date(text: str) -> str:
    dates = extract_dates(text)
    if dates:
        return dates[-1]
    return ""


def extract_dates(text: str) -> list[str]:
    matches: list[tuple[int, str]] = []
    for match in FULL_DATE_RE.finditer(text):
        matches.append((match.start(), f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"))
    for match in ISO_DATE_RE.finditer(text):
        matches.append((match.start(), f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"))
    matches.sort(key=lambda item: item[0])
    return [value for _, value in matches]


def choose_document_date(*sources: tuple[str, str]) -> tuple[str, str]:
    for label, text in sources:
        dates = extract_dates(text)
        if dates:
            return dates[-1], label
    return "", ""


def context_for(text: str, start: int, end: int, radius: int = 42) -> str:
    snippet = text[max(0, start - radius) : min(len(text), end + radius)]
    return normalize_spaces(snippet)


def date_candidates_from_pages(page_texts: list[tuple[str, str]]) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    for page_idx, (label, text) in enumerate(page_texts, 1):
        for pattern in (FULL_DATE_RE, ISO_DATE_RE):
            for match in pattern.finditer(text):
                if pattern is FULL_DATE_RE:
                    value = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
                else:
                    value = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
                ctx = context_for(text, match.start(), match.end())
                candidates.append(
                    {
                        "page_index": str(page_idx),
                        "page_label": label,
                        "date": value,
                        "context": ctx,
                    }
                )
    return candidates


def looks_like_attachment_page(text: str) -> bool:
    compact = re.sub(r"\s+", "", text)
    head = compact[:80]
    return bool(
        re.search(r"^(附件|附表|附录|名单|项目表|资金表|绩效目标|申报表)", head)
        or re.search(r"附件\d*[:：]", head)
    )


def find_outer_boundary(page_texts: list[tuple[str, str]]) -> int | None:
    for idx, (_label, text) in enumerate(page_texts, 1):
        if idx <= 1:
            continue
        compact = re.sub(r"\s+", "", text)
        head = compact[:160]
        if looks_like_attachment_page(text):
            return idx
        if re.search(r"^(转发)?(国务院|国家文物局|内蒙古自治区|自治区|鄂尔多斯市|伊金霍洛旗).{0,35}(通知|意见|办法|函|请示|报告)", head):
            return idx
        if re.search(r"^.{0,30}(文件)$", head) and "关于" in head:
            return idx
    return None


def choose_document_date_from_pages(
    page_texts: list[tuple[str, str]],
    *,
    fallback_name: str,
) -> tuple[str, str, list[dict[str, str]], list[str]]:
    notes: list[str] = []
    candidates = date_candidates_from_pages(page_texts)
    if not candidates:
        name_dates = extract_dates(fallback_name)
        if name_dates:
            return name_dates[-1], "文件名", [{"page_index": "", "page_label": "文件名", "date": name_dates[-1], "context": fallback_name}], ["编制时间来自文件名，需核对是否为落款时间"]
        return "", "", [], []

    outer_boundary = find_outer_boundary(page_texts)

    pool = candidates
    if outer_boundary:
        before_outer_boundary = [c for c in candidates if int(c["page_index"]) < outer_boundary]
        if before_outer_boundary:
            pool = before_outer_boundary
            notes.append(f"检测到附件/被转发原文可能从第{outer_boundary}页开始，排序日期从当前文件范围内日期候选中选取")

    def score(candidate: dict[str, str]) -> tuple[int, int]:
        ctx = candidate["context"]
        page_index = int(candidate["page_index"])
        s = 0
        if page_index == 1:
            s += 2
        if outer_boundary and page_index >= outer_boundary:
            s -= 8
        if re.search(r"(人民政府|文化和旅游局|文物局|财政局|办公室|委员会|管理所)", ctx):
            s += 4
        if re.search(r"(印发|落款|签发|日期)", ctx):
            s += 1
        if re.search(r"(转发|现将|请认真贯彻|结合实际|贯彻执行)", ctx):
            s += 2
        if re.search(r"(附件|附表|名单|项目|资金|绩效|申报|联系电话|联系人|抄送|原文|原通知)", ctx):
            s -= 4
        return s, int(candidate["page_index"])

    chosen = sorted(pool, key=score)[-1]
    source = f"{chosen['page_label']} OCR/文本"
    if len(candidates) > 1:
        compact = "；".join(f"{c['page_label']}={c['date']}" for c in candidates[:12])
        notes.append(f"日期候选: {compact}")
        if len(candidates) > 12:
            notes.append(f"另有{len(candidates) - 12}个日期候选未列出")
    return chosen["date"], source, candidates, notes


def extract_issuer(text: str, title: str) -> str:
    for source in (text, title):
        match = ISSUER_RE.search(source)
        if match:
            return safe_name(match.group(1), limit=60)
    if "关于" in title:
        prefix = title.split("关于", 1)[0]
        if 2 <= len(prefix) <= 40 and any(key in prefix for key in ["局", "政府", "办公室", "委员会", "管理所", "财政"]):
            return safe_name(prefix, limit=60)
    return ""


def extract_issuer_current(text: str, title: str) -> str:
    if "关于" in title:
        prefix = title.split("关于", 1)[0]
        if 2 <= len(prefix) <= 40 and any(key in prefix for key in ["局", "政府", "办公室", "委员会", "管理所", "财政"]):
            return safe_name(prefix, limit=60)
    search_area = text[:350]
    if re.search(r"(按照|依据|根据|参照|贯彻|落实)", search_area[:120]):
        return ""
    match = ISSUER_RE.search(search_area)
    if match:
        return safe_name(match.group(1), limit=60)
    return ""


def title_from_name(name: str, docno: str) -> str:
    title = Path(name).stem
    title = re.sub(r"^\d+[-.、_\s]*", "", title)
    title = title.strip()

    # Common pattern: 文号-正式题名
    if "-" in title:
        left, right = title.split("-", 1)
        if extract_docno(left) != "待核" or re.search(r"\d{4}.*\d+号", left):
            title = right

    if docno != "待核":
        title = title.replace(docno, "")
    title = re.sub(r"^[\s，。、.．-]+", "", title)
    title = re.sub(r"[（(【\[]?[\u4e00-\u9fffA-Za-z]{1,16}(?:发|字|函|报|文|政|办|办法|文旅发|文物发)\s*[〔【\[\(（]?\s*\d{4}\s*[〕】\]\)）]?\s*\d+\s*号?[】\]）)]?$", "", title)
    title = re.sub(r"^[\u4e00-\u9fffA-Za-z]{1,16}(?:发|字|函|报|文|政|办|办法|文旅发|文物发)\s*[〔【\[\(（]?\s*\d{4}\s*[〕】\]\)）]?\s*\d+\s*号?\s*", "", title)
    title = normalize_spaces(title)
    return safe_name(title or name)


def build_record(
    item: SourceItem,
    *,
    seq: int,
    input_order: int,
    out_images_root: Path,
    project: str,
    dpi: int,
    ocr_date_pages: bool,
    ocr_dir: Path,
    remove_blank_pages: bool,
    dark_ratio_threshold: float,
    stddev_threshold: float,
    include_docno: bool,
) -> BuiltRecord:
    source = Path(item.source_path)
    notes = list(item.issues)
    first_text = ""
    last_text = ""
    blank_removed = 0
    page_labels: list[str] = []
    ocr_unavailable = False

    preliminary_text = item.source_name
    if item.source_kind == "pdf":
        first_text = pdf_page_text(source, 1)
        preliminary_text = f"{item.source_name} {first_text}"
    docno, docno_source = extract_docno_current(preliminary_text, source_name=item.source_name)
    title_docno = docno
    if not include_docno:
        # Legal/reference/paper volumes use the same exe package shape but do not fill 文号.
        title_docno = extract_docno(item.source_name)
    standard_title = title_from_name(item.source_name, title_docno)
    folder_title = safe_name(f"{seq}.{standard_title}", limit=140)
    image_folder = out_images_root / project / folder_title

    if item.source_kind == "pdf":
        pages, render_issues, blank_removed, original_page_numbers = render_pdf(
            source,
            image_folder,
            dpi=dpi,
            remove_blank_pages=remove_blank_pages,
            dark_ratio_threshold=dark_ratio_threshold,
            stddev_threshold=stddev_threshold,
        )
        notes.extend(render_issues)
        page_labels = [f"第{n}页" for n in original_page_numbers]
    elif item.source_kind in {"image_folder", "single_image"}:
        source_pages = [Path(p) for p in item.page_sources]
        pages, image_issues, blank_removed, source_page_names = copy_images_as_png(
            source_pages,
            image_folder,
            remove_blank_pages=remove_blank_pages,
            dark_ratio_threshold=dark_ratio_threshold,
            stddev_threshold=stddev_threshold,
        )
        notes.extend(image_issues)
        page_labels = source_page_names
    else:
        raise ValueError(f"Unsupported source kind: {item.source_kind}")

    page_texts: list[tuple[str, str]] = []
    if item.source_kind == "pdf":
        for idx, label in enumerate(page_labels):
            text = ""
            match = re.search(r"第(\d+)页", label)
            if match:
                text = pdf_page_text(source, int(match.group(1)))
            if ocr_date_pages and idx < len(pages) and not ocr_unavailable:
                ocr_text, ocr_error = ocr_image(pages[idx], ocr_dir=ocr_dir)
                if ocr_error:
                    notes.append(f"{label} OCR失败: {ocr_error}")
                    if "OCR服务连接失败" in ocr_error:
                        notes.append("已跳过后续逐页OCR")
                        ocr_unavailable = True
                if ocr_text:
                    text = ocr_text if not text else f"{text} {ocr_text}"
            page_texts.append((label, normalize_spaces(text)))
    elif item.source_kind in {"image_folder", "single_image"}:
        for idx, label in enumerate(page_labels):
            text = ""
            if ocr_date_pages and idx < len(pages) and not ocr_unavailable:
                ocr_text, ocr_error = ocr_image(pages[idx], ocr_dir=ocr_dir)
                if ocr_error:
                    notes.append(f"{label} OCR失败: {ocr_error}")
                    if "OCR服务连接失败" in ocr_error:
                        notes.append("已跳过后续逐页OCR")
                        ocr_unavailable = True
                text = ocr_text
            page_texts.append((label, normalize_spaces(text)))

    if page_texts:
        first_text = page_texts[0][1]
        last_text = page_texts[-1][1]

    combined_text = normalize_spaces(f"{item.source_name} {' '.join(text for _label, text in page_texts)}")
    if include_docno and docno == "待核":
        docno, docno_source = extract_docno_current(combined_text, source_name=item.source_name)
    if include_docno and docno == "待核":
        notes.append("文号待核")
    if not include_docno:
        docno = ""
    title_docno = docno if include_docno else extract_docno(item.source_name)
    standard_title = title_from_name(item.source_name, title_docno)
    folder_title_final = safe_name(f"{seq}.{standard_title}", limit=140)
    if folder_title_final != folder_title:
        final_folder = out_images_root / project / folder_title_final
        if final_folder.exists():
            shutil.rmtree(final_folder)
        image_folder.rename(final_folder)
        image_folder = final_folder
        folder_title = folder_title_final

    date, date_source, date_candidates, date_notes = choose_document_date_from_pages(
        page_texts,
        fallback_name=item.source_name,
    )
    notes.extend(date_notes)
    if not date:
        notes.append("编制时间待核")
    elif date_source == "文件名":
        notes.append("编制时间来自文件名，需核对是否为落款时间")
    issuer = extract_issuer_current(combined_text, standard_title)
    if not issuer:
        notes.append("编制单位待核")

    return BuiltRecord(
        seq=seq,
        input_order=input_order,
        source_path=str(source),
        source_kind=item.source_kind,
        original_name=item.source_name,
        title=standard_title,
        folder_title=folder_title,
        docno=docno,
        issuer=issuer,
        date=date,
        date_source=date_source,
        date_candidates=[f"{c['page_label']}={c['date']} {c['context']}" for c in date_candidates],
        page_count=len(pages),
        output_folder=str(image_folder),
        notes=dedupe(notes),
        first_page_text=first_text[:800],
        last_page_text=last_text[:800],
        blank_pages_removed=blank_removed,
    )


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def reorder_records_by_date(records: list[BuiltRecord], *, image_root: Path, project: str, start_seq: int) -> list[BuiltRecord]:
    sorted_records = sorted(records, key=lambda r: (r.date or "9999-99-99", r.input_order))
    temp_pairs: list[tuple[BuiltRecord, Path, Path]] = []
    for idx, record in enumerate(sorted_records, 1):
        old_folder = Path(record.output_folder)
        tmp_folder = image_root / project / f"__tmp_reorder_{idx:04d}"
        if tmp_folder.exists():
            shutil.rmtree(tmp_folder)
        if old_folder.exists():
            old_folder.rename(tmp_folder)
        temp_pairs.append((record, old_folder, tmp_folder))

    for offset, (record, _old_folder, tmp_folder) in enumerate(temp_pairs):
        new_seq = start_seq + offset
        new_title = safe_name(f"{new_seq}.{record.title}", limit=140)
        new_folder = image_root / project / new_title
        if new_folder.exists():
            shutil.rmtree(new_folder)
        if tmp_folder.exists():
            tmp_folder.rename(new_folder)
        record.seq = new_seq
        record.folder_title = new_title
        record.output_folder = str(new_folder)
    return sorted_records


def excel_date_text(value: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", value or "")
    if not match:
        return value or ""
    return f"{int(match.group(1))}年{int(match.group(2))}月{int(match.group(3))}日"


def main_table_text(value: str) -> str:
    if not value or value == "待核":
        return ""
    return value


def write_workbook(path: Path, records: list[BuiltRecord], *, project: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for record in records:
        ws.append(
            [
                project,
                "",
                main_table_text(record.docno),
                main_table_text(record.issuer),
                record.folder_title,
                excel_date_text(record.date),
                "",
                "",
                record.page_count,
                "",
                "",
                "",
            ]
        )
    widths = [10, 10, 24, 28, 70, 18, 18, 18, 10, 55, 18, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_review(path: Path, records: list[BuiltRecord], unsupported: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "复核清单"
    headers = [
        "序号",
        "原输入顺序",
        "来源类型",
        "原名称",
        "来源路径",
        "输出题名",
        "文号",
        "编制单位",
        "编制时间",
        "日期来源",
        "张数",
        "删除空白页数",
        "问题/待核",
        "日期候选",
        "首页文本摘录",
        "末页文本摘录",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for record in records:
        ws.append(
            [
                record.seq,
                record.input_order,
                record.source_kind,
                record.original_name,
                record.source_path,
                record.folder_title,
                record.docno,
                record.issuer,
                record.date,
                record.date_source,
                record.page_count,
                record.blank_pages_removed,
                "；".join(record.notes),
                "；".join(record.date_candidates[:20]),
                record.first_page_text[:300],
                record.last_page_text[:300],
            ]
        )
    for idx, width in enumerate([8, 12, 14, 38, 80, 70, 24, 28, 18, 24, 10, 12, 60, 90, 70, 70], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("未处理输入")
    ws2.append(["路径", "原因"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in unsupported:
        ws2.append([row.get("path", ""), row.get("reason", "")])
    ws2.column_dimensions["A"].width = 90
    ws2.column_dimensions["B"].width = 40
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def validate_package(out_dir: Path, records: list[BuiltRecord], *, project: str) -> list[str]:
    issues: list[str] = []
    workbook = out_dir / "administration_data_chinese.xlsx"
    if not workbook.exists():
        issues.append("缺少 administration_data_chinese.xlsx")
    for record in records:
        folder = out_dir / "images" / project / record.folder_title
        if not folder.exists():
            issues.append(f"缺少图片文件夹: {record.folder_title}")
            continue
        pages = sorted(folder.glob("page_*.png"), key=lambda p: natural_key(p.name))
        if len(pages) != record.page_count:
            issues.append(f"图片数量不一致: {record.folder_title} 张数={record.page_count} 实际={len(pages)}")
        expected_names = [f"page_{idx:03d}.png" for idx in range(1, len(pages) + 1)]
        actual_names = [p.name for p in pages]
        if actual_names != expected_names:
            issues.append(f"页图命名不连续: {record.folder_title}")
    return issues


def write_csv_manifest(path: Path, records: list[BuiltRecord]) -> None:
    fields = [
        "seq",
        "input_order",
        "source_kind",
        "source_path",
        "original_name",
        "folder_title",
        "docno",
        "issuer",
        "date",
        "date_source",
        "page_count",
        "blank_pages_removed",
        "notes",
        "date_candidates",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "seq": record.seq,
                    "input_order": record.input_order,
                    "source_kind": record.source_kind,
                    "source_path": record.source_path,
                    "original_name": record.original_name,
                    "folder_title": record.folder_title,
                    "docno": record.docno,
                    "issuer": record.issuer,
                    "date": record.date,
                    "date_source": record.date_source,
                    "page_count": record.page_count,
                    "blank_pages_removed": record.blank_pages_removed,
                    "notes": "；".join(record.notes),
                    "date_candidates": "；".join(record.date_candidates),
                }
            )


def write_note(path: Path, records: list[BuiltRecord], unsupported: list[dict[str, str]], validation_issues: list[str]) -> None:
    needs_review = [record for record in records if record.notes]
    lines = [
        "# 行政资料 exe 可跑材料包生成说明",
        "",
        f"- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 处理资料数: {len(records)}",
        f"- 需复核资料数: {len(needs_review)}",
        f"- 未处理输入数: {len(unsupported)}",
        f"- 删除空白页数: {sum(record.blank_pages_removed for record in records)}",
        "",
        "## 输出内容",
        "",
        "- `administration_data_chinese.xlsx`",
        "- `images/1/<序号.标准题名>/page_001.png...`",
        "- `administration_review.xlsx`",
        "- `audit.json`",
        "- `source_manifest.csv`",
        "",
        "## 说明",
        "",
        "- v1.0 只处理人工给定的候选资料，不判断行政资料全集是否完整。",
        "- 输出顺序按识别出的当前文件落款时间由早到晚排序；未识别日期的资料排在最后。",
        "- PDF或图片文件夹中的空白页会先删除，Excel `张数` 使用删除空白页后的页图数量。",
        "- 文号、日期、单位无法确认时已标入备注和复核清单。",
        "- Excel `题名` 与 `images/1/` 下文件夹名必须完全一致，本次已自动校验。",
        "",
        "## 校验结果",
        "",
    ]
    if validation_issues:
        lines.extend(f"- {issue}" for issue in validation_issues)
    else:
        lines.append("- Excel 行与图片目录数量校验通过。")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build exe-ready text-volume archive package from PDFs and image folders.")
    parser.add_argument("inputs", nargs="+", help="PDF files, image-page folders, or parent folders.")
    parser.add_argument("--output", required=True, help="Output package directory.")
    parser.add_argument(
        "--volume-type",
        choices=["administration", "legal", "reference", "paper"],
        default="administration",
        help="Volume type. administration fills 文号 when reliable; legal/reference/paper leave 文号 blank. Default: administration.",
    )
    parser.add_argument("--project", default="1", help="Excel 项目 value and images/<project>/ level. Default: 1.")
    parser.add_argument("--start-seq", type=int, default=1, help="First sequence number. Default: 1.")
    parser.add_argument("--dpi", type=int, default=180, help="PDF render DPI. Default: 180.")
    parser.add_argument("--ocr-first-page", action="store_true", help="Deprecated alias; OCR date pages is now enabled by default.")
    parser.add_argument("--no-ocr-date-pages", action="store_true", help="Do not OCR every retained page for date candidates; use embedded PDF text and filenames only.")
    parser.add_argument("--ocr-dir", default="/Users/drevan01/Desktop/OCR", help="Local OCR directory. Default: /Users/drevan01/Desktop/OCR.")
    parser.add_argument("--keep-blank-pages", action="store_true", help="Keep blank pages instead of removing them.")
    parser.add_argument("--blank-dark-ratio-threshold", type=float, default=0.00005, help="Blank-page dark-pixel ratio threshold. Default: 0.00005.")
    parser.add_argument("--blank-stddev-threshold", type=float, default=0.5, help="Blank-page grayscale stddev threshold. Default: 0.5.")
    parser.add_argument("--overwrite", action="store_true", help="Delete output directory before building.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    out_dir = Path(args.output).expanduser().resolve()
    if out_dir.exists() and args.overwrite:
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    input_paths = [Path(value).expanduser() for value in args.inputs]
    items, unsupported = discover_inputs(input_paths)
    if not items:
        print("No supported PDF or image-folder inputs found.", file=sys.stderr)
        return 2

    image_root = out_dir / "images"
    include_docno = args.volume_type == "administration"
    records: list[BuiltRecord] = []
    for offset, item in enumerate(items):
        seq = args.start_seq + offset
        records.append(
            build_record(
                item,
                seq=seq,
                input_order=offset + 1,
                out_images_root=image_root,
                project=args.project,
                dpi=args.dpi,
                ocr_date_pages=not args.no_ocr_date_pages,
                ocr_dir=Path(args.ocr_dir).expanduser(),
                remove_blank_pages=not args.keep_blank_pages,
                dark_ratio_threshold=args.blank_dark_ratio_threshold,
                stddev_threshold=args.blank_stddev_threshold,
                include_docno=include_docno,
            )
        )
    records = reorder_records_by_date(records, image_root=image_root, project=args.project, start_seq=args.start_seq)

    write_workbook(out_dir / "administration_data_chinese.xlsx", records, project=args.project)
    write_review(out_dir / "administration_review.xlsx", records, unsupported)
    write_csv_manifest(out_dir / "source_manifest.csv", records)
    validation_issues = validate_package(out_dir, records, project=args.project)
    audit = {
        "output": str(out_dir),
        "project": args.project,
        "volume_type": args.volume_type,
        "records": [asdict(record) for record in records],
        "unsupported": unsupported,
        "validation_issues": validation_issues,
    }
    (out_dir / "audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    write_note(out_dir / "生成说明.md", records, unsupported, validation_issues)
    print(json.dumps({"output": str(out_dir), "records": len(records), "validation_issues": validation_issues}, ensure_ascii=False, indent=2))
    return 1 if validation_issues else 0


if __name__ == "__main__":
    raise SystemExit(main())

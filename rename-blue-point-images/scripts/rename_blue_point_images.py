#!/usr/bin/env python3
"""Create contact sheets, detect units, rotate photos, and safely rename images."""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import uuid
from pathlib import Path


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}
POINT_LABEL_RE = r"A\d{3}B\d{3}"
DEFAULT_LABEL_RE = r"^(?:U\d{2}-)?A\d{3}B\d{3}$"
FINAL_LABEL_RE = r"^U\d{2}-A\d{3}B\d{3}$"


def natural_key(path: Path) -> tuple[int, object]:
    stem = path.stem
    if stem.isdigit():
        return (0, int(stem))
    return (1, path.name.lower())


def image_files(folder: Path) -> list[Path]:
    return sorted(
        (p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS),
        key=natural_key,
    )


def contiguous_groups(indices, gap: int = 3) -> list[tuple[int, int]]:
    values = [int(v) for v in indices]
    if not values:
        return []
    groups: list[tuple[int, int]] = []
    start = previous = values[0]
    for value in values[1:]:
        if value - previous <= gap:
            previous = value
        else:
            groups.append((start, previous))
            start = previous = value
    groups.append((start, previous))
    return groups


def require_image_libs():
    try:
        from PIL import Image
        import numpy as np
    except ImportError as exc:
        raise RuntimeError("This command requires Pillow and NumPy") from exc
    return Image, np


def connected_components(mask):
    import numpy as np

    height, width = mask.shape
    seen = np.zeros_like(mask, dtype=bool)
    components = []

    ys, xs = np.where(mask)
    for y_value, x_value in zip(ys, xs):
        y = int(y_value)
        x = int(x_value)
        if seen[y, x]:
            continue

        stack = [(y, x)]
        seen[y, x] = True
        comp_xs = []
        comp_ys = []

        while stack:
            current_y, current_x = stack.pop()
            comp_xs.append(current_x)
            comp_ys.append(current_y)
            for delta_y, delta_x in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                next_y = current_y + delta_y
                next_x = current_x + delta_x
                if (
                    0 <= next_y < height
                    and 0 <= next_x < width
                    and mask[next_y, next_x]
                    and not seen[next_y, next_x]
                ):
                    seen[next_y, next_x] = True
                    stack.append((next_y, next_x))

        area = len(comp_xs)
        min_x = min(comp_xs)
        max_x = max(comp_xs)
        min_y = min(comp_ys)
        max_y = max(comp_ys)
        box_width = max_x - min_x + 1
        box_height = max_y - min_y + 1
        components.append(
            {
                "area": area,
                "min_x": min_x,
                "max_x": max_x,
                "min_y": min_y,
                "max_y": max_y,
                "width": box_width,
                "height": box_height,
                "center_x": sum(comp_xs) / area,
                "center_y": sum(comp_ys) / area,
                "fill": area / (box_width * box_height),
                "ratio": box_width / box_height,
            }
        )

    return components


def detect_tk_label_orientation(path: Path) -> dict[str, object]:
    Image, np = require_image_libs()
    image = Image.open(path).convert("RGB")
    image.thumbnail((1200, 1200))
    array = np.array(image)
    height, width = array.shape[:2]
    mask = (array[:, :, 0] > 185) & (array[:, :, 1] > 185) & (array[:, :, 2] > 185)

    candidates = []
    for component in connected_components(mask):
        area = component["area"]
        box_area = component["width"] * component["height"]
        if area < 1200 or area > height * width * 0.12:
            continue
        if box_area < 1800:
            continue

        ratio = component["ratio"]
        if 1.25 <= ratio <= 2.4 or 0.4 <= ratio <= 0.85:
            shape_score = 1.0
        else:
            shape_score = 0.35

        center_x = component["center_x"]
        center_y = component["center_y"]
        edge_distance = min(center_x, width - center_x, center_y, height - center_y)
        edge_score = max(0.05, 1 - edge_distance / (max(width, height) / 2))
        score = area * shape_score * edge_score
        candidates.append((score, component))

    if not candidates:
        raise ValueError("No likely white TK label was detected")

    candidates.sort(key=lambda item: item[0], reverse=True)
    score, label = candidates[0]
    if label["width"] >= label["height"]:
        rotation = 0 if label["center_y"] < height / 2 else 180
    else:
        rotation = 90 if label["center_x"] > width / 2 else 270

    return {
        "rotation": rotation,
        "score": score,
        "label_x": label["center_x"],
        "label_y": label["center_y"],
        "label_width": label["width"],
        "label_height": label["height"],
    }


def detect_blue_point(array) -> tuple[float, float, float]:
    import math

    red = array[:, :, 0]
    green = array[:, :, 1]
    blue = array[:, :, 2]
    mask = (blue > 150) & (red < 80) & (green < 100)
    candidates = []

    for component in connected_components(mask):
        if component["area"] < 200 or component["width"] < 10 or component["height"] < 10:
            continue
        score = component["area"] * component["fill"] / (1 + abs(math.log(component["ratio"])))
        candidates.append((score, component))

    if not candidates:
        raise ValueError("No saturated blue point marker was detected")

    candidates.sort(key=lambda item: item[0], reverse=True)
    best = candidates[0][1]
    confidence = min(1.0, candidates[0][0] / 5000)
    return best["center_x"], best["center_y"], confidence


def detect_unit_grid(array) -> tuple[float, float, float]:
    import numpy as np

    height, width = array.shape[:2]
    red = array[:, :, 0]
    green = array[:, :, 1]
    blue = array[:, :, 2]

    blue_marker = (blue > 150) & (red < 80) & (green < 100)
    purple = (red > 90) & (blue > 120) & (green < 100) & ~blue_marker

    column_counts = purple.sum(axis=0)
    row_counts = purple.sum(axis=1)
    x_groups = contiguous_groups(np.where(column_counts > height * 0.35)[0], gap=4)
    y_groups = contiguous_groups(np.where(row_counts > width * 0.35)[0], gap=4)

    x_centers = [(start + end) / 2 for start, end in x_groups if width * 0.15 < (start + end) / 2 < width * 0.85]
    y_centers = [(start + end) / 2 for start, end in y_groups if height * 0.20 < (start + end) / 2 < height * 0.90]

    if not x_centers:
        raise ValueError("Could not detect the central unit divider")
    if len(y_centers) < 2:
        raise ValueError("Could not detect the two horizontal unit dividers")

    x_mid = min(x_centers, key=lambda value: abs(value - width / 2))
    y_top = min(y_centers, key=lambda value: abs(value - height * 0.38))
    remaining = [value for value in y_centers if abs(value - y_top) > 30]
    if not remaining:
        raise ValueError("Could not detect the lower horizontal unit divider")
    y_bottom = min(remaining, key=lambda value: abs(value - height * 0.68))
    if y_bottom < y_top:
        y_top, y_bottom = y_bottom, y_top
    return x_mid, y_top, y_bottom


def classify_unit(point_x: float, point_y: float, x_mid: float, y_top: float, y_bottom: float) -> str:
    if point_y < y_top:
        return "U05" if point_x < x_mid else "U06"
    if point_y < y_bottom:
        return "U03" if point_x < x_mid else "U04"
    return "U01" if point_x < x_mid else "U02"


def detect_unit_for_image(path: Path) -> dict[str, object]:
    Image, np = require_image_libs()
    array = np.array(Image.open(path).convert("RGB"))
    point_x, point_y, confidence = detect_blue_point(array)
    x_mid, y_top, y_bottom = detect_unit_grid(array)
    unit = classify_unit(point_x, point_y, x_mid, y_top, y_bottom)
    boundary_distance = min(abs(point_x - x_mid), abs(point_y - y_top), abs(point_y - y_bottom))
    return {
        "unit": unit,
        "x": point_x,
        "y": point_y,
        "confidence": confidence,
        "boundary_distance": boundary_distance,
    }


def label_from_target(target_name: str) -> str:
    stem = Path(target_name).stem
    if "-" in stem:
        stem = stem.split("-", 1)[1]
    if not re.fullmatch(POINT_LABEL_RE, stem):
        raise ValueError(f"Target must contain a point label like A031B017: {target_name}")
    return stem


def read_unit_workbook(path: Path) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("--unit-workbook requires openpyxl") from exc

    workbook = load_workbook(path, data_only=True)
    mapping: dict[str, str] = {}
    for sheet in workbook.worksheets:
        headers = None
        header_row = None
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if "勘探单元" in values and "标准孔编号" in values:
                headers = values
                header_row = row_index
                break
        if headers is None or header_row is None:
            continue

        unit_index = headers.index("勘探单元")
        label_index = headers.index("标准孔编号")
        last_unit = ""
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            unit = row[unit_index] if unit_index < len(row) else None
            label = row[label_index] if label_index < len(row) else None
            unit_text = "" if unit is None else str(unit).strip()
            label_text = "" if label is None else str(label).strip()
            if unit_text:
                last_unit = unit_text
            if label_text and re.fullmatch(POINT_LABEL_RE, label_text):
                if not last_unit:
                    raise ValueError(f"Missing unit for {label_text} in {path}")
                mapping[label_text] = last_unit
    if not mapping:
        raise ValueError(f"No 标准孔编号 -> 勘探单元 rows found in {path}")
    return mapping


def build_auto_unit_rows(
    folder: Path,
    rows: list[tuple[str, str]],
    unit_workbook: Path | None = None,
) -> list[tuple[str, str]]:
    unit_map = read_unit_workbook(unit_workbook) if unit_workbook else None
    converted = []
    for source_name, target_name in rows:
        source = folder / source_name
        if not source.exists():
            raise FileNotFoundError(source)
        point_label = label_from_target(target_name)
        if unit_map is not None:
            if point_label not in unit_map:
                raise ValueError(f"{point_label} was not found in {unit_workbook}")
            unit = unit_map[point_label]
        else:
            detected = detect_unit_for_image(source)
            unit = str(detected["unit"])
        suffix = Path(target_name).suffix or source.suffix
        converted.append((source_name, f"{unit}-{point_label}{suffix}"))
    return converted


def make_contact_sheet(args: argparse.Namespace) -> int:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("contact-sheet requires Pillow. Install pillow or inspect images individually.", file=sys.stderr)
        return 2

    folder = Path(args.folder).expanduser().resolve()
    files = image_files(folder)
    if not files:
        print(f"No image files found in {folder}", file=sys.stderr)
        return 1

    thumb_w, thumb_h = args.thumb_width, args.thumb_height
    label_h = 28
    padding = 14
    columns = max(1, args.columns)
    rows = (len(files) + columns - 1) // columns
    cell_w = thumb_w + padding * 2
    cell_h = thumb_h + label_h + padding * 2
    sheet = Image.new("RGB", (cell_w * columns, cell_h * rows), "white")
    draw = ImageDraw.Draw(sheet)

    try:
        font = ImageFont.load_default()
    except Exception:
        font = None

    for index, path in enumerate(files):
        image = Image.open(path).convert("RGB")
        image.thumbnail((thumb_w, thumb_h))
        x0 = (index % columns) * cell_w
        y0 = (index // columns) * cell_h
        draw.text((x0 + padding, y0 + 8), path.name, fill=(0, 0, 0), font=font)
        x = x0 + padding + (thumb_w - image.width) // 2
        y = y0 + label_h + padding + (thumb_h - image.height) // 2
        sheet.paste(image, (x, y))

    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(output)
    print(output)
    return 0


def orient_photos(args: argparse.Namespace) -> int:
    Image, _np = require_image_libs()
    folder = Path(args.folder).expanduser().resolve()
    files = image_files(folder)
    if not files:
        print(f"No image files found in {folder}", file=sys.stderr)
        return 1

    overrides = parse_rotation_overrides(args.rotate)
    plans = []
    errors = []
    for path in files:
        try:
            if path.name in overrides:
                detected = {
                    "rotation": overrides[path.name],
                    "score": 0,
                    "label_x": 0,
                    "label_y": 0,
                    "label_width": 0,
                    "label_height": 0,
                    "manual": True,
                }
            else:
                detected = detect_tk_label_orientation(path)
            plans.append((path, detected))
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")

    if errors:
        print("Orientation detection failed:", file=sys.stderr)
        for error in errors:
            print(error, file=sys.stderr)
        return 1

    action = "APPLY" if args.apply else "DRY RUN"
    print(f"{action}: {len(plans)} photo orientation check(s)")
    for path, detected in plans:
        rotation = detected["rotation"]
        status = "keep" if rotation == 0 else f"rotate_ccw_{rotation}"
        if detected.get("manual"):
            print(f"{path.name}: {status} manual")
        else:
            print(
                f"{path.name}: {status} "
                f"label=({detected['label_x']:.0f},{detected['label_y']:.0f}) "
                f"box={detected['label_width']}x{detected['label_height']}"
            )

    if not args.apply:
        print("No files changed. Re-run with --apply to rotate files.")
        return 0

    for path, detected in plans:
        rotation = detected["rotation"]
        if rotation == 0:
            continue
        image = Image.open(path).convert("RGB")
        rotated = image.rotate(rotation, expand=True)
        temp = path.with_name(f".{path.stem}.orient-{uuid.uuid4().hex}{path.suffix}")
        save_kwargs = {}
        if path.suffix.lower() in {".jpg", ".jpeg"}:
            save_kwargs.update({"quality": args.quality, "subsampling": 0, "optimize": True})
        rotated.save(temp, **save_kwargs)
        os.replace(temp, path)

    print("Done")
    return 0


def parse_rotation_overrides(values: list[str] | None) -> dict[str, int]:
    overrides: dict[str, int] = {}
    for value in values or []:
        if ":" not in value:
            raise ValueError(f"Rotation override must be filename:degrees, got {value}")
        name, degrees_text = value.split(":", 1)
        try:
            degrees = int(degrees_text)
        except ValueError as exc:
            raise ValueError(f"Invalid rotation degrees in {value}") from exc
        if degrees not in {0, 90, 180, 270}:
            raise ValueError(f"Rotation override must be one of 0, 90, 180, 270: {value}")
        overrides[name] = degrees
    return overrides


def detect_units(args: argparse.Namespace) -> int:
    folder = Path(args.folder).expanduser().resolve()
    rows = None
    if args.mapping:
        try:
            rows = read_mapping(Path(args.mapping).expanduser().resolve())
        except Exception as exc:
            print(f"Mapping read failed:\n{exc}", file=sys.stderr)
            return 1

    files = [folder / source for source, _target in rows] if rows is not None else image_files(folder)
    if not files:
        print(f"No image files found in {folder}", file=sys.stderr)
        return 1

    print("source,unit,label,target,x,y,boundary_distance,confidence")
    for path in files:
        try:
            detected = detect_unit_for_image(path)
            if rows is None:
                if re.fullmatch(POINT_LABEL_RE, path.stem):
                    point_label = path.stem
                elif re.fullmatch(FINAL_LABEL_RE, path.stem):
                    point_label = path.stem.split("-", 1)[1]
                else:
                    point_label = ""
            else:
                target_name = dict(rows)[path.name]
                point_label = label_from_target(target_name)
            target = f"{detected['unit']}-{point_label}{path.suffix}" if point_label else ""
            print(
                f"{path.name},{detected['unit']},{point_label},{target},"
                f"{detected['x']:.0f},{detected['y']:.0f},"
                f"{detected['boundary_distance']:.0f},{detected['confidence']:.2f}"
            )
        except Exception as exc:
            print(f"{path.name},ERROR,,,,,,{exc}", file=sys.stderr)
            return 1
    return 0


def read_mapping(path: Path) -> list[tuple[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or "source" not in reader.fieldnames or "target" not in reader.fieldnames:
            raise ValueError("Mapping CSV must have source,target columns")
        rows: list[tuple[str, str]] = []
        for line_no, row in enumerate(reader, start=2):
            source = (row.get("source") or "").strip()
            target = (row.get("target") or "").strip()
            if not source and not target:
                continue
            if not source or not target:
                raise ValueError(f"Line {line_no}: source and target are both required")
            rows.append((source, target))
    return rows


def validate_mapping(
    folder: Path,
    rows: list[tuple[str, str]],
    label_re: re.Pattern[str],
) -> list[tuple[Path, Path]]:
    if not rows:
        raise ValueError("Mapping CSV has no rows")

    seen_sources: set[str] = set()
    seen_targets: set[str] = set()
    operations: list[tuple[Path, Path]] = []
    errors: list[str] = []

    for source_name, target_name in rows:
        source = folder / source_name
        target = folder / target_name
        target_label = target.stem

        if source_name in seen_sources:
            errors.append(f"Duplicate source in mapping: {source_name}")
        if target_name in seen_targets:
            errors.append(f"Duplicate target in mapping: {target_name}")
        seen_sources.add(source_name)
        seen_targets.add(target_name)

        if source.parent != folder or target.parent != folder:
            errors.append(f"Nested paths are not allowed: {source_name} -> {target_name}")
        if not source.exists():
            errors.append(f"Source does not exist: {source_name}")
        if target.exists() and source.resolve() != target.resolve():
            errors.append(f"Target already exists: {target_name}")
        if source.suffix.lower() not in IMAGE_EXTENSIONS:
            errors.append(f"Source is not a supported image type: {source_name}")
        if target.suffix.lower() not in IMAGE_EXTENSIONS:
            errors.append(f"Target is not a supported image type: {target_name}")
        if not label_re.match(target_label):
            errors.append(f"Target label does not match expected pattern: {target_name}")

        operations.append((source, target))

    if errors:
        raise ValueError("\n".join(errors))
    return operations


def apply_renames(operations: list[tuple[Path, Path]]) -> None:
    token = f".rename-blue-point-{uuid.uuid4().hex}"
    staged: list[tuple[Path, Path, Path]] = []

    try:
        for source, target in operations:
            temp = source.with_name(source.name + token)
            source.rename(temp)
            staged.append((temp, source, target))
        for temp, _source, target in staged:
            temp.rename(target)
    except Exception:
        for temp, source, _target in reversed(staged):
            if temp.exists() and not source.exists():
                temp.rename(source)
        raise


def rename_from_mapping(args: argparse.Namespace) -> int:
    folder = Path(args.folder).expanduser().resolve()
    mapping = Path(args.mapping).expanduser().resolve()
    label_re = re.compile(args.pattern)

    try:
        rows = read_mapping(mapping)
        if args.auto_unit:
            workbook = Path(args.unit_workbook).expanduser().resolve() if args.unit_workbook else None
            rows = build_auto_unit_rows(folder, rows, workbook)
        operations = validate_mapping(folder, rows, label_re)
    except Exception as exc:
        print(f"Validation failed:\n{exc}", file=sys.stderr)
        return 1

    action = "APPLY" if args.apply else "DRY RUN"
    print(f"{action}: {len(operations)} rename operation(s)")
    for source, target in operations:
        print(f"{source.name} -> {target.name}")

    if args.apply:
        apply_renames(operations)
        print("Done")
    else:
        print("No files changed. Re-run with --apply to rename files.")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    contact = subparsers.add_parser("contact-sheet", help="Create a labeled contact sheet")
    contact.add_argument("folder")
    contact.add_argument("--output", required=True)
    contact.add_argument("--columns", type=int, default=3)
    contact.add_argument("--thumb-width", type=int, default=420)
    contact.add_argument("--thumb-height", type=int, default=620)
    contact.set_defaults(func=make_contact_sheet)

    orient = subparsers.add_parser("orient-photos", help="Rotate standard-hole photos so the TK label is upright at top")
    orient.add_argument("folder")
    orient.add_argument("--rotate", action="append", help="Manual override as filename:counterclockwise_degrees")
    orient.add_argument("--quality", type=int, default=95)
    orient.add_argument("--apply", action="store_true")
    orient.set_defaults(func=orient_photos)

    detect = subparsers.add_parser("detect-units", help="Detect blue point unit cells and propose final targets")
    detect.add_argument("folder")
    detect.add_argument("--mapping")
    detect.set_defaults(func=detect_units)

    rename = subparsers.add_parser("rename", help="Validate and optionally apply a CSV rename mapping")
    rename.add_argument("folder")
    rename.add_argument("--mapping", required=True)
    rename.add_argument("--pattern", default=DEFAULT_LABEL_RE)
    rename.add_argument("--auto-unit", action="store_true", help="Prefix each target with the detected Uxx unit")
    rename.add_argument("--unit-workbook", help="Read 标准孔编号 -> 勘探单元 mapping from 标准孔坐标.xlsx")
    rename.add_argument("--apply", action="store_true")
    rename.set_defaults(func=rename_from_mapping)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())

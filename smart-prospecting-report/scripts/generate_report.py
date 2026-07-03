#!/usr/bin/env python3
"""考古调查、勘探报告模板填充工具。

用法：
  python3 generate_report.py init-input 过程资料/新项目资料
  python3 generate_report.py generate 过程资料/新项目资料 --output 生成报告
"""

from __future__ import annotations

import argparse
import copy
import io
import json
import re
import shutil
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from PIL import Image


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"w": W_NS, "r": R_NS, "rel": REL_NS}
ET.register_namespace("w", W_NS)
ET.register_namespace("r", R_NS)


KV_SHEETS = {"项目概况", "人员信息"}
TABLE_SHEETS = {"红线坐标", "勘探单元", "标准孔与剖线", "遗迹信息", "图片清单"}
REQUIRED_SHEETS = KV_SHEETS | TABLE_SHEETS

FIELD_ALIASES = {
    "项目名称": ["项目名称", "工程名称"],
    "建设单位": ["建设单位", "委托单位"],
    "勘探单位": ["勘探单位", "报告单位", "编制单位"],
    "项目位置": ["项目位置", "位置", "建设地点", "所在地"],
    "项目面积": ["项目面积", "用地面积", "总用地面积"],
    "调查面积": ["调查面积"],
    "勘探面积": ["勘探面积"],
    "勘探时间": ["勘探时间", "工作时间"],
    "报告年月": ["报告年月", "报告日期", "年月"],
    "遗迹结论": ["遗迹结论", "遗迹", "遗迹现象", "是否发现遗迹"],
    "经度": ["经度", "东经"],
    "纬度": ["纬度", "北纬"],
    "地块现状": ["地块现状", "项目地块现状", "现状描述"],
    "建设内容": ["建设内容", "项目内容"],
}

REQUIRED_FIELDS = ["项目名称", "建设单位", "勘探单位", "项目位置", "项目面积", "勘探面积", "勘探时间", "遗迹结论"]
TEMPLATE_LIBRARY_DIR = Path("基础信息")
DEFAULT_TEMPLATE_DIRS = [Path("旧模板"), TEMPLATE_LIBRARY_DIR / "旧模板"]
DEFAULT_OUTPUT_DIR = Path("生成报告")
DEFAULT_PROCESS_DIR = Path("过程资料")

PREFIX_FIELDS = {
    "项目名称：": "项目名称",
    "建设单位：": "建设单位",
    "项目位置：": "项目位置",
    "项目面积：": "项目面积",
    "调查面积：": "调查面积",
    "勘探面积：": "勘探面积",
    "勘探时间：": "勘探时间",
    "遗迹现象：": "遗迹结论",
    "遗迹：": "遗迹结论",
}

TABLE_CAPTIONS = {
    "红线坐标": ["红线四至坐标"],
    "勘探单元": ["勘探单元坐标"],
    "标准孔与剖线": ["标准孔坐标", "剖面探孔记录表"],
    "遗迹信息": ["考古勘探遗迹登记表", "考古勘探遗迹坐标表"],
}

IMAGE_EXT_TO_FORMAT = {
    ".jpg": "JPEG",
    ".jpeg": "JPEG",
    ".png": "PNG",
}


@dataclass
class CheckReport:
    info: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def add_info(self, msg: str) -> None:
        self.info.append(msg)

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)

    def add_error(self, msg: str) -> None:
        self.errors.append(msg)

    def text(self) -> str:
        parts = ["生成检查报告", ""]
        for title, rows in [("错误", self.errors), ("警告", self.warnings), ("信息", self.info)]:
            parts.append(f"## {title}")
            if rows:
                parts.extend(f"- {row}" for row in rows)
            else:
                parts.append("- 无")
            parts.append("")
        return "\n".join(parts)


@dataclass
class ProjectData:
    fields: dict[str, str]
    tables: dict[str, list[dict[str, str]]]

    def get(self, key: str, default: str = "") -> str:
        return clean_value(self.fields.get(key, default))


def clean_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", "", clean_value(value).replace("：", "").replace(":", ""))


def canonical_key(key: str) -> str:
    normalized = normalize_key(key)
    for canonical, aliases in FIELD_ALIASES.items():
        if normalized in {normalize_key(a) for a in aliases}:
            return canonical
    return clean_value(key)


def truthy_relic(value: str) -> bool:
    text = clean_value(value)
    if not text:
        return False
    if any(token in text for token in ["未发现", "无遗迹", "无文化遗存", "否", "无"]):
        return False
    return any(token in text for token in ["有", "发现", "灰坑", "墓", "遗迹", "文化遗存", "是"])


def parse_kv_sheet(ws) -> dict[str, str]:
    data: dict[str, str] = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return data

    headers = [clean_value(v) for v in rows[0]]
    if len([h for h in headers if h]) >= 2 and "字段" not in headers[:2]:
        for idx, header in enumerate(headers):
            if header and len(rows) > 1:
                data[canonical_key(header)] = clean_value(rows[1][idx])
        return data

    for row in rows[1:] if "字段" in headers else rows:
        if len(row) < 2:
            continue
        key, value = clean_value(row[0]), clean_value(row[1])
        if key:
            data[canonical_key(key)] = value
    return data


def parse_table_sheet(ws) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = 0
    while header_idx < len(rows) and not any(clean_value(v) for v in rows[header_idx]):
        header_idx += 1
    if header_idx >= len(rows):
        return []
    headers = [clean_value(v) for v in rows[header_idx]]
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        item = {headers[i]: clean_value(row[i]) if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        if any(item.values()):
            out.append(item)
    return out


def load_project_data(input_dir: Path, report: CheckReport) -> ProjectData:
    workbook_path = input_dir / "项目信息.xlsx"
    if not workbook_path.exists():
        report.add_error(f"缺少输入文件：{workbook_path}")
        return ProjectData({}, {})

    wb = load_workbook(workbook_path, data_only=True)
    missing_sheets = sorted(REQUIRED_SHEETS - set(wb.sheetnames))
    if missing_sheets:
        report.add_error("Excel 缺少工作表：" + "、".join(missing_sheets))

    fields: dict[str, str] = {}
    tables: dict[str, list[dict[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in KV_SHEETS:
            fields.update(parse_kv_sheet(wb[sheet_name]))
        elif sheet_name in TABLE_SHEETS:
            tables[sheet_name] = parse_table_sheet(wb[sheet_name])

    for required in REQUIRED_FIELDS:
        if not clean_value(fields.get(required)):
            report.add_error(f"必填字段为空：{required}")

    if truthy_relic(fields.get("遗迹结论", "")) and not tables.get("遗迹信息"):
        report.add_warning("遗迹结论显示发现遗迹，但“遗迹信息”工作表为空；将保留模板中的遗迹段落。")

    return ProjectData(fields, tables)


def select_template(template_dir: Path, project: ProjectData, report: CheckReport) -> Path | None:
    unit = project.get("勘探单位")
    has_relics = truthy_relic(project.get("遗迹结论"))
    candidates = sorted(template_dir.glob("*.docx"))
    matched = []
    for path in candidates:
        name = path.name
        unit_ok = unit and unit in name
        relic_ok = ("有遗迹" in name) if has_relics else ("无遗迹" in name)
        if unit_ok and relic_ok:
            matched.append(path)
    if len(matched) == 1:
        report.add_info(f"选择模板：{matched[0].name}")
        return matched[0]
    if not matched:
        report.add_error(f"没有找到匹配模板：勘探单位={unit or '空'}，是否有遗迹={has_relics}")
        return None
    report.add_error("匹配到多个模板：" + "；".join(p.name for p in matched))
    return None


def resolve_template_dir(template_dir: Path | None, report: CheckReport | None = None) -> Path:
    if template_dir is not None and template_dir.exists():
        return template_dir
    candidates = []
    if template_dir is not None:
        candidates.append(template_dir)
    candidates.extend(DEFAULT_TEMPLATE_DIRS)
    for candidate in candidates:
        if candidate.exists() and any(candidate.glob("*.docx")):
            return candidate
    message = "未找到模板目录；请确认存在 `旧模板/` 或 `基础信息/旧模板/`，或使用 --templates 指定。"
    if report is not None:
        report.add_error(message)
    return template_dir or DEFAULT_TEMPLATE_DIRS[-1]


def qn(tag: str) -> str:
    prefix, local = tag.split(":")
    return f"{{{NS[prefix]}}}{local}"


def text_nodes(elem: ET.Element) -> list[ET.Element]:
    return [node for node in elem.iter(qn("w:t"))]


def element_text(elem: ET.Element) -> str:
    pieces = []
    for node in elem.iter():
        if node.tag == qn("w:t") and node.text:
            pieces.append(node.text)
        elif node.tag == qn("w:tab"):
            pieces.append(" ")
    return "".join(pieces).strip()


def set_element_text(elem: ET.Element, value: str) -> None:
    nodes = text_nodes(elem)
    if not nodes:
        p = elem.find(".//w:p", NS)
        if p is None:
            p = ET.SubElement(elem, qn("w:p"))
        r = ET.SubElement(p, qn("w:r"))
        t = ET.SubElement(r, qn("w:t"))
        t.text = value
        return
    nodes[0].text = value
    for node in nodes[1:]:
        node.text = ""


def extract_template_fields(root: ET.Element) -> dict[str, str]:
    found: dict[str, str] = {}
    for p in root.findall(".//w:p", NS):
        text = element_text(p).replace("[]", "")
        for prefix, field in PREFIX_FIELDS.items():
            if text.startswith(prefix) and field not in found:
                found[field] = text[len(prefix) :].strip()
    paragraphs = [element_text(p).replace("[]", "") for p in root.findall(".//w:p", NS) if element_text(p)]
    if paragraphs:
        first = paragraphs[0]
        if "考古调查" in first:
            found.setdefault("项目名称", first.split("考古调查")[0].strip())
        else:
            found.setdefault("项目名称", first.strip())
    for paragraph in paragraphs[:8]:
        if re.fullmatch(r"\d{4}年\d{1,2}月", paragraph):
            found.setdefault("报告年月", paragraph)
            break
    return found


def make_replacement_map(old: dict[str, str], project: ProjectData) -> dict[str, str]:
    replacements: dict[str, str] = {}
    for field in ["项目名称", "建设单位", "项目位置", "项目面积", "调查面积", "勘探面积", "勘探时间", "报告年月", "遗迹结论"]:
        old_value = clean_value(old.get(field))
        new_value = project.get(field)
        if old_value and new_value and old_value != new_value:
            replacements[old_value] = new_value
    return replacements


def replace_text_in_part(root: ET.Element, replacements: dict[str, str], project: ProjectData) -> int:
    changed = 0
    for elem in root.findall(".//w:p", NS) + root.findall(".//w:tc", NS):
        original = element_text(elem).replace("[]", "")
        if not original:
            continue
        new_text = original
        for prefix, field in PREFIX_FIELDS.items():
            value = project.get(field)
            if value and original.startswith(prefix):
                new_text = prefix + value
                break
        else:
            for old, new in sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True):
                if old and old in new_text:
                    new_text = new_text.replace(old, new)

        if new_text != original:
            set_element_text(elem, new_text)
            changed += 1
    return changed


def iter_body_blocks(root: ET.Element) -> list[ET.Element]:
    body = root.find("w:body", NS)
    return list(body) if body is not None else []


def find_table_after_caption(root: ET.Element, captions: Iterable[str]) -> ET.Element | None:
    captions = list(captions)
    seen = False
    for block in iter_body_blocks(root):
        if block.tag == qn("w:p"):
            text = element_text(block)
            if any(caption in text for caption in captions):
                seen = True
        elif seen and block.tag == qn("w:tbl"):
            return block
    return None


def table_rows(tbl: ET.Element) -> list[ET.Element]:
    return tbl.findall("w:tr", NS)


def row_cells(tr: ET.Element) -> list[ET.Element]:
    return tr.findall("w:tc", NS)


def replace_table_rows(tbl: ET.Element, records: list[dict[str, str]]) -> None:
    if not records:
        return
    rows = table_rows(tbl)
    if not rows:
        return
    template_row = rows[-1]
    header_rows = 1 if len(rows) > 1 else 0
    for old_row in rows[header_rows:]:
        tbl.remove(old_row)

    headers = list(records[0].keys())
    for record in records:
        new_row = copy.deepcopy(template_row)
        cells = row_cells(new_row)
        for idx, cell in enumerate(cells):
            value = record.get(headers[idx], "") if idx < len(headers) else ""
            set_element_text(cell, value)
        tbl.append(new_row)


def update_tables(root: ET.Element, project: ProjectData, report: CheckReport) -> None:
    for sheet_name, captions in TABLE_CAPTIONS.items():
        records = project.tables.get(sheet_name, [])
        if not records:
            continue
        tbl = find_table_after_caption(root, captions)
        if tbl is None:
            report.add_warning(f"未在模板中找到“{sheet_name}”对应表格，已跳过。")
            continue
        replace_table_rows(tbl, records)
        report.add_info(f"已更新表格：{sheet_name}（{len(records)} 行）")


def paragraph_index_by_text(body: ET.Element, predicate) -> int | None:
    for idx, child in enumerate(list(body)):
        if child.tag == qn("w:p") and predicate(element_text(child).replace("[]", "")):
            return idx
    return None


def remove_relic_blocks(root: ET.Element, keep_count: int, report: CheckReport) -> None:
    body = root.find("w:body", NS)
    if body is None:
        return
    blocks = list(body)
    starts: list[tuple[int, int]] = []
    for idx, block in enumerate(blocks):
        if block.tag != qn("w:p"):
            continue
        match = re.match(r"^(\d+)[.、].*H(\d+)\b", element_text(block).replace("[]", ""))
        if match:
            starts.append((idx, int(match.group(2))))
    if not starts:
        return
    end_idx = paragraph_index_by_text(body, lambda t: t.startswith("五、资料整理") or t.startswith("六、结论"))
    if end_idx is None:
        end_idx = len(blocks)

    to_remove: list[ET.Element] = []
    for pos, (start, h_num) in enumerate(starts):
        next_start = starts[pos + 1][0] if pos + 1 < len(starts) else end_idx
        if h_num > keep_count:
            to_remove.extend(blocks[start:next_start])
    for block in to_remove:
        if block in list(body):
            body.remove(block)
    if to_remove:
        report.add_info(f"已删除多余遗迹段落组：{len(to_remove)} 个 XML 块")


def update_relic_sections(root: ET.Element, project: ProjectData, report: CheckReport) -> None:
    records = project.tables.get("遗迹信息", [])
    if not truthy_relic(project.get("遗迹结论")) or not records:
        return
    if len(records) > 30:
        report.add_warning("遗迹数量超过旧模板 30 处；已填充前 30 处，其余请人工补充或扩展模板。")
    usable = records[:30]
    def relic_description(idx: int, record: dict[str, str]) -> str:
        relic_id = clean_value(record.get("编号") or record.get("遗迹编号") or f"H{idx}")
        relic_type = clean_value(record.get("类型") or record.get("遗迹类型") or "灰坑")
        description = clean_value(record.get("描述") or record.get("文字描述"))
        if description:
            return description
        unit = clean_value(record.get("勘探单元"))
        coord = clean_value(record.get("坐标") or record.get("中点坐标"))
        shape = clean_value(record.get("形制"))
        size = clean_value(record.get("尺寸"))
        depth = clean_value(record.get("深度") or record.get("开口深度"))
        remains = clean_value(record.get("包含物"))
        lead = f"{relic_type}{relic_id}"
        if unit:
            lead += f"位于勘探单元{unit}"
        if coord:
            lead += f"，遗迹中点坐标:{coord}"
        tail = "，".join(v for v in [shape, size, depth, remains] if v)
        return lead + ("。" if not tail else f"。{tail}。")

    for p in root.findall(".//w:p", NS):
        text = element_text(p).replace("[]", "")
        if not text:
            continue
        new_text = text
        title_match = re.match(r"^(\d+)[.、].*H(\d+)\b", text)
        if title_match:
            idx = int(title_match.group(2))
            if idx <= len(usable):
                record = usable[idx - 1]
                relic_id = clean_value(record.get("编号") or record.get("遗迹编号") or f"H{idx}")
                relic_type = clean_value(record.get("类型") or record.get("遗迹类型") or "灰坑")
                new_text = f"{idx}.{relic_type}{relic_id}"
        else:
            desc_match = re.match(r"^灰坑H(\d+)位于", text)
            if desc_match:
                idx = int(desc_match.group(1))
                if idx <= len(usable):
                    record = usable[idx - 1]
                    new_text = relic_description(idx, record)
        if new_text != text:
            set_element_text(p, new_text)

    remove_relic_blocks(root, len(usable), report)
    report.add_info(f"已处理遗迹段落：{len(usable)} 处")


def media_targets_from_rels(rels_root: ET.Element) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for rel in rels_root.findall("rel:Relationship", NS):
        target = rel.attrib.get("Target", "")
        if target.startswith("media/"):
            out.append((rel.attrib.get("Id", ""), "word/" + target))

    def image_number(item: tuple[str, str]) -> int:
        match = re.search(r"image(\d+)", item[1])
        return int(match.group(1)) if match else 10**9

    return sorted(out, key=image_number)


def image_bytes_for_target(source: Path, target_name: str, report: CheckReport) -> bytes | None:
    suffix = Path(target_name).suffix.lower()
    if suffix not in IMAGE_EXT_TO_FORMAT:
        report.add_warning(f"模板图片格式暂不支持自动替换：{target_name}")
        return None
    try:
        with Image.open(source) as im:
            if suffix in {".jpg", ".jpeg"} and im.mode in {"RGBA", "LA", "P"}:
                im = im.convert("RGB")
            bio = io.BytesIO()
            im.save(bio, format=IMAGE_EXT_TO_FORMAT[suffix])
            return bio.getvalue()
    except Exception as exc:
        report.add_warning(f"图片转换失败：{source.name} -> {target_name}（{exc}）")
        return None


def update_images(zip_entries: dict[str, bytes], input_dir: Path, project: ProjectData, report: CheckReport) -> None:
    rows = project.tables.get("图片清单", [])
    if not rows:
        return
    rels_name = "word/_rels/document.xml.rels"
    if rels_name not in zip_entries:
        report.add_warning("模板缺少 document.xml.rels，无法替换图片。")
        return
    rels_root = ET.fromstring(zip_entries[rels_name])
    targets = media_targets_from_rels(rels_root)
    image_dir = input_dir / "图片"
    replaced = 0
    for idx, row in enumerate(rows, start=1):
        filename = clean_value(row.get("文件名") or row.get("图片文件名"))
        if not filename:
            continue
        order_text = clean_value(row.get("模板图片序号") or row.get("排序"))
        target_idx = int(order_text) if order_text.isdigit() else idx
        if target_idx < 1 or target_idx > len(targets):
            report.add_warning(f"图片序号超出模板范围：{filename} -> {target_idx}")
            continue
        source = image_dir / filename
        if not source.exists():
            report.add_warning(f"图片缺失：{source}")
            continue
        target = targets[target_idx - 1][1]
        data = image_bytes_for_target(source, target, report)
        if data is None:
            continue
        zip_entries[target] = data
        replaced += 1
    if replaced:
        report.add_info(f"已替换图片：{replaced} 张")


def load_docx_entries(template_path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(template_path) as zf:
        return {name: zf.read(name) for name in zf.namelist()}


def write_docx_entries(entries: dict[str, bytes], output_path: Path) -> None:
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in entries.items():
            zf.writestr(name, data)


def mark_toc_fields_for_update(entries: dict[str, bytes], root: ET.Element | None = None) -> int:
    """Ask Word to refresh TOC fields while preserving TOC font/paragraph styles."""
    if root is None:
        root = ET.fromstring(entries["word/document.xml"])
    dirty_count = 0
    for para in root.findall(".//w:p", NS):
        instr = "".join(node.text or "" for node in para.findall(".//w:instrText", NS))
        if not re.search(r"\bTOC\b", instr):
            continue
        for fld_char in para.findall(".//w:fldChar", NS):
            if fld_char.get(qn("w:fldCharType")) == "begin":
                if fld_char.get(qn("w:dirty")) != "true":
                    fld_char.set(qn("w:dirty"), "true")
                    dirty_count += 1
                break
    entries["word/document.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    settings_name = "word/settings.xml"
    if settings_name in entries:
        settings_root = ET.fromstring(entries[settings_name])
    else:
        settings_root = ET.Element(qn("w:settings"))
    update_fields = settings_root.find("w:updateFields", NS)
    if update_fields is None:
        update_fields = ET.SubElement(settings_root, qn("w:updateFields"))
    update_fields.set(qn("w:val"), "true")
    entries[settings_name] = ET.tostring(settings_root, encoding="utf-8", xml_declaration=True)
    return dirty_count


def safe_filename(name: str) -> str:
    return re.sub(r'[/:*?"<>|\\]+', "_", name).strip() or "考古调查勘探工作报告"


def is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def ensure_not_in_template_library(path: Path, purpose: str) -> None:
    if is_relative_to(path, TEMPLATE_LIBRARY_DIR):
        raise ValueError(f"{purpose}不能位于模板库 `{TEMPLATE_LIBRARY_DIR}` 内；请放入 `{DEFAULT_PROCESS_DIR}` 或 `{DEFAULT_OUTPUT_DIR}`。")


def generate_report(input_dir: Path, output_dir: Path, template_dir: Path | None) -> tuple[Path | None, Path]:
    report = CheckReport()
    project = load_project_data(input_dir, report)
    template_dir = resolve_template_dir(template_dir, report)
    try:
        ensure_not_in_template_library(output_dir, "生成输出目录")
    except ValueError as exc:
        report.add_error(str(exc))
        output_dir = DEFAULT_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    project_label = safe_filename(project.get("项目名称") or input_dir.name)
    check_path = output_dir / f"{project_label}-生成检查报告.txt"

    template_path = None if report.errors else select_template(template_dir, project, report)
    if report.errors or template_path is None:
        check_path.write_text(report.text(), encoding="utf-8")
        return None, check_path

    entries = load_docx_entries(template_path)
    doc_root = ET.fromstring(entries["word/document.xml"])
    old_fields = extract_template_fields(doc_root)
    replacements = make_replacement_map(old_fields, project)

    changed_parts = 0
    for name in list(entries.keys()):
        if re.match(r"word/(document|header\d+|footer\d+)\.xml$", name):
            root = ET.fromstring(entries[name])
            changed = replace_text_in_part(root, replacements, project)
            if name == "word/document.xml":
                update_tables(root, project, report)
                update_relic_sections(root, project, report)
                doc_root = root
            entries[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            changed_parts += changed
    update_images(entries, input_dir, project, report)
    report.add_info(f"已替换文本段落/单元格：{changed_parts} 处")
    toc_fields_marked = mark_toc_fields_for_update(entries, doc_root)
    report.add_info(f"目录域已标记为打开时自动更新：TOC域 {toc_fields_marked} 处；目录文字可更新，需保留原字体格式")

    output_name = f"{project_label}-考古调查勘探工作报告.docx"
    output_path = output_dir / output_name
    write_docx_entries(entries, output_path)
    report.add_info(f"已生成报告：{output_path}")
    check_path.write_text(report.text(), encoding="utf-8")
    return output_path, check_path


def write_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "项目概况"
    ws.append(["字段", "值", "说明"])
    for row in [
        ["项目名称", "示例项目", "必填"],
        ["建设单位", "示例建设单位", "必填"],
        ["勘探单位", "北京卓凡文博技术有限公司", "必须与旧模板中的单位名称一致"],
        ["项目位置", "内蒙古自治区鄂尔多斯市伊金霍洛旗", "必填"],
        ["项目面积", "100000平方米", "必填"],
        ["调查面积", "100000平方米", "可空"],
        ["勘探面积", "80000平方米", "必填"],
        ["勘探时间", "2026年5月1日至2026年5月10日", "必填"],
        ["报告年月", "2026年5月", "可空"],
        ["遗迹结论", "未发现文化遗存", "填“灰坑3个”等会选择有遗迹模板"],
        ["经度", "109°31′27″", "可空"],
        ["纬度", "39°34′38″", "可空"],
        ["地块现状", "地势平坦，土色黄褐，植被以沙生植物为主", "可空"],
    ]:
        ws.append(row)

    ws = wb.create_sheet("人员信息")
    ws.append(["字段", "值", "说明"])
    ws.append(["说明", "默认不替换公司人员信息", "公司、人员、资质附件由匹配到的旧模板原样保留"])

    ws = wb.create_sheet("红线坐标")
    ws.append(["角点", "X", "Y"])
    for row in [
        ["西南角", "37372298.938", "4383074.828"],
        ["东南角", "37373096.938", "4383074.828"],
        ["东北角", "37373096.938", "4384118.018"],
        ["西北角", "37372298.938", "4384118.018"],
    ]:
        ws.append(row)

    for sheet, headers in [
        ("勘探单元", ["单元编号", "西南角X", "西南角Y", "东北角X", "东北角Y", "面积", "说明"]),
        ("标准孔与剖线", ["编号", "位置", "地层描述", "图件"]),
        ("遗迹信息", ["编号", "类型", "勘探单元", "坐标", "形制", "尺寸", "深度", "包含物", "描述"]),
        ("图片清单", ["排序", "模板图片序号", "文件名", "插入位置", "图题"]),
    ]:
        ws = wb.create_sheet(sheet)
        ws.append(headers)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(12, min(36, max(len(clean_value(c.value)) for c in col) + 2))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def init_input(output_dir: Path) -> None:
    ensure_not_in_template_library(output_dir, "过程资料目录")
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "图片").mkdir(exist_ok=True)
    (output_dir / "附件").mkdir(exist_ok=True)
    (output_dir / "图件").mkdir(exist_ok=True)
    (output_dir / "勘探记录").mkdir(exist_ok=True)
    write_sample_workbook(output_dir / "项目信息.xlsx")
    readme = """# 项目资料包

请填写 `项目信息.xlsx`，并把报告图片放入 `图片/`。

生成命令：

```bash
python3 ../generate_report.py generate . --output 输出
```
"""
    (output_dir / "README.md").write_text(readme, encoding="utf-8")


def run_regression(template_dir: Path | None) -> int:
    report = CheckReport()
    template_dir = resolve_template_dir(template_dir, report)
    units = sorted({p.name.split("—")[0] for p in template_dir.glob("*.docx") if "—" in p.name})
    errors = []
    for unit in units:
        for relic_text, expected in [("未发现文化遗存", "无遗迹"), ("灰坑1个", "有遗迹")]:
            project = ProjectData(
                {
                    "项目名称": "测试项目",
                    "建设单位": "测试建设单位",
                    "勘探单位": unit,
                    "项目位置": "测试位置",
                    "项目面积": "1平方米",
                    "勘探面积": "1平方米",
                    "勘探时间": "2026年1月1日至2026年1月2日",
                    "遗迹结论": relic_text,
                },
                {},
            )
            report = CheckReport()
            selected = select_template(template_dir, project, report)
            if selected is None or expected not in selected.name:
                errors.append(f"{unit}/{expected} 模板选择失败：{selected}")
    if errors:
        print("\n".join(errors), file=sys.stderr)
        return 1
    print(f"模板选择回归测试通过：{len(units) * 2}/{len(units) * 2}")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="考古调查、勘探报告自动生成器")
    sub = parser.add_subparsers(dest="command", required=True)

    init_parser = sub.add_parser("init-input", help="创建项目资料包模板")
    init_parser.add_argument("output_dir", type=Path, nargs="?", default=DEFAULT_PROCESS_DIR / "新项目资料")

    gen_parser = sub.add_parser("generate", help="根据项目资料包生成报告")
    gen_parser.add_argument("input_dir", type=Path)
    gen_parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_DIR)
    gen_parser.add_argument("--templates", type=Path, default=None)

    test_parser = sub.add_parser("self-test", help="运行模板选择回归测试")
    test_parser.add_argument("--templates", type=Path, default=None)

    args = parser.parse_args(argv)
    if args.command == "init-input":
        try:
            init_input(args.output_dir)
        except ValueError as exc:
            print(str(exc), file=sys.stderr)
            return 2
        print(f"已创建项目资料包：{args.output_dir}")
        return 0
    if args.command == "generate":
        output_path, check_path = generate_report(args.input_dir, args.output, args.templates)
        print(f"检查报告：{check_path}")
        if output_path is None:
            print("生成失败，请查看检查报告。", file=sys.stderr)
            return 2
        print(f"生成报告：{output_path}")
        return 0
    if args.command == "self-test":
        return run_regression(args.templates)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())

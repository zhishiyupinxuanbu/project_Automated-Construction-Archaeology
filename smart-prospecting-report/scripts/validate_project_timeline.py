#!/usr/bin/env python3
"""Validate archaeological project timeline consistency before generation.

This script is a hard gate. When it finds a conflict, callers must stop and let
the user resolve the source materials instead of silently adjusting dates.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - reported at runtime for xlsx inputs
    load_workbook = None


OCR_ROOT = Path("/Users/drevan01/Desktop/OCR")
OCR_CLIENT = OCR_ROOT / "scripts" / "ocr_image.sh"
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".heic", ".tif", ".tiff", ".bmp", ".webp"}
TEXT_SUFFIXES = {".txt", ".md"}
DOC_SUFFIXES = {".docx", ".txt", ".md"}


@dataclass(frozen=True)
class TimelineIssue:
    code: str
    message: str
    severity: str = "error"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_date_text(text: str) -> list[date]:
    found: list[date] = []
    for match in re.finditer(r"(?<!\d)(20\d{2})[年./-]\s*(\d{1,2})[月./-]\s*(\d{1,2})日?(?!\d)", text):
        year, month, day = (int(part) for part in match.groups())
        try:
            found.append(date(year, month, day))
        except ValueError:
            continue
    return found


def first_date(*values: object) -> date | None:
    for value in values:
        dates = parse_date_text(clean(value))
        if dates:
            return dates[0]
    return None


def date_range_from_fields(fields: dict[str, str]) -> tuple[date | None, date | None]:
    start = first_date(
        fields.get("开始日期"),
        fields.get("计划开始日期"),
        fields.get("勘探开始日期"),
        fields.get("planned_start_date"),
    )
    end = first_date(
        fields.get("结束日期"),
        fields.get("计划结束日期"),
        fields.get("勘探结束日期"),
        fields.get("planned_end_date"),
    )
    if start and end:
        return start, end

    for key in ["勘探时间", "计划勘探时间", "prospecting_time", "planned_prospecting_time"]:
        text = clean(fields.get(key))
        dates = parse_date_text(text)
        if len(dates) >= 2:
            return start or dates[0], end or dates[1]
    return start, end


def request_date_from_fields(fields: dict[str, str]) -> date | None:
    return first_date(
        fields.get("请示日期"),
        fields.get("打请示日期"),
        fields.get("发文日期"),
        fields.get("request_date"),
        fields.get("issue_date"),
    )


def reply_date_from_fields(fields: dict[str, str]) -> date | None:
    return first_date(fields.get("回函日期"), fields.get("复函日期"), fields.get("reply_date"))


def investigation_end_from_fields(fields: dict[str, str]) -> date | None:
    return first_date(
        fields.get("文物调查结束日期"),
        fields.get("调查结束日期"),
        fields.get("investigation_end_date"),
    )


def read_xlsx_fields(path: Path) -> dict[str, str]:
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法读取 Excel 时间字段。")
    wb = load_workbook(path, read_only=True, data_only=True)
    fields: dict[str, str] = {}
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, values_only=True):
                if len(row) < 2:
                    continue
                key = clean(row[0])
                value = clean(row[1])
                if key and value and key not in fields:
                    fields[key] = value
    finally:
        wb.close()
    return fields


def read_json_fields(path: Path) -> dict[str, str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"JSON 顶层必须是对象：{path}")
    return {str(key): clean(value) for key, value in data.items()}


def docx_text(path: Path) -> str:
    try:
        with ZipFile(path) as zf:
            parts = []
            for name in ["word/document.xml", *sorted(n for n in zf.namelist() if n.startswith("word/header"))]:
                if name not in zf.namelist():
                    continue
                xml = zf.read(name).decode("utf-8", errors="ignore")
                parts.append(re.sub(r"<[^>]+>", "", xml))
            return "\n".join(parts)
    except (BadZipFile, OSError):
        return ""


def file_text(path: Path) -> str:
    if path.suffix.lower() == ".docx":
        return docx_text(path)
    if path.suffix.lower() in TEXT_SUFFIXES:
        return path.read_text(encoding="utf-8", errors="ignore")
    return path.stem


def infer_fields_from_project_dir(project_dir: Path) -> dict[str, str]:
    fields: dict[str, str] = {}
    if not project_dir or not project_dir.exists():
        return fields
    for path in sorted(project_dir.rglob("*")):
        if not path.is_file() or path.name.startswith(".") or path.suffix.lower() not in DOC_SUFFIXES:
            continue
        text = f"{path.name}\n{file_text(path)}"
        dates = parse_date_text(text)
        if not dates:
            continue
        newest = max(dates)
        name = path.name
        if ("文物调查" in name or ("调查报告" in name and "勘探" not in name)) and "文物调查结束日期" not in fields:
            fields["文物调查结束日期"] = newest.isoformat()
        if ("回函" in name or "复函" in name) and "回函日期" not in fields:
            fields["回函日期"] = newest.isoformat()
        if "勘探请示" in name or ("申请开展" in name and "勘探" in name and "请示" in name):
            fields.setdefault("请示日期", newest.isoformat())
        if "勘探计划" in name:
            dates_sorted = sorted(dates)
            fields.setdefault("开始日期", dates_sorted[0].isoformat())
            fields.setdefault("结束日期", dates_sorted[-1].isoformat())
        if "勘探报告" in name:
            dates_sorted = sorted(dates)
            fields.setdefault("报告开始日期", dates_sorted[0].isoformat())
            fields.setdefault("报告结束日期", dates_sorted[-1].isoformat())
    return fields


def collect_strings(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(collect_strings(item))
        return result
    if isinstance(value, dict):
        result = []
        for item in value.values():
            result.extend(collect_strings(item))
        return result
    return []


def cached_ocr_text(image: Path, extra_cache_dirs: list[Path] | None = None) -> str:
    candidates = [
        image.with_suffix(".ocr.txt"),
        image.with_suffix(image.suffix + ".txt"),
        image.with_suffix(".txt"),
        image.with_suffix(".ocr.json"),
        OCR_ROOT / "outputs" / f"{image.stem}.ocr.json",
    ]
    for cache_dir in extra_cache_dirs or []:
        candidates.extend(
            [
                cache_dir / f"{image.name}.ocr.txt",
                cache_dir / f"{image.stem}.ocr.txt",
                cache_dir / f"{image.name}.txt",
                cache_dir / f"{image.stem}.txt",
                cache_dir / f"{image.name}.ocr.json",
                cache_dir / f"{image.stem}.ocr.json",
            ]
        )
    parts = [image.stem]
    for candidate in candidates:
        if not candidate.exists():
            continue
        if candidate.suffix == ".json":
            try:
                data = json.loads(candidate.read_text(encoding="utf-8"))
                parts.extend(collect_strings(data))
            except Exception:
                continue
        else:
            parts.append(candidate.read_text(encoding="utf-8", errors="ignore"))
    return "\n".join(parts)


def run_ocr(image: Path, cache_dir: Path) -> str:
    if not OCR_CLIENT.exists():
        return ""
    cache_dir.mkdir(parents=True, exist_ok=True)
    output = cache_dir / f"{image.stem}.ocr.json"
    result = subprocess.run(
        [str(OCR_CLIENT), str(image), "--output", str(output)],
        cwd=str(OCR_ROOT),
        check=False,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    if result.returncode != 0:
        return ""
    return cached_ocr_text(image, [cache_dir]) + "\n" + output.read_text(encoding="utf-8", errors="ignore")


def image_files(photos_dir: Path) -> list[Path]:
    if not photos_dir.exists():
        return []
    return sorted(
        path
        for path in photos_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES and not path.name.startswith(".")
    )


def extract_docx_images(docx_path: Path, output_dir: Path) -> list[Path]:
    images: list[Path] = []
    if not docx_path.exists():
        return images
    try:
        with ZipFile(docx_path) as zf:
            for name in zf.namelist():
                if not name.startswith("word/media/"):
                    continue
                suffix = Path(name).suffix.lower()
                if suffix not in IMAGE_SUFFIXES:
                    continue
                target = output_dir / docx_path.stem / Path(name).name
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(zf.read(name))
                images.append(target)
    except (BadZipFile, OSError):
        return []
    return images


def validate_image_paths(
    images: list[Path],
    start: date,
    end: date,
    *,
    require_photos: bool = False,
    ocr_images: bool = False,
    ocr_cache_dir: Path | None = None,
    extra_cache_dirs: list[Path] | None = None,
) -> list[TimelineIssue]:
    issues: list[TimelineIssue] = []
    if not images and require_photos:
        issues.append(TimelineIssue("PHOTO_MISSING", "未找到可校验的图片。"))
        return issues
    cache_dir = ocr_cache_dir or (Path.cwd() / "过程资料" / "时间校验OCR缓存")
    for image in images:
        text = cached_ocr_text(image, extra_cache_dirs)
        dates = parse_date_text(text)
        if not dates and ocr_images:
            text = f"{text}\n{run_ocr(image, cache_dir)}"
            dates = parse_date_text(text)
        if not dates:
            issues.append(
                TimelineIssue(
                    "PHOTO_WATERMARK_DATE_UNDETECTED",
                    f"图片未识别到水印日期，建议用户人工核对照片时间：{image}",
                    severity="warning",
                )
            )
            continue
        for watermark_date in dates:
            if watermark_date < start or watermark_date > end:
                issues.append(
                    TimelineIssue(
                        "PHOTO_WATERMARK_OUT_OF_RANGE",
                        f"图片水印日期不在勘探时间范围内：{image}；水印={watermark_date.isoformat()}；范围={start.isoformat()}至{end.isoformat()}",
                    )
                )
    return issues


def validate_photo_watermarks(
    photos_dir: Path | None,
    start: date | None,
    end: date | None,
    *,
    require_photos: bool = False,
    ocr_images: bool = False,
    ocr_cache_dir: Path | None = None,
) -> list[TimelineIssue]:
    issues: list[TimelineIssue] = []
    if not start or not end:
        return issues
    if not photos_dir:
        if require_photos:
            issues.append(TimelineIssue("PHOTOS_DIR_MISSING", "未提供外业照片目录，无法校验图片水印时间。"))
        return issues
    images = image_files(photos_dir)
    return validate_image_paths(
        images,
        start,
        end,
        require_photos=require_photos,
        ocr_images=ocr_images,
        ocr_cache_dir=ocr_cache_dir,
    )


def validate_docx_image_watermarks(
    docx_paths: list[Path],
    start: date | None,
    end: date | None,
    *,
    require_photos: bool = False,
    ocr_images: bool = False,
    ocr_cache_dir: Path | None = None,
) -> list[TimelineIssue]:
    if not start or not end:
        return []
    issues: list[TimelineIssue] = []
    with tempfile.TemporaryDirectory() as temp:
        temp_dir = Path(temp)
        all_images: list[Path] = []
        cache_dirs: list[Path] = []
        for docx_path in docx_paths:
            all_images.extend(extract_docx_images(docx_path, temp_dir))
            cache_dirs.extend([docx_path.with_suffix(".docx.ocr"), docx_path.with_suffix(".ocr")])
        issues.extend(
            validate_image_paths(
                all_images,
                start,
                end,
                require_photos=require_photos,
                ocr_images=ocr_images,
                ocr_cache_dir=ocr_cache_dir,
                extra_cache_dirs=cache_dirs,
            )
        )
    return issues



def validate_timeline(
    fields: dict[str, str],
    *,
    photos_dir: Path | None = None,
    docx_paths: list[Path] | None = None,
    require_request_sequence: bool = False,
    require_photos: bool = False,
    ocr_images: bool = False,
    ocr_cache_dir: Path | None = None,
) -> list[TimelineIssue]:
    issues: list[TimelineIssue] = []
    request_date = request_date_from_fields(fields)
    reply_date = reply_date_from_fields(fields)
    investigation_end = investigation_end_from_fields(fields)
    start, end = date_range_from_fields(fields)

    if require_request_sequence:
        if not investigation_end:
            issues.append(TimelineIssue("INVESTIGATION_END_MISSING", "缺少文物调查工作结束日期，不能生成勘探请示。"))
        if not reply_date:
            issues.append(TimelineIssue("REPLY_DATE_MISSING", "缺少文物回函/复函日期，不能生成勘探请示。"))
        if not request_date:
            issues.append(TimelineIssue("REQUEST_DATE_MISSING", "缺少勘探请示日期，不能校验请示与计划勘探时间。"))
        if not start:
            issues.append(TimelineIssue("PLAN_START_MISSING", "缺少计划勘探开始日期，不能校验请示后 2-3 天开工规则。"))

    if start and end and start > end:
        issues.append(TimelineIssue("DATE_RANGE_REVERSED", f"勘探开始日期晚于结束日期：{start.isoformat()} > {end.isoformat()}"))

    if request_date and investigation_end and request_date <= investigation_end:
        issues.append(
            TimelineIssue(
                "REQUEST_BEFORE_INVESTIGATION_END",
                f"勘探请示日期必须在文物调查工作结束之后：请示={request_date.isoformat()}；调查结束={investigation_end.isoformat()}",
            )
        )
    if request_date and reply_date and request_date < reply_date:
        issues.append(
            TimelineIssue(
                "REQUEST_BEFORE_REPLY",
                f"勘探请示日期不得早于企业取得文物回函/复函日期：请示={request_date.isoformat()}；回函/复函={reply_date.isoformat()}",
            )
        )
    if request_date and start:
        delta = (start - request_date).days
        if delta not in {2, 3}:
            issues.append(
                TimelineIssue(
                    "REQUEST_TO_START_DELTA",
                    f"计划勘探开始日期应在勘探请示日期后 2-3 天：请示={request_date.isoformat()}；开始={start.isoformat()}；间隔={delta}天",
                )
            )

    issues.extend(
        validate_photo_watermarks(
            photos_dir,
            start,
            end,
            require_photos=require_photos and not docx_paths,
            ocr_images=ocr_images,
            ocr_cache_dir=ocr_cache_dir,
        )
    )
    if docx_paths:
        issues.extend(
            validate_docx_image_watermarks(
                docx_paths,
                start,
                end,
                require_photos=require_photos,
                ocr_images=ocr_images,
                ocr_cache_dir=ocr_cache_dir,
            )
        )
    return issues


def merge_fields(*sources: dict[str, str]) -> dict[str, str]:
    merged: dict[str, str] = {}
    for source in sources:
        for key, value in source.items():
            if clean(value) and key not in merged:
                merged[key] = clean(value)
    return merged


def error_issues(issues: list[TimelineIssue]) -> list[TimelineIssue]:
    return [issue for issue in issues if issue.severity != "warning"]


def warning_issues(issues: list[TimelineIssue]) -> list[TimelineIssue]:
    return [issue for issue in issues if issue.severity == "warning"]


def print_timeline_result(issues: list[TimelineIssue]) -> int:
    errors = error_issues(issues)
    warnings = warning_issues(issues)
    watermark_warnings = [issue for issue in warnings if issue.code == "PHOTO_WATERMARK_DATE_UNDETECTED"]
    other_warnings = [issue for issue in warnings if issue.code != "PHOTO_WATERMARK_DATE_UNDETECTED"]
    if errors:
        print("时间校验失败，已停止生成。请用户先处理以下冲突：", file=sys.stderr)
        for issue in errors:
            print(f"- [{issue.code}] {issue.message}", file=sys.stderr)
        if warnings:
            print("同时建议人工复核以下照片时间：", file=sys.stderr)
            for issue in other_warnings:
                print(f"- [{issue.code}] {issue.message}", file=sys.stderr)
            if watermark_warnings:
                print(
                    f"- [PHOTO_WATERMARK_DATE_UNDETECTED] 有 {len(watermark_warnings)} 张图片未识别到水印日期，建议用户人工核对照片时间。",
                    file=sys.stderr,
                )
        return 2
    if warnings:
        print("时间校验通过，但建议人工复核以下照片时间：", file=sys.stderr)
        for issue in other_warnings:
            print(f"- [{issue.code}] {issue.message}", file=sys.stderr)
        if watermark_warnings:
            print(
                f"- [PHOTO_WATERMARK_DATE_UNDETECTED] 有 {len(watermark_warnings)} 张图片未识别到水印日期，建议用户人工核对照片时间。",
                file=sys.stderr,
            )
        return 0
    print("时间校验通过。")
    return 0


def should_require_request_sequence(doc_type: str, fields: dict[str, str]) -> bool:
    joined = doc_type + " " + " ".join(clean(value) for value in fields.values())
    return (
        "勘探请示" in joined
        or "勘探计划备案请示" in joined
        or ("申请开展" in joined and "勘探" in joined and "请示" in joined)
        or ("勘探" in joined and "计划" in joined and "请示" in joined)
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="校验文物调查、回函/复函、勘探请示、勘探计划/报告及照片水印时间。")
    parser.add_argument("--form", type=Path, help="智能报告/计划信息填报表 xlsx。")
    parser.add_argument("--project-json", type=Path, help="公文生成 project.json。")
    parser.add_argument("--project-dir", type=Path, help="项目源资料目录，用于补充读取文物调查报告、回函/复函、勘探计划/报告日期。")
    parser.add_argument("--photos-dir", type=Path, help="外业照片目录。")
    parser.add_argument("--docx", type=Path, action="append", default=[], help="计划或报告 Word，用于抽取内嵌图片并校验水印日期；可重复传入。")
    parser.add_argument("--doc-type", default="", help="当前拟生成文种，例如 勘探请示、勘探计划、勘探报告。")
    parser.add_argument("--require-request-sequence", action="store_true", help="强制要求调查结束、回函/复函、请示日期和计划开始日期齐全并符合顺序。")
    parser.add_argument("--require-photos", action="store_true", help="强制要求校验外业照片水印日期。")
    parser.add_argument("--ocr-images", action="store_true", help="缓存或文件名未识别到水印日期时调用本地 Paddle OCR。")
    parser.add_argument("--ocr-cache-dir", type=Path, default=None, help="本次 OCR JSON 缓存目录。")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    field_sources: list[dict[str, str]] = []
    if args.form:
        field_sources.append(read_xlsx_fields(args.form.resolve()))
    if args.project_json:
        field_sources.append(read_json_fields(args.project_json.resolve()))
    if args.project_dir:
        field_sources.append(infer_fields_from_project_dir(args.project_dir.resolve()))
    fields = merge_fields(*field_sources)
    require_request_sequence = args.require_request_sequence or should_require_request_sequence(args.doc_type, fields)
    issues = validate_timeline(
        fields,
        photos_dir=args.photos_dir.resolve() if args.photos_dir else None,
        docx_paths=[path.resolve() for path in args.docx],
        require_request_sequence=require_request_sequence,
        require_photos=args.require_photos or bool(args.docx),
        ocr_images=args.ocr_images or bool(args.docx),
        ocr_cache_dir=args.ocr_cache_dir.resolve() if args.ocr_cache_dir else None,
    )
    return print_timeline_result(issues)


if __name__ == "__main__":
    raise SystemExit(main())

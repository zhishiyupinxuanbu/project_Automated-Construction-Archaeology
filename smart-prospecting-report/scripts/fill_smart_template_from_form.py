#!/usr/bin/env python3
"""Fill the smart archaeology report template from the workbook form.

Inputs stay in 过程资料; generated DOCX and checks default to the user's Desktop.
The template/source library 基础信息 is intentionally never written.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import os
import re
import subprocess
import urllib.error
import urllib.request
import zipfile
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image

try:
    import region_overview_agent
except ImportError:  # pragma: no cover - only used when imported from unusual paths
    region_overview_agent = None

try:
    import create_manual_form_from_project as manual_form_builder
except ImportError:  # pragma: no cover - only used when imported from unusual paths
    manual_form_builder = None


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_DIR.parent
ASSETS_DIR = SKILL_ROOT / "assets"
ROOT = Path.cwd().resolve()
PROCESS_DIR = ROOT / "过程资料"
CORE_DIR = ASSETS_DIR
REPORT_DIR = Path.home() / "Desktop"
TEMPLATE_LIBRARY_DIR = ROOT / "基础信息"
SMART_TEMPLATE_DIR = CORE_DIR / "templates"
REPORT_PLAN_TEMPLATE_DIR = SMART_TEMPLATE_DIR / "reports"
DATA_ORGANIZATION_WORK_PHOTO_LIBRARY = CORE_DIR / "photo-libraries" / "资料整理工作照-照片库"
DATA_ORGANIZATION_WORK_PHOTO_ROTATION_STATE = DATA_ORGANIZATION_WORK_PHOTO_LIBRARY / "资料整理工作照轮换状态.json"
COMPANY_PERSONNEL_PACKAGE_DIR = CORE_DIR / "company-personnel-library"


def find_smart_template(filename: str) -> Path:
    direct_candidates = [
        REPORT_PLAN_TEMPLATE_DIR / filename,
        SMART_TEMPLATE_DIR / filename,
    ]
    for path in direct_candidates:
        if path.exists():
            return path
    matches = sorted(
        path
        for path in REPORT_PLAN_TEMPLATE_DIR.rglob(filename)
        if path.is_file() and not path.name.startswith("._") and not path.name.startswith(".~")
    )
    if not matches:
        matches = sorted(
            path
            for path in SMART_TEMPLATE_DIR.rglob(filename)
            if path.is_file() and not path.name.startswith("._") and not path.name.startswith(".~")
        )
    return matches[0] if matches else direct_candidates[0]


DEFAULT_TEMPLATE = find_smart_template("智能报告生成基准模板_北京卓凡文博技术有限公司_无遗迹.docx")
DEFAULT_FORM = PROCESS_DIR / "智能报告信息填报表.xlsx"
REGION_OVERVIEW_MODEL = "doubao-seed-2-0-lite-260215"
REGION_OVERVIEW_KEYS = [
    "项目所在地旗县地理位置概况",
    "项目所在地旗县行政区划与社会经济概况",
    "项目所在地旗县气候条件",
    "项目所在地旗县历史沿革",
]
MATCH_ERROR_TEXT = "匹配错误"
IMAGE_INSERT_ERROR_TEXT = "插入错误"
MISSING_FIELD_TEXT = "未填入"
NO_DOCUMENT_NUMBER_VALUES = {"", "无", "无文号", "未见文号", "未识别", "未识别文号", "无正式文号"}

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
WP_NS = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
PIC_NS = "http://schemas.openxmlformats.org/drawingml/2006/picture"
NS = {"w": W_NS, "r": R_NS, "rel": REL_NS, "ct": CT_NS, "wp": WP_NS, "a": A_NS, "pic": PIC_NS}
ET.register_namespace("w", W_NS)
ET.register_namespace("r", R_NS)
ET.register_namespace("wp", WP_NS)
ET.register_namespace("a", A_NS)
ET.register_namespace("pic", PIC_NS)

KV_SHEETS = {"项目基础信息", "项目区域概况", "文物概况", "勘探参数", "现场限制", "人员构成", "自动生成字段"}
TABLE_SHEETS = {
    "红线坐标",
    "勘探单元",
    "剖线地层堆积",
    "标准孔",
    "遗迹记录",
    "遗迹坐标",
    "文物范围遗迹统计",
    "图件照片目录",
    "图片清单",
}
REQUIRED_FIELDS = ["项目名称", "建设单位", "项目位置", "项目面积"]
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
FIELDWORK_OPTIONAL_PHOTO_KEYS = {
    "实地踏查照",
    "走访调查照",
    "项目地块现状照",
    "勘探单元布置照",
    "布设探孔照",
    "普探工作照",
    "取样记录照",
    "勘探后局部照",
    "勘探后航拍照",
    "信息采集照",
    "资料整理工作照",
    "文献资料收集与整理工作照",
}
RECORD_SCOPED_IMAGE_KEYS = {
    "标准孔位置图",
    "标准孔土样照",
    "遗迹土样照",
    "遗迹现场照",
    "遗迹平、剖面图",
}
SINGLE_EXPANSION_IMAGE_KEYS = {
    "项目地块位置示意图",
    "项目位置图",
    "文献资料收集与整理工作照",
    "资料整理工作照",
}
IMAGE_MAX_WIDTH_EMU = int(6.2 * 914400)
IMAGE_MAX_HEIGHT_EMU = int(8.6 * 914400)
EMU_PER_CM = 360000
PHOTO_WIDTH_CM = 14.63
COMPANY_LICENSE_HEIGHT_CM = 14.63
COMPANY_QUALIFICATION_HEIGHT_CM = 14.63
PERSONNEL_ID_CARD_WIDTH_CM = 10.42
PERSONNEL_CERTIFICATE_WIDTH_CM = 14.63
PERSONNEL_CONTRACT_WIDTH_CM = 14.63
PERSONNEL_SOCIAL_SECURITY_WIDTH_CM = 14.63
PERSONNEL_SOCIAL_SECURITY_LANDSCAPE_HEIGHT_CM = 14.63
COMPANY_SEAL_WIDTH_CM = 4.00
COMPANY_SEAL_VERTICAL_OFFSET_CM = -2.40
DRAWING_SECTION_WIDTH_CM = 24.50
DRAWING_STANDARD_WIDTH_CM = 14.63
DRAWING_ROTATED_HEIGHT_CM = 14.63
RELIC_PLAN_SECTION_WIDTH_CM = 14.63
STANDARD_HOLE_IMAGE_WIDTH_CM = 12.40
STANDARD_HOLE_PHOTO_MAX_HEIGHT_CM = 9.80
APPENDIX_TABLE_ROW_HEIGHT_CM = 0.62
RELIC_ID_PREFIXES = "HMFGYK"
RELIC_ID_RE = rf"[{RELIC_ID_PREFIXES}]\s*0*\d+"
CHINESE_NUMERAL = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}
CHINESE_ORDINAL = {
    1: "一",
    2: "二",
    3: "三",
    4: "四",
    5: "五",
    6: "六",
    7: "七",
    8: "八",
    9: "九",
    10: "十",
    11: "十一",
    12: "十二",
    13: "十三",
    14: "十四",
    15: "十五",
    16: "十六",
    17: "十七",
    18: "十八",
    19: "十九",
    20: "二十",
}


@dataclass
class ImageSlot:
    path: Path
    caption: str = ""
    source: str = ""
    key: str = ""


COMPANY_ATTACHMENT_EXTENSIONS = IMAGE_EXTENSIONS
PERSONNEL_SET_ALIASES = {"人员1": "人员信息1", "人员2": "人员信息2"}
PERSONNEL_ATTACHMENT_ROLE_ORDER = ["领队", "技师"]
PERSONNEL_TEXT_ROLE_ALIASES = {
    "领队": ["项目负责"],
    "现场负责人": ["现场负责", "现场负责人"],
    "测绘人员": ["测绘员", "测绘人员"],
}

PROJECT_OVERVIEW_LABELS = [
    "项目名称",
    "建设单位",
    "勘探单位",
    "项目位置",
    "项目面积",
    "调查面积",
    "勘探面积",
    "勘探时间",
    "遗迹",
    "项目负责",
    "现场负责",
    "技师",
    "报告执笔",
    "测绘员",
    "资料员",
    "校核",
]

PROJECT_OVERVIEW_FIELD_MAP = {
    "遗迹": "遗迹结论",
    "现场负责": "现场负责",
}

REGION_OVERVIEW_PRESETS = {
    "伊金霍洛旗": {
        "项目所在地旗县地理位置概况": (
            "伊金霍洛旗位于内蒙古自治区西南部、鄂尔多斯市中南部，地处鄂尔多斯高原腹地。"
            "旗域东西长、南北宽均具有一定跨度，地貌以丘陵沟壑、梁滩相间地貌及沙地过渡地貌为主，"
            "项目所在的伊金霍洛镇位于旗域中部偏南，是旗内重要的历史文化与文旅资源集中区域。"
        ),
        "项目所在地旗县行政区划与社会经济概况": (
            "伊金霍洛旗为鄂尔多斯市下辖旗，旗人民政府驻阿勒腾席热镇，现辖多个镇及相关园区。"
            "近年来，当地经济以能源、现代服务、文旅融合和新能源产业等为重要支撑，区域交通和城镇基础设施条件较为完善。"
        ),
        "项目所在地旗县气候条件": (
            "伊金霍洛旗属温带大陆性气候，气候特点表现为干旱少雨、日照较强、昼夜温差和季节温差较大，"
            "冬季寒冷干燥，夏季炎热少雨，风沙天气相对较多。区域年降水量约350毫米，年日照时数较长。"
        ),
        "项目所在地旗县历史沿革": (
            "伊金霍洛为蒙古语，意为“圣主的陵园”，因境内成吉思汗陵园而得名。"
            "今旗境历史上曾分属上郡、西河、朔方等郡县及后世相关地方建置。"
            "清代鄂尔多斯地区设盟旗制度后，区域行政隶属逐渐稳定；1959年经批准定名为伊金霍洛旗，沿用至今。"
        ),
    }
}


def qn(tag: str) -> str:
    prefix, local = tag.split(":")
    return f"{{{NS[prefix]}}}{local}"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\ufeff", "")
    text = re.sub(r"[\u200b\u200c\u200d\u2060]", "", text)
    return text.strip()


def format_elevation(value) -> str:
    text = clean(value)
    if not text:
        return ""
    normalized = re.sub(r"\s+", "", text).replace(",", "").replace("，", "")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", normalized):
        return text
    return f"{float(normalized):.2f}"


def format_table_value(header: str, value) -> str:
    if "高程" in clean(header):
        return format_elevation(value)
    return clean(value)


def report_error_marker(kind: str, reason: str) -> str:
    reason = clean(reason) or "原因待复核"
    return f"【{kind}：{reason}】"


def report_field_value(key: str, value: str) -> str:
    text = clean(value)
    if text.startswith("【") and any(marker in text for marker in (MATCH_ERROR_TEXT, IMAGE_INSERT_ERROR_TEXT, MISSING_FIELD_TEXT)):
        return text
    if text == MISSING_FIELD_TEXT:
        return report_error_marker(MISSING_FIELD_TEXT, f"{key}缺少信息")
    if text == MATCH_ERROR_TEXT:
        return report_error_marker(MATCH_ERROR_TEXT, f"{key}未能可靠匹配")
    if text == IMAGE_INSERT_ERROR_TEXT:
        return report_error_marker(IMAGE_INSERT_ERROR_TEXT, f"{key}未能可靠插入")
    return text


def infer_image_missing_reason(key: str) -> str:
    if any(word in key for word in ("照", "照片")):
        return "未在外业成果中匹配到对应照片"
    if any(word in key for word in ("图", "示意图", "剖面", "剖线")):
        return "未在内业成果中匹配到对应图件"
    return "未在执行资料中匹配到对应图片"


def image_insert_error_marker(key: str, reason: str | None = None) -> str:
    detail = f"{reason or infer_image_missing_reason(key)}：{key}"
    return report_error_marker(IMAGE_INSERT_ERROR_TEXT, detail)


def table_insert_error_marker(table_name: str, reason: str | None = None) -> str:
    detail = f"{reason or '未在内业成果表格中匹配到对应数据'}：{table_name}"
    return report_error_marker(IMAGE_INSERT_ERROR_TEXT, detail)


def safe_filename(name: str) -> str:
    return re.sub(r'[/:*?"<>|\\]+', "_", clean(name)) or "未命名项目"


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def ensure_writable_target(path: Path, purpose: str) -> None:
    if is_relative_to(path, TEMPLATE_LIBRARY_DIR):
        raise ValueError(f"{purpose}不能位于模板库 `基础信息/` 内。")


def parse_kv_sheet(ws) -> dict[str, str]:
    data: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        key = clean(row[0])
        if key:
            value = clean(row[1])
            if len(row) > 3:
                for extra in row[3:]:
                    extra_value = clean(extra)
                    if extra_value:
                        value = extra_value
            data[key] = value
    return data


def parse_table_sheet(ws) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {
            headers[idx]: format_table_value(headers[idx], row[idx] if idx < len(row) else "")
            for idx in range(len(headers))
            if headers[idx]
        }
        if any(record.values()):
            records.append(record)
    return records


def normalize_personnel_set(value: str | None) -> str:
    text = clean(value)
    return PERSONNEL_SET_ALIASES.get(text, text) or "人员信息1"


def split_person_names(value: str) -> list[str]:
    text = clean(value)
    if not text:
        return []
    text = re.sub(r"等\s*\d*\s*人?", "", text)
    return [item.strip() for item in re.split(r"[、,，/；;\\s]+", text) if item.strip()]


def company_package_dir(company: str) -> Path | None:
    company = clean(company)
    if not company:
        return None
    direct = COMPANY_PERSONNEL_PACKAGE_DIR / company
    if direct.exists():
        return direct
    if not COMPANY_PERSONNEL_PACKAGE_DIR.exists():
        return None
    for path in COMPANY_PERSONNEL_PACKAGE_DIR.iterdir():
        if path.is_dir() and (company in path.name or path.name in company):
            return path
    return None


def company_workbook_path(company_dir: Path) -> Path | None:
    candidates = sorted(company_dir.glob("*_资料卡.xlsx"))
    return candidates[0] if candidates else None


def read_personnel_sheet(wb, personnel_set: str) -> dict[str, str]:
    sheet = normalize_personnel_set(personnel_set)
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    values: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        key = clean(row[0])
        value = clean(row[1])
        if key and value:
            values[key] = value
            for alias in PERSONNEL_TEXT_ROLE_ALIASES.get(key, []):
                values.setdefault(alias, value)
    return values


def personnel_attachment_order(personnel_values: dict[str, str]) -> list[str]:
    names: list[str] = []
    for role in PERSONNEL_ATTACHMENT_ROLE_ORDER:
        for name in split_person_names(personnel_values.get(role, "")):
            if name not in names:
                names.append(name)
    if names:
        return names
    for role in ["项目负责", "技师", "报告执笔", "测绘员", "资料员", "现场负责"]:
        for name in split_person_names(personnel_values.get(role, "")):
            if name not in names:
                names.append(name)
    return names


def attachment_sort_key(record: dict[str, str]) -> tuple[int, int, str]:
    category = clean(record.get("资料类别"))
    name = clean(record.get("资料名称")) + clean(record.get("文件名"))
    if category == "身份证":
        if "正面" in name:
            return (0, 0, name)
        if "背面" in name:
            return (0, 1, name)
    match = re.search(r"第\s*(\d+)\s*页", name)
    page = int(match.group(1)) if match else 0
    category_order = {"身份证": 0, "证书": 1, "劳动合同": 2, "社保": 3}
    return (category_order.get(category, 9), page, name)


def read_personnel_attachment_records(wb, company_dir: Path, personnel_set: str) -> dict[tuple[str, str], list[Path]]:
    if "人员附件清单" not in wb.sheetnames:
        return {}
    ws = wb["人员附件清单"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [clean(value) for value in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {
            headers[idx]: clean(row[idx] if idx < len(row) else "")
            for idx in range(len(headers))
            if headers[idx]
        }
        if not any(record.values()):
            continue
        if normalize_personnel_set(record.get("人员套组")) != normalize_personnel_set(personnel_set):
            continue
        path = resolve_personnel_attachment_path(company_dir, record)
        if not (path.exists() and path.suffix.lower() in IMAGE_EXTENSIONS):
            continue
        record["_path"] = str(path)
        records.append(record)
    out: dict[tuple[str, str], list[Path]] = defaultdict(list)
    for record in sorted(records, key=attachment_sort_key):
        out[(record.get("姓名", ""), record.get("资料类别", ""))].append(Path(record["_path"]))
    return out


def resolve_personnel_attachment_path(company_dir: Path, record: dict[str, str]) -> Path:
    relative_path = record.get("相对路径", "")
    candidates: list[Path] = []
    if relative_path:
        candidates.append(company_dir / relative_path)
    personnel_library = company_dir / "人员信息库"
    person_name = clean(record.get("姓名"))
    category = clean(record.get("资料类别"))
    filename = clean(record.get("文件名"))
    if person_name and category and filename:
        candidates.append(personnel_library / person_name / category / filename)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0] if candidates else company_dir / relative_path


def find_company_attachment(company_dir: Path, key: str) -> Path | None:
    wanted_keys = [normalize_key(key)]
    if normalize_key(key) in {"公司印章", "印章", "公章"}:
        wanted_keys = ["公司印章", "公章", "印章"]
    for info_dir_name in ["公司信息库", "公司信息"]:
        info_dir = company_dir / info_dir_name
        if not info_dir.exists():
            continue
        for path in sorted(info_dir.iterdir()):
            if path.is_file() and path.suffix.lower() in COMPANY_ATTACHMENT_EXTENSIONS:
                stem_key = normalize_key(path.stem)
                if any(wanted in stem_key for wanted in wanted_keys):
                    return path
    return None


def load_company_personnel_context(
    fields: dict[str, str],
    personnel_set: str | None = None,
) -> tuple[dict[str, object], list[str]]:
    company = clean(fields.get("勘探单位") or fields.get("公司全称"))
    notes: list[str] = []
    company_dir = company_package_dir(company)
    if company_dir is None:
        return {}, [f"未找到公司资料包：{company or '未填写勘探单位'}"]
    workbook = company_workbook_path(company_dir)
    if workbook is None:
        return {}, [f"未找到公司资料卡：{company_dir}"]
    wb = load_workbook(workbook, data_only=True)
    personnel_set = normalize_personnel_set(personnel_set or fields.get("人员套组") or "人员信息1")
    personnel_values = read_personnel_sheet(wb, personnel_set)
    if not personnel_values:
        notes.append(f"未找到人员套组工作表：{personnel_set}")
    fields["公司全称"] = company_dir.name
    fields["勘探单位"] = company_dir.name
    for key, value in personnel_values.items():
        if key == "套组名称":
            continue
        fields.setdefault(key, value)
        if value:
            fields[key] = value
    order = personnel_attachment_order(personnel_values)
    attachments = read_personnel_attachment_records(wb, company_dir, personnel_set)
    notes.append(f"公司人员资料库：{company_dir.name} / {personnel_set}；附件人员顺序：{'、'.join(order) if order else '未识别'}")
    return {
        "company_dir": company_dir,
        "personnel_set": personnel_set,
        "personnel_values": personnel_values,
        "attachment_order": order,
        "attachments": attachments,
    }, notes


def resolve_image_path(value: str, *roots: Path | None) -> Path | None:
    text = clean(value)
    if not text:
        return None
    path = Path(text).expanduser()
    candidates = [path] if path.is_absolute() else []
    for root in roots:
        if root is not None:
            candidates.append(root / path)
    for candidate in candidates:
        if candidate.exists() and candidate.suffix.lower() in IMAGE_EXTENSIONS:
            return candidate
    name_key = normalize_key(Path(text).name)
    stem_key = normalize_key(Path(text).stem)
    loose_stem_key = normalize_image_stem(Path(text).stem)
    for root in roots:
        if root is None or not root.exists():
            continue
        for candidate in root.rglob("*"):
            if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            if normalize_key(candidate.name) == name_key or normalize_key(candidate.stem) == stem_key:
                return candidate
        if loose_stem_key:
            for candidate in root.rglob("*"):
                if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
                    continue
                if normalize_image_stem(candidate.stem) == loose_stem_key:
                    return candidate
    return None


def normalize_image_stem(value: str) -> str:
    text = normalize_key(value)
    text = re.sub(r"[\(（]\s*\d+\s*[\)）]$", "", text)
    return text


def find_image_by_keywords(root: Path, keywords: list[str]) -> Path | None:
    if not root.exists():
        return None
    normalized = [(keyword, normalize_key(keyword)) for keyword in keywords]
    best: tuple[int, Path] | None = None
    for candidate in root.rglob("*"):
        if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        rel = normalize_key(str(candidate.relative_to(root)))
        score = sum(1 for _, keyword in normalized if keyword and keyword in rel)
        if score and (best is None or score > best[0] or (score == best[0] and str(candidate) < str(best[1]))):
            best = (score, candidate)
    return best[1] if best else None


def find_image_by_exact_stems(root: Path, stems: list[str]) -> Path | None:
    if not root.exists():
        return None
    wanted = {normalize_key(stem) for stem in stems if stem}
    for candidate in sorted(root.rglob("*"), key=lambda path: str(path)):
        if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        if normalize_key(candidate.stem) in wanted:
            return candidate
    return None


def sorted_images_under(root: Path, keywords: list[str] | None = None) -> list[Path]:
    if not root.exists():
        return []
    images = [
        path
        for path in root.rglob("*")
        if path.suffix.lower() in IMAGE_EXTENSIONS and "_normalized" not in path.parts
    ]
    if keywords:
        keys = [normalize_key(item) for item in keywords]
        images = [path for path in images if any(key in normalize_key(str(path.relative_to(root))) for key in keys)]

    def sort_key(path: Path) -> tuple[int, str]:
        match = re.search(r"(\d+)", path.stem)
        return (int(match.group(1)) if match else 9999, str(path))

    return sorted(images, key=sort_key)


def data_organization_work_photos() -> list[Path]:
    return sorted_images_under(DATA_ORGANIZATION_WORK_PHOTO_LIBRARY)


def first_data_organization_work_photo() -> Path | None:
    images = sorted_images_under(DATA_ORGANIZATION_WORK_PHOTO_LIBRARY)
    return images[0] if images else None


def select_rotating_data_organization_work_photo() -> tuple[Path | None, str]:
    images = data_organization_work_photos()
    if not images:
        return None, ""
    last_file = ""
    if DATA_ORGANIZATION_WORK_PHOTO_ROTATION_STATE.exists():
        try:
            state = json.loads(DATA_ORGANIZATION_WORK_PHOTO_ROTATION_STATE.read_text(encoding="utf-8"))
            last_file = clean(state.get("last_file"))
        except (OSError, json.JSONDecodeError):
            last_file = ""
    names = [path.name for path in images]
    if last_file in names:
        index = (names.index(last_file) + 1) % len(images)
    else:
        index = 0
    selected = images[index]
    DATA_ORGANIZATION_WORK_PHOTO_ROTATION_STATE.write_text(
        json.dumps(
            {
                "last_file": selected.name,
                "next_file": images[(index + 1) % len(images)].name,
                "library": str(DATA_ORGANIZATION_WORK_PHOTO_LIBRARY),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return selected, f"{index + 1}/{len(images)}"


def partition_order_value(label: str) -> int:
    match = re.search(r"勘探区域\s*([A-ZＡ-Ｚ一二三四五六七八九十]+)区?", label)
    if not match:
        return 9999
    token = match.group(1).upper()
    token = token.translate(str.maketrans("ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ", "ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    if len(token) == 1 and "A" <= token <= "Z":
        return ord(token) - ord("A") + 1
    return CHINESE_NUMERAL.get(token, 9999)


def sorted_partitioned_images_under(root: Path, keywords: list[str] | None = None) -> list[Path]:
    images = sorted_images_under(root, keywords)

    def sort_key(path: Path) -> tuple[int, int, str]:
        label = partition_label_for_path(path, root)
        number_match = re.search(r"(\d+)", path.stem)
        return (
            partition_order_value(label),
            int(number_match.group(1)) if number_match else 9999,
            str(path),
        )

    return sorted(images, key=sort_key)


def relic_sort_key(value: str) -> tuple[str, int, str]:
    text = clean(value).upper().replace("墓葬", "M").replace("灰坑", "H").replace("房址", "F").replace("窑址", "Y").replace("窑坑", "Y").replace("活土坑", "K")
    match = re.search(r"([A-Z]+)\s*0*(\d+)", text)
    if match:
        prefix = match.group(1)
        prefix_order = {"H": "01H", "M": "02M", "Y": "03Y", "F": "04F", "K": "05K", "G": "06G"}.get(prefix, prefix)
        return (prefix_order, int(match.group(2)), text)
    return (text, 999999, text)


def canonical_relic_id(value: str) -> str:
    text = clean(value).upper().replace("墓葬", "M").replace("灰坑", "H").replace("房址", "F").replace("窑址", "Y").replace("窑坑", "Y").replace("活土坑", "K")
    match = re.search(r"([A-Z]+)\s*0*(\d+)", text)
    return f"{match.group(1)}{int(match.group(2))}" if match else clean(value)


def relic_type_from_id(relic_id: str) -> str:
    key = canonical_relic_id(relic_id).upper()
    if key.startswith("H"):
        return "灰坑"
    if key.startswith("M"):
        return "墓葬"
    if key.startswith("F"):
        return "房址"
    if key.startswith("Y"):
        return "窑址"
    if key.startswith("K"):
        return "活土坑"
    if key.startswith("G"):
        return "沟"
    return "遗迹"


def strip_relic_group_prefix(value: str) -> str:
    return re.sub(r"^(?:灰坑|墓葬|遗迹)", "", clean(value), flags=re.I)


def image_matches_relic(path: Path, relic_id: str) -> bool:
    stem = canonical_relic_id(path.stem)
    wanted = canonical_relic_id(relic_id)
    if stem == wanted:
        return True
    return False


def first_relic_image(
    root: Path | None,
    relic_id: str,
) -> Path | None:
    if root is None or not root.exists():
        return None
    images = sorted_images_under(root)
    for path in images:
        if image_matches_relic(path, relic_id):
            return path
    return None


def first_direct_relic_image(root: Path | None, relic_id: str) -> Path | None:
    if root is None or not root.exists():
        return None
    images = sorted(
        path
        for path in root.iterdir()
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS and "_normalized" not in path.parts
    )
    for path in images:
        if image_matches_relic(path, relic_id):
            return path
    return None


def first_relic_image_in_roots(
    roots: list[Path],
    subfolder: str,
    relic_id: str,
    *,
    allow_direct_root: bool = False,
) -> Path | None:
    for root in roots:
        image = first_relic_image(root / subfolder, relic_id)
        if image:
            return image
        for folder in root.iterdir():
            if not folder.is_dir():
                continue
            image = first_relic_image(folder / subfolder, relic_id)
            if image:
                return image
    if allow_direct_root:
        for root in roots:
            image = first_direct_relic_image(root, relic_id)
            if image:
                return image
    return None


def sorted_project_location_maps(drawings_dir: Path) -> list[Path]:
    location_dir = drawings_dir / "1.位置图"
    search_root = location_dir if location_dir.exists() else drawings_dir
    images = [
        path
        for path in search_root.rglob("*")
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS and is_project_location_map(path)
    ]

    return sorted(images, key=lambda path: (project_location_admin_level(path), normalize_key(path.stem)))


def unit_layout_map_sort_key(path: Path) -> tuple[int, int, str]:
    stem = normalize_key(path.stem).upper()
    if "总" in stem or not re.search(r"(勘探区域|分区|[A-ZＡ-Ｚ]区|T\d+)", stem):
        return (0, 0, str(path))
    letter_match = re.search(r"(?:勘探区域|分区)?\s*([A-ZＡ-Ｚ])\s*区?", stem)
    if letter_match:
        token = letter_match.group(1).translate(str.maketrans("ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ", "ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        return (1, ord(token) - ord("A") + 1, str(path))
    t_match = re.search(r"T\s*0*(\d+)", stem)
    if t_match:
        return (2, int(t_match.group(1)), str(path))
    number_match = re.search(r"(?:分区|勘探区域)\s*0*(\d+)", stem)
    if number_match:
        return (3, int(number_match.group(1)), str(path))
    return (9, 9999, str(path))


def sorted_unit_layout_maps(root: Path) -> list[Path]:
    return sorted(sorted_images_under(root), key=unit_layout_map_sort_key)


def project_location_admin_level(path: Path) -> int:
    stem = path.stem
    if any(word in stem for word in ["自治区", "省", "直辖市"]):
        return 1
    if any(word in stem for word in ["盟", "市", "地区", "州"]):
        return 2
    if any(word in stem for word in ["旗", "县", "区"]):
        return 3
    if any(word in stem for word in ["镇", "乡", "村", "苏木", "街道"]):
        return 4
    return 9


def sorted_relic_distribution_maps(relic_distribution_dir: Path) -> list[Path]:
    images = sorted_images_under(relic_distribution_dir)

    def sort_key(path: Path) -> tuple[int, int, str]:
        stem = path.stem
        local_match = re.search(r"局部\s*([0-9一二三四五六七八九十]+)", stem)
        if not local_match:
            return (0, 0, normalize_key(stem))
        token = local_match.group(1)
        number = int(token) if token.isdigit() else CHINESE_NUMERAL.get(token, 999)
        return (1, number, normalize_key(stem))

    return sorted(images, key=sort_key)


def first_existing_dir(*candidates: Path | None) -> Path | None:
    for candidate in candidates:
        if candidate and candidate.exists() and candidate.is_dir():
            return candidate
    return None


def find_workbook_by_names(root: Path, names: list[str]) -> Path | None:
    if not root.exists():
        return None
    wanted = {normalize_key(name) for name in names}
    wanted_stems = {normalize_key(Path(name).stem) for name in names}
    loose_matches: list[Path] = []
    for candidate in root.rglob("*.xlsx"):
        if candidate.name.startswith(".~"):
            continue
        candidate_name = normalize_key(candidate.name)
        candidate_stem = normalize_key(candidate.stem)
        if candidate_name in wanted or candidate_stem in wanted:
            return candidate
        if any(stem and stem in candidate_stem for stem in wanted_stems):
            loose_matches.append(candidate)
    if loose_matches:
        return sorted(loose_matches, key=lambda path: str(path))[0]
    return None


def default_caption_from_path(path: Path) -> str:
    caption = re.sub(r"_norm$", "", path.stem)
    return caption


def standard_probe_number(record: dict[str, str]) -> str:
    existing = clean(record.get("勘探编号"))
    if existing:
        return existing
    unit = clean(record.get("勘探单元") or record.get("勘探单元显示"))
    standard_code = clean(record.get("标准孔编号"))
    if unit and standard_code:
        return f"{unit}-{standard_code}"
    return standard_code


def normalize_tk_code(value: object) -> str:
    text = clean(value).upper()
    match = re.search(r"TK\s*0*(\d+)", text, flags=re.I)
    if match:
        return f"TK{int(match.group(1))}"
    if re.fullmatch(r"\d+", text):
        return f"TK{int(text)}"
    return text


def tk_code_aliases(value: object) -> list[str]:
    normalized = normalize_tk_code(value)
    number = first_number(normalized)
    if not number:
        return [normalized] if normalized else []
    aliases = [f"TK{number}", f"TK{number:02d}", f"TK{number:03d}"]
    out: list[str] = []
    for alias in aliases:
        if alias not in out:
            out.append(alias)
    return out


def standard_hole_code(record: dict[str, str]) -> str:
    existing = normalize_tk_code(record.get("探孔编号"))
    if existing:
        return existing
    photo = clean(record.get("标准孔土样照"))
    if photo:
        stem = Path(photo).stem.upper()
        normalized = normalize_tk_code(stem)
        if normalized.startswith("TK"):
            return normalized
    return ""


def partition_label_for_path(path: Path, photos_dir: Path | None) -> str:
    if photos_dir is None:
        return ""
    try:
        parts = path.resolve().relative_to(photos_dir.resolve()).parts
    except ValueError:
        return ""
    for part in parts:
        match = re.search(r"勘探区域\s*([A-ZＡ-Ｚ一二三四五六七八九十]+)区?", part)
        if match:
            token = match.group(1).upper().translate(
                str.maketrans("ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ", "ABCDEFGHIJKLMNOPQRSTUVWXYZ")
            )
            return f"勘探区域{token}区"
    return ""


def has_partitioned_photo_dirs(photos_dir: Path | None) -> bool:
    if photos_dir is None or not photos_dir.exists():
        return False
    return any(partition_label_for_path(path, photos_dir) for path in photos_dir.iterdir() if path.is_dir())


def append_partition_to_caption(caption: str, path: Path, photos_dir: Path | None, *, enabled: bool) -> str:
    text = clean(caption)
    if not enabled:
        return text
    label = partition_label_for_path(path, photos_dir)
    if not label or label in text:
        return text
    return f"{text}（{label}）" if text else f"（{label}）"


def file_content_signature(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def unique_photo_paths_for_caption(
    paths: list[Path],
    *,
    partitioned: bool,
    photos_dir: Path | None = None,
) -> list[tuple[Path, bool]]:
    unique: list[tuple[Path, str]] = []
    seen: set[str] = set()
    for path in paths:
        try:
            signature = file_content_signature(path)
        except OSError:
            signature = f"path:{path.resolve()}"
        if signature in seen:
            continue
        seen.add(signature)
        unique.append((path, signature))
    kept_partitions = {partition_label_for_path(path, photos_dir) for path, _ in unique}
    add_partition_suffix = partitioned and len({label for label in kept_partitions if label}) > 1
    return [(path, add_partition_suffix) for path, _ in unique]


def is_project_location_map(path: Path) -> bool:
    stem = default_caption_from_path(path)
    return bool(re.fullmatch(r"项目地块在.+位置示意图", stem))


def infer_image_source(path: Path, drawings_dir: Path | None, photos_dir: Path | None) -> str:
    if drawings_dir and is_relative_to(path, drawings_dir):
        return "drawing"
    if photos_dir and is_relative_to(path, photos_dir):
        return "photo"
    return "form"


def add_image_slot(
    slots: dict[str, deque[ImageSlot]],
    key: str,
    path: Path | None,
    caption: str = "",
    *,
    source: str = "",
    drawings_dir: Path | None = None,
    photos_dir: Path | None = None,
) -> None:
    if path is not None:
        resolved = path.resolve()
        slots[key].append(ImageSlot(resolved, caption, source or infer_image_source(resolved, drawings_dir, photos_dir), key))


def relic_image_slot_key(base_key: str, relic_id: str) -> str:
    return f"{base_key}:{canonical_relic_id(relic_id)}"


def build_image_slots(
    tables: dict[str, list[dict[str, str]]],
    form: Path,
    drawings_dir: Path | None,
    photos_dir: Path | None = None,
) -> dict[str, deque[ImageSlot]]:
    slots: dict[str, deque[ImageSlot]] = defaultdict(deque)
    image_root = form.parent / "图片"
    search_roots = [image_root, drawings_dir, photos_dir]
    standard_position_dir = drawings_dir / "图纸" / "单个标准孔" if drawings_dir else None
    standard_photo_dir = photos_dir / "12.标准孔照" if photos_dir else None
    if photos_dir and (photos_dir / "11.标准孔照").exists():
        standard_photo_dir = photos_dir / "11.标准孔照"

    for row in tables.get("图片清单", []):
        key = clean(row.get("图片位") or row.get("插入位置"))
        filename = clean(row.get("文件名") or row.get("图片文件名"))
        caption = clean(row.get("图题"))
        add_image_slot(
            slots,
            key,
            resolve_image_path(filename, *search_roots),
            caption,
            drawings_dir=drawings_dir,
            photos_dir=photos_dir,
        )

    for row in tables.get("剖线地层堆积", []):
        add_image_slot(
            slots,
            "剖线图",
            resolve_image_path(row.get("剖线图", ""), *search_roots),
            clean(row.get("图题")),
            drawings_dir=drawings_dir,
            photos_dir=photos_dir,
        )
    for row in tables.get("标准孔", []):
        probe_number = standard_probe_number(row)
        hole_code = standard_hole_code(row)
        position_roots = [standard_position_dir, *search_roots]
        photo_roots = [standard_photo_dir, *search_roots]
        position_path = (
            resolve_image_path(f"{probe_number}.jpg", *position_roots)
            or resolve_image_path(row.get("标准孔位置图", ""), *position_roots)
            or resolve_image_path(f"{clean(row.get('序号'))}.jpg", *position_roots)
        )
        caption = clean(row.get("图题") or probe_number or row.get("标准孔编号"))
        if probe_number:
            add_image_slot(
                slots,
                f"标准孔位置图:{probe_number}",
                position_path,
                caption,
                drawings_dir=drawings_dir,
                photos_dir=photos_dir,
            )
        soil_path = resolve_image_path(row.get("标准孔土样照", ""), *photo_roots)
        if soil_path is None:
            for alias in tk_code_aliases(hole_code):
                soil_path = resolve_image_path(f"{alias}.jpg", *photo_roots)
                if soil_path is not None:
                    break
        if probe_number:
            add_image_slot(
                slots,
                f"标准孔土样照:{probe_number}",
                soil_path,
                clean(row.get("图题") or hole_code or row.get("标准孔编号")),
                drawings_dir=drawings_dir,
                photos_dir=photos_dir,
            )

    if drawings_dir and drawings_dir.exists():
        relic_distribution_dir = first_existing_dir(
            drawings_dir / "遗迹分布示意图",
            drawings_dir / "图纸" / "遗迹分布示意图",
        )
        if relic_distribution_dir:
            for path in sorted_relic_distribution_maps(relic_distribution_dir):
                add_image_slot(slots, "遗迹分布示意图", path, default_caption_from_path(path), source="drawing")
        unit_layout_dir = first_existing_dir(
            drawings_dir / "勘探单元布设示意图",
            drawings_dir / "图纸" / "勘探单元布设示意图",
        )
        if unit_layout_dir:
            for path in sorted_unit_layout_maps(unit_layout_dir):
                add_image_slot(slots, "勘探单元布设示意图", path, default_caption_from_path(path), source="drawing")
            if slots.get("勘探单元布设示意图"):
                slots["勘探单元布置示意图"].extend(copy.deepcopy(slots["勘探单元布设示意图"]))
                slots["划分勘探单元图"].extend(copy.deepcopy(slots["勘探单元布设示意图"]))
        project_location_maps = sorted_project_location_maps(drawings_dir)
        for path in sorted_project_location_maps(drawings_dir):
            add_image_slot(slots, "项目地块位置示意图", path, default_caption_from_path(path), source="drawing")
            level_alias = {
                1: "项目地块在省级行政区位置示意图",
                2: "项目地块在地级市位置示意图",
                3: "项目地块在旗县区位置示意图",
                4: "项目地块在镇乡位置示意图",
            }.get(project_location_admin_level(path))
            if level_alias:
                add_image_slot(slots, level_alias, path, default_caption_from_path(path), source="drawing")
        auto_map = {
            "项目地块在内蒙古自治区位置示意图": ["位置图", "内蒙古自治区"],
            "项目地块在省级行政区位置示意图": ["位置图", "内蒙古自治区"],
            "项目地块在地级市位置示意图": ["位置图"],
            "项目地块在旗县区位置示意图": ["位置图"],
            "项目地块卫星图": ["位置图", "卫星图"],
            "项目红线四至坐标图": ["红线四至坐标"],
            "红线四至坐标图": ["红线四至坐标"],
            "红线图": ["项目红线"],
            "项目红线图": ["项目红线"],
            "项目勘探区域图": ["勘探区域"],
            "项目勘探区域示意图": ["勘探区域"],
            "项目勘探分区示意图": ["勘探分区"],
            "划分勘探单元图": ["勘探单元", "布置"],
            "勘探单元布置示意图": ["勘探单元", "布设"],
            "勘探单元布设示意图": ["勘探单元", "布设"],
            "勘探布孔示意图": ["探孔", "布"],
            "探孔布设示意图": ["探孔", "布"],
            "探孔布置示意图": ["探孔", "布置"],
            "剖面位置图": ["剖线", "位置图"],
            "剖线位置示意图": ["剖线", "位置"],
            "标准探孔位置示意图": ["探孔", "布"],
        }
        level_fallbacks = {
            "项目地块在省级行政区位置示意图": [path for path in project_location_maps if project_location_admin_level(path) == 1],
            "项目地块在地级市位置示意图": [path for path in project_location_maps if project_location_admin_level(path) == 2],
            "项目地块在旗县区位置示意图": [path for path in project_location_maps if project_location_admin_level(path) == 3],
            "项目地块在镇乡位置示意图": [path for path in project_location_maps if project_location_admin_level(path) == 4],
        }
        for key, keywords in auto_map.items():
            if not slots.get(key):
                exact = find_image_by_exact_stems(drawings_dir, [key])
                fallback = level_fallbacks.get(key, [None])[0] if level_fallbacks.get(key) else None
                add_image_slot(slots, key, exact or fallback or find_image_by_keywords(drawings_dir, keywords), source="drawing")
        alias_map = {
            "项目位置图": "项目地块位置示意图",
            "勘探分区示意图": "项目勘探分区示意图",
            "项目红线图": "红线图",
        }
        for alias, canonical in alias_map.items():
            if not slots.get(alias) and slots.get(canonical):
                slots[alias].extend(copy.deepcopy(slots[canonical]))
        if not slots.get("剖线图"):
            for path in sorted_images_under(drawings_dir, ["地层堆积剖线图", "地层堆积剖面图"]):
                caption = default_caption_from_path(path)
                add_image_slot(slots, "剖线图", path, caption, source="drawing")
                add_image_slot(slots, caption, path, caption, source="drawing")
                add_image_slot(slots, caption.replace("剖线图", "剖面图"), path, caption.replace("剖线图", "剖面图"), source="drawing")
                add_image_slot(slots, caption.replace("剖面图", "剖线图"), path, caption.replace("剖面图", "剖线图"), source="drawing")

    if photos_dir and photos_dir.exists():
        partitioned_photos = has_partitioned_photo_dirs(photos_dir)
        photo_map = {
            "实地踏查照": ["实地踏查照"],
            "走访调查照": ["走访调查照"],
            "项目地块现状照": ["项目地块现状照", "地块现状照"],
            "勘探单元布置照": ["勘探单元布置照"],
            "布设探孔照": ["布设探孔照"],
            "普探工作照": ["普探工作照"],
            "取样记录照": ["取样记录照"],
            "勘探后局部照": ["勘探后局部照"],
            "勘探后航拍照": ["勘探后航拍照", "航拍", "dji"],
            "信息采集照": ["信息采集照"],
            "资料整理工作照": ["资料整理工作照", "资料整理照"],
            "文献资料收集与整理工作照": ["文献资料收集与整理工作照", "文献资料收集"],
        }
        for key, keywords in photo_map.items():
            if not slots.get(key):
                paths = sorted_partitioned_images_under(photos_dir, keywords)
                for path, add_suffix in unique_photo_paths_for_caption(
                    paths,
                    partitioned=partitioned_photos,
                    photos_dir=photos_dir,
                ):
                    caption = append_partition_to_caption(key, path, photos_dir, enabled=add_suffix)
                    add_image_slot(slots, key, path, caption, source="photo")
        relic_roots = sorted((p for p in photos_dir.rglob("13.遗迹") if p.is_dir()), key=lambda p: str(p))
        if not relic_roots:
            relic_image = find_image_by_keywords(photos_dir, ["遗迹"])
            relic_roots = [relic_image.parent] if relic_image else []
        single_relic_dir = first_existing_dir(
            drawings_dir / "单个遗迹" if drawings_dir else None,
            drawings_dir / "图纸" / "单个遗迹" if drawings_dir else None,
            drawings_dir / "单个遗址" if drawings_dir else None,
            drawings_dir / "图纸" / "单个遗址" if drawings_dir else None,
        )
        for record in tables.get("遗迹记录") or tables.get("遗迹", []):
            relic_id = record.get("遗迹编号", "")
            if not relic_id:
                continue
            normalized_relic_id = canonical_relic_id(relic_id)
            soil = first_relic_image_in_roots(
                relic_roots,
                "土样照",
                normalized_relic_id,
                allow_direct_root=True,
            )
            site = first_relic_image_in_roots(
                relic_roots,
                "现场照",
                normalized_relic_id,
                allow_direct_root=False,
            )
            plan = first_relic_image(single_relic_dir, relic_id)
            add_image_slot(
                slots,
                relic_image_slot_key("遗迹土样照", normalized_relic_id),
                soil,
                f"{normalized_relic_id}土样照",
                source="photo",
            )
            add_image_slot(
                slots,
                relic_image_slot_key("遗迹现场照", normalized_relic_id),
                site,
                f"{normalized_relic_id}现场照",
                source="photo",
            )
            add_image_slot(
                slots,
                relic_image_slot_key("遗迹平、剖面图", normalized_relic_id),
                plan,
                f"{normalized_relic_id}平、剖面图",
                source="drawing",
            )

    return slots


def add_data_organization_work_photo_library_slot(slots: dict[str, deque[ImageSlot]]) -> list[str]:
    key = "资料整理工作照"
    slots.pop(key, None)
    library_photo, rotation_position = select_rotating_data_organization_work_photo()
    if library_photo is None:
        return ["资料整理工作照照片库为空：未插入资料整理工作照"]
    add_image_slot(slots, key, library_photo, key, source="photo")
    return [f"资料整理工作照照片库轮换：{rotation_position}，本次使用 {library_photo}"]


def split_people_count(text: str) -> int:
    text = clean(text)
    if not text:
        return 0
    number = re.search(r"(\d+)\s*人", text)
    if number:
        return int(number.group(1))
    parts = [item for item in re.split(r"[、,，;\s]+", text) if item]
    return len(parts)


def first_number(text: str) -> int | None:
    match = re.search(r"(\d+)", clean(text))
    return int(match.group(1)) if match else None


def natural_unit_sort_key(unit: str) -> tuple[str, int, str]:
    text = clean(unit)
    match = re.fullmatch(r"([A-Za-z\u4e00-\u9fa5]*)(\d+)", text)
    if match:
        return (match.group(1), int(match.group(2)), text)
    return (text, 0, text)


def abbreviate_unit_numbers(units: list[str]) -> str:
    normalized = sorted({clean(unit) for unit in units if clean(unit)}, key=natural_unit_sort_key)
    if not normalized:
        return ""

    match_data = [re.fullmatch(r"([A-Za-z\u4e00-\u9fa5]*)(\d+)", unit) for unit in normalized]
    if all(match_data):
        prefixes = {match.group(1) for match in match_data if match}
        numbers = [int(match.group(2)) for match in match_data if match]
        if (
            len(prefixes) == 1
            and numbers == list(range(numbers[0], numbers[-1] + 1))
            and len(normalized) >= 4
        ):
            return f"{normalized[0]}、{normalized[1]}……{normalized[-1]}"

    return "、".join(normalized)


def report_month_from_end_date(end_date: str) -> str:
    text = clean(end_date)
    match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", text)
    if match:
        return f"{match.group(1)}年{int(match.group(2))}月"
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.]\d{1,2}", text)
    if match:
        return f"{match.group(1)}年{int(match.group(2))}月"
    return ""


def infer_county(fields: dict[str, str]) -> str:
    explicit = clean(fields.get("项目所在地旗县"))
    text = explicit or clean(fields.get("项目位置"))
    if not text:
        return ""
    token_pattern = r"([^\s省自治区市盟州地区,，、；;：:()（）]{2,12}%s)"
    for suffix in ["旗", "县", "区"]:
        matches = re.findall(token_pattern % suffix, text)
        matches = [item for item in matches if item not in {"自治区", "开发区", "工业区", "园区"}]
        if matches:
            return matches[-1]
    city_matches = re.findall(token_pattern % "市", text)
    return city_matches[-1] if city_matches else explicit


def calculate_probe_max(unit_spec: str, spacing: str, multiplier: int) -> str:
    unit = first_number(unit_spec)
    interval = first_number(spacing)
    if not unit or not interval:
        return ""
    # Existing reports number positions as odd tails: 400m/2m -> 399, 400m/1m -> 799.
    return str(unit // interval * multiplier - 1)


def normalize_area_value(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    match = re.search(r"([0-9]+(?:[,.，][0-9]+)?)", text)
    if not match:
        return text
    number = match.group(1).replace(",", "").replace("，", "")
    if any(unit in text for unit in ["㎡", "平方米", "m2", "m²"]):
        return f"{number}平方米"
    return text


def normalize_review_title(value: str) -> str:
    title = clean(value)
    if not title or "：" in title and title.startswith("【"):
        return title
    title = title.strip("《》")
    return f"《{title}》"


def normalize_review_docno(value: str) -> str:
    docno = clean(value).strip("，,；;。 ")
    return "" if docno in NO_DOCUMENT_NUMBER_VALUES else docno


def normalize_cultural_relic_review_fields(fields: dict[str, str], notes: list[str]) -> None:
    if fields.get("文物概况类型") != "有文物审查意见":
        return
    title = fields.get("文物审查意见文件名", "")
    normalized_title = normalize_review_title(title)
    if normalized_title and normalized_title != title:
        fields["文物审查意见文件名"] = normalized_title
        notes.append("文物审查意见文件名=已按书名号格式规范")
    docno = fields.get("文物审查意见文号", "")
    normalized_docno = normalize_review_docno(docno)
    if normalized_docno != docno:
        fields["文物审查意见文号"] = normalized_docno
        notes.append("文物审查意见文号=无文号时不写入报告正文")


def derive_fields(fields: dict[str, str], tables: dict[str, list[dict[str, str]]]) -> list[str]:
    notes: list[str] = []

    for area_key in ["项目面积", "调查面积", "勘探面积"]:
        normalized_area = normalize_area_value(fields.get(area_key, ""))
        if normalized_area and normalized_area != fields.get(area_key):
            fields[area_key] = normalized_area
            notes.append(f"{area_key}={normalized_area}（面积单位规范化）")

    if not fields.get("调查面积") and fields.get("项目面积"):
        fields["调查面积"] = fields["项目面积"]
        notes.append(f"调查面积={fields['调查面积']}（默认等于项目面积）")

    county = infer_county(fields)
    if county and not fields.get("项目所在地旗县"):
        fields["项目所在地旗县"] = county
        notes.append(f"项目所在地旗县={county}（由项目位置识别）")
    preset = REGION_OVERVIEW_PRESETS.get(county)
    if preset:
        filled = []
        for key, value in preset.items():
            if not fields.get(key):
                fields[key] = value
                filled.append(key)
        if filled:
            notes.append(f"已自动补入区域概况：{county}（{len(filled)}项）")

    if fields.get("文物概况类型") == "有文物审查意见":
        review_path = Path(fields.get("回函文件路径", "")).expanduser() if fields.get("回函文件路径") else None
        needs_review_text = not fields.get("文物审查意见结论")
        if manual_form_builder and review_path and review_path.exists() and needs_review_text:
            try:
                review_text = ""
                if review_path.suffix.lower() == ".pdf":
                    review_text = manual_form_builder.read_pdf_text(review_path)
                elif review_path.suffix.lower() == ".docx":
                    review_text = "\n".join(manual_form_builder.read_docx_texts(review_path))
                if not fields.get("文物审查意见文号"):
                    docno = manual_form_builder.extract_document_number(review_path.name) or manual_form_builder.extract_document_number(review_text)
                    if docno:
                        fields["文物审查意见文号"] = docno
                        notes.append("文物审查意见文号=由回函文件路径自动识别")
                if not fields.get("文物审查意见结论") and review_text:
                    conclusion = manual_form_builder.extract_review_conclusion(review_text)
                    if conclusion:
                        fields["文物审查意见结论"] = conclusion
                        notes.append("文物审查意见结论=由回函正文自动识别")
            except Exception as exc:
                notes.append(f"文物审查意见回函自动读取失败：{exc}")
        if not fields.get("文物审查意见结论") and fields.get("涉及文物名称及情况"):
            fields["文物审查意见结论"] = f"该项目选址范围涉及{fields['涉及文物名称及情况']}"
            notes.append("文物审查意见结论=由涉及文物名称及情况自动生成")
        normalize_cultural_relic_review_fields(fields, notes)
        missing_relic_doc_fields = [
            key
            for key in ["文物审查意见文件名", "文物审查意见结论"]
            if not fields.get(key)
        ]
        if missing_relic_doc_fields:
            fields["文物概况类型"] = "文物审查意见待补"
            fields["文物概况待补说明"] = report_error_marker(
                MATCH_ERROR_TEXT,
                "文物审查意见信息不完整，需人工复核回函",
            )
            for key in missing_relic_doc_fields:
                fields[key] = report_error_marker(
                    MATCH_ERROR_TEXT,
                    f"{key}缺少或OCR未能可靠识别",
                )
            notes.append("文物审查意见信息不完整，已在生成稿中标记为匹配错误。")

    unit_spec_raw = fields.get("勘探单元规格", "")
    if re.fullmatch(r"\d+", unit_spec_raw):
        fields["勘探单元规格"] = f"{unit_spec_raw}米×{unit_spec_raw}米"
        notes.append(f"勘探单元规格={fields['勘探单元规格']}（由纯数字规范化）")

    if not fields.get("报告年月"):
        report_month = report_month_from_end_date(fields.get("结束日期", "")) or report_month_from_end_date(fields.get("勘探时间", ""))
        if report_month:
            fields["报告年月"] = report_month
            notes.append(f"报告年月={report_month}（由结束日期推导）")

    unit_records = tables.get("勘探单元", [])
    if unit_records:
        units = sorted(
            {record.get("勘探单元", "") for record in unit_records if record.get("勘探单元")},
            key=natural_unit_sort_key,
        )
        if units:
            if not fields.get("勘探单元数量"):
                fields["勘探单元数量"] = str(len(units))
            fields["勘探单元编号范围"] = abbreviate_unit_numbers(units)
            notes.append(f"勘探单元数量={fields['勘探单元数量']}，编号={fields['勘探单元编号范围']}")

    if not fields.get("技师数量"):
        count = split_people_count(fields.get("技师", ""))
        if count:
            fields["技师数量"] = str(count)
            notes.append(f"技师数量={count}")

    entry_count = first_number(fields.get("进场人数", ""))
    worker_count = first_number(fields.get("探工数量", "")) or split_people_count(fields.get("探工人员描述", ""))
    if entry_count and entry_count >= 8:
        worker_count = entry_count - 8
    elif worker_count:
        entry_count = worker_count + 8
        notes.append(f"进场人数={entry_count}（由探工数量{worker_count}+8回推）")
    else:
        notes.append("进场人数缺失：请在人工填写表中填写进场人数或探工数量。")
    if worker_count:
        fields["探工数量"] = str(worker_count)
        if not fields.get("踏查探工数量"):
            fields["踏查探工数量"] = str(worker_count)
    if entry_count:
        fields["进场人数"] = str(entry_count)
        fields["每日在场人员下限"] = str(max(entry_count - 5, 0))
    if worker_count or entry_count:
        notes.append(f"探工数量={fields.get('探工数量')}，进场人数={fields.get('进场人数')}，每日在场人员下限={fields.get('每日在场人员下限')}")

    if not fields.get("踏查技师数量") and fields.get("技师数量"):
        fields["踏查技师数量"] = fields["技师数量"]

    unit_spec = fields.get("勘探单元规格", "")
    normal_spacing = fields.get("普探孔距", "2米×2米")
    intensive_spacing = fields.get("重点勘探孔距", "1米×1米")
    for key, spacing, multiplier in [
        ("普探列最大编号", normal_spacing, 2),
        ("普探行最大编号", normal_spacing, 2),
        ("重点勘探列最大编号", intensive_spacing, 2),
        ("重点勘探行最大编号", intensive_spacing, 2),
    ]:
        if not fields.get(key):
            value = calculate_probe_max(unit_spec, spacing, multiplier)
            if value:
                fields[key] = value
    if unit_spec:
        notes.append(
            "探孔编号尾号：普探A/B"
            f"{fields.get('普探列最大编号', '')}，重点A/B{fields.get('重点勘探列最大编号', '')}"
        )

    section_records = tables.get("剖线地层堆积", [])
    standard_records = tables.get("标准孔", [])
    fields.setdefault("剖线数量", str(len(section_records)) if section_records else "")
    fields.setdefault("剖面数量", fields.get("剖线数量", ""))
    fields.setdefault("标准孔数量", str(len(standard_records)) if standard_records else "")
    if section_records:
        for idx, record in enumerate(section_records, start=1):
            record.setdefault("序号", str(idx))
            record.setdefault("X坐标", "")
            record.setdefault("Y坐标", "")
            record.setdefault("高程", "")
    if standard_records:
        for idx, record in enumerate(standard_records, start=1):
            record.setdefault("序号", str(idx))

    if not fields.get("勘探成果综合结论"):
        fields["勘探成果综合结论"] = build_conclusion(section_records, standard_records, fields)
        if fields["勘探成果综合结论"]:
            notes.append("已根据剖线地层和标准孔描述生成结论占位内容。")

    return notes


def mark_required_generation_errors(fields: dict[str, str], notes: list[str]) -> list[str]:
    marked: list[str] = []
    for key in REGION_OVERVIEW_KEYS:
        if not fields.get(key):
            fields[key] = report_error_marker(MATCH_ERROR_TEXT, f"{key}缺少信息或区域概况生成失败")
            marked.append(key)
    if fields.get("文物概况类型") == "有文物审查意见":
        for key in ["文物审查意见文件名", "文物审查意见结论"]:
            if not fields.get(key):
                fields[key] = report_error_marker(MATCH_ERROR_TEXT, f"{key}缺少或OCR未能可靠识别")
                marked.append(key)
    if marked:
        notes.append("已在生成稿中标记匹配错误：" + "、".join(marked))
    return marked


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        if key and key not in os.environ:
            os.environ[key] = value.strip().strip('"').strip("'")


def build_region_overview_prompt(fields: dict[str, str]) -> str:
    county = fields.get("项目所在地旗县") or infer_county(fields) or "项目所在地旗县"
    return f"""
你是考古调查勘探工作报告撰写助手。请为“项目区域概况”生成正式报告正文，只围绕“{county}”这个旗县级行政区，不结合任何具体项目、矿山、用地红线或建设内容。风格参照考古勘探报告基准模板：事实性、稳健、客观，避免宣传腔，避免无法核实的具体数据。

    已有人工参考内容仅可作为本地兜底参考；正式调用区域概况 agent 时不得发送人工填写表参考内容。如参考内容包含具体项目、项目用地或矿区内容，必须忽略：
- 地理位置：{fields.get('项目所在地旗县地理位置概况')}
- 行政区划与社会经济：{fields.get('项目所在地旗县行政区划与社会经济概况')}
- 气候条件：{fields.get('项目所在地旗县气候条件')}
- 历史沿革：{fields.get('项目所在地旗县历史沿革')}

生成要求：
1. 只输出 JSON 对象，不要 Markdown。
    2. JSON 必须包含四个键：{', '.join(REGION_OVERVIEW_KEYS)}。
3. 每个字段为可直接填入 Word 的中文正文，不要加小标题。
    4. 地理位置概况按旧报告范文的厚度写成 2-3 个自然段，合计约 550-650 个汉字：先写{county}的行政隶属、地理方位、邻接关系、地貌位置和面积；再写名称由来、交通区位、自然资源、历史文化资源或生态资源。不要写具体项目位置。
    5. 行政区划与社会经济概况约 330-400 个汉字，覆盖{county}的镇级行政区划、主导产业、产业结构、资源开发、城镇基础、交通条件和居民生产生活概况。不要写荣誉名单、宣传口号或旅游广告。
    6. 气候条件只写{county}的地形地貌、气候类型、气温、降水、风沙、蒸发、日照等指标，约 330-420 个汉字，数据无法稳妥核实时使用“约、一般、呈现”等稳健表述，不写作业建议。
    7. 历史沿革约 1250-1400 个汉字，按时间顺序分 8-11 段写作；段落之间用换行分隔。优先按秦汉、魏晋南北朝、隋唐、辽金元、明清、民国、中华人民共和国成立后等阶段组织，重点写行政隶属、建置沿革、族群活动和区域交通/军政地位。不要编造不可核验的文号、批复号或具体考古结论。
    8. 四个字段合计控制在 2000-2400 个汉字左右。
9. 以上限制只针对 API/agent 生成的旗县概况字段，不影响模板中固定的“{{项目名称}}坐落于{{项目位置}}，该项目面积为{{项目面积}}，地理坐标：{{项目地理坐标}}。”开头句；agent 生成字段本身不得出现“本项目”“该项目”“项目区”“项目所在地”“项目用地”“项目红线”“建设内容”“矿区范围”“坐落于”等把文字指向具体项目的表述。
10. 输出正文应像报告正文，不要像百科词条堆砌；自然资源和社会经济内容要融入段落，不要列清单。
"""


def extract_region_json(text: str) -> dict[str, str]:
    match = re.search(r"\{.*\}", text, flags=re.S)
    if not match:
        raise ValueError("OpenAI response did not contain a JSON object.")
    data = json.loads(match.group(0))
    missing = [key for key in REGION_OVERVIEW_KEYS if not clean(data.get(key))]
    if missing:
        raise ValueError(f"OpenAI response missing fields: {', '.join(missing)}")
    return {key: clean(data.get(key)) for key in REGION_OVERVIEW_KEYS}


def call_region_overview_api(fields: dict[str, str]) -> tuple[dict[str, str], str, str, list[str], dict[str, str]]:
    if region_overview_agent is None:
        raise RuntimeError("region_overview_agent.py 未能加载。")
    result = region_overview_agent.generate_region_overview(fields, strict=False)
    generated = result.get("fields", {})
    missing = [key for key in REGION_OVERVIEW_KEYS if not clean(generated.get(key))]
    if missing:
        raise RuntimeError(f"区域概况 agent 响应缺少字段：{', '.join(missing)}")
    provider = clean(result.get("provider")) or "region-overview-api"
    model = clean(result.get("model"))
    warnings = [clean(item) for item in result.get("warnings", []) if clean(item)]
    research = {str(key): clean(value) for key, value in (result.get("research") or {}).items()}
    return {key: clean(generated.get(key)) for key in REGION_OVERVIEW_KEYS}, provider, model, warnings, research


def project_region_overview_dir(fields: dict[str, str]) -> Path:
    project_label = safe_filename(fields.get("项目名称")) or "未命名项目"
    return PROCESS_DIR / project_label / "区域概况生成"


def load_cached_region_overview(fields: dict[str, str]) -> tuple[dict[str, str], Path] | None:
    project_label = safe_filename(fields.get("项目名称")) or "未命名项目"
    cache_path = project_region_overview_dir(fields) / f"{project_label}_区域概况_API生成.json"
    if not cache_path.exists():
        return None
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    fields_data = data.get("fields") if isinstance(data.get("fields"), dict) else data
    generated = {key: clean(fields_data.get(key)) for key in REGION_OVERVIEW_KEYS}
    if all(generated.values()):
        return generated, cache_path
    return None


def enrich_region_overview_with_api(
    fields: dict[str, str],
    output: Path,
    *,
    skip: bool = False,
    fail_on_error: bool = False,
) -> list[str]:
    if skip:
        return ["项目区域概况API生成：已按参数跳过"]
    load_env_file(ROOT / ".env.local")
    load_env_file(CORE_DIR / ".env.local")
    notes: list[str] = []
    try:
        generated, provider, model, warnings, research = call_region_overview_api(fields)
    except Exception as exc:
        cached = load_cached_region_overview(fields)
        if cached is not None:
            generated, cache_path = cached
            fields.update(generated)
            return [
                "项目区域概况API生成：本次 API 重试失败，已使用本项目既有 API 生成缓存"
                f"（{cache_path}）；失败原因：{exc}"
            ]
        missing_region_keys = []
        for key in REGION_OVERVIEW_KEYS:
            if not fields.get(key):
                fields[key] = report_error_marker(MATCH_ERROR_TEXT, f"{key}缺少信息且区域概况API生成失败")
                missing_region_keys.append(key)
        notes.append(f"项目区域概况API生成失败：{exc}")
        if fail_on_error:
            if missing_region_keys:
                notes.append("项目区域概况API生成失败，缺失字段已在生成稿中标记为匹配错误。")
            else:
                notes.append("项目区域概况API生成失败，已保留填报表中的区域概况内容。")
        return notes
    fields.update(generated)
    out_dir = project_region_overview_dir(fields)
    out_dir.mkdir(parents=True, exist_ok=True)
    project_label = safe_filename(fields.get("项目名称")) or "未命名项目"
    out_path = out_dir / f"{project_label}_区域概况_API生成.json"
    artifact = {
        "provider": provider,
        "model": model,
        "fields": generated,
        "research": research,
        "warnings": warnings,
    }
    out_path.write_text(json.dumps(artifact, ensure_ascii=False, indent=2), encoding="utf-8")
    notes.append(
        f"项目区域概况API生成：已通过 {provider}{f'/{model}' if model else ''} 更新地理位置、气候条件、历史沿革"
        f"（{out_path}）"
    )
    notes.extend(f"项目区域概况API生成警告：{warning}" for warning in warnings)
    return notes


def build_conclusion(section_records: list[dict[str, str]], standard_records: list[dict[str, str]], fields: dict[str, str]) -> str:
    descriptions = []
    for record in section_records:
        descriptions.append(record.get("剖线地层描述", ""))
    for record in standard_records:
        descriptions.append(record.get("标准孔地层描述", ""))
    text = "；".join(item for item in descriptions if item)
    if not text:
        return ""
    layers = []
    for name in ["表土层", "耕土层", "扰土层", "沙土层", "粉砂层", "黏土层", "生土层"]:
        if name in text:
            layers.append(name)
    layer_text = "、".join(dict.fromkeys(layers)) if layers else "表土层、下部自然堆积层及生土层"
    conclusion = fields.get("遗迹结论") or "未发现古代文化遗存"
    return (
        "综合考古调查与勘探结果，在本项目已完成勘探区域内，"
        f"{conclusion}。通过勘探情况可知，勘探区域内主要地层包括{layer_text}。"
        "各层土色、土质及包含物情况以剖线记录和标准孔记录为准，整体反映本项目范围内地层堆积情况。"
    )


CONCLUSION_SAFETY_TEXT = (
    "鉴于地下文物埋藏的复杂性和考古勘探的局限性，项目在建设过程中如发现文物遗存，"
    "建设单位应立即停工，保护好现场，及时报告文物行政管理部门，待文物遗存得到妥善处理，"
    "经文物主管部门同意后，工程方可继续施工，从而确保地下文物安全和建设工程的顺利进行。"
)


def sentence_text(text: str) -> str:
    text = normalize_generated_punctuation(clean(text))
    if text and text[-1] not in "。！？；":
        text += "。"
    return text


def conclusion_relic_sentence(fields: dict[str, str]) -> str:
    explicit = clean(fields.get("结论文物遗迹认定"))
    if explicit:
        return sentence_text(explicit)
    relic = clean(fields.get("遗迹结论"))
    if not relic or re.search(r"未发现|无", relic):
        return "通过考古调查及勘探，认定在勘探区域内未发现文物及遗迹现象。"
    if relic.startswith("发现") or relic.startswith("未发现"):
        return sentence_text(f"通过考古调查及勘探，认定在勘探区域内{relic}")
    return sentence_text(f"通过考古调查及勘探，认定在勘探区域内{relic}")


def build_conclusion_paragraphs(fields: dict[str, str]) -> list[str]:
    first = sentence_text(fields.get("勘探成果综合结论"))
    second = conclusion_relic_sentence(fields) + CONCLUSION_SAFETY_TEXT
    return [item for item in [first, second] if item]


def conclusion_layer_summary(fields: dict[str, str]) -> str:
    explicit = clean(fields.get("根据地层堆积情况进行总结"))
    if explicit:
        return explicit.rstrip("。")
    text = clean(fields.get("勘探成果综合结论"))
    if not text:
        return report_error_marker(MATCH_ERROR_TEXT, "缺少勘探成果综合结论，无法生成地层堆积总结")
    text = re.sub(r"^通过(?:勘探(?:工作)?情况|勘探)可知[，,]?", "", text)
    text = re.sub(r"^项目?勘探区域地层堆积(?:序列)?(?:较为)?清晰[，,]层位关系明确[，,]?", "", text)
    text = re.sub(r"^勘探区域地层堆积(?:序列)?(?:较为)?清晰[，,]层位关系明确[，,]?", "", text)
    text = re.sub(r"^自上而下依次为[:：]", "", text)
    return text.rstrip("。")


def partition_stratigraphy_summary(fields: dict[str, str]) -> str:
    explicit = clean(fields.get("分区地层逐区总结"))
    if explicit:
        return explicit
    text = clean(fields.get("勘探成果综合结论"))
    if re.search(r"[A-ZＡ-Ｚ][区區]地层", text):
        return text
    detail = clean(fields.get("勘探分区逐项说明"))
    if re.search(r"[A-ZＡ-Ｚ][区區].*地层", detail):
        return detail
    if text:
        return text.rstrip("。")
    return report_error_marker(MATCH_ERROR_TEXT, "缺少分区地层逐区总结")


def refresh_conclusion_derived_fields(fields: dict[str, str]) -> None:
    fields["根据地层堆积情况进行总结"] = conclusion_layer_summary(fields)
    fields["分区地层逐区总结"] = partition_stratigraphy_summary(fields)
    fields["结论文物遗迹认定"] = conclusion_relic_sentence(fields).rstrip("。")


def load_form(path: Path) -> tuple[dict[str, str], dict[str, list[dict[str, str]]], list[str]]:
    wb = load_workbook(path, data_only=True)
    fields: dict[str, str] = {}
    tables: dict[str, list[dict[str, str]]] = {}
    notes: list[str] = []
    for sheet in wb.sheetnames:
        if sheet in KV_SHEETS:
            for key, value in parse_kv_sheet(wb[sheet]).items():
                if value or key not in fields:
                    fields[key] = value
        elif sheet in TABLE_SHEETS:
            tables[sheet] = parse_table_sheet(wb[sheet])
    notes.extend(derive_fields(fields, tables))
    refresh_conclusion_derived_fields(fields)
    return fields, tables, notes


def read_redline_coordinate_table(drawings_dir: Path | None) -> tuple[list[dict[str, str]], str]:
    if drawings_dir is None:
        return [], ""
    candidates = [
        drawings_dir / "表格" / "红线四至坐标.xlsx",
        drawings_dir / "表格" / "红线四至坐标表.xlsx",
        drawings_dir / "表格" / "四至范围坐标.xlsx",
        drawings_dir / "表格" / "四至范围坐标表.xlsx",
        drawings_dir / "红线四至坐标.xlsx",
        drawings_dir / "红线四至坐标表.xlsx",
        drawings_dir / "四至范围坐标.xlsx",
        drawings_dir / "四至范围坐标表.xlsx",
    ]
    source = next((path for path in candidates if path.exists() and not path.name.startswith(".~")), None)
    if source is None:
        source = find_workbook_by_names(
            drawings_dir,
            ["红线四至坐标.xlsx", "红线四至坐标表.xlsx", "四至范围坐标.xlsx", "四至范围坐标表.xlsx"],
        )
    if source is None:
        return [], ""
    wb = load_workbook(source, data_only=True)
    best_records: list[dict[str, str]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        headers: dict[str, int] = {}
        for row_idx, row in enumerate(rows):
            values = [clean(value) for value in row]
            if not any(values):
                continue
            normalized = [normalize_key(value) for value in values]
            if not headers and (
                any(value in {"角点", "点号", "名称", "界点", "拐点"} for value in values)
                and any("x坐标" in value or "平面坐标x" in value for value in normalized)
                and any("y坐标" in value or "平面坐标y" in value for value in normalized)
            ):
                for col_idx, value in enumerate(values):
                    key = normalize_key(value)
                    if value in {"角点", "点号", "名称", "界点", "拐点"}:
                        headers["corner"] = col_idx
                    elif "x坐标" in key or "平面坐标x" in key:
                        headers["x"] = col_idx
                    elif "y坐标" in key or "平面坐标y" in key:
                        headers["y"] = col_idx
                continue
            if headers:
                corner = clean(row[headers["corner"]] if len(row) > headers["corner"] else "")
                x_value = clean(row[headers["x"]] if len(row) > headers["x"] else "")
                y_value = clean(row[headers["y"]] if len(row) > headers["y"] else "")
            else:
                corner = clean(row[0] if len(row) > 0 else "")
                x_value = clean(row[1] if len(row) > 1 else "")
                y_value = clean(row[2] if len(row) > 2 else "")
                if row_idx == 0 and not (x_value and y_value):
                    continue
            if not (corner and x_value and y_value):
                continue
            if "坐标" in corner and not re.search(r"(东|西|南|北|角|点|[A-Z]\d*)", corner, flags=re.I):
                continue
            if not (re.search(r"\d", x_value) and re.search(r"\d", y_value)):
                continue
            best_records.append({"角点": corner, "X坐标": x_value, "Y坐标": y_value})
        if len(best_records) >= 4:
            break
    return best_records, str(source)


def read_project_land_range_coordinate_table(drawings_dir: Path | None) -> tuple[list[dict[str, object]], str]:
    if drawings_dir is None or not drawings_dir.exists():
        return [], ""
    search_roots = [
        drawings_dir,
        drawings_dir / "3.内业成果",
        drawings_dir / "内业成果",
        drawings_dir / "3.制图成果",
        drawings_dir / "制图成果",
        drawings_dir.parent / "3.内业成果",
        drawings_dir.parent / "内业成果",
        drawings_dir.parent / "3.制图成果",
        drawings_dir.parent / "制图成果",
    ]
    unique_roots: list[Path] = []
    seen_roots: set[Path] = set()
    for root in search_roots:
        normalized = root.resolve() if root.exists() else root
        if normalized in seen_roots:
            continue
        seen_roots.add(normalized)
        unique_roots.append(root)
    candidates: list[Path] = []
    for root in unique_roots:
        candidates.extend([
            root / "表格" / "项目用地范围坐标表.xlsx",
            root / "项目用地范围坐标表.xlsx",
        ])
    source = next((path for path in candidates if path.exists() and not path.name.startswith(".~")), None)
    if source is None:
        for root in unique_roots:
            source = find_workbook_by_names(root, ["项目用地范围坐标表.xlsx"]) if root.exists() else None
            if source is not None:
                break
    if source is None:
        return [], ""
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, object]] = []
    max_cols = 0
    source_row_to_output_row: dict[int, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells = [clean(cell.value) for cell in row]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        max_cols = max(max_cols, len(cells))
        styles = []
        for cell in row[:len(cells)]:
            styles.append({
                "bold": bool(cell.font.bold),
                "horizontal": cell.alignment.horizontal or "center",
                "vertical": cell.alignment.vertical or "center",
            })
        source_row_to_output_row[row_idx] = len(rows)
        rows.append({"__cells": cells, "__styles": styles, "__source_row": row_idx})
    for record in rows:
        cells = list(record["__cells"])
        cells.extend([""] * (max_cols - len(cells)))
        record["__cells"] = cells
        styles = list(record.get("__styles", []))
        styles.extend([{"bold": False, "horizontal": "center", "vertical": "center"}] * (max_cols - len(styles)))
        record["__styles"] = styles
    merges: list[dict[str, int]] = []
    for merged_range in ws.merged_cells.ranges:
        kept_rows = [row_idx for row_idx in range(merged_range.min_row, merged_range.max_row + 1) if row_idx in source_row_to_output_row]
        if not kept_rows:
            continue
        row_start = source_row_to_output_row[kept_rows[0]]
        row_end = source_row_to_output_row[kept_rows[-1]]
        col_start = merged_range.min_col - 1
        col_end = min(merged_range.max_col, max_cols) - 1
        if col_start > col_end:
            continue
        if row_start == row_end and col_start == col_end:
            continue
        merges.append({
            "row_start": row_start,
            "row_end": row_end,
            "col_start": col_start,
            "col_end": col_end,
        })
    if rows:
        rows[0]["__merges"] = merges
    return rows, str(source)


def read_unit_coordinate_table(drawings_dir: Path | None) -> tuple[list[dict[str, str]], str]:
    if drawings_dir is None:
        return [], ""
    candidates = [
        drawings_dir / "表格" / "勘探单元坐标.xlsx",
        drawings_dir / "勘探单元坐标.xlsx",
        drawings_dir / "3.勘探单元" / "勘探单元.xlsx",
    ]
    source = next((path for path in candidates if path.exists()), None)
    if source is None:
        source = find_workbook_by_names(drawings_dir, ["勘探单元坐标.xlsx", "勘探单元.xlsx"])
    if source is None:
        return [], ""
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[dict[str, str]] = []
    current_unit = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit = clean(row[0] if len(row) > 0 else "")
        if unit:
            current_unit = unit
        corner = clean(row[1] if len(row) > 1 else "")
        x_value = clean(row[2] if len(row) > 2 else "")
        y_value = clean(row[3] if len(row) > 3 else "")
        if current_unit and corner and x_value and y_value:
            records.append(
                {
                    "勘探单元": current_unit,
                    "勘探单元显示": current_unit if corner == "西南角" else "",
                    "角点": corner,
                    "X坐标": x_value,
                    "Y坐标": y_value,
                }
            )
    return records, str(source)


def read_standard_coordinate_table(drawings_dir: Path | None) -> tuple[list[dict[str, str]], str]:
    if drawings_dir is None:
        return [], ""
    candidates = [
        drawings_dir / "表格" / "标准孔坐标.xlsx",
        drawings_dir / "标准孔坐标.xlsx",
    ]
    source = next((path for path in candidates if path.exists() and not path.name.startswith(".~")), None)
    if source is None:
        source = find_workbook_by_names(drawings_dir, ["标准孔坐标.xlsx"])
    if source is None:
        return [], ""
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[dict[str, str]] = []
    current_unit = ""
    header_row_idx = 2
    headers: list[str] = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), start=1):
        values = [clean(value) for value in row]
        if "标准孔编号" in values and ("X坐标" in values or "Y坐标" in values):
            header_row_idx = idx
            headers = values
            break
    col = {header: idx for idx, header in enumerate(headers) if header}
    unit_idx = col.get("勘探单元", 0)
    probe_idx = col.get("探孔编号")
    standard_idx = col.get("标准孔编号", 2 if probe_idx is not None else 1)
    x_idx = col.get("X坐标", 3 if probe_idx is not None else 2)
    y_idx = col.get("Y坐标", 4 if probe_idx is not None else 3)
    elevation_idx = col.get("高程", 5 if probe_idx is not None else 4)
    row_number = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        unit = clean(row[unit_idx] if len(row) > unit_idx else "")
        if unit:
            current_unit = unit
        standard_code = clean(row[standard_idx] if len(row) > standard_idx else "")
        if current_unit and standard_code:
            row_number += 1
        hole_code = normalize_tk_code(row[probe_idx] if probe_idx is not None and len(row) > probe_idx else f"TK{row_number}")
        x_value = clean(row[x_idx] if len(row) > x_idx else "")
        y_value = clean(row[y_idx] if len(row) > y_idx else "")
        elevation = format_elevation(row[elevation_idx] if len(row) > elevation_idx else "")
        if current_unit and hole_code and standard_code and x_value and y_value:
            probe_number = f"{current_unit}-{standard_code}"
            records.append(
                {
                    "勘探单元": current_unit,
                    "勘探单元显示": current_unit,
                    "探孔编号": hole_code,
                    "标准孔编号": standard_code,
                    "勘探编号": probe_number,
                    "X坐标": x_value,
                    "Y坐标": y_value,
                    "高程": elevation,
                    "探孔坐标": f"X：{x_value}，Y：{y_value}，高程：{elevation}",
                    "标准孔位置图": f"{probe_number}.jpg",
                    "标准孔土样照": f"{hole_code}.jpg",
                }
            )
    return records, str(source)


def read_section_coordinate_tables(drawings_dir: Path | None) -> tuple[dict[str, list[dict[str, str]]], str]:
    if drawings_dir is None:
        return {}, ""
    candidates = [
        drawings_dir / "表格" / "剖线坐标.xlsx",
        drawings_dir / "剖线坐标.xlsx",
    ]
    source = next((path for path in candidates if path.exists() and not path.name.startswith(".~")), None)
    if source is None:
        source = find_workbook_by_names(drawings_dir, ["剖线坐标.xlsx", "剖面坐标.xlsx"])
    if source is None:
        return {}, ""
    wb = load_workbook(source, data_only=True)
    out: dict[str, list[dict[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_id = "A-A′" if sheet_name.strip().upper() == "A" else "B-B′" if sheet_name.strip().upper() == "B" else sheet_name
        records: list[dict[str, str]] = []
        current_unit = ""
        for row in ws.iter_rows(min_row=4, values_only=True):
            unit = clean(row[0] if len(row) > 0 else "")
            if unit:
                current_unit = unit
            col_a = clean(row[1] if len(row) > 1 else "")
            row_b = clean(row[2] if len(row) > 2 else "")
            x_value = clean(row[3] if len(row) > 3 else "")
            y_value = clean(row[4] if len(row) > 4 else "")
            elevation = format_elevation(row[5] if len(row) > 5 else "")
            if current_unit and col_a and row_b and x_value and y_value:
                records.append(
                    {
                        "剖线编号": section_id,
                        "勘探单元": current_unit,
                        "勘探单元显示": current_unit,
                        "列A": col_a,
                        "行B": row_b,
                        "X坐标": x_value,
                        "Y坐标": y_value,
                        "高程": elevation,
                    }
                )
        out[section_id] = records
    return out, str(source)


def read_docx_paragraphs(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as docx:
        root = ET.fromstring(docx.read("word/document.xml"))
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", NS):
        text = element_text(para).strip()
        if text:
            paragraphs.append(text)
    return paragraphs


def find_site_record_docx(photos_dir: Path | None, *, prefer_summary: bool = True) -> Path | None:
    if photos_dir is None:
        return None
    if prefer_summary:
        candidates = [
            photos_dir / "项目现场记录-益民.docx",
            photos_dir / "项目现场记录.docx",
            photos_dir / "17.汇总" / "项目现场记录.docx",
            photos_dir / "15.汇总" / "项目现场记录.docx",
        ]
    else:
        candidates = [
            photos_dir / "17.汇总" / "项目现场记录.docx",
            photos_dir / "15.汇总" / "项目现场记录.docx",
        ]
    source = next((path for path in candidates if path.exists() and not path.name.startswith(".~")), None)
    if source is None and photos_dir.exists():
        record_candidates = {
            path
            for pattern in ("项目现场记录*.docx", "*项目现场记录*.docx", "*现场记录*.docx")
            for path in photos_dir.rglob(pattern)
        }
        source = next(
            (
                path
                for path in sorted(
                    record_candidates,
                    key=lambda item: ((len(item.parts) if prefer_summary else -len(item.parts)), str(item)),
                )
                if not path.name.startswith(".~") and "模板" not in path.name
            ),
            None,
        )
    return source


def normalize_site_record_field_value(key: str, value: str) -> str:
    text = re.sub(r"[（(]\s*有/无\s*[)）]", "", clean(value)).strip()
    if key == "项目建设内容":
        match = re.search(r"建设内容(?:用途)?为\s*([^。；;\n]+)", text)
        if match:
            return match.group(1).strip()
    if key in {"项目面积", "调查面积", "勘探面积"}:
        match = re.match(r"^([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)?$", text)
        if match:
            return match.group(1).replace(",", "").replace("，", "") + "平方米"
        match = re.search(r"([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)", text)
        if match:
            return match.group(1).replace(",", "").replace("，", "") + "平方米"
    if key == "勘探单元规格" and re.fullmatch(r"\d+", text):
        return f"{text}米×{text}米"
    if key == "勘探时间":
        return re.sub(r"\s+", "", text)
    return text


def read_site_record_fields(photos_dir: Path | None) -> tuple[dict[str, str], str]:
    source = find_site_record_docx(photos_dir, prefer_summary=True)
    if source is None:
        return {}, ""
    paragraphs = read_docx_paragraphs(source)
    fields: dict[str, str] = {}
    label_map = {
        "项目名称": "项目名称",
        "项目名字": "项目名称",
        "建设单位": "建设单位",
        "勘探单位": "勘探单位",
        "项目位置": "项目位置",
        "项目地点": "项目位置",
        "项目地理坐标": "项目地理坐标",
        "地理位置坐标经纬度（中心点）": "项目地理坐标",
        "地理位置坐标经纬度": "项目地理坐标",
        "项目面积": "项目面积",
        "调查面积": "调查面积",
        "考古调查面积": "调查面积",
        "项目建设内容": "项目建设内容",
        "项目概况": "项目建设内容",
        "勘探面积": "勘探面积",
        "实际勘探面积": "勘探面积",
        "考古勘探面积": "勘探面积",
        "勘探时间": "勘探时间",
        "开始日期": "开始日期",
        "结束日期": "结束日期",
        "工作天数": "工作天数",
        "遗迹结论": "遗迹结论",
        "回函信息": "回函信息",
        "项目地块情况": "项目地块情况",
        "地貌": "项目地块情况",
        "是否存在不可勘探区域": "是否存在不可勘探区域",
        "不可勘探原因": "不可勘探原因",
        "进场人数": "进场人数",
        "勘探单元数量": "勘探单元数量",
        "勘探单元规格": "勘探单元规格",
        "是否存在勘探分区": "是否存在勘探分区",
        "勘探分区数量": "勘探分区数量",
        "剖线数量": "剖线数量",
        "标准孔数量": "标准孔数量",
        "遗迹数量": "遗迹数量",
        "勘探分区原因": "勘探分区原因",
    }
    for text in paragraphs:
        match = re.match(r"^([^：:]{2,24})[：:]\s*(.+)$", text)
        if not match:
            continue
        label = clean(match.group(1))
        value = clean(match.group(2))
        key = label_map.get(label)
        if key and value:
            fields[key] = normalize_site_record_field_value(key, value)

    joined = "\n".join(paragraphs)
    if "勘探面积" not in fields:
        match = re.search(r"实际勘探面积\s*([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)", joined)
        if match:
            fields["勘探面积"] = normalize_site_record_field_value("勘探面积", match.group(1))
    if "探孔总数" not in fields:
        match = re.search(r"探孔总数[:：]?\s*([0-9,，]+)\s*个?", joined)
        if match:
            fields["探孔总数"] = match.group(1).replace(",", "").replace("，", "")
    if "剖线数量" not in fields:
        match = re.search(r"勘探剖线[:：]?\s*([0-9]+)\s*条", joined)
        if match:
            fields["剖线数量"] = match.group(1)
    if "探工数量" not in fields:
        match = re.search(r"勘探人数[:：]?\s*([0-9]+)", joined)
        if match:
            fields["探工数量"] = match.group(1)
    match = re.search(r"结论[:：]\s*([^\n]+)", joined)
    if match:
        conclusion = clean(match.group(1))
        if conclusion and "勘探成果综合结论" not in fields:
            fields["勘探成果综合结论"] = conclusion
        if "遗迹结论" not in fields and re.search(r"未发现(?:文化遗存|遗迹|人工扰动)|无人工扰动", conclusion):
            fields["遗迹结论"] = "未发现文化遗存"

    if "勘探时间" in fields and ("开始日期" not in fields or "结束日期" not in fields):
        match = re.search(
            r"([0-9]{4})年([0-9]{1,2})月([0-9]{1,2})日[—\-至到]+([0-9]{4})?年?([0-9]{1,2})月([0-9]{1,2})日",
            fields["勘探时间"],
        )
        if match:
            start_year, start_month, start_day, end_year, end_month, end_day = match.groups()
            end_year = end_year or start_year
            fields.setdefault("开始日期", f"{int(start_year)}年{int(start_month)}月{int(start_day)}日")
            fields.setdefault("结束日期", f"{int(end_year)}年{int(end_month)}月{int(end_day)}日")
            if "工作天数" not in fields:
                try:
                    from datetime import date

                    start = date(int(start_year), int(start_month), int(start_day))
                    end = date(int(end_year), int(end_month), int(end_day))
                    fields["工作天数"] = str((end - start).days + 1)
                except ValueError:
                    pass

    detail_lines: list[str] = []
    in_partition_detail = False
    for text in paragraphs:
        if text.startswith("各分区情况说明"):
            in_partition_detail = True
            tail = text.split("：", 1)[1].strip() if "：" in text else ""
            if tail:
                detail_lines.append(tail)
            continue
        if in_partition_detail and re.match(r"^(剖线记录|标准孔记录|遗迹记录|灰坑描述|墓葬描述)$", text):
            break
        if in_partition_detail:
            detail_lines.append(text)
    if detail_lines:
        fields["勘探分区逐项说明"] = "\n".join(detail_lines)
    return fields, str(source)


def read_section_descriptions_from_site_record(photos_dir: Path | None) -> tuple[list[dict[str, str]], str]:
    source = find_site_record_docx(photos_dir, prefer_summary=False)
    if source is None:
        return [], ""

    records: list[dict[str, str]] = []
    current_number = ""
    current_lines: list[str] = []
    in_section_records = False

    def flush_current() -> None:
        nonlocal current_number, current_lines
        if current_number and current_lines:
            records.append(
                {
                    "剖线编号": current_number,
                    "剖线地层描述": "".join(current_lines),
                }
            )
        current_number = ""
        current_lines = []

    pending_texts: list[str] = []
    for text in read_docx_paragraphs(source):
        if "剖线记录" in text:
            in_section_records = True
            tail = re.split(r"剖线记录[:：]?", text, maxsplit=1)[-1].strip()
            if tail and tail != text:
                pending_texts.append(tail)
            continue
        if not in_section_records:
            continue
        pending_texts.append(text)

    for text in pending_texts:
        if "标准孔记录" in text:
            flush_current()
            break
        match = re.match(
            r"^[剖刨]线\s*([A-Za-zＡ-Ｚａ-ｚ](?:-[A-Za-zＡ-Ｚａ-ｚ][′'’]?)?)\s*(?:地层堆积情况)?\s*[:：]?\s*(.*)$",
            text,
        )
        if match:
            flush_current()
            current_number = normalize_section_number(match.group(1))
            if "-" not in current_number and len(current_number) == 1:
                current_number = f"{current_number}-{current_number}′"
            tail = clean(match.group(2))
            if tail:
                current_lines.append(tail)
            continue
        if current_number:
            current_lines.append(text)
    flush_current()

    normalize_sequential_section_records(records)
    return records, str(source)


def normalize_sequential_section_records(records: list[dict[str, str]]) -> None:
    section_sequence = ["A-A′", "B-B′", "C-C′", "D-D′"]
    for idx, record in enumerate(records):
        number = section_number_key(record.get("剖线编号") or record.get("剖面编号") or "")
        description = clean(record.get("剖线地层描述") or record.get("剖面地层描述"))
        match = re.match(r"^([0-9]+)\s*[：:]\s*(.+)$", description, flags=re.S)
        if number in {"P-P′", "P"} and match:
            seq_index = int(match.group(1)) - 1
            if 0 <= seq_index < len(section_sequence):
                record["剖线编号"] = section_sequence[seq_index]
                record["剖线地层描述"] = match.group(2).strip()
        elif number in {"P-P′", "P"} and idx < len(section_sequence):
            record["剖线编号"] = section_sequence[idx]


def read_standard_descriptions_from_site_record(photos_dir: Path | None) -> tuple[dict[str, str], str]:
    source = find_site_record_docx(photos_dir, prefer_summary=False)
    if source is None:
        return {}, ""
    paragraphs = read_docx_paragraphs(source)
    start_idx = next((idx for idx, text in enumerate(paragraphs) if "标准孔记录" in text), None)
    if start_idx is None:
        return {}, str(source)

    standard_paragraphs: list[str] = []
    start_tail = re.split(r"标准孔记录[:：]?", paragraphs[start_idx], maxsplit=1)[-1].strip()
    if start_tail and start_tail != paragraphs[start_idx]:
        standard_paragraphs.append(start_tail)
    for para in paragraphs[start_idx + 1 :]:
        if re.match(r"^(遗迹记录|灰坑记录|墓葬记录|房址记录|窑址记录|灰坑描述|墓葬描述|遗迹描述|房址描述|窑址描述)$", para):
            break
        standard_paragraphs.append(para)

    text = "".join(standard_paragraphs)
    records: dict[str, str] = {}
    matches = list(re.finditer(r"TK\s*0*(\d{1,3})", text, flags=re.I))
    for idx, match in enumerate(matches):
        key = normalize_tk_code(match.group(0))
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        description = text[start:end].strip(" ：:")
        description = re.sub(r"^[（(]\s*U\d+\s*[-－—]\s*[A-Z]\d+[A-Z]\d+\s*[）)]", "", description).strip(" ：:")
        if description:
            records[key] = description
            for alias in tk_code_aliases(key):
                records[alias] = description
    return records, str(source)


def read_relic_location_map(drawings_dir: Path | None) -> tuple[dict[str, str], str]:
    if drawings_dir is None:
        return {}, ""
    source = find_workbook_by_names(drawings_dir, ["文物内外遗迹统计表.xlsx"])
    if source is None:
        return {}, ""
    wb = load_workbook(source, data_only=True)
    mapping: dict[str, str] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean(value) for value in rows[0]]
        try:
            location_idx = headers.index("遗址位置")
            relic_idx = headers.index("遗迹")
        except ValueError:
            continue
        for row in rows[1:]:
            relic_id = canonical_relic_id(row[relic_idx] if relic_idx < len(row) else "")
            location = clean(row[location_idx] if location_idx < len(row) else "")
            if relic_id:
                mapping[relic_id] = location
    return mapping, str(source)


def read_relic_unit_map(drawings_dir: Path | None) -> tuple[dict[str, str], str]:
    if drawings_dir is None:
        return {}, ""
    source = find_workbook_by_names(drawings_dir, ["勘探单元遗迹统计表.xlsx"])
    if source is None:
        return {}, ""
    wb = load_workbook(source, data_only=True)
    mapping: dict[str, str] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean(value) for value in rows[0]]
        try:
            unit_idx = headers.index("勘探单元")
            relic_idx = headers.index("遗迹")
        except ValueError:
            continue
        current_unit = ""
        for row in rows[1:]:
            unit = clean(row[unit_idx] if unit_idx < len(row) else "")
            if unit:
                current_unit = unit
            relic_id = canonical_relic_id(row[relic_idx] if relic_idx < len(row) else "")
            if relic_id and current_unit:
                mapping[relic_id] = current_unit
    return mapping, str(source)


def read_relic_coordinate_table(drawings_dir: Path | None) -> tuple[list[dict[str, str]], str]:
    if drawings_dir is None:
        return [], ""
    source = find_workbook_by_names(drawings_dir, ["遗迹坐标.xlsx"])
    if source is None:
        return [], ""
    wb = load_workbook(source, data_only=True)
    records: list[dict[str, str]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean(value) for value in rows[0]]
        for row in rows[1:]:
            record = {
                headers[idx]: clean(row[idx]) if idx < len(row) else ""
                for idx in range(len(headers))
                if headers[idx]
            }
            name = clean(record.get("名称") or record.get("遗迹编号") or record.get("界点序号"))
            relic_id = canonical_relic_id(name)
            point_match = re.search(r"(?:-|点)\s*0*(\d+)", name, flags=re.I)
            point_number = point_match.group(1) if point_match else clean(record.get("界点序号"))
            if not relic_id or not record.get("平面坐标X") or not record.get("平面坐标Y"):
                continue
            records.append(
                {
                    "遗迹编号": relic_id,
                    "界点序号": f"{relic_id}点{point_number}" if point_number else name,
                    "平面坐标X": record.get("平面坐标X", ""),
                    "平面坐标Y": record.get("平面坐标Y", ""),
                    "经度": record.get("经度", ""),
                    "纬度": record.get("纬度", ""),
                }
            )
    return sorted(records, key=lambda row: (*relic_sort_key(row.get("遗迹编号", "")), first_number(row.get("界点序号", "")) or 0)), str(source)


def parse_relic_measurements(description: str) -> dict[str, str]:
    text = clean(description)
    out: dict[str, str] = {}
    shape = re.search(r"平面呈([^，。]+)", text)
    if shape:
        out["形制"] = shape.group(1)
    width_matches = re.findall(r"(?:东西|南北)(?:长|宽)\s*约?\s*([0-9.]+)\s*米", text)
    if width_matches:
        out["长度（米）"] = width_matches[0]
        if len(width_matches) > 1:
            out["宽度（米）"] = width_matches[1]
    area = re.search(r"面积\s*约?\s*([0-9.]+)\s*平方米", text)
    if area:
        out["面积（平方米）"] = area.group(1)
    open_layer = re.search(r"开口于([^，。]+)", text)
    if open_layer:
        out["开口层位"] = open_layer.group(1)
    mouth_depth = re.search(r"开口距地表\s*([0-9.]+)\s*米", text)
    if mouth_depth:
        out["口深（米）"] = mouth_depth.group(1)
    bottom_depth = re.search(r"底距地表\s*([0-9.]+)\s*米", text)
    if bottom_depth:
        out["底深（米）"] = bottom_depth.group(1)
    contents = re.search(r"(?:坑内|内填)见?([^。；;，]*(?:、[^。；;，]+)*)", text)
    if contents:
        out["内部填充与包含物"] = contents.group(1).strip("，。")
    return out


def read_relic_descriptions_from_site_record(
    photos_dir: Path | None,
    unit_map: dict[str, str],
    location_map: dict[str, str],
) -> tuple[list[dict[str, str]], str]:
    if photos_dir is None:
        return [], ""
    source = find_site_record_docx(photos_dir, prefer_summary=False)
    if source is None:
        return [], ""
    paragraphs = read_docx_paragraphs(source)
    records_by_id: dict[str, dict[str, str]] = {}
    idx = 0
    while idx < len(paragraphs):
        title = clean(paragraphs[idx])
        inline_match = re.match(
            rf"^(?:灰坑|墓葬|遗迹|房址|窑址|窑坑|活土坑|[\u4e00-\u9fff、，,\s]*?(?:灰坑|墓葬|房址|窑址|窑坑|活土坑)?)?\s*({RELIC_ID_RE})\s*[：:]\s*(.+)$",
            title,
            flags=re.I,
        )
        if inline_match:
            relic_id = canonical_relic_id(inline_match.group(1))
            description_body = clean(inline_match.group(2))
            description = description_body if relic_id in description_body else f"{relic_id}：{description_body}"
            record = {
                "遗迹编号": relic_id,
                "遗迹类型": relic_type_from_id(relic_id),
                "遗迹描述": description,
                "遗迹土样描述": "",
                "勘探单元": unit_map.get(relic_id, ""),
                "遗址位置": location_map.get(relic_id, ""),
            }
            record.update(parse_relic_measurements(description))
            records_by_id[relic_id] = record
            idx += 1
            continue
        match = re.fullmatch(
            rf"(?:灰坑|墓葬|遗迹|房址|窑址|窑坑|活土坑|[\u4e00-\u9fff、，,\s]*?(?:灰坑|墓葬|房址|窑址|窑坑|活土坑)?)?\s*({RELIC_ID_RE})",
            title,
            flags=re.I,
        )
        if not match:
            idx += 1
            continue
        relic_id = canonical_relic_id(match.group(1))
        if idx + 1 >= len(paragraphs):
            idx += 1
            continue
        description = clean(paragraphs[idx + 1])
        if relic_id not in description:
            idx += 1
            continue
        soil_description = ""
        if idx + 2 < len(paragraphs) and "土样描述" in paragraphs[idx + 2]:
            soil_description = clean(paragraphs[idx + 2])
        record = {
            "遗迹编号": relic_id,
            "遗迹类型": relic_type_from_id(relic_id),
            "遗迹描述": description,
            "遗迹土样描述": soil_description,
            "勘探单元": unit_map.get(relic_id, ""),
            "遗址位置": location_map.get(relic_id, ""),
        }
        record.update(parse_relic_measurements(description))
        records_by_id[relic_id] = record
        idx += 3
    records = [records_by_id[key] for key in sorted(records_by_id, key=relic_sort_key)]
    for idx, record in enumerate(records, start=1):
        record["序号"] = str(idx)
    return records, str(source)


def normalize_section_number(value: str) -> str:
    text = clean(value)
    text = text.replace("'", "′").replace("’", "′")
    text = text.replace("－", "-").replace("—", "-").replace("～", "-")
    return text.upper()


def import_external_tables(tables: dict[str, list[dict[str, str]]], drawings_dir: Path | None) -> list[str]:
    notes: list[str] = []
    redline_records, redline_source = read_redline_coordinate_table(drawings_dir)
    if redline_records and not tables.get("红线坐标"):
        tables["红线坐标"] = redline_records
        notes.append(f"已从制图成果表格导入红线坐标：{len(redline_records)} 条（{redline_source}）")
    land_range_records, land_range_source = read_project_land_range_coordinate_table(drawings_dir)
    if land_range_records:
        tables["项目用地范围坐标表"] = land_range_records
        notes.append(f"已从内业成果表格导入项目用地范围坐标表：{len(land_range_records)} 行（{land_range_source}）")
    unit_records, unit_source = read_unit_coordinate_table(drawings_dir)
    if unit_records and not tables.get("勘探单元"):
        tables["勘探单元"] = unit_records
        notes.append(f"已从制图成果表格导入勘探单元坐标：{len(unit_records)} 条（{unit_source}）")
    standard_records, standard_source = read_standard_coordinate_table(drawings_dir)
    if standard_records and (
        not tables.get("标准孔")
        or not any(record.get("探孔编号") for record in tables.get("标准孔", []))
    ):
        tables["标准孔"] = standard_records
        notes.append(f"已从制图成果表格导入标准孔坐标：{len(standard_records)} 条（{standard_source}）")
    section_coordinate_tables, section_coordinate_source = read_section_coordinate_tables(drawings_dir)
    if section_coordinate_tables:
        tables["剖线坐标_AA"] = section_coordinate_tables.get("A-A′", [])
        tables["剖线坐标_BB"] = section_coordinate_tables.get("B-B′", [])
        total = len(tables["剖线坐标_AA"]) + len(tables["剖线坐标_BB"])
        notes.append(f"已从制图成果表格导入剖线坐标：{total} 条（{section_coordinate_source}）")
    relic_location_map, relic_location_source = read_relic_location_map(drawings_dir)
    if relic_location_map:
        tables["遗迹位置映射"] = [{"遗迹编号": key, "遗址位置": value} for key, value in relic_location_map.items()]
        if not tables.get("文物范围遗迹统计"):
            tables["文物范围遗迹统计"] = [
                {"遗址位置": value, "遗迹编号": key, "备注": ""}
                for key, value in sorted(relic_location_map.items(), key=lambda item: relic_sort_key(item[0]))
            ]
        notes.append(f"已从制图成果表格导入遗迹内外位置：{len(relic_location_map)} 条（{relic_location_source}）")
    relic_unit_map, relic_unit_source = read_relic_unit_map(drawings_dir)
    if relic_unit_map:
        tables["遗迹单元映射"] = [{"遗迹编号": key, "勘探单元": value} for key, value in relic_unit_map.items()]
        notes.append(f"已从制图成果表格导入遗迹所属勘探单元：{len(relic_unit_map)} 条（{relic_unit_source}）")
    relic_coordinates, relic_coordinate_source = read_relic_coordinate_table(drawings_dir)
    if relic_coordinates and not tables.get("遗迹坐标"):
        tables["遗迹坐标"] = relic_coordinates
        notes.append(f"已从制图成果表格导入遗迹坐标：{len(relic_coordinates)} 条（{relic_coordinate_source}）")
    return notes


def fill_coordinate_basepoint(fields: dict[str, str], tables: dict[str, list[dict[str, str]]], notes: list[str]) -> None:
    if fields.get("坐标基点X") and fields.get("坐标基点Y"):
        return
    for record in tables.get("红线坐标", []):
        corner = clean(record.get("角点"))
        if "西南" not in corner:
            continue
        x_value = clean(record.get("X坐标"))
        y_value = clean(record.get("Y坐标"))
        if not x_value or not y_value:
            continue
        if not fields.get("坐标基点X"):
            fields["坐标基点X"] = x_value
        if not fields.get("坐标基点Y"):
            fields["坐标基点Y"] = y_value
        notes.append(f"坐标基点=西南角（X={fields['坐标基点X']}，Y={fields['坐标基点Y']}，由四至范围坐标导入）")
        return


def import_site_record_tables(fields: dict[str, str], tables: dict[str, list[dict[str, str]]], photos_dir: Path | None) -> list[str]:
    notes: list[str] = []
    site_fields, site_source = read_site_record_fields(photos_dir)
    filled_fields = []
    filled_partition_long_fields = []
    for key, value in site_fields.items():
        if value and not fields.get(key):
            fields[key] = value
            filled_fields.append(key)
            if key in {"勘探分区原因", "勘探分区逐项说明"}:
                filled_partition_long_fields.append(key)
    if filled_fields:
        notes.append(f"已从项目现场记录导入基础/分区字段：{len(filled_fields)} 项（{site_source}）")
    if filled_partition_long_fields:
        notes.append("已从项目现场记录导入分区长文本：" + "、".join(filled_partition_long_fields))
    if has_partition_field(fields) and not (fields.get("勘探分区原因") and fields.get("勘探分区逐项说明")):
        missing = [key for key in ["勘探分区原因", "勘探分区逐项说明"] if not fields.get(key)]
        notes.append(
            "有分区长文本待复核："
            + "、".join(missing)
            + " 未从现场记录中识别到；生成稿会清空对应占位符，需人工补充或修正现场记录。"
        )
    section_records, section_source = read_section_descriptions_from_site_record(photos_dir)
    if section_records:
        existing_sections = tables.get("剖线地层堆积", [])
        had_existing_sections = bool(existing_sections)
        imported_by_key = {section_number_key(record.get("剖线编号", "")): record for record in section_records}
        filled_sections = 0
        for record in existing_sections:
            if record.get("剖线地层描述"):
                continue
            imported = imported_by_key.get(section_number_key(record.get("剖线编号", "")))
            if imported and imported.get("剖线地层描述"):
                record["剖线地层描述"] = imported["剖线地层描述"]
                filled_sections += 1
        existing_keys = {section_number_key(record.get("剖线编号", "")) for record in existing_sections}
        appended_sections = [
            record for record in section_records
            if section_number_key(record.get("剖线编号", "")) not in existing_keys
        ]
        if appended_sections:
            existing_sections.extend(appended_sections)
            filled_sections += len(appended_sections)
        if not had_existing_sections:
            tables["剖线地层堆积"] = section_records
            filled_sections = len(section_records)
        if filled_sections:
            notes.append(f"已从项目现场记录导入剖线地层描述：{filled_sections} 条（{section_source}）")
    standard_descriptions, standard_source = read_standard_descriptions_from_site_record(photos_dir)
    filled = 0
    for record in tables.get("标准孔", []):
        if record.get("标准孔地层描述"):
            continue
        description = None
        for alias in tk_code_aliases(record.get("探孔编号", "")) or tk_code_aliases(standard_hole_code(record)):
            description = standard_descriptions.get(alias)
            if description:
                break
        if description:
            record["标准孔地层描述"] = description
            filled += 1
    if filled:
        notes.append(f"已从项目现场记录导入标准孔地层描述：{filled} 条（{standard_source}）")
    unit_map = {row.get("遗迹编号", ""): row.get("勘探单元", "") for row in tables.get("遗迹单元映射", [])}
    location_map = {row.get("遗迹编号", ""): row.get("遗址位置", "") for row in tables.get("遗迹位置映射", [])}
    relic_records, relic_source = read_relic_descriptions_from_site_record(photos_dir, unit_map, location_map)
    if relic_records and not (tables.get("遗迹记录") or tables.get("遗迹")):
        tables["遗迹记录"] = relic_records
        notes.append(f"已从项目现场记录导入遗迹描述：{len(relic_records)} 条（{relic_source}）")
    elif tables.get("遗迹记录") and not tables.get("遗迹"):
        tables["遗迹"] = tables["遗迹记录"]
    return notes


def has_partition_field(fields: dict[str, str]) -> bool:
    partition_value = clean(fields.get("是否存在勘探分区"))
    partition_count = first_number(fields.get("勘探分区数量", ""))
    return partition_value in {"有", "是", "存在", "yes", "YES"} or bool(partition_count)


def refresh_section_derived_fields(fields: dict[str, str], tables: dict[str, list[dict[str, str]]], notes: list[str]) -> None:
    unit_records = tables.get("勘探单元", [])
    if unit_records:
        units = sorted(
            {record.get("勘探单元", "") for record in unit_records if record.get("勘探单元")},
            key=natural_unit_sort_key,
        )
        if units:
            if not fields.get("勘探单元数量"):
                fields["勘探单元数量"] = str(len(units))
            fields["勘探单元编号范围"] = abbreviate_unit_numbers(units)
            notes.append(f"勘探单元数量={fields['勘探单元数量']}，编号={fields['勘探单元编号范围']}（由勘探单元坐标导入）")
    section_records = tables.get("剖线地层堆积", [])
    if section_records:
        normalize_sequential_section_records(section_records)
        for idx, record in enumerate(section_records, start=1):
            record.setdefault("序号", str(idx))
            record.setdefault("X坐标", "")
            record.setdefault("Y坐标", "")
            record.setdefault("高程", "")
        if not fields.get("剖线数量"):
            fields["剖线数量"] = str(len(section_records))
    if fields.get("剖线数量") and not fields.get("剖面数量"):
        fields["剖面数量"] = fields["剖线数量"]
    standard_records = tables.get("标准孔", [])
    if standard_records:
        for idx, record in enumerate(standard_records, start=1):
            record.setdefault("序号", str(idx))
        fields["标准孔数量"] = str(len(standard_records))
    relic_records = tables.get("遗迹记录") or tables.get("遗迹", [])
    if relic_records:
        fields["遗迹数量"] = str(len(relic_records))
        if not fields.get("遗迹结论"):
            fields["遗迹结论"] = f"发现遗迹{len(relic_records)}处"
    if not fields.get("勘探成果综合结论"):
        conclusion = build_conclusion(section_records, tables.get("标准孔", []), fields)
        if conclusion:
            fields["勘探成果综合结论"] = conclusion
            notes.append("已根据外部剖线地层描述生成结论占位内容。")


def text_nodes(elem: ET.Element) -> list[ET.Element]:
    return [node for node in elem.iter(qn("w:t"))]


def element_text(elem: ET.Element) -> str:
    pieces: list[str] = []
    for node in elem.iter():
        if node.tag == qn("w:t") and node.text:
            pieces.append(node.text)
        elif node.tag == qn("w:tab"):
            pieces.append(" ")
    return "".join(pieces)


def set_element_text(elem: ET.Element, value: str) -> None:
    nodes = text_nodes(elem)
    if not nodes:
        return
    nodes[0].text = value
    if value.startswith(" ") or value.endswith(" "):
        nodes[0].set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    for node in nodes[1:]:
        node.text = ""


def set_paragraph_lines(elem: ET.Element, lines: list[str]) -> None:
    nodes = text_nodes(elem)
    if not nodes:
        set_element_text(elem, "\n".join(lines))
        return
    first = nodes[0]
    parent = find_element_parent(elem, first)
    if parent is None:
        set_element_text(elem, "\n".join(lines))
        return
    first.text = lines[0] if lines else ""
    for node in nodes[1:]:
        node.text = ""
    insert_at = list(parent).index(first) + 1
    for line in lines[1:]:
        br = ET.Element(qn("w:br"))
        parent.insert(insert_at, br)
        insert_at += 1
        text_node = ET.Element(qn("w:t"))
        text_node.text = line
        parent.insert(insert_at, text_node)
        insert_at += 1


def set_paragraph_font(elem: ET.Element, font_name: str, half_points: str) -> None:
    for run in elem.findall(".//w:r", NS):
        rpr = run.find("w:rPr", NS)
        if rpr is None:
            rpr = ET.Element(qn("w:rPr"))
            run.insert(0, rpr)
        rfonts = rpr.find("w:rFonts", NS)
        if rfonts is None:
            rfonts = ET.SubElement(rpr, qn("w:rFonts"))
        for attr in ("ascii", "hAnsi", "eastAsia", "cs"):
            rfonts.set(qn(f"w:{attr}"), font_name)
        sz = rpr.find("w:sz", NS)
        if sz is None:
            sz = ET.SubElement(rpr, qn("w:sz"))
        sz.set(qn("w:val"), half_points)
        sz_cs = rpr.find("w:szCs", NS)
        if sz_cs is None:
            sz_cs = ET.SubElement(rpr, qn("w:szCs"))
        sz_cs.set(qn("w:val"), half_points)


def find_element_parent(root: ET.Element, target: ET.Element) -> ET.Element | None:
    for parent in root.iter():
        if target in list(parent):
            return parent
    return None


def insert_paragraphs_after(root: ET.Element, template_para: ET.Element, texts: list[str]) -> int:
    parent = find_element_parent(root, template_para)
    if parent is None:
        return 0
    idx = list(parent).index(template_para)
    inserted = 0
    for text in texts:
        if not clean(text):
            continue
        clone = copy.deepcopy(template_para)
        set_element_text(clone, clean(text))
        parent.insert(idx + 1 + inserted, clone)
        inserted += 1
    parent.remove(template_para)
    return inserted


def split_partition_detail_text(text: str) -> list[str]:
    text = clean(text)
    if not text:
        return []
    explicit = [item.strip() for item in re.split(r"[\r\n]+", text) if item.strip()]
    if len(explicit) > 1:
        return explicit
    parts = re.split(r"(?=(?:[A-ZＡ-Ｚ]|[一二三四五六七八九十]+|第[一二三四五六七八九十]+)[区區]\s*[:：])", text)
    return [part.strip() for part in parts if part.strip()]


def set_cell_text(cell: ET.Element, value: str) -> None:
    nodes = text_nodes(cell)
    if nodes:
        set_element_text(cell, value)
        return
    para = cell.find("w:p", NS)
    if para is None:
        para = ET.SubElement(cell, qn("w:p"))
    run = ET.SubElement(para, qn("w:r"))
    text = ET.SubElement(run, qn("w:t"))
    text.text = value


def set_cell_lines(cell: ET.Element, lines: list[str]) -> None:
    paragraphs = cell.findall("w:p", NS)
    para = paragraphs[0] if paragraphs else ET.SubElement(cell, qn("w:p"))
    for extra in paragraphs[1:]:
        cell.remove(extra)
    ppr = para.find("w:pPr", NS)
    for child in list(para):
        if child is not ppr:
            para.remove(child)
    for idx, line in enumerate(lines):
        if idx:
            br_run = ET.SubElement(para, qn("w:r"))
            ET.SubElement(br_run, qn("w:br"))
        run = ET.SubElement(para, qn("w:r"))
        text = ET.SubElement(run, qn("w:t"))
        text.text = line


def clone_paragraph_with_text(template_para: ET.Element, text: str) -> ET.Element:
    para = copy.deepcopy(template_para)
    set_element_text(para, text)
    return para


def ensure_child(parent: ET.Element, tag: str, *, first: bool = False) -> ET.Element:
    child = parent.find(tag, NS)
    if child is None:
        child = ET.Element(qn(tag))
        if first:
            parent.insert(0, child)
        else:
            parent.append(child)
    return child


def ensure_table_all_borders(tbl: ET.Element) -> None:
    def set_border(node: ET.Element) -> None:
        node.set(qn("w:val"), "single")
        node.set(qn("w:sz"), "4")
        node.set(qn("w:space"), "0")
        node.set(qn("w:color"), "000000")

    tbl_pr = ensure_child(tbl, "w:tblPr", first=True)
    borders = tbl_pr.find("w:tblBorders", NS)
    if borders is None:
        borders = ET.Element(qn("w:tblBorders"))
        tbl_pr.append(borders)
    for edge in ["top", "left", "bottom", "right", "insideH", "insideV"]:
        tag = f"w:{edge}"
        node = borders.find(tag, NS)
        if node is None:
            node = ET.SubElement(borders, qn(tag))
        set_border(node)
    for cell in tbl.findall(".//w:tc", NS):
        tc_pr = ensure_child(cell, "w:tcPr", first=True)
        tc_borders = tc_pr.find("w:tcBorders", NS)
        if tc_borders is None:
            tc_borders = ET.Element(qn("w:tcBorders"))
            tc_pr.append(tc_borders)
        for edge in ["top", "left", "bottom", "right"]:
            node = tc_borders.find(f"w:{edge}", NS)
            if node is None:
                node = ET.SubElement(tc_borders, qn(f"w:{edge}"))
            set_border(node)


def normalize_all_table_borders(root: ET.Element) -> int:
    tables = root.findall(".//w:tbl", NS)
    for tbl in tables:
        ensure_table_all_borders(tbl)
    return len(tables)


def set_cell_vertical_merge(cell: ET.Element, value: str | None) -> None:
    tc_pr = ensure_child(cell, "w:tcPr", first=True)
    v_merges = tc_pr.findall("w:vMerge", NS)
    v_merge = v_merges[0] if v_merges else ET.SubElement(tc_pr, qn("w:vMerge"))
    for extra in v_merges[1:]:
        tc_pr.remove(extra)
    if value is None:
        v_merge.attrib.pop(qn("w:val"), None)
    else:
        v_merge.set(qn("w:val"), value)


def clear_cell_vertical_merge(cell: ET.Element) -> None:
    tc_pr = cell.find("w:tcPr", NS)
    if tc_pr is None:
        return
    for v_merge in list(tc_pr.findall("w:vMerge", NS)):
        tc_pr.remove(v_merge)


def merge_unit_number_cells(rows: list[ET.Element], records: list[dict[str, str]]) -> None:
    previous_unit = ""
    active_unit = ""
    for row, record in zip(rows, records):
        cells = row.findall("w:tc", NS)
        if not cells:
            continue
        unit = clean(record.get("勘探单元")) or clean(record.get("勘探单元显示")) or active_unit
        display_unit = clean(record.get("勘探单元显示")) or unit
        if unit and unit != previous_unit:
            set_cell_vertical_merge(cells[0], "restart")
            set_cell_text(cells[0], display_unit)
            previous_unit = unit
            active_unit = unit
        elif unit:
            set_cell_vertical_merge(cells[0], None)
            set_cell_text(cells[0], "")
            active_unit = unit
        else:
            clear_cell_vertical_merge(cells[0])


def render_text(
    template: str,
    fields: dict[str, str],
    row: dict[str, str] | None = None,
    *,
    preserve_image_tokens: bool = False,
) -> str:
    values = dict(fields)
    if row:
        values.update(row)
    for key, value in list(values.items()):
        if "高程" in clean(key):
            values[key] = format_elevation(value)
    text = template
    image_tokens: dict[str, str] = {}
    protected_tokens: dict[str, str] = {}
    for idx, token in enumerate(re.findall(r"\{\{(?:公司附件:[^{}]+|人员附件:[^{}]+|公司印章)}}", text)):
        marker = f"__PROTECTED_ATTACHMENT_TOKEN_{idx}__"
        protected_tokens[marker] = token
        text = text.replace(token, marker, 1)
    if preserve_image_tokens:
        for idx, token in enumerate(re.findall(r"\{\{图:[^{}]+}}", text)):
            marker = f"__IMAGE_TOKEN_{idx}__"
            image_tokens[marker] = token
            text = text.replace(token, marker, 1)
    if not values.get("项目建设内容"):
        text = re.sub(r"项目建设内容为\{\{项目建设内容}}。?", "", text)
    if not values.get("不可勘探原因"):
        text = re.sub(r"因项目用地范围内部分区域存在\{\{不可勘探原因}}，暂不具备勘探条件，", "", text)
    if not values.get("项目地理坐标"):
        text = re.sub(r"，?地理坐标：\{\{项目地理坐标}}。?", "。", text)
    text = text.replace("{{自动图号}}", "")
    text = re.sub(r"\{\{可选图:[^{}]+}}", "", text)
    for key, value in values.items():
        text = text.replace("{{" + key + "}}", report_field_value(key, value))
    text = re.sub(
        r"\{\{(?!IF_|ENDIF|#|/|图:)([^{}]+)}}",
        lambda match: report_error_marker(MATCH_ERROR_TEXT, f"未找到字段：{clean(match.group(1))}"),
        text,
    )
    text = re.sub(r"\{\{[^{}]+}}", "", text)
    for marker, token in image_tokens.items():
        text = text.replace(marker, token)
    for marker, token in protected_tokens.items():
        text = text.replace(marker, token)
    text = normalize_generated_punctuation(text)
    return text


def normalize_generated_punctuation(text: str) -> str:
    text = re.sub(r"。{2,}", "。", text)
    text = re.sub(r"，{2,}", "，", text)
    text = re.sub(r"；{2,}", "；", text)
    text = re.sub(r"、{2,}", "、", text)
    text = re.sub(r"：{2,}", "：", text)
    text = re.sub(r"！？", "！", text)
    text = re.sub(r"？！", "？", text)
    text = re.sub(r"，。", "。", text)
    text = re.sub(r"；。", "。", text)
    text = re.sub(r"、。", "。", text)
    return text


def clear_run_formatting_artifacts(rpr: ET.Element | None) -> None:
    if rpr is None:
        return
    for tag in ["w:spacing", "w:kern", "w:w", "w:scale", "w:fitText"]:
        for node in list(rpr.findall(tag, NS)):
            rpr.remove(node)


def ensure_rpr(run: ET.Element) -> ET.Element:
    rpr = run.find("w:rPr", NS)
    if rpr is None:
        rpr = ET.Element(qn("w:rPr"))
        run.insert(0, rpr)
    return rpr


def set_run_style(run: ET.Element, *, bold: bool, size_half_points: str) -> None:
    rpr = ensure_rpr(run)
    clear_run_formatting_artifacts(rpr)
    for tag in ["w:b", "w:bCs"]:
        for node in list(rpr.findall(tag, NS)):
            rpr.remove(node)
    if bold:
        ET.SubElement(rpr, qn("w:b"))
        ET.SubElement(rpr, qn("w:bCs"))
    for tag in ["w:sz", "w:szCs"]:
        node = rpr.find(tag, NS)
        if node is None:
            node = ET.SubElement(rpr, qn(tag))
        node.set(qn("w:val"), size_half_points)


def set_run_font(run: ET.Element, *, font: str, size_half_points: str) -> None:
    rpr = ensure_rpr(run)
    clear_run_formatting_artifacts(rpr)
    fonts = rpr.find("w:rFonts", NS)
    if fonts is None:
        fonts = ET.SubElement(rpr, qn("w:rFonts"))
    for attr in ["w:ascii", "w:hAnsi", "w:eastAsia", "w:cs"]:
        fonts.set(qn(attr), font)
    for tag in ["w:sz", "w:szCs"]:
        node = rpr.find(tag, NS)
        if node is None:
            node = ET.SubElement(rpr, qn(tag))
        node.set(qn("w:val"), size_half_points)


ERROR_MARKER_PATTERN = re.compile(
    rf"【(?:{MATCH_ERROR_TEXT}|{IMAGE_INSERT_ERROR_TEXT}|{MISSING_FIELD_TEXT})：[^】]+】"
    rf"|(?:{MATCH_ERROR_TEXT}|{IMAGE_INSERT_ERROR_TEXT}|{MISSING_FIELD_TEXT})"
)


def make_text_run_like(source: ET.Element, text: str, *, marker: bool) -> ET.Element:
    run = ET.Element(qn("w:r"))
    source_rpr = source.find("w:rPr", NS)
    if source_rpr is not None:
        run.append(copy.deepcopy(source_rpr))
    text_node = ET.SubElement(run, qn("w:t"))
    if text[:1].isspace() or text[-1:].isspace():
        text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text_node.text = text
    if marker:
        rpr = ensure_rpr(run)
        clear_run_formatting_artifacts(rpr)
        for tag in ["w:b", "w:bCs"]:
            if rpr.find(tag, NS) is None:
                ET.SubElement(rpr, qn(tag))
        color = rpr.find("w:color", NS)
        if color is None:
            color = ET.SubElement(rpr, qn("w:color"))
        color.set(qn("w:val"), "000000")
        highlight = rpr.find("w:highlight", NS)
        if highlight is None:
            highlight = ET.SubElement(rpr, qn("w:highlight"))
        highlight.set(qn("w:val"), "red")
    return run


def apply_error_marker_style(root: ET.Element) -> int:
    """Style only the bracketed report-visible error marker, not the whole sentence."""
    styled = 0
    for parent in root.iter():
        children = list(parent)
        for child in children:
            if child.tag != qn("w:r"):
                continue
            text_nodes = child.findall("w:t", NS)
            text = "".join(node.text or "" for node in text_nodes)
            if not text or not ERROR_MARKER_PATTERN.search(text):
                continue
            new_runs: list[ET.Element] = []
            pos = 0
            for match in ERROR_MARKER_PATTERN.finditer(text):
                if match.start() > pos:
                    new_runs.append(make_text_run_like(child, text[pos:match.start()], marker=False))
                new_runs.append(make_text_run_like(child, match.group(0), marker=True))
                styled += 1
                pos = match.end()
            if pos < len(text):
                new_runs.append(make_text_run_like(child, text[pos:], marker=False))
            insert_at = list(parent).index(child)
            parent.remove(child)
            for offset, new_run in enumerate(new_runs):
                parent.insert(insert_at + offset, new_run)
    return styled


def set_table_text_font(tbl: ET.Element, *, font: str, size_half_points: str) -> None:
    for run in tbl.findall(".//w:r", NS):
        if run.find(".//w:t", NS) is not None:
            set_run_font(run, font=font, size_half_points=size_half_points)


def set_table_row_height(tbl: ET.Element, *, height_cm: float) -> None:
    height_twips = str(round(height_cm * 567))
    for row in tbl.findall("w:tr", NS):
        tr_pr = ensure_child(row, "w:trPr", first=True)
        heights = tr_pr.findall("w:trHeight", NS)
        if heights:
            height_node = heights[0]
            for extra in heights[1:]:
                tr_pr.remove(extra)
        else:
            height_node = ET.SubElement(tr_pr, qn("w:trHeight"))
        height_node.set(qn("w:val"), height_twips)
        height_node.set(qn("w:hRule"), "exact")


def normalize_appendix_table_format(root: ET.Element) -> int:
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    in_appendix_tables = False
    changed = 0
    for block in list(body):
        if block.tag == qn("w:p"):
            text = element_text(block).strip()
            if re.match(r"^附表[一二三四五六七八九十\d]", text):
                in_appendix_tables = True
            elif re.match(r"^附图[一二三四五六七八九十\d]", text):
                in_appendix_tables = False
        elif block.tag == qn("w:tbl") and in_appendix_tables:
            set_table_text_font(block, font="宋体", size_half_points="21")
            set_table_row_height(block, height_cm=APPENDIX_TABLE_ROW_HEIGHT_CM)
            changed += 1
    return changed


def make_text_run(text: str, template_run: ET.Element | None, *, bold: bool, size_half_points: str) -> ET.Element:
    run = copy.deepcopy(template_run) if template_run is not None else ET.Element(qn("w:r"))
    for child in list(run):
        if child.tag != qn("w:rPr"):
            run.remove(child)
    set_run_style(run, bold=bold, size_half_points=size_half_points)
    t = ET.SubElement(run, qn("w:t"))
    t.text = text
    if text.startswith(" ") or text.endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    return run


def normalize_project_overview_format(root: ET.Element, fields: dict[str, str] | None = None) -> int:
    """Format project-overview label/value runs without changing template-owned text.

    Field replacement is intentionally limited to explicit {{字段名}} tokens in
    apply_plain_fields(). This pass only preserves the visible text after the
    label colon and applies the required run styling.
    """
    changed = 0
    for para in root.findall(".//w:p", NS):
        text = element_text(para).strip()
        for label in PROJECT_OVERVIEW_LABELS:
            prefix = f"{label}："
            if not text.startswith(prefix):
                continue
            value = text[len(prefix) :]
            runs = para.findall("w:r", NS)
            template_run = runs[0] if runs else None
            for run in runs:
                para.remove(run)
            para.append(make_text_run(prefix, template_run, bold=True, size_half_points="32"))
            para.append(make_text_run(clean(value), template_run, bold=False, size_half_points="30"))
            changed += 1
            break
    return changed


def apply_conditionals(text: str, fields: dict[str, str]) -> str:
    kind = fields.get("文物概况类型", "")
    if kind == "文物审查意见待补" and (
        "{{IF_有文物审查意见}}" in text
        or "{{IF_无回函_未发现文物}}" in text
        or "{{IF_无回函_涉及文物}}" in text
    ):
        return fields.get("文物概况待补说明", "")
    has_limited_area = fields.get("是否存在不可勘探区域", "") == "是" or bool(fields.get("不可勘探原因", ""))
    has_relic = "未发现" not in fields.get("遗迹结论", "") and "无" not in fields.get("遗迹结论", "")
    enabled = {
        "IF_有文物审查意见": kind == "有文物审查意见",
        "IF_无回函_未发现文物": kind == "无回函且未发现文物",
        "IF_无回函_涉及文物": kind == "无回函但涉及文物",
        "IF_文物审查意见待补": kind == "文物审查意见待补",
        "IF_存在不可勘探区域": has_limited_area,
        "IF_无遗迹": not has_relic,
        "IF_有遗迹": has_relic,
    }
    for name, keep in enabled.items():
        pattern = re.compile(r"\{\{" + re.escape(name) + r"}}(.*?)\{\{ENDIF}}", re.S)
        text = pattern.sub(lambda match: match.group(1) if keep else "", text)
    return text


def apply_plain_fields(root: ET.Element, fields: dict[str, str]) -> int:
    changed = 0
    for para in list(root.findall(".//w:p", NS)):
        original = element_text(para)
        if "{{勘探分区逐项说明}}" not in original and "{{分区地层逐区总结}}" not in original:
            continue
        key = "勘探分区逐项说明" if "{{勘探分区逐项说明}}" in original else "分区地层逐区总结"
        lines = split_partition_detail_text(fields.get(key, ""))
        if len(lines) > 1 and original.strip() == "{{" + key + "}}":
            changed += insert_paragraphs_after(root, para, lines)
    for elem in root.findall(".//w:p", NS) + root.findall(".//w:tc", NS):
        original = element_text(elem)
        if "{{" not in original:
            continue
        new_text = apply_conditionals(original, fields)
        new_text = render_text(new_text, fields)
        if new_text != original:
            set_element_text(elem, new_text)
            changed += 1
    return changed


def rewrite_conclusion_section(root: ET.Element, fields: dict[str, str]) -> int:
    paragraphs = build_conclusion_paragraphs(fields)
    if not paragraphs:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    start_idx = end_idx = None
    for idx, block in enumerate(blocks):
        if block.tag != qn("w:p"):
            continue
        text = element_text(block).strip()
        if text.startswith("五、结论"):
            start_idx = idx
            break
    if start_idx is None:
        return 0
    for idx in range(start_idx + 1, len(blocks)):
        block = blocks[idx]
        if block.tag == qn("w:p") and element_text(block).strip().startswith("附表一"):
            end_idx = idx
            break
    if end_idx is None or end_idx <= start_idx + 1:
        return 0

    old_content = blocks[start_idx + 1 : end_idx]
    nonempty_paragraphs = [
        block for block in old_content if block.tag == qn("w:p") and element_text(block).strip()
    ]
    if len(nonempty_paragraphs) >= 3:
        changed = 0
        replacements = {
            "保护好现场。及时报告": "保护好现场，及时报告",
            "文物主管部门问意后": "文物主管部门同意后",
            "文物及遭迹现象": "文物及遗迹现象",
            "自上而下依次为:": "自上而下依次为：",
        }
        for block in nonempty_paragraphs:
            text = element_text(block)
            new_text = text
            for old, new in replacements.items():
                new_text = new_text.replace(old, new)
            if new_text != text:
                set_element_text(block, new_text)
                changed += 1
        return changed

    para_templates = [block for block in old_content if block.tag == qn("w:p") and element_text(block).strip()]
    template = para_templates[0] if para_templates else blocks[start_idx]
    for block in old_content:
        if block in list(body):
            body.remove(block)

    insert_at = list(body).index(blocks[start_idx]) + 1
    for paragraph in paragraphs:
        clone = copy.deepcopy(template)
        set_element_text(clone, paragraph)
        body.insert(insert_at, clone)
        insert_at += 1
    return len(paragraphs)


def split_history_overview_text(text: str) -> list[str]:
    text = clean(text)
    if not text:
        return []
    explicit = [item.strip() for item in re.split(r"[\r\n]+", text) if item.strip()]
    if len(explicit) > 1:
        return explicit
    if len(text) < 320:
        return [text]

    sentences = [item for item in re.split(r"(?<=。)", text) if item]
    if len(sentences) <= 1:
        return [text]
    split_markers = (
        "魏晋",
        "三国",
        "隋唐",
        "宋",
        "辽",
        "金",
        "元",
        "明清",
        "清代",
        "近现代",
        "新中国",
        "额济纳旗境内",
        "由于区域",
    )
    paragraphs: list[str] = []
    current = ""
    for sentence in sentences:
        stripped = sentence.strip()
        should_break = (
            bool(current)
            and (
                len(current) >= 220
                or any(stripped.startswith(marker) for marker in split_markers)
            )
        )
        if should_break:
            paragraphs.append(current)
            current = stripped
        else:
            current += stripped
    if current:
        paragraphs.append(current)

    merged: list[str] = []
    for paragraph in paragraphs:
        if merged and len(paragraph) < 70 and len(merged[-1]) < 180:
            merged[-1] += paragraph
        else:
            merged.append(paragraph)
    return merged


def split_region_overview_text(text: str, *, history: bool = False) -> list[str]:
    if history:
        return split_history_overview_text(text)
    text = clean(text)
    if not text:
        return []
    explicit = [item.strip() for item in re.split(r"[\r\n]+", text) if item.strip()]
    if len(explicit) > 1:
        return explicit
    if len(text) < 360:
        return [text]
    sentences = [item.strip() for item in re.split(r"(?<=。)", text) if item.strip()]
    if len(sentences) <= 1:
        return [text]
    parts: list[str] = []
    current = ""
    for sentence in sentences:
        if current and len(current) >= 180:
            parts.append(current)
            current = sentence
        else:
            current += sentence
    if current:
        parts.append(current)
    return parts


def split_region_overview_paragraphs(root: ET.Element, fields: dict[str, str]) -> int:
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    changed = 0
    for key in REGION_OVERVIEW_KEYS:
        value = clean(fields.get(key))
        if not value:
            continue
        parts = split_region_overview_text(value, history=key == "项目所在地旗县历史沿革")
        if len(parts) <= 1:
            continue
        blocks = list(body)
        for idx, block in enumerate(blocks):
            if block.tag != qn("w:p") or element_text(block).strip() != value:
                continue
            set_element_text(block, parts[0])
            insert_at = idx + 1
            for part in parts[1:]:
                body.insert(insert_at, clone_paragraph_with_text(block, part))
                insert_at += 1
            changed += len(parts)
            break
    return changed


def split_history_overview_paragraphs(root: ET.Element, fields: dict[str, str]) -> int:
    history = clean(fields.get("项目所在地旗县历史沿革"))
    parts = split_history_overview_text(history)
    if len(parts) <= 1:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    changed = 0
    for idx, block in enumerate(blocks):
        if block.tag != qn("w:p") or element_text(block).strip() != history:
            continue
        set_element_text(block, parts[0])
        insert_at = idx + 1
        for part in parts[1:]:
            body.insert(insert_at, clone_paragraph_with_text(block, part))
            insert_at += 1
        changed += len(parts)
        break
    return changed


def body_blocks(root: ET.Element) -> tuple[ET.Element | None, list[ET.Element]]:
    body = root.find("w:body", NS)
    return body, list(body) if body is not None else []


def remove_element(root: ET.Element, target: ET.Element) -> bool:
    for parent in root.iter():
        children = list(parent)
        if target in children:
            parent.remove(target)
            return True
    return False


def replace_repeating_block(root: ET.Element, marker: str, records: list[dict[str, str]], fields: dict[str, str]) -> int:
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    start_token = "{{#" + marker + "}}"
    end_token = "{{/" + marker + "}}"
    start_idx = end_idx = None
    for idx, block in enumerate(blocks):
        if block.tag == qn("w:p") and start_token in element_text(block):
            start_idx = idx
        if start_idx is not None and block.tag == qn("w:p") and end_token in element_text(block):
            end_idx = idx
            break
    if start_idx is None or end_idx is None:
        return 0
    template_blocks = blocks[start_idx + 1 : end_idx]
    insertion_index = start_idx
    for block in blocks[start_idx : end_idx + 1]:
        body.remove(block)
    for record in records:
        for block in template_blocks:
            clone = copy.deepcopy(block)
            for elem in clone.findall(".//w:p", NS) + clone.findall(".//w:tc", NS):
                text = element_text(elem)
                if "{{" in text:
                    set_element_text(elem, render_text(text, fields, record, preserve_image_tokens=True))
            body.insert(insertion_index, clone)
            insertion_index += 1
    return len(records)


def expand_section_description_placeholders(
    root: ET.Element, records: list[dict[str, str]], fields: dict[str, str]
) -> int:
    if not records:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    blocks = list(body)
    for idx, block in enumerate(blocks[:-1]):
        if block.tag != qn("w:p"):
            continue
        title_text = element_text(block)
        if not re.search(r"剖[线面]\s*\{\{剖[线面]编号}}\s*地层堆积情况", title_text):
            continue
        desc_block = blocks[idx + 1]
        desc_text = element_text(desc_block)
        if desc_block.tag != qn("w:p") or not re.search(r"\{\{剖[线面]地层描述}}", desc_text):
            continue

        insertion_index = list(body).index(block)
        body.remove(block)
        body.remove(desc_block)
        for record in records:
            section_label = record.get("剖线编号") or record.get("剖面编号") or ""
            section_description = record.get("剖线地层描述") or record.get("剖面地层描述") or ""
            title_clone = copy.deepcopy(block)
            desc_clone = copy.deepcopy(desc_block)
            replacement_title = re.sub(r"\{\{剖[线面]编号}}", section_label, title_text)
            replacement_desc = re.sub(r"\{\{剖[线面]地层描述}}", section_description, desc_text)
            set_element_text(title_clone, render_text(replacement_title, fields, record))
            set_element_text(desc_clone, render_text(replacement_desc, fields, record))
            body.insert(insertion_index, title_clone)
            body.insert(insertion_index + 1, desc_clone)
            body.insert(insertion_index + 2, make_page_break_paragraph())
            insertion_index += 3
        return len(records)
    return 0


def expand_standard_detail_tables(root: ET.Element, records: list[dict[str, str]], fields: dict[str, str]) -> int:
    if not records:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    start_idx = end_idx = None
    for idx, block in enumerate(blocks):
        if block.tag == qn("w:p") and "{{#标准孔详情}}" in element_text(block):
            start_idx = idx
        if start_idx is not None and block.tag == qn("w:p") and "{{/标准孔详情}}" in element_text(block):
            end_idx = idx
            break
    if start_idx is None or end_idx is None:
        return 0

    template_tbl_idx = None
    for idx in range(start_idx - 1, max(-1, start_idx - 8), -1):
        block = blocks[idx]
        if block.tag == qn("w:tbl") and "勘探编号" in element_text(block) and "土样照" in element_text(block):
            template_tbl_idx = idx
            break
    if template_tbl_idx is None:
        return replace_repeating_block(root, "标准孔详情", records, fields)

    heading_idx = template_tbl_idx - 1 if template_tbl_idx > 0 and blocks[template_tbl_idx - 1].tag == qn("w:p") else template_tbl_idx
    heading_template = copy.deepcopy(blocks[heading_idx]) if blocks[heading_idx].tag == qn("w:p") else ET.Element(qn("w:p"))
    table_template = copy.deepcopy(blocks[template_tbl_idx])
    remove_start = heading_idx
    remove_end = end_idx
    insertion_index = remove_start
    for block in blocks[remove_start : remove_end + 1]:
        body.remove(block)

    for record in records:
        probe_number = standard_probe_number(record)
        heading = copy.deepcopy(heading_template)
        set_element_text(heading, f"（{record.get('序号', '')}）勘探单元{record.get('勘探单元', '')}")
        table = copy.deepcopy(table_template)
        rows = table.findall("w:tr", NS)
        values = {
            "勘探编号": probe_number,
            "探孔坐标": record.get("探孔坐标")
            or f"X：{record.get('X坐标', '')}，Y：{record.get('Y坐标', '')}，高程：{format_elevation(record.get('高程', ''))}",
            "地层堆积情况": record.get("标准孔地层描述", ""),
            "位置图": f"{{{{图:标准孔位置图:{probe_number}}}}}",
            "土样照": f"{{{{图:标准孔土样照:{probe_number}}}}}",
        }
        for row in rows:
            cells = row.findall("w:tc", NS)
            if len(cells) < 2:
                continue
            label = element_text(cells[0]).strip()
            if label in values:
                if label == "地层堆积情况":
                    set_cell_lines(cells[1], split_standard_description(values[label]))
                else:
                    set_cell_text(cells[1], values[label])
        set_table_text_font(table, font="宋体", size_half_points="24")
        body.insert(insertion_index, heading)
        body.insert(insertion_index + 1, table)
        body.insert(insertion_index + 2, make_page_break_paragraph())
        insertion_index += 3
    return len(records)


def render_relic_description(record: dict[str, str]) -> str:
    text = clean(record.get("遗迹描述"))
    unit = clean(record.get("勘探单元"))
    if unit:
        text = re.sub(r"位于勘探单元（后补）", f"位于勘探单元{unit}", text)
        text = re.sub(r"位于勘探单元\s*[（(]\s*([UＵ]\s*0*\d+)\s*[）)]", r"位于勘探单元\1", text)
        text = re.sub(r"位于勘探单元(?=[,，])", f"位于勘探单元{unit}", text)
    else:
        text = re.sub(r"位于勘探单元\s*[（(]\s*([UＵ]\s*0*\d+)\s*[）)]", r"位于勘探单元\1", text)
    text = text.replace("GPS坐标", "坐标")
    return text


def scope_relic_image_placeholders(text: str, relic_id: str) -> str:
    scoped_keys = {"遗迹土样照", "遗迹现场照", "遗迹平、剖面图"}
    normalized_relic_id = canonical_relic_id(relic_id)

    def replace_token(match: re.Match[str]) -> str:
        key = clean(match.group(1))
        if key in scoped_keys:
            return f"{{{{图:{relic_image_slot_key(key, normalized_relic_id)}}}}}"
        return match.group(0)

    return re.sub(r"\{\{图:([^{}]+)}}", replace_token, text)


def expand_relic_detail_paragraphs(root: ET.Element, records: list[dict[str, str]], fields: dict[str, str]) -> int:
    if not records:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    start_idx = end_idx = next_heading_idx = None
    heading_re = re.compile(rf"^\d+\s*[、.．]\s*(?:灰坑|墓葬|遗迹|房址|窑址|窑坑|活土坑)?\s*{RELIC_ID_RE}", flags=re.I)
    for idx, block in enumerate(blocks):
        if block.tag == qn("w:p") and heading_re.match(element_text(block).strip()):
            start_idx = idx
            break
    if start_idx is None:
        return 0
    for idx in range(start_idx + 1, len(blocks)):
        if blocks[idx].tag == qn("w:p") and heading_re.match(element_text(blocks[idx]).strip()):
            next_heading_idx = idx
            break
    if next_heading_idx is None:
        return 0
    for idx in range(start_idx + 1, len(blocks)):
        text = element_text(blocks[idx]).strip()
        if blocks[idx].tag == qn("w:p") and (text.startswith("五、结论") or text.startswith("附表一")):
            end_idx = idx
            break
    if end_idx is None:
        return 0

    group_templates = [copy.deepcopy(block) for block in blocks[start_idx:next_heading_idx]]

    for block in blocks[start_idx:end_idx]:
        body.remove(block)

    insertion_index = start_idx
    for idx, record in enumerate(records, start=1):
        relic_id = record.get("遗迹编号", "")
        relic_type = record.get("遗迹类型") or relic_type_from_id(relic_id)
        description_done = False
        for template in group_templates:
            block = copy.deepcopy(template)
            text = element_text(block).strip()
            if block.tag == qn("w:p") and heading_re.match(text):
                set_element_text(block, f"{idx}、{relic_type} {relic_id}")
            elif (
                block.tag == qn("w:p")
                and not description_done
                and re.match(rf"^(?:灰坑|墓葬|遗迹|房址|窑址|窑坑|活土坑)?{RELIC_ID_RE}位于", text, flags=re.I)
            ):
                set_element_text(block, render_relic_description(record))
                description_done = True
            elif block.tag == qn("w:p") and "{{图:" in text:
                scoped_text = scope_relic_image_placeholders(text, relic_id)
                if scoped_text != text:
                    set_element_text(block, scoped_text)
            body.insert(insertion_index, block)
            insertion_index += 1
    return len(records)


def fill_placeholder_table(root: ET.Element, title: str, records: list[dict[str, str]], fields: dict[str, str]) -> int:
    if not records:
        return 0
    for tbl in root.findall(".//w:tbl", NS):
        table_text = element_text(tbl)
        if title not in table_text or "{{角点}}" not in table_text:
            continue
        rows = tbl.findall("w:tr", NS)
        template_row = None
        first_placeholder_row_idx = None
        first_placeholder_child_idx = None
        for idx, row in enumerate(rows):
            row_text = element_text(row)
            if "{{角点}}" in row_text and "{{X坐标}}" in row_text and "{{Y坐标}}" in row_text:
                template_row = row
                first_placeholder_row_idx = idx
                first_placeholder_child_idx = list(tbl).index(row)
                break
        if template_row is None or first_placeholder_row_idx is None or first_placeholder_child_idx is None:
            return 0
        for row in rows[first_placeholder_row_idx:]:
            tbl.remove(row)
        insert_at = first_placeholder_child_idx
        inserted_rows: list[ET.Element] = []
        for record in records:
            clone = copy.deepcopy(template_row)
            for cell in clone.findall(".//w:tc", NS):
                cell_text = element_text(cell)
                if cell_text.strip() == "" and record.get("勘探单元显示"):
                    set_cell_text(cell, record["勘探单元显示"])
                elif "{{" in cell_text:
                    rendered = render_text(cell_text, fields, record)
                    set_cell_text(cell, rendered)
            tbl.insert(insert_at, clone)
            inserted_rows.append(clone)
            insert_at += 1
        ensure_table_all_borders(tbl)
        if title == "勘探单元坐标":
            merge_unit_number_cells(inserted_rows, records)
        return len(records)
    return 0


def fill_unit_coordinate_tables(root: ET.Element, records: list[dict[str, str]], fields: dict[str, str]) -> int:
    if not records:
        return 0
    filled_rows = 0
    for tbl in root.findall(".//w:tbl", NS):
        if "勘探单元坐标" not in element_text(tbl):
            continue
        rows = tbl.findall("w:tr", NS)
        template_row = None
        first_data_idx = None
        first_data_child_idx = None
        for idx, row in enumerate(rows):
            row_text = element_text(row)
            cells = row.findall("w:tc", NS)
            cell_values = [element_text(cell).strip() for cell in cells]
            has_placeholder = "{{角点}}" in row_text and "{{X坐标}}" in row_text and "{{Y坐标}}" in row_text
            has_coordinate_row = any(value in {"西南角", "东南角", "东北角", "西北角"} for value in cell_values)
            has_unit_cell = bool(cell_values and re.fullmatch(r"U\d+", cell_values[0]))
            if has_placeholder or has_coordinate_row or has_unit_cell:
                template_row = row
                first_data_idx = idx
                first_data_child_idx = list(tbl).index(row)
                break
        if template_row is None or first_data_idx is None or first_data_child_idx is None:
            continue

        for row in rows[first_data_idx:]:
            tbl.remove(row)

        inserted_rows: list[ET.Element] = []
        insert_at = first_data_child_idx
        for record in records:
            clone = copy.deepcopy(template_row)
            cells = clone.findall("w:tc", NS)
            if len(cells) >= 4:
                set_cell_text(cells[0], record.get("勘探单元显示", ""))
                set_cell_text(cells[1], render_text("{{角点}}", fields, record))
                set_cell_text(cells[2], render_text("{{X坐标}}", fields, record))
                set_cell_text(cells[3], render_text("{{Y坐标}}", fields, record))
            else:
                for cell in clone.findall(".//w:tc", NS):
                    cell_text = element_text(cell)
                    if cell_text.strip() == "" and record.get("勘探单元显示"):
                        set_cell_text(cell, record["勘探单元显示"])
                    elif "{{" in cell_text:
                        set_cell_text(cell, render_text(cell_text, fields, record))
            tbl.insert(insert_at, clone)
            inserted_rows.append(clone)
            insert_at += 1

        ensure_table_all_borders(tbl)
        merge_unit_number_cells(inserted_rows, records)
        filled_rows += len(records)
    return filled_rows


def fill_standard_coordinate_tables(root: ET.Element, records: list[dict[str, str]]) -> int:
    if not records:
        return 0
    filled_rows = 0
    for tbl in root.findall(".//w:tbl", NS):
        table_text = element_text(tbl)
        if "标准孔坐标" not in table_text or "标准孔编号" not in table_text or "X坐标" not in table_text:
            continue
        rows = tbl.findall("w:tr", NS)
        data_idx = None
        for idx, row in enumerate(rows):
            cells = row.findall("w:tc", NS)
            values = [element_text(cell).strip() for cell in cells]
            if values and re.fullmatch(r"U\d+", values[0]) and len(values) >= 5:
                data_idx = idx
                break
        if data_idx is None:
            continue
        template_row = copy.deepcopy(rows[data_idx])
        child_idx = list(tbl).index(rows[data_idx])
        for row in rows[data_idx:]:
            tbl.remove(row)
        inserted_rows: list[ET.Element] = []
        insert_at = child_idx
        for record in records:
            clone = copy.deepcopy(template_row)
            cells = clone.findall("w:tc", NS)
            if len(cells) >= 5:
                unit = record.get("勘探单元", "")
                set_cell_text(cells[0], unit)
                set_cell_text(cells[1], record.get("标准孔编号", ""))
                set_cell_text(cells[2], record.get("X坐标", ""))
                set_cell_text(cells[3], record.get("Y坐标", ""))
                set_cell_text(cells[4], format_elevation(record.get("高程", "")))
            tbl.insert(insert_at, clone)
            inserted_rows.append(clone)
            insert_at += 1
            filled_rows += 1
        ensure_table_all_borders(tbl)
        merge_unit_number_cells(inserted_rows, records)
    return filled_rows


def relics_in_cultural_range(records: list[dict[str, str]]) -> bool:
    for record in records:
        location = clean(record.get("遗址位置"))
        if location and location not in {"外", "范围外", "项目外"}:
            return True
    return False


def fill_relic_coordinate_tables(root: ET.Element, records: list[dict[str, str]], relic_records: list[dict[str, str]]) -> int:
    if not records:
        return 0
    location_by_id = {record.get("遗迹编号", ""): record.get("遗址位置", "") for record in relic_records}
    for record in records:
        record.setdefault("遗迹位置", location_by_id.get(record.get("遗迹编号", ""), ""))
    candidates: list[ET.Element] = []
    for tbl in root.findall(".//w:tbl", NS):
        table_text = element_text(tbl)
        if "遗迹编号" in table_text and "界点序号" in table_text and "平面坐标X" in table_text:
            candidates.append(tbl)
    if not candidates:
        return 0
    target = candidates[0]
    for tbl in candidates:
        if "遗迹位置" not in element_text(tbl):
            target = tbl
            break
    for tbl in candidates:
        if tbl is not target:
            remove_element(root, tbl)
    rows = target.findall("w:tr", NS)
    if len(rows) < 2:
        return 0
    header_cells = [element_text(cell).strip() for cell in rows[0].findall("w:tc", NS)]
    template_row = copy.deepcopy(rows[1])
    child_idx = list(target).index(rows[1])
    for row in rows[1:]:
        target.remove(row)
    previous_relic = ""
    previous_location = ""
    insert_at = child_idx
    for record in records:
        clone = copy.deepcopy(template_row)
        cells = clone.findall("w:tc", NS)
        for idx, cell in enumerate(cells):
            header = header_cells[idx] if idx < len(header_cells) else ""
            value = record.get(header, "")
            if header == "遗迹编号":
                current = record.get("遗迹编号", "")
                if current and current != previous_relic:
                    set_cell_vertical_merge(cell, "restart")
                    value = current
                elif current:
                    set_cell_vertical_merge(cell, None)
                    value = ""
            elif header == "遗迹位置":
                current = record.get("遗迹位置", "")
                if current and current != previous_location:
                    set_cell_vertical_merge(cell, "restart")
                    value = current
                elif current:
                    set_cell_vertical_merge(cell, None)
                    value = ""
            set_cell_text(cell, format_table_value(header, value))
        target.insert(insert_at, clone)
        insert_at += 1
        previous_relic = record.get("遗迹编号", previous_relic)
        previous_location = record.get("遗迹位置", previous_location)
    ensure_table_all_borders(target)
    return len(records)


def fill_relic_registration_tables(root: ET.Element, records: list[dict[str, str]]) -> int:
    if not records:
        return 0
    use_cultural_range_template = relics_in_cultural_range(records)
    candidates: list[ET.Element] = []
    for tbl in root.findall(".//w:tbl", NS):
        table_text = element_text(tbl)
        if "遗迹编号" in table_text and "遗迹类型" in table_text and "内部填充与包含物" in table_text:
            candidates.append(tbl)
    if not candidates:
        return 0
    target = None
    for tbl in candidates:
        has_location = "遗址位置" in element_text(tbl)
        if has_location == use_cultural_range_template:
            target = tbl
            break
    if target is None:
        target = candidates[0 if use_cultural_range_template else -1]
    for tbl in candidates:
        if tbl is not target:
            remove_element(root, tbl)
    rows = target.findall("w:tr", NS)
    if len(rows) < 2:
        return 0
    header_cells = [element_text(cell).strip() for cell in rows[0].findall("w:tc", NS)]
    template_row = copy.deepcopy(rows[1])
    child_idx = list(target).index(rows[1])
    for row in rows[1:]:
        target.remove(row)
    insert_at = child_idx
    for idx, record in enumerate(records, start=1):
        clone = copy.deepcopy(template_row)
        cells = clone.findall("w:tc", NS)
        values = dict(record)
        values["序号"] = str(idx)
        for cell_idx, cell in enumerate(cells):
            header = header_cells[cell_idx] if cell_idx < len(header_cells) else ""
            set_cell_text(cell, format_table_value(header, values.get(header, "")))
        target.insert(insert_at, clone)
        insert_at += 1
    ensure_table_all_borders(target)
    return len(records)


def section_layer_text_specs(section_id: str, layer_count: int) -> tuple[list[str], list[str]]:
    key = section_number_key(section_id)
    three_layers = layer_count >= 3
    if key == "A-A′":
        contains = (
            [
                "黄褐色沙土，土质较疏松，含植物根系",
                "黄褐色沙土，土质较疏松，含少量水",
                "黄褐色沙质生土层，土壤含大量水，滑铲不上",
            ]
            if three_layers
            else [
                "黄褐色沙土，土质较疏松，含植物根系",
                "黄褐色沙质生土层，土壤含大量水，滑铲不上",
            ]
        )
    else:
        contains = (
            [
                "黄褐色沙土，土质较疏松，含少量植物根系",
                "褐色沙土，土质较疏松",
                "褐色沙质生土层，土壤含大量水，滑铲不上",
            ]
            if three_layers
            else [
                "黄褐色沙土，土质较疏松，含少量植物根系",
                "褐色沙质生土层，土壤含大量水，滑铲不上",
            ]
        )
    nature = ["表土层", "沙土层", "生土层"] if three_layers else ["表土层", "生土层"]
    return contains, nature


def find_section_drawing_pdf(drawings_dir: Path | None, section_id: str) -> Path | None:
    if drawings_dir is None or not drawings_dir.exists():
        return None
    key = section_number_key(section_id).replace("′", "'")
    candidates = []
    for path in drawings_dir.rglob("*.pdf"):
        name = path.name.replace("′", "'").replace("’", "'")
        if key in name and ("地层堆积剖面" in name or "地层堆积剖线" in name):
            candidates.append(path)
    return sorted(candidates, key=lambda path: str(path))[0] if candidates else None


def parse_section_layer_thickness_rows(
    lines: list[str],
    *,
    label_line_index: int,
    record_count: int,
    layer_count: int,
) -> list[list[str]]:
    if label_line_index < 0 or label_line_index >= len(lines):
        return []
    label_matches = list(re.finditer(r"A\s*0*\d+\s*B\s*0*\d+", lines[label_line_index], flags=re.I))
    if len(label_matches) < record_count:
        return []
    label_centers = [(match.start() + match.end()) / 2 for match in label_matches[:record_count]]
    values_by_record: list[list[tuple[int, str]]] = [[] for _ in range(record_count)]
    for line_idx, line in enumerate(lines[label_line_index + 1 :], start=label_line_index + 1):
        for match in re.finditer(r"([0-9]+(?:\.[0-9]+)?)\s*m", line, flags=re.I):
            number = float(match.group(1))
            if number >= 5:
                continue
            center = (match.start() + match.end()) / 2
            record_idx = min(range(record_count), key=lambda idx: abs(label_centers[idx] - center))
            values_by_record[record_idx].append((line_idx, f"{number:.2f}"))
    out: list[list[str]] = []
    required = layer_count - 1
    for record_values in values_by_record:
        if len(record_values) < required:
            return []
        ordered = sorted(record_values, key=lambda item: item[0])
        out.append([value for _, value in ordered[:required]])
    return out


def extract_section_layer_thicknesses_from_pdf(
    drawings_dir: Path | None,
    section_id: str,
    records: list[dict[str, str]],
) -> tuple[list[list[str]], int]:
    pdf = find_section_drawing_pdf(drawings_dir, section_id)
    if pdf is None or not records:
        return [], 0
    try:
        result = subprocess.run(
            ["pdftotext", str(pdf), "-", "-layout"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=15,
        )
    except Exception:
        return [], 0

    lines = result.stdout.splitlines()
    label_count = 0
    label_line_index = -1
    for line in lines:
        matches = list(re.finditer(r"A\s*0*\d+\s*B\s*0*\d+", line, flags=re.I))
        if matches:
            label_count = len(matches)
            label_line_index = lines.index(line)
            break
    if label_count < len(records):
        return [], 0

    drawing_text = "\n".join(lines[label_line_index + 1 :])
    has_middle_layer = bool(re.search(r"(沙土层|淤积层|扰土层)", drawing_text))
    layer_count = 3 if has_middle_layer else 2
    values_by_record = parse_section_layer_thickness_rows(
        lines,
        label_line_index=label_line_index,
        record_count=len(records),
        layer_count=layer_count,
    )
    if not values_by_record:
        return [], 0
    return values_by_record, layer_count


def format_depth(value: float) -> str:
    if abs(value) < 1e-9:
        return "0"
    return f"{value:.2f}"


def build_section_probe_layer_rows(
    *,
    unit: str,
    record: dict[str, str],
    layer_thicknesses: list[str],
    contains: list[str],
    nature: list[str],
    previous_unit: str,
) -> list[list[str]]:
    row_values: list[list[str]] = []
    depth = 0.0
    for idx, layer_nature in enumerate(nature):
        is_last_natural_layer = idx == len(nature) - 1 or "生土层" in layer_nature
        thickness = "—" if is_last_natural_layer else layer_thicknesses[idx]
        row_values.append(
            [
                unit if idx == 0 and unit != previous_unit else "",
                record.get("列A", "") if idx == 0 else "",
                record.get("行B", "") if idx == 0 else "",
                record.get("X坐标", "") if idx == 0 else "",
                record.get("Y坐标", "") if idx == 0 else "",
                format_elevation(record.get("高程", "")) if idx == 0 else "",
                str(idx + 1).translate(str.maketrans("123456789", "①②③④⑤⑥⑦⑧⑨")),
                format_depth(depth),
                thickness,
                contains[idx] if idx < len(contains) else "",
                layer_nature,
            ]
        )
        if thickness != "—":
            depth += float(thickness)
    return row_values


def normalize_section_probe_table_headers(tbl: ET.Element) -> int:
    changed = 0
    rows = tbl.findall("w:tr", NS)
    if len(rows) < 2:
        return 0
    header_row = rows[1]
    tr_pr = ensure_child(header_row, "w:trPr", first=True)
    heights = tr_pr.findall("w:trHeight", NS)
    height_node = heights[0] if heights else ET.SubElement(tr_pr, qn("w:trHeight"))
    for extra in heights[1:]:
        tr_pr.remove(extra)
    height_node.set(qn("w:val"), "900")
    height_node.set(qn("w:hRule"), "atLeast")
    for cell in rows[1].findall("w:tc", NS):
        text = element_text(cell).strip()
        compact = re.sub(r"\s+", "", text)
        if compact in {"距离地表深度", "距离地表深度（单位：米）"}:
            set_cell_lines(cell, ["距离地表深度", "（单位：米）"])
            changed += 1
        elif compact in {"厚度", "厚度（单位：米）"}:
            set_cell_lines(cell, ["厚度", "（单位：米）"])
            changed += 1
    return changed


def normalize_all_section_probe_table_headers(root: ET.Element) -> int:
    changed = 0
    for tbl in root.findall(".//w:tbl", NS):
        text = element_text(tbl)
        if "普通探孔编号" in text and "距离地表深度" in text and "堆积和包含物" in text:
            changed += normalize_section_probe_table_headers(tbl)
    return changed


def mark_section_probe_table_layer_error(tbl: ET.Element, rows: list[ET.Element], data_idx: int, message: str) -> None:
    if len(rows) <= data_idx:
        return
    template_row = copy.deepcopy(rows[data_idx])
    child_idx = list(tbl).index(rows[data_idx])
    for row in rows[data_idx:]:
        tbl.remove(row)
    cells = template_row.findall("w:tc", NS)
    values = [""] * len(cells)
    if values:
        values[0] = "错误"
    if len(values) > 6:
        values[6] = "错误"
    target_idx = 9 if len(values) > 9 else max(len(values) - 1, 0)
    if values:
        values[target_idx] = message
    for cell, value in zip(cells, values):
        set_cell_text(cell, value)
    tbl.insert(child_idx, template_row)
    ensure_table_all_borders(tbl)


def fill_section_probe_table(
    root: ET.Element,
    section_id: str,
    records: list[dict[str, str]],
    fields: dict[str, str],
    drawings_dir: Path | None = None,
) -> int:
    candidate_tables = []
    for tbl in root.findall(".//w:tbl", NS):
        text = element_text(tbl)
        if "普通探孔编号" in text and "距离地表深度" in text and "堆积和包含物" in text:
            candidate_tables.append(tbl)
    table_index = 0 if section_number_key(section_id) == "A-A′" else 1
    if len(candidate_tables) <= table_index:
        return 0
    tbl = candidate_tables[table_index]
    rows = tbl.findall("w:tr", NS)
    if len(rows) < 6:
        return 0
    set_cell_text(rows[0].findall("w:tc", NS)[1], fields.get("项目名称", ""))
    normalize_section_probe_table_headers(tbl)
    data_idx = 3
    if not records:
        mark_section_probe_table_layer_error(
            tbl,
            rows,
            data_idx,
            f"【插入错误：缺少{section_number_key(section_id)}剖线坐标记录，无法生成剖面探孔记录表】",
        )
        return 0
    layer_thicknesses, drawing_layer_count = extract_section_layer_thicknesses_from_pdf(drawings_dir, section_id, records)
    if not layer_thicknesses:
        mark_section_probe_table_layer_error(
            tbl,
            rows,
            data_idx,
            f"【插入错误：未能从{section_number_key(section_id)}地层堆积剖面图可靠提取逐孔层厚】",
        )
        return 0
    layer_count = drawing_layer_count
    template_rows = [copy.deepcopy(row) for row in rows[data_idx : data_idx + layer_count]]
    if len(template_rows) < layer_count:
        mark_section_probe_table_layer_error(
            tbl,
            rows,
            data_idx,
            f"【插入错误：{section_number_key(section_id)}剖面探孔记录表模板行不足，无法按剖面图层数生成】",
        )
        return 0
    child_idx = list(tbl).index(rows[data_idx])
    for row in rows[data_idx:]:
        tbl.remove(row)
    contains, nature = section_layer_text_specs(section_id, layer_count)
    insert_at = child_idx
    filled = 0
    previous_unit = ""
    inserted_rows_with_units: list[tuple[ET.Element, str]] = []
    for idx, record in enumerate(records):
        record_thicknesses = layer_thicknesses[idx] if idx < len(layer_thicknesses) else layer_thicknesses[-1]
        unit = record.get("勘探单元", "")
        row_values = build_section_probe_layer_rows(
            unit=unit,
            record=record,
            layer_thicknesses=record_thicknesses,
            contains=contains,
            nature=nature,
            previous_unit=previous_unit,
        )
        previous_unit = unit
        for layer_idx in range(layer_count):
            clone = copy.deepcopy(template_rows[layer_idx])
            cells = clone.findall("w:tc", NS)
            for cell, value in zip(cells, row_values[layer_idx]):
                set_cell_text(cell, value)
            tbl.insert(insert_at, clone)
            inserted_rows_with_units.append((clone, unit))
            insert_at += 1
            filled += 1
    ensure_table_all_borders(tbl)
    previous_merge_unit = ""
    for row, unit in inserted_rows_with_units:
        cells = row.findall("w:tc", NS)
        if not cells:
            continue
        if unit and unit != previous_merge_unit:
            set_cell_vertical_merge(cells[0], "restart")
            set_cell_text(cells[0], unit)
            previous_merge_unit = unit
        elif unit:
            set_cell_vertical_merge(cells[0], None)
            set_cell_text(cells[0], "")
    return filled


def find_table_by_terms(root: ET.Element, terms: list[str], occurrence: int = 0) -> ET.Element | None:
    matched = 0
    for tbl in root.findall(".//w:tbl", NS):
        text = element_text(tbl)
        if all(term in text for term in terms):
            if matched == occurrence:
                return tbl
            matched += 1
    return None


def insert_error_before_element(root: ET.Element, target: ET.Element, message: str) -> bool:
    parent = find_element_parent(root, target)
    if parent is None:
        return False
    idx = list(parent).index(target)
    parent.insert(idx, make_plain_paragraph(message))
    return True


def mark_table_insertion_errors(
    root: ET.Element,
    tables: dict[str, list[dict[str, str]]],
    *,
    redline_table_rows: int,
    unit_table_rows: int,
    standard_coordinate_table_rows: int,
    relic_coordinate_table_rows: int,
    relic_registration_table_rows: int,
    section_aa_probe_rows: int,
    section_bb_probe_rows: int,
) -> list[str]:
    checks = [
        {
            "name": "红线四至坐标表",
            "terms": ["四至范围坐标"],
            "source_count": len(tables.get("红线坐标", [])),
            "inserted_count": redline_table_rows,
            "required": True,
        },
        {
            "name": "勘探单元坐标表",
            "terms": ["勘探单元坐标"],
            "source_count": len(tables.get("勘探单元", [])),
            "inserted_count": unit_table_rows,
            "required": True,
        },
        {
            "name": "标准孔坐标表",
            "terms": ["标准孔坐标", "标准孔编号"],
            "source_count": len(tables.get("标准孔", [])),
            "inserted_count": standard_coordinate_table_rows,
            "required": True,
        },
        {
            "name": "A-A′剖面探孔记录表",
            "terms": ["普通探孔编号", "距离地表深度", "堆积和包含物"],
            "occurrence": 0,
            "source_count": len(tables.get("剖线坐标_AA", [])),
            "inserted_count": section_aa_probe_rows,
            "required": bool(tables.get("剖线坐标_AA")),
        },
        {
            "name": "B-B′剖面探孔记录表",
            "terms": ["普通探孔编号", "距离地表深度", "堆积和包含物"],
            "occurrence": 1,
            "source_count": len(tables.get("剖线坐标_BB", [])),
            "inserted_count": section_bb_probe_rows,
            "required": bool(tables.get("剖线坐标_BB")),
        },
        {
            "name": "遗迹坐标表",
            "terms": ["遗迹编号", "界点序号", "平面坐标X"],
            "source_count": len(tables.get("遗迹坐标", [])),
            "inserted_count": relic_coordinate_table_rows,
            "required": bool(tables.get("遗迹坐标")),
        },
        {
            "name": "遗迹登记表",
            "terms": ["遗迹编号", "遗迹类型", "内部填充与包含物"],
            "source_count": len(tables.get("遗迹记录") or tables.get("遗迹", [])),
            "inserted_count": relic_registration_table_rows,
            "required": bool(tables.get("遗迹记录") or tables.get("遗迹", [])),
        },
    ]
    messages: list[str] = []
    for check in checks:
        source_count = int(check["source_count"])
        inserted_count = int(check["inserted_count"])
        required = bool(check["required"])
        if not required and source_count == 0:
            continue
        if source_count > 0 and inserted_count >= source_count:
            continue
        if source_count == 0:
            reason = "未在内业成果表格中匹配到对应数据"
        elif inserted_count == 0:
            reason = f"源表已读取{source_count}条，但报告表格未成功替换"
        else:
            reason = f"表格插入不完整，源表{source_count}条，报告插入{inserted_count}条"
        message = table_insert_error_marker(str(check["name"]), reason)
        target = find_table_by_terms(root, list(check["terms"]), int(check.get("occurrence", 0)))
        if target is not None and insert_error_before_element(root, target, message):
            messages.append(f"{check['name']}：{reason}")
    return messages


def next_relationship_id(rels_root: ET.Element) -> str:
    used = []
    for rel in rels_root.findall("rel:Relationship", NS):
        rid = rel.get("Id", "")
        match = re.fullmatch(r"rId(\d+)", rid)
        if match:
            used.append(int(match.group(1)))
    return f"rId{(max(used) if used else 0) + 1}"


def ensure_content_type_defaults(root: ET.Element) -> None:
    required = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png"}
    existing: set[str] = set()
    for node in root.findall("ct:Default", NS):
        extension = node.get("Extension", "").lower()
        if extension in required:
            node.set("Extension", extension)
            node.set("ContentType", required[extension])
            existing.add(extension)
    for extension, content_type in required.items():
        if extension not in existing:
            node = ET.Element(f"{{{CT_NS}}}Default")
            node.set("Extension", extension)
            node.set("ContentType", content_type)
            root.insert(0, node)


def cm_to_emu(value: float) -> int:
    return int(value * EMU_PER_CM)


def is_section_drawing(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path.stem}"
    if "遗迹平、剖面图" in text or "平、剖面图" in text or "平剖面图" in text:
        return False
    if "剖面位置图" in text or "剖线位置示意图" in text:
        return False
    return "剖线图" in text or "剖面图" in text or "地层堆积剖线" in text or "地层堆积剖面" in text


def is_standard_hole_image(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path}"
    return "标准孔位置图" in text or "标准孔土样照" in text or "单个标准孔" in text or "12.标准孔照" in text


def is_standard_hole_photo(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path}"
    return "标准孔土样照" in text or "12.标准孔照" in text


def is_relic_plan_section_image(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path.stem}"
    return "遗迹平、剖面图" in text or "平、剖面图" in text or "平剖面图" in text


def is_single_relic_drawing(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path}"
    return is_relic_plan_section_image(slot) or "单个遗迹" in text or "单个遗址" in text


def is_project_parcel_photo_drawing(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path.stem}"
    return "项目地块照" in text or "项目地块现状照" in text or "地块现状照" in text


def is_project_location_overview(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path.stem}"
    return "项目地块位置示意图" in text or bool(re.search(r"项目地块在.+位置示意图", text))


def attachment_display_width_cm(slot: ImageSlot) -> float | None:
    key = clean(slot.key)
    if key == "人员附件:身份证":
        return PERSONNEL_ID_CARD_WIDTH_CM
    if key == "人员附件:证书":
        return PERSONNEL_CERTIFICATE_WIDTH_CM
    if key == "人员附件:劳动合同":
        return PERSONNEL_CONTRACT_WIDTH_CM
    if key == "人员附件:社保" and not is_landscape_social_security_attachment(slot):
        return PERSONNEL_SOCIAL_SECURITY_WIDTH_CM
    return None


def is_landscape_social_security_attachment(slot: ImageSlot) -> bool:
    if clean(slot.key) != "人员附件:社保":
        return False
    try:
        with Image.open(slot.path) as image:
            width_px, height_px = image.size
    except Exception:
        return False
    return width_px > height_px


def attachment_display_height_cm(slot: ImageSlot) -> float | None:
    key = clean(slot.key)
    if key == "公司附件:营业执照":
        return COMPANY_LICENSE_HEIGHT_CM
    if key.startswith("公司附件:") and "资质" in normalize_key(key):
        return COMPANY_QUALIFICATION_HEIGHT_CM
    if is_landscape_social_security_attachment(slot):
        return PERSONNEL_SOCIAL_SECURITY_LANDSCAPE_HEIGHT_CM
    return None


def is_attachment_image_slot(slot: ImageSlot) -> bool:
    key = clean(slot.key)
    return key.startswith("公司附件:") or key.startswith("人员附件:")


def company_attachment_requires_page_break(key: str) -> bool:
    normalized = normalize_key(key)
    return "营业执照" in normalized or "资质" in normalized


def personnel_attachment_requires_page_break(category: str) -> bool:
    return clean(category) in {"证书", "社保"}


def is_rotated_standard_drawing(slot: ImageSlot) -> bool:
    text = f"{slot.key} {slot.caption} {slot.path.stem}"
    markers = [
        "项目地块卫星图",
        "地块卫星图",
        "红线四至坐标图",
        "项目红线四至坐标图",
        "红线图",
        "勘探区域示意图",
        "项目勘探区域示意图",
        "勘探分区示意图",
        "项目勘探分区示意图",
        "勘探单元布设示意图",
        "勘探单元布置示意图",
        "划分勘探单元图",
        "探孔布设示意图",
        "探孔布置示意图",
        "勘探布孔示意图",
        "剖线位置示意图",
        "剖面位置图",
        "标准探孔位置示意图",
        "遗迹分布示意图",
    ]
    return any(marker in text for marker in markers)


def slot_display_extent(slot: ImageSlot, width_px: int, height_px: int) -> tuple[int, int]:
    ratio = width_px / height_px if height_px else 1
    attachment_height = attachment_display_height_cm(slot)
    if attachment_height is not None:
        height = cm_to_emu(attachment_height)
        width = int(height * ratio)
        return width, height
    attachment_width = attachment_display_width_cm(slot)
    if attachment_width is not None:
        width = cm_to_emu(attachment_width)
        height = int(width / ratio)
        return width, height
    if is_standard_hole_image(slot):
        max_width = cm_to_emu(STANDARD_HOLE_IMAGE_WIDTH_CM)
        width = max_width
        height = int(width / ratio)
        max_height = cm_to_emu(STANDARD_HOLE_PHOTO_MAX_HEIGHT_CM)
        if height > max_height:
            height = max_height
            width = int(height * ratio)
        return width, height
    if slot.source == "photo":
        width = cm_to_emu(PHOTO_WIDTH_CM)
        height = int(width / ratio)
        return width, height
    if is_single_relic_drawing(slot):
        width = cm_to_emu(RELIC_PLAN_SECTION_WIDTH_CM)
        height = int(width / ratio)
        return width, height
    if slot.source == "drawing" and is_section_drawing(slot):
        width = cm_to_emu(DRAWING_SECTION_WIDTH_CM)
        height = int(width / ratio)
        return width, height
    if slot.source == "drawing" and is_project_parcel_photo_drawing(slot):
        width = cm_to_emu(DRAWING_STANDARD_WIDTH_CM)
        height = int(width / ratio)
        return width, height
    if slot.source == "drawing":
        if image_needs_left_rotation(slot):
            height = cm_to_emu(DRAWING_ROTATED_HEIGHT_CM)
            width = int(height * ratio)
            return width, height
        width = cm_to_emu(DRAWING_STANDARD_WIDTH_CM)
        height = int(width / ratio)
        return width, height
    width = IMAGE_MAX_WIDTH_EMU
    height = int(width / ratio)
    if height > IMAGE_MAX_HEIGHT_EMU:
        height = IMAGE_MAX_HEIGHT_EMU
        width = int(height * ratio)
    return width, height


def emu_to_cm(value: int) -> float:
    return value / EMU_PER_CM


def assert_close_cm(actual_emu: int, expected_cm: float, label: str, tolerance: float = 0.02) -> None:
    actual_cm = emu_to_cm(actual_emu)
    if abs(actual_cm - expected_cm) > tolerance:
        raise AssertionError(f"{label}: expected {expected_cm:.2f} cm, got {actual_cm:.2f} cm")


def self_test_image_rules() -> None:
    def slot(key: str, caption: str = "", source: str = "drawing", stem: str | None = None) -> ImageSlot:
        name = stem or caption or key
        return ImageSlot(Path(f"/tmp/{name}.jpg"), caption, source, key)

    width, height = slot_display_extent(slot("实地踏查照", source="photo"), 3000, 2000)
    assert not image_needs_left_rotation(slot("实地踏查照", source="photo"))
    assert_close_cm(width, 14.63, "外业照片宽度")

    location = slot("项目地块位置示意图", "项目地块在内蒙古自治区位置示意图")
    width, height = slot_display_extent(location, 3000, 2000)
    assert not image_needs_left_rotation(location)
    assert_close_cm(width, 14.63, "位置示意图宽度")

    rotated_keys = [
        "项目地块卫星图",
        "红线四至坐标图",
        "项目勘探区域示意图",
        "项目勘探分区示意图",
        "勘探单元布设示意图",
        "探孔布设示意图",
        "剖线位置示意图",
        "标准探孔位置示意图",
        "遗迹分布示意图",
    ]
    for key in rotated_keys:
        test_slot = slot(key)
        assert image_needs_left_rotation(test_slot), f"{key}: should rotate left"
        width, height = slot_display_extent(test_slot, 2000, 3000)
        assert_close_cm(height, 14.63, f"{key}高度")

    standard = slot("标准孔位置图", "U01-A001B001", stem="U01-A001B001")
    width, height = slot_display_extent(standard, 3000, 2000)
    assert not image_needs_left_rotation(standard)
    assert emu_to_cm(width) <= 12.40 + 0.02 and emu_to_cm(height) <= 9.80 + 0.02

    relic_plan = slot("遗迹平、剖面图", "H1平、剖面图", stem="H1平、剖面图")
    width, height = slot_display_extent(relic_plan, 3000, 2000)
    assert not image_needs_left_rotation(relic_plan)
    assert_close_cm(width, 14.63, "单个遗迹平剖面图宽度")

    section = slot("剖线图", "A-A′剖线图")
    width, height = slot_display_extent(section, 3000, 1000)
    assert not image_needs_left_rotation(section)
    assert_close_cm(width, 24.50, "剖线/剖面图宽度")

    single_queue = defaultdict(deque)
    single_queue["项目地块卫星图"].append(slot("项目地块卫星图"))
    assert take_image_slots(single_queue, "项目地块卫星图", expand_all=False)
    assert not take_image_slots(single_queue, "项目地块卫星图", expand_all=False), "单张图片不得重复消费"

    multi_queue = defaultdict(deque)
    multi_queue["遗迹分布示意图"].extend([
        slot("遗迹分布示意图", "遗迹分布示意图"),
        slot("遗迹分布示意图", "遗迹分布示意图局部1"),
    ])
    assert len(take_image_slots(multi_queue, "遗迹分布示意图", expand_all=True)) == 2
    assert not take_image_slots(multi_queue, "遗迹分布示意图", expand_all=True), "批量展开图片不得重复消费"

    standard_record = {
        "勘探单元": "U02",
        "标准孔编号": "A332B232",
        "标准孔土样照": "/tmp/TK77.jpg",
    }
    assert standard_probe_number(standard_record) == "U02-A332B232"
    assert standard_hole_code(standard_record) == "TK77"
    assert normalize_tk_code("TK01") == "TK1"
    assert normalize_tk_code("1") == "TK1"
    assert tk_code_aliases("TK1") == ["TK1", "TK01", "TK001"]


def image_extension(path: Path, data: bytes | None = None) -> str:
    if data is not None:
        header = data[:16]
    else:
        with path.open("rb") as fh:
            header = fh.read(16)
    if header.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if header.startswith(b"\xff\xd8"):
        return ".jpg"
    suffix = path.suffix.lower()
    return ".jpg" if suffix == ".jpeg" else suffix


def image_slot_signature(slot: ImageSlot) -> str:
    return hashlib.sha256(slot.path.read_bytes()).hexdigest()


def keep_unseen_image_slots(
    slots: list[ImageSlot],
    inserted_signatures: set[str],
) -> list[tuple[ImageSlot, str]]:
    kept: list[tuple[ImageSlot, str]] = []
    seen = set(inserted_signatures)
    for slot in slots:
        signature = image_slot_signature(slot)
        if signature in seen:
            continue
        kept.append((slot, signature))
        seen.add(signature)
    return kept


def image_needs_left_rotation(slot: ImageSlot) -> bool:
    return (
        slot.source == "drawing"
        and is_rotated_standard_drawing(slot)
        and not is_section_drawing(slot)
        and not is_standard_hole_image(slot)
        and not is_single_relic_drawing(slot)
        and not is_project_parcel_photo_drawing(slot)
        and not is_project_location_overview(slot)
    )


def image_exif_orientation(path: Path) -> int:
    try:
        with Image.open(path) as image:
            return int(image.getexif().get(274, 1) or 1)
    except Exception:
        return 1


def image_display_rotation_degrees(slot: ImageSlot) -> int:
    key = clean(slot.key)
    if key == "公司附件:营业执照" or (key.startswith("公司附件:") and "资质" in normalize_key(key)):
        return 270
    if is_landscape_social_security_attachment(slot):
        return 270
    if image_needs_left_rotation(slot):
        return 270
    orientation = image_exif_orientation(slot.path)
    if orientation == 8:
        return 270
    if orientation == 6:
        return 90
    if orientation == 3:
        return 180
    return 0


def image_display_dimensions(path: Path) -> tuple[int, int]:
    with Image.open(path) as image:
        width_px, height_px = image.size
        orientation = int(image.getexif().get(274, 1) or 1)
    if orientation in {6, 8}:
        return height_px, width_px
    return width_px, height_px


def prepare_image_data(slot: ImageSlot) -> tuple[bytes, str, int, int]:
    original_data = slot.path.read_bytes()
    width_px, height_px = image_display_dimensions(slot.path)
    return original_data, image_extension(slot.path, original_data), width_px, height_px


def make_image_paragraph(
    rid: str,
    name: str,
    width: int,
    height: int,
    doc_pr_id: int,
    *,
    border: bool = False,
    rotate_left: bool = False,
    rotation_degrees: int | None = None,
    alt_text: str = "",
) -> ET.Element:
    para = ET.Element(qn("w:p"))
    ppr = ET.SubElement(para, qn("w:pPr"))
    jc = ET.SubElement(ppr, qn("w:jc"))
    jc.set(qn("w:val"), "center")
    run = ET.SubElement(para, qn("w:r"))
    drawing = ET.SubElement(run, qn("w:drawing"))
    inline = ET.SubElement(drawing, qn("wp:inline"))
    extent = ET.SubElement(inline, qn("wp:extent"))
    extent.set("cx", str(width))
    extent.set("cy", str(height))
    effect_extent = ET.SubElement(inline, qn("wp:effectExtent"))
    for side in ["l", "t", "r", "b"]:
        effect_extent.set(side, "0")
    doc_pr = ET.SubElement(inline, qn("wp:docPr"))
    doc_pr.set("id", str(doc_pr_id))
    doc_pr.set("name", name)
    if alt_text:
        doc_pr.set("descr", alt_text)
    c_nv = ET.SubElement(inline, qn("wp:cNvGraphicFramePr"))
    locks = ET.SubElement(c_nv, qn("a:graphicFrameLocks"))
    locks.set("noChangeAspect", "1")
    graphic = ET.SubElement(inline, qn("a:graphic"))
    graphic_data = ET.SubElement(graphic, qn("a:graphicData"))
    graphic_data.set("uri", PIC_NS)
    pic = ET.SubElement(graphic_data, qn("pic:pic"))
    nv_pic_pr = ET.SubElement(pic, qn("pic:nvPicPr"))
    c_nv_pr = ET.SubElement(nv_pic_pr, qn("pic:cNvPr"))
    c_nv_pr.set("id", "0")
    c_nv_pr.set("name", name)
    if alt_text:
        c_nv_pr.set("descr", alt_text)
    ET.SubElement(nv_pic_pr, qn("pic:cNvPicPr"))
    blip_fill = ET.SubElement(pic, qn("pic:blipFill"))
    blip = ET.SubElement(blip_fill, qn("a:blip"))
    blip.set(qn("r:embed"), rid)
    stretch = ET.SubElement(blip_fill, qn("a:stretch"))
    ET.SubElement(stretch, qn("a:fillRect"))
    sp_pr = ET.SubElement(pic, qn("pic:spPr"))
    xfrm = ET.SubElement(sp_pr, qn("a:xfrm"))
    if rotation_degrees is None:
        rotation_degrees = 270 if rotate_left else 0
    if rotation_degrees:
        xfrm.set("rot", str((rotation_degrees % 360) * 60000))
    off = ET.SubElement(xfrm, qn("a:off"))
    off.set("x", "0")
    off.set("y", "0")
    ext = ET.SubElement(xfrm, qn("a:ext"))
    ext.set("cx", str(width))
    ext.set("cy", str(height))
    prst = ET.SubElement(sp_pr, qn("a:prstGeom"))
    prst.set("prst", "rect")
    ET.SubElement(prst, qn("a:avLst"))
    if border:
        line = ET.SubElement(sp_pr, qn("a:ln"))
        line.set("w", "12700")
        solid_fill = ET.SubElement(line, qn("a:solidFill"))
        srgb = ET.SubElement(solid_fill, qn("a:srgbClr"))
        srgb.set("val", "000000")
    return para


def make_floating_image_paragraph(
    rid: str,
    name: str,
    width: int,
    height: int,
    doc_pr_id: int,
    *,
    vertical_offset: int = 0,
    alt_text: str = "",
) -> ET.Element:
    para = ET.Element(qn("w:p"))
    ppr = ET.SubElement(para, qn("w:pPr"))
    jc = ET.SubElement(ppr, qn("w:jc"))
    jc.set(qn("w:val"), "center")
    spacing = ET.SubElement(ppr, qn("w:spacing"))
    spacing.set(qn("w:before"), "0")
    spacing.set(qn("w:after"), "0")
    spacing.set(qn("w:line"), "1")
    spacing.set(qn("w:lineRule"), "exact")
    run = ET.SubElement(para, qn("w:r"))
    drawing = ET.SubElement(run, qn("w:drawing"))
    anchor = ET.SubElement(drawing, qn("wp:anchor"))
    anchor.set("distT", "0")
    anchor.set("distB", "0")
    anchor.set("distL", "0")
    anchor.set("distR", "0")
    anchor.set("simplePos", "0")
    anchor.set("relativeHeight", "251659264")
    anchor.set("behindDoc", "0")
    anchor.set("locked", "0")
    anchor.set("layoutInCell", "1")
    anchor.set("allowOverlap", "1")
    simple_pos = ET.SubElement(anchor, qn("wp:simplePos"))
    simple_pos.set("x", "0")
    simple_pos.set("y", "0")
    position_h = ET.SubElement(anchor, qn("wp:positionH"))
    position_h.set("relativeFrom", "column")
    ET.SubElement(position_h, qn("wp:align")).text = "center"
    position_v = ET.SubElement(anchor, qn("wp:positionV"))
    position_v.set("relativeFrom", "paragraph")
    ET.SubElement(position_v, qn("wp:posOffset")).text = str(vertical_offset)
    extent = ET.SubElement(anchor, qn("wp:extent"))
    extent.set("cx", str(width))
    extent.set("cy", str(height))
    effect_extent = ET.SubElement(anchor, qn("wp:effectExtent"))
    for side in ["l", "t", "r", "b"]:
        effect_extent.set(side, "0")
    ET.SubElement(anchor, qn("wp:wrapNone"))
    doc_pr = ET.SubElement(anchor, qn("wp:docPr"))
    doc_pr.set("id", str(doc_pr_id))
    doc_pr.set("name", name)
    if alt_text:
        doc_pr.set("descr", alt_text)
    c_nv = ET.SubElement(anchor, qn("wp:cNvGraphicFramePr"))
    locks = ET.SubElement(c_nv, qn("a:graphicFrameLocks"))
    locks.set("noChangeAspect", "1")
    graphic = ET.SubElement(anchor, qn("a:graphic"))
    graphic_data = ET.SubElement(graphic, qn("a:graphicData"))
    graphic_data.set("uri", PIC_NS)
    pic = ET.SubElement(graphic_data, qn("pic:pic"))
    nv_pic_pr = ET.SubElement(pic, qn("pic:nvPicPr"))
    c_nv_pr = ET.SubElement(nv_pic_pr, qn("pic:cNvPr"))
    c_nv_pr.set("id", "0")
    c_nv_pr.set("name", name)
    if alt_text:
        c_nv_pr.set("descr", alt_text)
    ET.SubElement(nv_pic_pr, qn("pic:cNvPicPr"))
    blip_fill = ET.SubElement(pic, qn("pic:blipFill"))
    blip = ET.SubElement(blip_fill, qn("a:blip"))
    blip.set(qn("r:embed"), rid)
    stretch = ET.SubElement(blip_fill, qn("a:stretch"))
    ET.SubElement(stretch, qn("a:fillRect"))
    sp_pr = ET.SubElement(pic, qn("pic:spPr"))
    xfrm = ET.SubElement(sp_pr, qn("a:xfrm"))
    off = ET.SubElement(xfrm, qn("a:off"))
    off.set("x", "0")
    off.set("y", "0")
    ext = ET.SubElement(xfrm, qn("a:ext"))
    ext.set("cx", str(width))
    ext.set("cy", str(height))
    prst = ET.SubElement(sp_pr, qn("a:prstGeom"))
    prst.set("prst", "rect")
    ET.SubElement(prst, qn("a:avLst"))
    return para


def make_page_break_paragraph() -> ET.Element:
    para = ET.Element(qn("w:p"))
    run = ET.SubElement(para, qn("w:r"))
    br = ET.SubElement(run, qn("w:br"))
    br.set(qn("w:type"), "page")
    return para


def make_section_break_paragraph(sect_pr: ET.Element | None = None) -> ET.Element:
    para = ET.Element(qn("w:p"))
    ppr = ET.SubElement(para, qn("w:pPr"))
    if sect_pr is None:
        sect_pr = ET.Element(qn("w:sectPr"))
        pg_sz = ET.SubElement(sect_pr, qn("w:pgSz"))
        pg_sz.set(qn("w:w"), "16838")
        pg_sz.set(qn("w:h"), "11906")
        pg_sz.set(qn("w:orient"), "landscape")
        pg_mar = ET.SubElement(sect_pr, qn("w:pgMar"))
        pg_mar.set(qn("w:top"), "1800")
        pg_mar.set(qn("w:right"), "1440")
        pg_mar.set(qn("w:bottom"), "1800")
        pg_mar.set(qn("w:left"), "1440")
        pg_mar.set(qn("w:header"), "851")
        pg_mar.set(qn("w:footer"), "992")
        pg_mar.set(qn("w:gutter"), "0")
        ET.SubElement(sect_pr, qn("w:cols")).set(qn("w:space"), "425")
        doc_grid = ET.SubElement(sect_pr, qn("w:docGrid"))
        doc_grid.set(qn("w:type"), "lines")
        doc_grid.set(qn("w:linePitch"), "312")
        doc_grid.set(qn("w:charSpace"), "0")
    ppr.append(copy.deepcopy(sect_pr))
    return para


def paragraph_sect_pr(para: ET.Element) -> ET.Element | None:
    return para.find(".//w:pPr/w:sectPr", NS)


def ensure_paragraph_properties(para: ET.Element) -> ET.Element:
    ppr = para.find("w:pPr", NS)
    if ppr is None:
        ppr = ET.Element(qn("w:pPr"))
        para.insert(0, ppr)
    return ppr


def make_section_break_continuous(sect_pr: ET.Element) -> None:
    break_type = sect_pr.find("w:type", NS)
    if break_type is None:
        break_type = ET.Element(qn("w:type"))
        sect_pr.insert(0, break_type)
    break_type.set(qn("w:val"), "continuous")


def is_landscape_sect_pr(sect_pr: ET.Element | None) -> bool:
    if sect_pr is None:
        return False
    pg_sz = sect_pr.find("w:pgSz", NS)
    return pg_sz is not None and pg_sz.get(qn("w:orient")) == "landscape"


def make_plain_paragraph(text: str) -> ET.Element:
    para = ET.Element(qn("w:p"))
    run = ET.SubElement(para, qn("w:r"))
    t = ET.SubElement(run, qn("w:t"))
    t.text = text
    return para


def make_table_cell(
    text: str,
    style: dict[str, object] | None = None,
    *,
    grid_span: int = 1,
    v_merge: str | None = None,
) -> ET.Element:
    style = style or {}
    cell = ET.Element(qn("w:tc"))
    tc_pr = ET.SubElement(cell, qn("w:tcPr"))
    ET.SubElement(tc_pr, qn("w:tcW")).set(qn("w:type"), "auto")
    if grid_span > 1:
        ET.SubElement(tc_pr, qn("w:gridSpan")).set(qn("w:val"), str(grid_span))
    if v_merge:
        v_merge_node = ET.SubElement(tc_pr, qn("w:vMerge"))
        if v_merge == "restart":
            v_merge_node.set(qn("w:val"), "restart")
    ET.SubElement(tc_pr, qn("w:vAlign")).set(qn("w:val"), clean(style.get("vertical")) or "center")
    para = ET.SubElement(cell, qn("w:p"))
    ppr = ET.SubElement(para, qn("w:pPr"))
    ET.SubElement(ppr, qn("w:jc")).set(qn("w:val"), clean(style.get("horizontal")) or "center")
    run = ET.SubElement(para, qn("w:r"))
    if style.get("bold"):
        rpr = ET.SubElement(run, qn("w:rPr"))
        ET.SubElement(rpr, qn("w:b"))
    t = ET.SubElement(run, qn("w:t"))
    if text[:1].isspace() or text[-1:].isspace():
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = text
    return cell


def make_word_table_from_raw_rows(records: list[dict[str, object]]) -> ET.Element:
    tbl = ET.Element(qn("w:tbl"))
    tbl_pr = ET.SubElement(tbl, qn("w:tblPr"))
    tbl_w = ET.SubElement(tbl_pr, qn("w:tblW"))
    tbl_w.set(qn("w:w"), "5000")
    tbl_w.set(qn("w:type"), "pct")
    tbl_jc = ET.SubElement(tbl_pr, qn("w:jc"))
    tbl_jc.set(qn("w:val"), "center")
    tbl_layout = ET.SubElement(tbl_pr, qn("w:tblLayout"))
    tbl_layout.set(qn("w:type"), "autofit")
    merges = records[0].get("__merges", []) if records else []
    if not isinstance(merges, list):
        merges = []

    def merge_for_position(row_idx: int, col_idx: int) -> dict[str, int] | None:
        for merge in merges:
            if not isinstance(merge, dict):
                continue
            if (
                int(merge.get("row_start", -1)) <= row_idx <= int(merge.get("row_end", -1))
                and int(merge.get("col_start", -1)) <= col_idx <= int(merge.get("col_end", -1))
            ):
                return merge
        return None

    max_cols = max((len(record.get("__cells", [])) for record in records), default=0)
    tbl_grid = ET.SubElement(tbl, qn("w:tblGrid"))
    for _ in range(max_cols):
        ET.SubElement(tbl_grid, qn("w:gridCol"))
    for row_idx, record in enumerate(records):
        row = ET.SubElement(tbl, qn("w:tr"))
        cells = record.get("__cells", [])
        if not isinstance(cells, list):
            cells = []
        styles = record.get("__styles", [])
        if not isinstance(styles, list):
            styles = []
        col_idx = 0
        while col_idx < len(cells):
            merge = merge_for_position(row_idx, col_idx)
            if merge and col_idx > int(merge.get("col_start", col_idx)):
                col_idx += 1
                continue
            grid_span = 1
            v_merge: str | None = None
            if merge:
                grid_span = int(merge.get("col_end", col_idx)) - int(merge.get("col_start", col_idx)) + 1
                if int(merge.get("row_end", row_idx)) > int(merge.get("row_start", row_idx)):
                    v_merge = "restart" if row_idx == int(merge.get("row_start", row_idx)) else "continue"
            style = styles[col_idx] if col_idx < len(styles) and isinstance(styles[col_idx], dict) else {}
            text = "" if v_merge == "continue" else clean(cells[col_idx])
            row.append(make_table_cell(text, style, grid_span=grid_span, v_merge=v_merge))
            col_idx += grid_span
    set_table_text_font(tbl, font="宋体", size_half_points="21")
    set_table_row_height(tbl, height_cm=APPENDIX_TABLE_ROW_HEIGHT_CM)
    ensure_table_all_borders(tbl)
    return tbl


def estimated_attachment_table_pages(records: list[dict[str, object]]) -> int:
    # A4 portrait with 0.62 cm row height leaves roughly 35-38 rows per page
    # after title/header margins. Use 35 as a conservative threshold estimate.
    return max(1, math.ceil(len(records) / 35))


def is_city_level_or_above(fields: dict[str, str]) -> bool | None:
    level = clean(fields.get("项目级别"))
    if not level:
        return None
    if "市级及市级以上" in level or "市级以上" in level or "市级" == level:
        return True
    if "市级以下" in level or "旗县" in level or "县级" in level:
        return False
    return None


def remove_previous_attachment_title(body: ET.Element, placeholder_idx: int, title: str) -> int:
    removed = 0
    idx = placeholder_idx - 1
    while idx >= 0:
        block = list(body)[idx]
        if block.tag != qn("w:p"):
            break
        text = element_text(block).strip()
        if not text:
            body.remove(block)
            removed += 1
            idx -= 1
            continue
        if title in text:
            body.remove(block)
            removed += 1
        break
    return removed


def insert_project_land_range_attachment_table(
    root: ET.Element,
    fields: dict[str, str],
    records: list[dict[str, object]],
) -> tuple[int, str]:
    body, _blocks = body_blocks(root)
    if body is None:
        return 0, "未找到正文 body"
    placeholder = "{{附件表:项目用地范围坐标表}}"
    targets = [block for block in list(body) if block.tag == qn("w:p") and placeholder in element_text(block)]
    if not targets:
        return 0, "模板未设置附件表占位符"
    city_level = is_city_level_or_above(fields)
    if city_level is None:
        for target in targets:
            set_element_text(target, report_error_marker(MATCH_ERROR_TEXT, "项目级别缺少，无法判断是否插入项目用地范围坐标表"))
        return 0, "项目级别缺少"
    if not records:
        for target in targets:
            set_element_text(target, image_insert_error_marker("项目用地范围坐标表", "未在内业成果/表格中匹配到项目用地范围坐标表.xlsx"))
        return 0, "项目用地范围坐标表缺失"
    estimated_pages = estimated_attachment_table_pages(records)
    if not city_level and estimated_pages > 30:
        for target in targets:
            idx = list(body).index(target)
            body.remove(target)
            remove_previous_attachment_title(body, idx, "项目用地范围坐标表")
        return 0, f"市级以下且项目用地范围坐标表估算 {estimated_pages} 页，已删除附件三"
    inserted = 0
    for pos, target in enumerate(targets):
        idx = list(body).index(target)
        body.remove(target)
        if pos == 0:
            tbl = make_word_table_from_raw_rows(records)
            body.insert(idx, tbl)
            inserted = len(records)
        else:
            remove_previous_attachment_title(body, idx, "项目用地范围坐标表")
    duplicate_note = f"，已删除重复附件三锚点 {len(targets) - 1} 处" if len(targets) > 1 else ""
    return inserted, f"项目用地范围坐标表已插入 {inserted} 行，估算 {estimated_pages} 页{duplicate_note}"


def remove_paragraph_section_properties(para: ET.Element) -> bool:
    ppr = para.find("w:pPr", NS)
    if ppr is None:
        return False
    removed = False
    for sect_pr in list(ppr.findall("w:sectPr", NS)):
        ppr.remove(sect_pr)
        removed = True
    return removed


def split_section_description(text: str) -> list[str]:
    cleaned = clean(text)
    if not cleaned:
        return []
    parts = re.split(
        r"(?=第[①②③④⑤⑥⑦⑧⑨⑩][^；。]{0,8}层[：:,，])|(?<!第)(?=[①②③④⑤⑥⑦⑧⑨⑩][^；。]{0,8}层[：:,，])|(?=第[①②③④⑤⑥⑦⑧⑨⑩]层以下)|(?<!第)(?=[①②③④⑤⑥⑦⑧⑨⑩]层以下)",
        cleaned,
    )
    lines = []
    for part in parts:
        line = part.strip()
        if not line:
            continue
        line = re.sub(r"^第([①②③④⑤⑥⑦⑧⑨⑩]层以下)", r"\1", line)
        lines.append(line)
    return lines


def normalize_layer_description(line: str) -> str:
    line = clean(line).strip()
    if line == "第":
        return ""
    line = re.sub(r"^第([①②③④⑤⑥⑦⑧⑨⑩])层[：:，,]\s*", r"\1层：", line)
    line = re.sub(r"^([①②③④⑤⑥⑦⑧⑨⑩])([^层：:，,]{1,8})层[：:，,]\s*", r"\1层：\2层，", line)
    line = re.sub(r"^([①②③④⑤⑥⑦⑧⑨⑩])层[：:，,]\s*", r"\1层：", line)
    line = re.sub(r"^(第)?([①②③④⑤⑥⑦⑧⑨⑩]层以下)[：:，,]?\s*为?", r"\2为", line)
    line = re.sub(r"：([^，；。]{1,12})：厚度约为", r"：\1，厚度约为", line)
    line = re.sub(r"：([^，；。]{1,12})，厚度", r"：\1，厚度", line)
    line = re.sub(r"^(①层：[^，；。]+，)(\d+(?:\.\d+)?m?-\d+(?:\.\d+)?m)", r"\1厚度约为\2", line)
    line = re.sub(r"^(②层：[^，；。]+，)(\d+(?:\.\d+)?m?-\d+(?:\.\d+)?m)", r"\1厚度约为\2", line)
    line = line.replace("：厚度", "：厚度")
    return line


def normalized_layer_lines(text: str) -> list[str]:
    return [line for line in (normalize_layer_description(line) for line in split_section_description(text)) if line]


def split_standard_description(text: str) -> list[str]:
    return normalized_layer_lines(re.sub(r"[\r\n]+", "", text))


def section_number_key(value: str) -> str:
    return normalize_section_number(value).replace("’", "′").replace("'", "′")


def arrange_section_descriptions_after_figures(
    root: ET.Element,
    records: list[dict[str, str]],
) -> int:
    if not records:
        return 0
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    by_number = {
        section_number_key(record.get("剖线编号") or record.get("剖面编号") or ""): record
        for record in records
    }
    if not by_number:
        return 0

    template_key = "__section_description_template__"
    existing: dict[str, tuple[ET.Element, list[ET.Element]]] = {}
    landscape_section_break: ET.Element | None = None
    blocks = list(body)
    idx = 0
    while idx < len(blocks):
        block = blocks[idx]
        text = element_text(block).strip()
        match = re.match(
            r"^剖[线面]\s*(?:([A-Z]-[A-Z][′'’]?)|\{\{剖[线面]编号}})\s*地层堆积情况[:：]",
            text,
        )
        if not match:
            idx += 1
            continue
        key = section_number_key(match.group(1)) if match.group(1) else template_key
        desc_blocks: list[ET.Element] = []
        stale_blocks: list[ET.Element] = [block]
        if idx > 0:
            previous_block = blocks[idx - 1]
            if (
                previous_block.tag == qn("w:p")
                and not element_text(previous_block).strip()
                and any(br.get(qn("w:type")) == "page" for br in previous_block.findall(".//w:br", NS))
            ):
                stale_blocks.insert(0, previous_block)
        cursor = idx + 1
        while cursor < len(blocks):
            next_text = element_text(blocks[cursor]).strip()
            if re.match(
                r"^剖[线面]\s*(?:[A-Z]-[A-Z][′'’]?|\{\{剖[线面]编号}})\s*地层堆积情况[:：]",
                next_text,
            ):
                break
            if next_text.startswith("共设置标准孔") or next_text.startswith("（二）"):
                break
            if next_text:
                desc_blocks.append(blocks[cursor])
            stale_blocks.append(blocks[cursor])
            cursor += 1
        for old in stale_blocks:
            sect_pr = paragraph_sect_pr(old)
            if landscape_section_break is None and is_landscape_sect_pr(sect_pr):
                landscape_section_break = copy.deepcopy(sect_pr)
        existing[key] = (block, desc_blocks)
        for old in stale_blocks:
            if old in list(body):
                body.remove(old)
        idx = cursor

    ordered_records = [
        record for record in records if section_number_key(record.get("剖线编号") or record.get("剖面编号") or "")
    ]
    generic_index = 0
    used_keys: set[str] = set()
    targets: list[tuple[ET.Element, str, dict[str, str]]] = []
    for block in list(body):
        caption = element_text(block).strip()
        match = re.match(r"^图\s*\d+\s+([A-Z]-[A-Z][′'’]?)\s*地层堆积剖[线面]图", caption)
        if match:
            key = section_number_key(match.group(1))
            record = by_number.get(key)
        elif re.match(r"^图\s*\d+\s+剖[线面]图$", caption):
            record = None
            while generic_index < len(ordered_records):
                candidate = ordered_records[generic_index]
                generic_index += 1
                candidate_key = section_number_key(candidate.get("剖线编号") or candidate.get("剖面编号") or "")
                if candidate_key not in used_keys:
                    record = candidate
                    key = candidate_key
                    break
        else:
            continue
        if not record:
            continue
        used_keys.add(key)
        targets.append((block, key, record))

    inserted = 0
    for target_index, (block, key, record) in enumerate(targets):
        insert_at = list(body).index(block) + 1
        body.insert(insert_at, make_page_break_paragraph())
        insert_at += 1
        section_label = record.get("剖线编号") or record.get("剖面编号") or key
        section_template = existing.get(key) or existing.get(template_key)
        title_para = copy.deepcopy((section_template or (make_plain_paragraph(""), []))[0])
        template_title = element_text(title_para)
        prefix = "剖面" if "剖面" in template_title else "剖线"
        title_text = f"{prefix}{section_label}地层堆积情况："
        set_element_text(title_para, title_text)
        set_paragraph_font(title_para, "宋体", "28")
        body.insert(insert_at, title_para)
        insert_at += 1
        template_desc = (section_template or (None, []))[1]
        desc_template = template_desc[0] if template_desc else make_plain_paragraph("")
        section_description = record.get("剖线地层描述") or record.get("剖面地层描述") or ""
        lines = normalized_layer_lines(section_description)
        if lines:
            para = copy.deepcopy(desc_template)
            set_paragraph_lines(para, lines)
            set_paragraph_font(para, "宋体", "28")
            body.insert(insert_at, para)
            insert_at += 1
        if target_index < len(targets) - 1:
            body.insert(insert_at, make_page_break_paragraph())
        else:
            body.insert(insert_at, make_section_break_paragraph(landscape_section_break))
        inserted += 1
    return inserted


def is_body_section_drawing_caption(text: str) -> bool:
    return bool(re.match(r"^图\s*\d+\s+.+地层堆积剖[线面]图$", clean(text)))


def is_section_drawing_boundary(text: str) -> bool:
    stripped = clean(text)
    return (
        is_body_section_drawing_caption(stripped)
        or stripped.startswith("（二）")
        or stripped.startswith("五、")
        or stripped.startswith("附表")
        or stripped.startswith("附图")
        or stripped.startswith("附件")
    )


def ensure_body_section_drawing_landscape_breaks(root: ET.Element) -> int:
    body, blocks = body_blocks(root)
    if body is None:
        return 0
    inserted = 0
    idx = 0
    while idx < len(blocks):
        block = blocks[idx]
        if block.tag != qn("w:p") or not is_body_section_drawing_caption(element_text(block)):
            idx += 1
            continue
        cursor = idx + 1
        has_landscape_break = False
        if cursor < len(blocks) and is_page_break_paragraph(blocks[cursor]):
            has_landscape_break = True
        while cursor < len(blocks):
            cursor_text = element_text(blocks[cursor]).strip()
            if cursor_text and is_section_drawing_boundary(cursor_text):
                break
            if blocks[cursor].tag == qn("w:p") and is_landscape_sect_pr(paragraph_sect_pr(blocks[cursor])):
                has_landscape_break = True
                break
            cursor += 1
        if not has_landscape_break:
            insert_at = list(body).index(block) + 1
            body.insert(insert_at, make_section_break_paragraph())
            inserted += 1
            blocks = list(body)
            idx = insert_at + 1
            continue
        idx += 1
    return inserted


def should_insert_image_for_paragraph(text: str) -> bool:
    stripped = text.strip()
    if stripped.startswith("目录"):
        return False
    return bool(
        re.search(r"\{\{图:[^{}]+}}", stripped)
        and (
            stripped.startswith("图")
            or stripped.startswith("附图")
            or re.fullmatch(r"\{\{图:[^{}]+}}", stripped)
        )
    )


def add_image_to_package(
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    slot: ImageSlot,
) -> tuple[str, str, int, int]:
    data, ext, width_px, height_px = prepare_image_data(slot)
    existing_numbers = []
    for name in entries:
        match = re.fullmatch(r"word/media/generated_image_(\d+)\.[a-z0-9]+", name)
        if match:
            existing_numbers.append(int(match.group(1)))
    idx = (max(existing_numbers) if existing_numbers else 0) + 1
    media_name = f"word/media/generated_image_{idx}{ext}"
    entries[media_name] = data
    ensure_content_type_defaults(content_types_root)
    rid = next_relationship_id(rels_root)
    rel = ET.SubElement(rels_root, f"{{{REL_NS}}}Relationship")
    rel.set("Id", rid)
    rel.set("Type", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image")
    rel.set("Target", f"media/generated_image_{idx}{ext}")
    return rid, media_name, width_px, height_px


def has_drawing(block: ET.Element) -> bool:
    return any(node.tag == qn("w:drawing") for node in block.iter())


def normalized_image_key(value: str) -> str:
    return normalize_key(value).replace("’", "'").replace("′", "'").replace("剖面", "剖线")


def image_key_exists(image_slots: dict[str, deque[ImageSlot]], key: str) -> bool:
    if key in image_slots:
        return True
    wanted = normalized_image_key(key)
    return any(normalized_image_key(candidate) == wanted for candidate in image_slots)


def pop_image_slot(image_slots: dict[str, deque[ImageSlot]], key: str) -> ImageSlot | None:
    queue = image_slots.get(key)
    if queue:
        return queue.popleft()
    wanted = normalized_image_key(key)
    for candidate, candidate_queue in image_slots.items():
        normalized = normalized_image_key(candidate)
        if normalized == wanted and candidate_queue:
            return candidate_queue.popleft()
    return None


def matching_image_queue(image_slots: dict[str, deque[ImageSlot]], key: str) -> deque[ImageSlot] | None:
    queue = image_slots.get(key)
    if queue:
        return queue
    wanted = normalized_image_key(key)
    for candidate, candidate_queue in image_slots.items():
        if normalized_image_key(candidate) == wanted and candidate_queue:
            return candidate_queue
    return None


def take_image_slots(image_slots: dict[str, deque[ImageSlot]], key: str, expand_all: bool) -> list[ImageSlot]:
    queue = matching_image_queue(image_slots, key)
    if not queue:
        return []
    if key in {"遗迹土样照", "遗迹现场照", "遗迹平、剖面图"}:
        return [queue.popleft()]
    if expand_all and len(queue) > 1:
        out = list(queue)
        queue.clear()
        return out
    return [queue.popleft()]


def matching_image_slots_snapshot(image_slots: dict[str, list[ImageSlot]], key: str) -> list[ImageSlot]:
    slots = image_slots.get(key)
    if slots:
        return list(slots)
    wanted = normalized_image_key(key)
    for candidate, candidate_slots in image_slots.items():
        if normalized_image_key(candidate) == wanted and candidate_slots:
            return list(candidate_slots)
    return []


def is_section_drawing_key(key: str) -> bool:
    normalized = normalized_image_key(key)
    if "剖线位置" in normalized or "剖面位置" in normalized:
        return False
    if "遗迹平剖线图" in normalized or "遗迹平剖面图" in normalized:
        return False
    return (
        "地层堆积剖线图" in normalized
        or "地层堆积剖面图" in normalized
        or bool(re.search(r"[A-Z]-[A-Z]['′’]?地层堆积剖线图", normalized))
    )


def take_appendix_image_slots(
    image_slots: dict[str, list[ImageSlot]],
    key: str,
    expand_all: bool,
) -> list[ImageSlot]:
    if is_section_drawing_key(key):
        section_slots = matching_image_slots_snapshot(image_slots, "剖线图")
        if section_slots:
            return section_slots
    slots = matching_image_slots_snapshot(image_slots, key)
    if not slots:
        return []
    if key in {"遗迹土样照", "遗迹现场照", "遗迹平、剖面图"}:
        return [slots[0]]
    if expand_all and len(slots) > 1:
        return slots
    return [slots[0]]


def find_placeholder_image_index(body: ET.Element, caption_index: int) -> int | None:
    blocks = list(body)
    for idx in range(caption_index - 1, max(-1, caption_index - 5), -1):
        block = blocks[idx]
        text = element_text(block).strip()
        if has_drawing(block) and not text:
            return idx
        if text and not text.startswith("图") and not text.startswith("附图"):
            break
    return None


def find_following_placeholder_image_index(body: ET.Element, caption_index: int) -> int | None:
    blocks = list(body)
    for idx in range(caption_index + 1, min(len(blocks), caption_index + 5)):
        block = blocks[idx]
        text = element_text(block).strip()
        if has_drawing(block) and not text:
            return idx
        if text and not text.startswith("附图"):
            break
    return None


def is_optional_empty_photo_key(key: str) -> bool:
    return key in FIELDWORK_OPTIONAL_PHOTO_KEYS


def remove_image_placeholder_block(body: ET.Element, caption_block: ET.Element, is_appendix: bool = False) -> None:
    blocks = list(body)
    if caption_block not in blocks:
        return
    caption_index = blocks.index(caption_block)
    placeholder_idx = (
        find_following_placeholder_image_index(body, caption_index)
        if is_appendix
        else find_placeholder_image_index(body, caption_index)
    )
    if placeholder_idx is not None:
        placeholder = list(body)[placeholder_idx]
        if placeholder is not caption_block and placeholder in list(body):
            body.remove(placeholder)
    if caption_block in list(body):
        body.remove(caption_block)


def apply_caption_text(text: str, token: str, caption: str, auto_number: int | None) -> str:
    text = text.replace(token, caption)
    if "{{自动图号}}" in text and auto_number is not None:
        text = text.replace("{{自动图号}}", str(auto_number))
    text = re.sub(r"\s+", " ", text).strip()
    return clean_caption_suffix(text)


def apply_image_insert_error_text(text: str, token: str, key: str, auto_number: int | None) -> str:
    return apply_caption_text(text, token, image_insert_error_marker(key), auto_number)


def clean_caption_suffix(text: str) -> str:
    match = re.match(r"^(图\s*(\d+)\s+.+?)(\2)$", text)
    if match:
        text = match.group(1)
    text = re.sub(r"^(图\s*\d+\s+.+(?:图|照))\d+$", r"\1", text)
    return text


def normalize_caption_suffixes(root: ET.Element) -> int:
    changed = 0
    for para in root.findall(".//w:p", NS):
        text = element_text(para)
        cleaned = clean_caption_suffix(re.sub(r"\s+", " ", text).strip())
        if cleaned != text.strip():
            set_element_text(para, cleaned)
            changed += 1
    return changed


def set_body_caption_style(para: ET.Element) -> None:
    ppr = para.find("w:pPr", NS)
    if ppr is None:
        ppr = ET.Element(qn("w:pPr"))
        para.insert(0, ppr)
    jc = ppr.find("w:jc", NS)
    if jc is None:
        jc = ET.SubElement(ppr, qn("w:jc"))
    jc.set(qn("w:val"), "center")
    runs = para.findall("w:r", NS)
    if not runs:
        return

    def style_caption_run(run: ET.Element) -> None:
        set_run_style(run, bold=False, size_half_points="20")
        rpr = ensure_rpr(run)
        fonts = rpr.find("w:rFonts", NS)
        if fonts is None:
            fonts = ET.SubElement(rpr, qn("w:rFonts"))
        for attr, value in {
            "w:ascii": "黑体",
            "w:hAnsi": "黑体",
            "w:eastAsia": "黑体",
            "w:cs": "黑体",
        }.items():
            fonts.set(qn(attr), value)

    if has_drawing(para):
        for run in runs:
            if run.find(".//w:t", NS) is not None:
                style_caption_run(run)
        return

    keep = runs[0]
    for run in runs[1:]:
        para.remove(run)
    style_caption_run(keep)


def renumber_figure_captions(root: ET.Element) -> int:
    changed = 0
    number = 1
    for para in root.findall(".//w:p", NS):
        text = re.sub(r"\s+", " ", element_text(para)).strip()
        if not text or text.startswith("目录") or text.startswith("附图"):
            continue
        match = re.match(r"^图\s*(?:\d+)?\s+(.+)$", text)
        if not match:
            continue
        new_text = f"图{number} {match.group(1).strip()}"
        if new_text != text:
            set_element_text(para, new_text)
            changed += 1
        set_body_caption_style(para)
        number += 1
    return changed


def renumber_appendix_captions(root: ET.Element) -> int:
    changed = 0
    number = 1
    for para in root.findall(".//w:p", NS):
        text = re.sub(r"\s+", " ", element_text(para)).strip()
        if not text.startswith("附图"):
            continue
        if re.search(r"\s+\d+$", text):
            continue
        match = re.match(r"^附图\s*([一二三四五六七八九十]+)\s*(.*)$", text)
        if not match:
            continue
        suffix = match.group(2).strip()
        new_text = f"附图{CHINESE_ORDINAL.get(number, str(number))}"
        if suffix:
            new_text += f" {suffix}"
        if new_text != text:
            set_element_text(para, new_text)
            changed += 1
        number += 1
    return changed


def mark_toc_fields_for_update(entries: dict[str, bytes], root: ET.Element | None = None) -> int:
    """Ask Word to refresh TOC fields while preserving TOC font/paragraph styles."""
    if root is None:
        root = ET.fromstring(entries["word/document.xml"])
    dirty_count = 0
    for para in root.findall(".//w:p", NS):
        instr = "".join(node.text or "" for node in para.findall(".//w:instrText", NS))
        if not re.search(r"\bTOC\b", instr):
            continue
        for fld_char in para.findall(".//w:fldChar", NS):
            if fld_char.get(qn("w:fldCharType")) == "begin":
                if fld_char.get(qn("w:dirty")) != "true":
                    fld_char.set(qn("w:dirty"), "true")
                    dirty_count += 1
                break
    entries["word/document.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    settings_name = "word/settings.xml"
    if settings_name in entries:
        settings_root = ET.fromstring(entries[settings_name])
    else:
        settings_root = ET.Element(qn("w:settings"))
    update_fields = settings_root.find("w:updateFields", NS)
    if update_fields is None:
        update_fields = ET.SubElement(settings_root, qn("w:updateFields"))
    update_fields.set(qn("w:val"), "true")
    entries[settings_name] = ET.tostring(settings_root, encoding="utf-8", xml_declaration=True)
    return dirty_count


def normalize_docx_with_python_docx(path: Path) -> bool:
    """Normalize package details when python-docx is available.

    The core generator writes OOXML directly. Some legacy templates contain
    package metadata that LibreOffice tolerates until new JPG attachments are
    added. Opening and saving once through python-docx normalizes those package
    details without changing media bytes.
    """
    try:
        from docx import Document  # type: ignore
    except Exception:
        return False
    try:
        doc = Document(path)
        doc.save(path)
        return True
    except Exception:
        return False


def insert_image_paragraph(
    body: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    slot: ImageSlot,
    insert_at: int,
    doc_pr_id: int,
) -> int:
    rid, media_name, width_px, height_px = add_image_to_package(entries, rels_root, content_types_root, slot)
    width, height = slot_display_extent(slot, width_px, height_px)
    image_para = make_image_paragraph(
        rid,
        Path(media_name).name,
        width,
        height,
        doc_pr_id,
        border=is_attachment_image_slot(slot) or (slot.source == "drawing" and is_section_drawing(slot)),
        rotation_degrees=image_display_rotation_degrees(slot),
        alt_text=slot.key,
    )
    body.insert(insert_at, image_para)
    return doc_pr_id + 1


def clear_paragraph_content(para: ET.Element) -> None:
    for child in list(para):
        if child.tag != qn("w:pPr"):
            para.remove(child)


def append_image_to_paragraph(
    para: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    slot: ImageSlot,
    doc_pr_id: int,
) -> int:
    rid, media_name, width_px, height_px = add_image_to_package(entries, rels_root, content_types_root, slot)
    width, height = slot_display_extent(slot, width_px, height_px)
    image_para = make_image_paragraph(
        rid,
        Path(media_name).name,
        width,
        height,
        doc_pr_id,
        border=is_attachment_image_slot(slot) or (slot.source == "drawing" and is_section_drawing(slot)),
        rotation_degrees=image_display_rotation_degrees(slot),
        alt_text=slot.key,
    )
    clear_paragraph_content(para)
    ppr = para.find("w:pPr", NS)
    if ppr is None:
        new_ppr = image_para.find("w:pPr", NS)
        if new_ppr is not None:
            para.insert(0, copy.deepcopy(new_ppr))
    for run in image_para.findall("w:r", NS):
        para.append(copy.deepcopy(run))
    return doc_pr_id + 1


def attachment_image_slot(path: Path, key: str) -> ImageSlot:
    return ImageSlot(path=path, caption=path.stem, source="photo", key=key)


def append_company_seal_to_paragraph(
    para: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    seal_path: Path,
    doc_pr_id: int,
) -> int:
    slot = attachment_image_slot(seal_path, "公司印章")
    rid, media_name, width_px, height_px = add_image_to_package(entries, rels_root, content_types_root, slot)
    ratio = width_px / height_px if height_px else 1
    width = cm_to_emu(COMPANY_SEAL_WIDTH_CM)
    height = int(width / ratio)
    image_para = make_floating_image_paragraph(
        rid,
        Path(media_name).name,
        width,
        height,
        doc_pr_id,
        vertical_offset=cm_to_emu(COMPANY_SEAL_VERTICAL_OFFSET_CM),
        alt_text="公司印章",
    )
    clear_paragraph_content(para)
    ppr = para.find("w:pPr", NS)
    if ppr is None:
        new_ppr = image_para.find("w:pPr", NS)
        if new_ppr is not None:
            para.insert(0, copy.deepcopy(new_ppr))
    for run in image_para.findall("w:r", NS):
        para.append(copy.deepcopy(run))
    return doc_pr_id + 1


def insert_company_seal_placeholders(
    root: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    context: dict[str, object],
    doc_pr_id: int,
) -> tuple[int, int, list[str]]:
    company_dir = context.get("company_dir")
    if not isinstance(company_dir, Path):
        return doc_pr_id, 0, []
    seal_path = find_company_attachment(company_dir, "公司印章")
    inserted = 0
    missing: list[str] = []
    for para in root.findall(".//w:p", NS):
        text = element_text(para).strip()
        if text != "{{公司印章}}":
            continue
        if seal_path is None:
            set_element_text(para, image_insert_error_marker("公司印章", "公司资料包未找到公司印章或公章"))
            missing.append("公司印章")
            continue
        doc_pr_id = append_company_seal_to_paragraph(
            para,
            entries,
            rels_root,
            content_types_root,
            seal_path,
            doc_pr_id,
        )
        inserted += 1
    return doc_pr_id, inserted, missing


def insert_company_attachment_placeholders(
    root: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    context: dict[str, object],
    doc_pr_id: int,
) -> tuple[int, int, list[str]]:
    company_dir = context.get("company_dir")
    if not isinstance(company_dir, Path):
        return doc_pr_id, 0, []
    body, _ = body_blocks(root)
    inserted = 0
    missing: list[str] = []
    for para in root.findall(".//w:p", NS):
        text = element_text(para).strip()
        match = re.fullmatch(r"\{\{公司附件:([^{}]+)}}", text)
        if not match:
            continue
        key = clean(match.group(1))
        path = find_company_attachment(company_dir, key)
        if path is None:
            set_element_text(para, image_insert_error_marker(f"公司附件:{key}", "公司资料包未找到对应附件"))
            missing.append(f"公司附件:{key}")
            continue
        doc_pr_id = append_image_to_paragraph(
            para,
            entries,
            rels_root,
            content_types_root,
            attachment_image_slot(path, f"公司附件:{key}"),
            doc_pr_id,
        )
        if company_attachment_requires_page_break(key) and body is not None:
            body_children = list(body)
            if para in body_children:
                body.insert(body_children.index(para) + 1, make_page_break_paragraph())
        inserted += 1
    return doc_pr_id, inserted, missing


def attachment_categories_for_group(group_name: str) -> list[str]:
    if group_name == "身份证证书":
        return ["身份证", "证书"]
    if group_name == "劳动合同社保":
        return ["劳动合同", "社保"]
    return []


def clone_attachment_text_block(template: ET.Element, person_index: int, person_name: str) -> ET.Element:
    clone = copy.deepcopy(template)
    text = element_text(clone)
    text = text.replace("{{附件人员序号}}", str(person_index)).replace("{{附件人员姓名}}", person_name)
    set_element_text(clone, text)
    return clone


def expand_personnel_attachment_groups(
    root: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    context: dict[str, object],
    doc_pr_id: int,
) -> tuple[int, int, list[str]]:
    body, _ = body_blocks(root)
    if body is None:
        return doc_pr_id, 0, []
    order = context.get("attachment_order")
    attachments = context.get("attachments")
    if not isinstance(order, list) or not isinstance(attachments, dict):
        return doc_pr_id, 0, []
    inserted = 0
    missing: list[str] = []
    while True:
        blocks = list(body)
        start_idx = end_idx = None
        group_name = ""
        for idx, block in enumerate(blocks):
            text = element_text(block).strip()
            match = re.fullmatch(r"\{\{#人员附件组:([^{}]+)}}", text)
            if match:
                start_idx = idx
                group_name = clean(match.group(1))
                break
        if start_idx is None:
            break
        end_marker = f"{{{{/人员附件组:{group_name}}}}}"
        for idx in range(start_idx + 1, len(blocks)):
            if element_text(blocks[idx]).strip() == end_marker:
                end_idx = idx
                break
        if end_idx is None:
            set_element_text(blocks[start_idx], report_error_marker(MATCH_ERROR_TEXT, f"人员附件组未闭合：{group_name}"))
            break
        template_blocks = blocks[start_idx + 1 : end_idx]
        for block in blocks[start_idx : end_idx + 1]:
            if block in list(body):
                body.remove(block)
        insert_at = start_idx
        categories = attachment_categories_for_group(group_name)
        for person_index, person_name in enumerate(order, start=1):
            for template_block in template_blocks:
                text = element_text(template_block).strip()
                category_match = re.fullmatch(r"\{\{人员附件:([^{}]+)}}", text)
                if category_match:
                    category = clean(category_match.group(1))
                    if category not in categories:
                        continue
                    paths = attachments.get((person_name, category), [])
                    if not paths:
                        body.insert(
                            insert_at,
                            make_plain_paragraph(image_insert_error_marker(f"{person_name}:{category}", "人员资料包未找到对应附件")),
                        )
                        insert_at += 1
                        missing.append(f"{person_name}:{category}")
                        continue
                    for path in paths:
                        doc_pr_id = insert_image_paragraph(
                            body,
                            entries,
                            rels_root,
                            content_types_root,
                            attachment_image_slot(path, f"人员附件:{category}"),
                            insert_at,
                            doc_pr_id,
                        )
                        insert_at += 1
                        inserted += 1
                    if personnel_attachment_requires_page_break(category):
                        body.insert(insert_at, make_page_break_paragraph())
                        insert_at += 1
                    continue
                clone = clone_attachment_text_block(template_block, person_index, person_name)
                body.insert(insert_at, clone)
                insert_at += 1
    return doc_pr_id, inserted, missing


def insert_company_personnel_assets(
    root: ET.Element,
    entries: dict[str, bytes],
    context: dict[str, object],
) -> tuple[int, list[str]]:
    rels_name = "word/_rels/document.xml.rels"
    rels_root = ET.fromstring(entries[rels_name]) if rels_name in entries else ET.Element(f"{{{REL_NS}}}Relationships")
    content_types_root = ET.fromstring(entries["[Content_Types].xml"])
    doc_pr_id = 8000
    doc_pr_id, seal_inserted, seal_missing = insert_company_seal_placeholders(
        root, entries, rels_root, content_types_root, context, doc_pr_id
    )
    doc_pr_id, company_inserted, company_missing = insert_company_attachment_placeholders(
        root, entries, rels_root, content_types_root, context, doc_pr_id
    )
    doc_pr_id, personnel_inserted, personnel_missing = expand_personnel_attachment_groups(
        root, entries, rels_root, content_types_root, context, doc_pr_id
    )
    entries[rels_name] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)
    entries["[Content_Types].xml"] = ET.tostring(content_types_root, encoding="utf-8", xml_declaration=True)
    return seal_inserted + company_inserted + personnel_inserted, seal_missing + company_missing + personnel_missing


def insert_table_cell_images(
    root: ET.Element,
    entries: dict[str, bytes],
    image_slots: dict[str, deque[ImageSlot]],
    inserted_signatures: set[str],
) -> tuple[int, list[str]]:
    rels_name = "word/_rels/document.xml.rels"
    rels_root = ET.fromstring(entries[rels_name])
    content_types_root = ET.fromstring(entries["[Content_Types].xml"])
    inserted = 0
    missing: list[str] = []
    doc_pr_id = 5000
    for cell in root.findall(".//w:tc", NS):
        for para in cell.findall("w:p", NS):
            text = element_text(para).strip()
            match = re.fullmatch(r"\{\{图:([^{}]+)}}", text)
            if not match:
                continue
            key = clean(match.group(1))
            slot = pop_image_slot(image_slots, key)
            if slot is None:
                if not is_optional_empty_photo_key(key) and not image_key_exists(image_slots, key):
                    missing.append(key)
                set_element_text(para, "" if is_optional_empty_photo_key(key) else image_insert_error_marker(key))
                continue
            signature = image_slot_signature(slot)
            if signature in inserted_signatures:
                set_element_text(para, image_insert_error_marker(key, "图片已被前文使用，未重复插入"))
                continue
            doc_pr_id = append_image_to_paragraph(para, entries, rels_root, content_types_root, slot, doc_pr_id)
            inserted_signatures.add(signature)
            inserted += 1
    entries[rels_name] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)
    entries["[Content_Types].xml"] = ET.tostring(content_types_root, encoding="utf-8", xml_declaration=True)
    return inserted, missing


def insert_appendix_page_break(body: ET.Element, image_index: int) -> None:
    blocks = list(body)
    cursor = image_index + 1
    while cursor < len(blocks) and is_empty_paragraph(blocks[cursor]) and not is_page_break_paragraph(blocks[cursor]):
        cursor += 1
    if cursor < len(blocks) and is_page_break_paragraph(blocks[cursor]):
        return
    body.insert(image_index + 1, make_page_break_paragraph())


def is_project_location_caption_text(text: str) -> bool:
    return bool(re.match(r"^图\s*\d+\s+项目地块在.+位置示意图$", clean(text)))


def complete_project_location_maps(
    body: ET.Element,
    entries: dict[str, bytes],
    rels_root: ET.Element,
    content_types_root: ET.Element,
    image_slots: dict[str, deque[ImageSlot]],
    inserted_signatures: set[str],
    doc_pr_id: int,
) -> tuple[int, int, int]:
    location_slots = list(image_slots.get("项目地块位置示意图", []))
    if not location_slots:
        return doc_pr_id, 0, 0

    captions = [block for block in list(body) if block.tag == qn("w:p") and is_project_location_caption_text(element_text(block))]
    if not captions:
        return doc_pr_id, 0, 0

    last_caption = captions[-1]
    insert_at = list(body).index(last_caption) + 1
    removed = 0
    while insert_at < len(list(body)):
        block = list(body)[insert_at]
        block_text = element_text(block).strip()
        if block_text or block.tag != qn("w:p"):
            break
        if has_drawing(block) or is_empty_paragraph(block):
            body.remove(block)
            removed += 1
            continue
        break

    added = 0
    missing_slots = [slot for slot, _ in keep_unseen_image_slots(location_slots, inserted_signatures)]
    for slot in missing_slots:
        caption = slot.caption or default_caption_from_path(slot.path)
        doc_pr_id = insert_image_paragraph(body, entries, rels_root, content_types_root, slot, insert_at, doc_pr_id)
        inserted_signatures.add(image_slot_signature(slot))
        insert_at += 1
        body.insert(insert_at, make_plain_paragraph(f"图0 {caption}"))
        insert_at += 1
        added += 1
    return doc_pr_id, added, removed


def mark_missing_project_location_map_errors(root: ET.Element, drawings_dir: Path | None) -> int:
    if drawings_dir is None or not drawings_dir.exists():
        return 0
    expected = [default_caption_from_path(path) for path in sorted_project_location_maps(drawings_dir)]
    if not expected:
        return 0
    body, _ = body_blocks(root)
    if body is None:
        return 0
    existing_text = "\n".join(element_text(block) for block in list(body))
    missing = [caption for caption in expected if caption and caption not in existing_text]
    if not missing:
        return 0

    captions = [
        block
        for block in list(body)
        if block.tag == qn("w:p") and is_project_location_caption_text(element_text(block))
    ]
    if captions:
        insert_at = list(body).index(captions[-1]) + 1
    else:
        insert_at = 0
        for idx, block in enumerate(list(body)):
            text = element_text(block)
            if "项目概况" in text or "项目用地红线" in text or "基于国家2000坐标系" in text:
                insert_at = idx + 1
                break
    for caption in missing:
        body.insert(
            insert_at,
            make_plain_paragraph(report_error_marker(IMAGE_INSERT_ERROR_TEXT, f"未插入必需位置图：{caption}")),
        )
        insert_at += 1
    return len(missing)


def is_page_break_paragraph(block: ET.Element) -> bool:
    return block.tag == qn("w:p") and not element_text(block).strip() and any(
        br.get(qn("w:type")) == "page" for br in block.findall(".//w:br", NS)
    )


def is_section_break_paragraph(block: ET.Element) -> bool:
    return block.tag == qn("w:p") and not element_text(block).strip() and paragraph_sect_pr(block) is not None


def ensure_page_break_after_block(body: ET.Element, block: ET.Element) -> int:
    blocks = list(body)
    if block not in blocks:
        return 0
    idx = blocks.index(block)
    if idx + 1 < len(blocks) and is_page_break_paragraph(blocks[idx + 1]):
        return 0
    body.insert(idx + 1, make_page_break_paragraph())
    return 1


def attachment_key_from_image_paragraph(block: ET.Element) -> str:
    doc_pr = block.find(".//wp:docPr", NS)
    if doc_pr is not None:
        key = clean(doc_pr.get("descr", ""))
        if key.startswith("公司附件:") or key.startswith("人员附件:"):
            return key
    c_nv_pr = block.find(".//pic:cNvPr", NS)
    if c_nv_pr is not None:
        key = clean(c_nv_pr.get("descr", ""))
        if key.startswith("公司附件:") or key.startswith("人员附件:"):
            return key
    return ""


def attachment_key_requires_page_break(key: str) -> bool:
    if key.startswith("公司附件:"):
        return company_attachment_requires_page_break(key.split(":", 1)[1])
    if key.startswith("人员附件:"):
        return personnel_attachment_requires_page_break(key.split(":", 1)[1])
    return False


def ensure_required_attachment_page_breaks(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    added = 0
    for block in list(body):
        key = attachment_key_from_image_paragraph(block)
        if key and attachment_key_requires_page_break(key):
            added += ensure_page_break_after_block(body, block)
    return added


def ensure_page_break_after_last_project_location_caption(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    captions = [
        block
        for block in list(body)
        if block.tag == qn("w:p") and is_project_location_caption_text(element_text(block))
    ]
    if not captions:
        return 0
    last_caption = captions[-1]
    blocks = list(body)
    idx = blocks.index(last_caption)
    cursor = idx + 1
    while cursor < len(blocks):
        block = blocks[cursor]
        if is_empty_paragraph(block) and not is_page_break_paragraph(block) and paragraph_sect_pr(block) is None:
            body.remove(block)
            blocks = list(body)
            continue
        break
    blocks = list(body)
    idx = blocks.index(last_caption)
    next_block = blocks[idx + 1] if idx + 1 < len(blocks) else None
    if next_block is not None and is_page_break_paragraph(next_block):
        return 0
    body.insert(idx + 1, make_page_break_paragraph())
    return 1


def is_survey_unit_numbering_paragraph(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text)
    return (
        "勘探单元以" in normalized
        and "U+数字" in normalized
        and "格式进行编号" in normalized
        and re.search(r"分别为U\d+", normalized) is not None
    )


def ensure_page_break_after_survey_unit_numbering(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    target = None
    for block in list(body):
        if block.tag != qn("w:p"):
            continue
        if is_survey_unit_numbering_paragraph(element_text(block)):
            target = block
            break
    if target is None:
        return 0
    return ensure_page_break_after_block(body, target)


def remove_extra_consecutive_page_breaks(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    removed = 0
    previous_was_break = False
    previous_was_section_break = False
    for block in list(body):
        if is_page_break_paragraph(block):
            idx = list(body).index(block)
            next_block = list(body)[idx + 1] if idx + 1 < len(list(body)) else None
            if previous_was_break or previous_was_section_break or (
                next_block is not None and is_section_break_paragraph(next_block)
            ):
                body.remove(block)
                removed += 1
                continue
            previous_was_break = True
            previous_was_section_break = False
        elif is_section_break_paragraph(block):
            previous_was_section_break = True
            previous_was_break = False
        elif previous_was_break and is_empty_paragraph(block):
            idx = list(body).index(block)
            next_block = list(body)[idx + 1] if idx + 1 < len(list(body)) else None
            if next_block is not None and is_page_break_paragraph(next_block):
                body.remove(block)
                removed += 1
                continue
        else:
            previous_was_break = False
            previous_was_section_break = False
    return removed


def remove_redundant_empty_section_break_pages(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    removed = 0
    captions = {
        "项目地块在白音察干镇位置示意图",
        "项目地块卫星图",
    }
    for block in list(body):
        text = element_text(block).strip()
        if not any(text.endswith(caption) for caption in captions):
            continue
        blocks = list(body)
        if block not in blocks:
            continue
        cursor = blocks.index(block) + 1
        empty_blocks: list[ET.Element] = []
        while cursor < len(blocks):
            candidate = blocks[cursor]
            candidate_text = element_text(candidate).strip()
            has_drawing = bool(candidate.findall(".//w:drawing", NS))
            if candidate_text or has_drawing:
                break
            sect_pr = paragraph_sect_pr(candidate)
            if sect_pr is not None:
                make_section_break_continuous(sect_pr)
                for empty in empty_blocks:
                    if empty in list(body):
                        body.remove(empty)
                        removed += 1
                break
            empty_blocks.append(candidate)
            cursor += 1
    return removed


def is_empty_paragraph(block: ET.Element) -> bool:
    return (
        block.tag == qn("w:p")
        and not element_text(block).strip()
        and block.find(".//w:drawing", NS) is None
        and block.find(".//w:pict", NS) is None
    )


def remove_stale_section_description_placeholders(root: ET.Element) -> int:
    body, _ = body_blocks(root)
    if body is None:
        return 0
    removed = 0
    for block in list(body):
        if block not in list(body) or block.tag != qn("w:p"):
            continue
        text = element_text(block).strip()
        if not re.fullmatch(r"剖[线面]地层堆积情况[:：]?", text):
            continue
        idx = list(body).index(block)
        body.remove(block)
        removed += 1
        while idx < len(list(body)):
            next_block = list(body)[idx]
            if not is_empty_paragraph(next_block):
                break
            # Empty template leftovers often hold page/section breaks; remove
            # them with the stale unnumbered section-description title.
            body.remove(next_block)
            removed += 1
    return removed


def insert_images(
    root: ET.Element,
    entries: dict[str, bytes],
    image_slots: dict[str, deque[ImageSlot]],
    inserted_signatures: set[str],
) -> tuple[int, list[str]]:
    body, blocks = body_blocks(root)
    if body is None:
        return 0, []

    appendix_image_slots = {key: list(queue) for key, queue in image_slots.items()}
    rels_name = "word/_rels/document.xml.rels"
    if rels_name in entries:
        rels_root = ET.fromstring(entries[rels_name])
    else:
        rels_root = ET.Element(f"{{{REL_NS}}}Relationships")
    content_types_root = ET.fromstring(entries["[Content_Types].xml"])

    inserted = 0
    replaced = 0
    missing: list[str] = []
    doc_pr_id = 1000
    last_figure_number = 0
    expanded_keys: set[tuple[str, str]] = set()
    appendix_inserted_signatures: set[str] = set()
    consumed_single_expansion_keys: set[tuple[str, str]] = set()
    used_slot_history: dict[tuple[str, str], list[ImageSlot]] = {}
    for block in list(blocks):
        if block.tag != qn("w:p"):
            continue
        text = element_text(block)
        token_match = re.search(r"\{\{图:([^{}]+)}}", text)
        if not token_match:
            number_match = re.match(r"\s*图\s*(\d+)", text)
            if number_match:
                last_figure_number = int(number_match.group(1))
            continue
        key = clean(token_match.group(1))
        if not should_insert_image_for_paragraph(text):
            set_element_text(block, text.replace(token_match.group(0), key))
            continue
        is_auto_number = "{{自动图号}}" in text
        expand_all = is_auto_number or text.strip().startswith("图")
        is_appendix = text.strip().startswith("附图")
        image_scope = "appendix" if is_appendix else "body"
        scoped_key = (image_scope, key)
        scoped_signatures = appendix_inserted_signatures if is_appendix else inserted_signatures
        if scoped_key in consumed_single_expansion_keys:
            remove_image_placeholder_block(body, block, is_appendix=is_appendix)
            continue
        slots = (
            take_appendix_image_slots(appendix_image_slots, key, expand_all=expand_all)
            if is_appendix
            else take_image_slots(image_slots, key, expand_all=expand_all)
        )
        slot_entries = keep_unseen_image_slots(slots, scoped_signatures)
        duplicate_only = bool(slots) and not slot_entries
        slots = [slot for slot, _ in slot_entries]
        auto_number = None
        if is_auto_number:
            last_figure_number += 1
            auto_number = last_figure_number
        else:
            number_match = re.match(r"\s*图\s*(\d+)", text)
            if number_match:
                last_figure_number = int(number_match.group(1))
        if not slots:
            if duplicate_only:
                remove_image_placeholder_block(body, block, is_appendix=is_appendix)
                continue
            if scoped_key in expanded_keys:
                remove_image_placeholder_block(body, block, is_appendix=is_appendix)
                continue
            if is_optional_empty_photo_key(key):
                remove_image_placeholder_block(body, block, is_appendix=is_appendix)
                continue
            if image_key_exists(image_slots, key) or scoped_key in used_slot_history:
                remove_image_placeholder_block(body, block, is_appendix=is_appendix)
                continue
            missing.append(key)
            set_element_text(block, apply_image_insert_error_text(text, token_match.group(0), key, auto_number))
            continue
        used_slot_history[scoped_key] = list(slots)
        if len(slots) > 1:
            expanded_keys.add(scoped_key)
        if key in SINGLE_EXPANSION_IMAGE_KEYS:
            consumed_single_expansion_keys.add(scoped_key)

        slot = slots[0]
        slot_signature = slot_entries[0][1]
        insert_at = list(body).index(block)
        placeholder_idx = (
            find_following_placeholder_image_index(body, insert_at)
            if is_appendix
            else find_placeholder_image_index(body, insert_at)
        )
        if placeholder_idx is not None:
            old_block = list(body)[placeholder_idx]
            body.remove(old_block)
            insert_idx = list(body).index(block) + 1 if is_appendix else placeholder_idx
            doc_pr_id = insert_image_paragraph(
                body, entries, rels_root, content_types_root, slot, insert_idx, doc_pr_id
            )
            scoped_signatures.add(slot_signature)
            if is_appendix:
                insert_appendix_page_break(body, insert_idx)
            replaced += 1
        else:
            insert_idx = insert_at + 1 if is_appendix else insert_at
            doc_pr_id = insert_image_paragraph(body, entries, rels_root, content_types_root, slot, insert_idx, doc_pr_id)
            scoped_signatures.add(slot_signature)
            if is_appendix:
                insert_appendix_page_break(body, insert_idx)
            inserted += 1
        caption = slot.caption or key
        set_element_text(block, apply_caption_text(text, token_match.group(0), caption, auto_number))
        if not is_appendix and not is_section_drawing_key(key):
            remove_paragraph_section_properties(block)
        if not is_appendix and key == "遗迹分布示意图":
            ensure_page_break_after_block(body, block)

        insertion_index = list(body).index(block) + 1
        if is_appendix:
            insertion_index += 2
        elif key == "遗迹分布示意图":
            insertion_index += 1
        for extra_slot, extra_signature in slot_entries[1:]:
            last_figure_number += 1
            caption_block = copy.deepcopy(block)
            extra_caption = extra_slot.caption or key
            set_element_text(caption_block, apply_caption_text(text, token_match.group(0), extra_caption, last_figure_number))
            if is_appendix:
                body.insert(insertion_index, caption_block)
                doc_pr_id = insert_image_paragraph(
                    body, entries, rels_root, content_types_root, extra_slot, insertion_index + 1, doc_pr_id
                )
                scoped_signatures.add(extra_signature)
                insert_appendix_page_break(body, insertion_index + 1)
                insertion_index += 3
            else:
                doc_pr_id = insert_image_paragraph(
                    body, entries, rels_root, content_types_root, extra_slot, insertion_index, doc_pr_id
                )
                scoped_signatures.add(extra_signature)
                body.insert(insertion_index + 1, caption_block)
                if key == "遗迹分布示意图":
                    body.insert(insertion_index + 2, make_page_break_paragraph())
                    insertion_index += 3
                else:
                    insertion_index += 2
            inserted += 1

    doc_pr_id, completed_location_maps, stale_location_blocks_removed = complete_project_location_maps(
        body,
        entries,
        rels_root,
        content_types_root,
        image_slots,
        inserted_signatures,
        doc_pr_id,
    )
    if completed_location_maps:
        inserted += completed_location_maps
    if stale_location_blocks_removed:
        replaced += stale_location_blocks_removed

    entries[rels_name] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)
    entries["[Content_Types].xml"] = ET.tostring(content_types_root, encoding="utf-8", xml_declaration=True)
    return inserted + replaced, missing


def fill_docx(
    template: Path,
    form: Path,
    output: Path,
    drawings_dir: Path | None = None,
    photos_dir: Path | None = None,
    skip_region_api: bool = True,
    personnel_set: str | None = None,
) -> tuple[Path, Path]:
    ensure_writable_target(output, "生成输出")
    fields, tables, notes = load_form(form)
    company_personnel_context, company_personnel_notes = load_company_personnel_context(fields, personnel_set)
    notes.extend(company_personnel_notes)
    notes.extend(import_external_tables(tables, drawings_dir))
    fill_coordinate_basepoint(fields, tables, notes)
    notes.extend(import_site_record_tables(fields, tables, photos_dir))
    refresh_section_derived_fields(fields, tables, notes)
    notes.extend(enrich_region_overview_with_api(fields, output, skip=skip_region_api, fail_on_error=not skip_region_api))
    marked_generation_errors = mark_required_generation_errors(fields, notes)
    check_path = output.with_name(f"{output.stem}-生成检查报告.txt")
    errors = [f"必填字段为空：{key}" for key in REQUIRED_FIELDS if not fields.get(key)]
    if errors:
        output.parent.mkdir(parents=True, exist_ok=True)
        check_path.write_text("生成检查报告\n\n## 错误\n" + "\n".join(f"- {item}" for item in errors), encoding="utf-8")
        return output, check_path
    image_slots = build_image_slots(tables, form, drawings_dir, photos_dir)
    notes.extend(add_data_organization_work_photo_library_slot(image_slots))

    with zipfile.ZipFile(template) as zf:
        entries = {name: zf.read(name) for name in zf.namelist()}
    root = ET.fromstring(entries["word/document.xml"])
    repeated = {
        "红线坐标": replace_repeating_block(root, "红线坐标", tables.get("红线坐标", []), fields),
        "勘探单元": replace_repeating_block(root, "勘探单元", tables.get("勘探单元", []), fields),
        "附表_勘探单元": replace_repeating_block(root, "附表_勘探单元", tables.get("勘探单元", []), fields),
        "剖线地层堆积": replace_repeating_block(root, "剖线地层堆积", tables.get("剖线地层堆积", []), fields),
        "剖线_AA": replace_repeating_block(root, "剖线_AA", tables.get("剖线地层堆积", []), fields),
        "剖线_BB": replace_repeating_block(root, "剖线_BB", tables.get("剖线地层堆积", []), fields),
        "标准孔坐标": replace_repeating_block(root, "标准孔坐标", tables.get("标准孔", []), fields),
        "标准孔详情": expand_standard_detail_tables(root, tables.get("标准孔", []), fields),
        "遗迹详情": expand_relic_detail_paragraphs(root, tables.get("遗迹记录") or tables.get("遗迹", []), fields),
    }
    if repeated["剖线地层堆积"] == 0:
        repeated["剖线地层堆积"] = expand_section_description_placeholders(
            root, tables.get("剖线地层堆积", []), fields
        )
    redline_table_rows = fill_placeholder_table(root, "四至范围坐标", tables.get("红线坐标", []), fields)
    unit_table_rows = fill_unit_coordinate_tables(root, tables.get("勘探单元", []), fields)
    standard_coordinate_table_rows = fill_standard_coordinate_tables(root, tables.get("标准孔", []))
    relic_coordinate_table_rows = fill_relic_coordinate_tables(
        root, tables.get("遗迹坐标", []), tables.get("遗迹记录") or tables.get("遗迹", [])
    )
    relic_registration_table_rows = fill_relic_registration_tables(root, tables.get("遗迹记录") or tables.get("遗迹", []))
    section_aa_probe_rows = fill_section_probe_table(root, "A-A′", tables.get("剖线坐标_AA", []), fields, drawings_dir)
    section_bb_probe_rows = fill_section_probe_table(root, "B-B′", tables.get("剖线坐标_BB", []), fields, drawings_dir)
    land_range_attachment_rows, land_range_attachment_note = insert_project_land_range_attachment_table(
        root,
        fields,
        tables.get("项目用地范围坐标表", []),
    )
    table_insert_errors = mark_table_insertion_errors(
        root,
        tables,
        redline_table_rows=redline_table_rows,
        unit_table_rows=unit_table_rows,
        standard_coordinate_table_rows=standard_coordinate_table_rows,
        relic_coordinate_table_rows=relic_coordinate_table_rows,
        relic_registration_table_rows=relic_registration_table_rows,
        section_aa_probe_rows=section_aa_probe_rows,
        section_bb_probe_rows=section_bb_probe_rows,
    )
    inserted_image_signatures: set[str] = set()
    company_personnel_images, company_personnel_missing = insert_company_personnel_assets(root, entries, company_personnel_context)
    table_images, table_missing_images = insert_table_cell_images(root, entries, image_slots, inserted_image_signatures)
    inserted_images, missing_images = insert_images(root, entries, image_slots, inserted_image_signatures)
    inserted_images += table_images + company_personnel_images
    missing_images.extend(table_missing_images)
    missing_images.extend(company_personnel_missing)
    missing_location_map_errors = mark_missing_project_location_map_errors(root, drawings_dir)
    section_descriptions_arranged = arrange_section_descriptions_after_figures(
        root, tables.get("剖线地层堆积", [])
    )
    section_landscape_breaks_added = ensure_body_section_drawing_landscape_breaks(root)
    changed = apply_plain_fields(root, fields)
    conclusion_paragraphs_rewritten = rewrite_conclusion_section(root, fields)
    region_paragraphs = split_region_overview_paragraphs(root, fields)
    overview_changed = normalize_project_overview_format(root, fields)
    caption_suffix_changed = normalize_caption_suffixes(root)
    figure_renumbered = renumber_figure_captions(root)
    appendix_renumbered = renumber_appendix_captions(root)
    appendix_tables_formatted = normalize_appendix_table_format(root)
    section_probe_headers_normalized = normalize_all_section_probe_table_headers(root)
    all_table_borders_normalized = normalize_all_table_borders(root)
    stale_section_placeholders_removed = remove_stale_section_description_placeholders(root)
    last_location_page_break_added = ensure_page_break_after_last_project_location_caption(root)
    survey_unit_page_break_added = ensure_page_break_after_survey_unit_numbering(root)
    extra_page_breaks_removed = remove_extra_consecutive_page_breaks(root)
    empty_section_break_pages_removed = remove_redundant_empty_section_break_pages(root)
    attachment_page_breaks_added = ensure_required_attachment_page_breaks(root)
    error_markers_styled = apply_error_marker_style(root)
    entries["word/document.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    toc_fields_marked = mark_toc_fields_for_update(entries, root)
    output.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in entries.items():
            zf.writestr(name, data)
    docx_normalized = normalize_docx_with_python_docx(output)
    if docx_normalized:
        notes.append("已通过 python-docx 规范化 DOCX 包结构。")

    unresolved = sorted(set(re.findall(r"\{\{[^{}]+}}", ET.tostring(root, encoding="unicode"))))
    lines = [
        "生成检查报告",
        "",
        "## 信息",
        f"- 输入表：{form}",
        f"- 智能模板：{template}",
        f"- 输出报告：{output}",
        f"- 普通字段替换段落/单元格：{changed}",
        f"- 结论区块重写：{conclusion_paragraphs_rewritten} 段",
        f"- 项目区域概况自动分段：{region_paragraphs} 段",
        f"- 项目概况格式修正：{overview_changed} 行",
        f"- 题注重复尾号清理：{caption_suffix_changed} 行",
        f"- 正文图号自动重排：{figure_renumbered} 行",
        f"- 附图编号自动重排：{appendix_renumbered} 行",
        f"- 附表表格格式统一：{appendix_tables_formatted} 个",
        f"- 剖面探孔记录表表头单位可见化：{section_probe_headers_normalized} 个",
        f"- 全文表格所有框线统一：{all_table_borders_normalized} 个",
        f"- 剖面地层空占位清理：{stale_section_placeholders_removed} 个",
        f"- 位置示意图末尾分页符补齐：{last_location_page_break_added} 个",
        f"- 勘探单元编号说明后分页符补齐：{survey_unit_page_break_added} 个",
        f"- 连续多余分页符清理：{extra_page_breaks_removed} 个",
        f"- 空白分节页清理：{empty_section_break_pages_removed} 个",
        f"- 附件图片分页符补齐：{attachment_page_breaks_added} 个",
        f"- 错误标记黑色加粗红色高亮：{error_markers_styled} 处",
        f"- 表格插入错误标记：{len(table_insert_errors)} 处",
        f"- 目录域更新：已标记打开时自动更新，TOC域 {toc_fields_marked} 处；目录文字可更新，需保留原字体格式",
        f"- 红线坐标生成：{repeated['红线坐标']} 条",
        f"- 红线四至坐标表填充：{redline_table_rows} 条",
        f"- 勘探单元生成：{repeated['勘探单元']} 条",
        f"- 勘探单元坐标表填充：{unit_table_rows} 条",
        f"- 勘探单元附表生成：{repeated['附表_勘探单元']} 条",
        f"- 剖线地层堆积生成：{repeated['剖线地层堆积']} 条",
        f"- 剖线图后文字说明排版：{section_descriptions_arranged} 条",
        f"- 剖线图横向分节补齐：{section_landscape_breaks_added} 处",
        f"- 剖线AA生成：{repeated['剖线_AA']} 条",
        f"- 剖线BB生成：{repeated['剖线_BB']} 条",
        f"- 标准孔坐标生成：{repeated['标准孔坐标']} 条",
        f"- 附表二标准孔坐标表填充：{standard_coordinate_table_rows} 条",
        f"- 附表三遗迹坐标表填充：{relic_coordinate_table_rows} 条",
        f"- 附表四遗迹登记表填充：{relic_registration_table_rows} 条",
        f"- A-A′剖面探孔记录表填充：{section_aa_probe_rows} 行",
        f"- B-B′剖面探孔记录表填充：{section_bb_probe_rows} 行",
        f"- 附件三项目用地范围坐标表：{land_range_attachment_rows} 行；{land_range_attachment_note}",
        f"- 标准孔详情生成：{repeated['标准孔详情']} 条",
        f"- 遗迹详情生成：{repeated['遗迹详情']} 条",
        f"- 自动插入图件：{inserted_images} 张",
        f"- 必插位置图正文错误标记：{missing_location_map_errors} 处",
        f"- 正文匹配错误标记：{len(marked_generation_errors)} 项",
    ]
    if drawings_dir:
        lines.append(f"- 制图成果目录：{drawings_dir}")
    if photos_dir:
        lines.append(f"- 外业照片目录：{photos_dir}")
    lines.extend(f"- {note}" for note in notes)
    if table_insert_errors:
        lines.extend(["", "## 表格待补"])
        lines.extend(f"- {item}" for item in table_insert_errors)
    if missing_images:
        lines.extend(["", "## 图件待补"])
        lines.extend(f"- {item}" for item in sorted(set(missing_images)))
    lines.extend(["", "## 未解析占位符"])
    lines.extend([f"- {item}" for item in unresolved] if unresolved else ["- 无"])
    check_path.write_text("\n".join(lines), encoding="utf-8")
    return output, check_path


def main() -> int:
    parser = argparse.ArgumentParser(description="从智能报告信息填报表生成报告")
    parser.add_argument("--self-test-image-rules", action="store_true", help="仅检查图片尺寸和旋转分类规则")
    parser.add_argument("--form", type=Path, default=DEFAULT_FORM)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--drawings-dir", type=Path, default=None, help="制图成果目录，可自动匹配并插入图件")
    parser.add_argument("--photos-dir", type=Path, default=None, help="外业照片目录，可自动匹配并插入照片")
    parser.add_argument("--personnel-set", default=None, help="公司资料包中的人员套组，例如 人员信息1 或 人员信息2")
    parser.add_argument("--use-region-api", action="store_true", help="允许调用外部 API 扩写项目区域概况")
    args = parser.parse_args()

    if args.self_test_image_rules:
        self_test_image_rules()
        print("图片尺寸和旋转规则自检通过")
        return 0

    output = args.output
    if output is None:
        fields, _, _ = load_form(args.form)
        output = REPORT_DIR / f"{safe_filename(fields.get('项目名称'))}-智能生成勘探报告.docx"
    try:
        report, check = fill_docx(
            args.template,
            args.form,
            output,
            args.drawings_dir,
            args.photos_dir,
            skip_region_api=not args.use_region_api,
            personnel_set=args.personnel_set,
        )
    except ValueError as exc:
        print(exc)
        return 2
    print(report)
    print(check)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

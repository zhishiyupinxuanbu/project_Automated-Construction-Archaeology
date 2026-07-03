#!/usr/bin/env python3
"""Orchestrate the smart archaeology report workflow.

This script is the first production entrypoint. It keeps the required human
confirmation points visible while making source discovery, template suggestion,
report generation, and audit output repeatable.
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

import create_manual_form_from_project as manual_builder
import fill_smart_template_from_form as smart_filler


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_DIR.parent
ASSETS_DIR = SKILL_ROOT / "assets"
ROOT = Path.cwd().resolve()
PROCESS_DIR = ROOT / "过程资料"
REPORT_DIR = Path.home() / "Desktop"
SMART_TEMPLATE_DIR = ASSETS_DIR / "templates"
FORM_TEMPLATE_DIR = SMART_TEMPLATE_DIR / "forms"
REPORT_PLAN_TEMPLATE_DIR = SMART_TEMPLATE_DIR / "reports"
COMPANY_PERSONNEL_PACKAGE_DIR = ASSETS_DIR / "company-personnel-library"
FORM_TEMPLATE = FORM_TEMPLATE_DIR / "人工填写表模板.xlsx"
LEGACY_TEMPLATE_DIRS = [ROOT / "旧模板", ROOT / "基础信息" / "旧模板"]
COMPANY_NAMES = [
    "北京卓凡文博技术有限公司",
    "内蒙古煊迹考古勘探有限公司",
    "内蒙古峰驰考古勘探有限公司",
    "三门峡市文物考古勘探有限公司",
    "河南燧火文物保护有限公司",
]
RELIC_STATES = ["无遗迹", "有遗迹", "有分区、有遗迹", "有分区、无遗迹"]
PLAN_STATES = ["无分区", "有分区"]
PERSONNEL_SETS = ["人员信息1", "人员信息2"]
DEFAULT_PERSONNEL_SET = "人员信息1"

SMART_SHEETS = {
    "项目基础信息",
    "项目区域概况",
    "文物概况",
    "勘探参数",
    "现场限制",
    "人员构成",
    "自动生成字段",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def first_number(value: object) -> int | None:
    match = re.search(r"\d+", clean(value))
    return int(match.group(0)) if match else None


def safe_filename(value: str) -> str:
    return re.sub(r'[/:*?"<>|\\]+', "_", clean(value)) or "未命名项目"


def project_process_dir(project_label: str) -> Path:
    return PROCESS_DIR / safe_filename(project_label)


def project_process_form_dir(project_label: str) -> Path:
    return project_process_dir(project_label) / "人工填写表"


def project_report_dir(project_label: str) -> Path:
    return REPORT_DIR


def status_line(label: str, value: object) -> str:
    return f"{label}: {value if value else '未找到'}"


def yes_no(value: bool) -> str:
    return "OK" if value else "缺失"


def docx_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(
        path
        for path in root.rglob("*.docx")
        if path.is_file() and not path.name.startswith(".~") and not path.name.startswith("._")
    )


def find_template_file(filename: str) -> Path:
    direct_candidates = [
        REPORT_PLAN_TEMPLATE_DIR / filename,
        SMART_TEMPLATE_DIR / filename,
        FORM_TEMPLATE_DIR / filename,
    ]
    for path in direct_candidates:
        if path.exists():
            return path
    matches = sorted(
        path
        for path in SMART_TEMPLATE_DIR.rglob(filename)
        if path.is_file() and not path.name.startswith(".~") and not path.name.startswith("._")
    )
    return matches[0] if matches else direct_candidates[0]


def available_personnel_sets(company: str) -> list[str]:
    company = clean(company)
    company_dir = COMPANY_PERSONNEL_PACKAGE_DIR / company
    if not company or not company_dir.exists():
        return []
    workbook = company_dir / f"{company}_资料卡.xlsx"
    found: set[str] = set()
    if workbook.exists():
        try:
            wb = load_workbook(workbook, read_only=True, data_only=True)
            found.update(sheet for sheet in PERSONNEL_SETS if sheet in wb.sheetnames)
            wb.close()
        except Exception:
            found = set()
    found.update(path.name for path in company_dir.iterdir() if path.is_dir() and path.name in PERSONNEL_SETS)
    return [value for value in PERSONNEL_SETS if value in found]


def company_uses_personnel_sets(company: str) -> bool:
    return bool(available_personnel_sets(company))


def effective_personnel_set(company: str, personnel_set: str | None = None) -> str | None:
    options = available_personnel_sets(company)
    if not options:
        return None
    value = clean(personnel_set) or DEFAULT_PERSONNEL_SET
    if value in options:
        return value
    if DEFAULT_PERSONNEL_SET in options:
        return DEFAULT_PERSONNEL_SET
    return options[0]


def path_matches_personnel_set(path: Path, company: str, personnel_set: str | None) -> bool:
    return True


def require_personnel_set_for_generation(company: str, personnel_set: str | None, explicit_template: Path | None = None) -> None:
    options = available_personnel_sets(company)
    if explicit_template is not None or not options:
        return
    if clean(personnel_set) in options:
        return
    option_text = " 或 ".join(f"`{value}`" for value in options)
    cli_text = " / ".join(options)
    raise ValueError(
        f"{company}资料库已分为 {option_text}。"
        f"生成报告/计划前必须先确认使用哪一套人员；请重新运行并添加 `--personnel-set {cli_text}` 中的一项。"
        f"平常默认推荐 `{effective_personnel_set(company)}`，但正式生成不能省略确认。"
    )


def require_template_confirmation(confirmed: bool, explicit_template: Path | None = None) -> None:
    if explicit_template is not None or confirmed:
        return
    raise ValueError(
        "生成前必须先把匹配到的模板完整路径告诉用户，并确认是否使用该模板。"
        "确认后请重新运行并添加 `--confirm-template`。"
    )


def find_dir(project_dir: Path, *parts: str) -> Path | None:
    direct = project_dir.joinpath(*parts)
    if direct.exists():
        return direct
    wanted = parts[-1]
    for path in project_dir.rglob(wanted):
        if path.is_dir():
            return path
    return None


def find_first_dir(project_dir: Path, candidates: list[tuple[str, ...]]) -> Path | None:
    for parts in candidates:
        found = find_dir(project_dir, *parts)
        if found:
            return found
    return None


def discover_source_dirs(project_dir: Path) -> dict[str, Path | None]:
    execute_dir = project_dir / "3.执行资料"
    return {
        "drawings_dir": find_first_dir(
            project_dir,
            [
                ("3.执行资料", "3.制图成果"),
                ("3.执行资料", "3.内业成果"),
                ("3.执行资料", "4.内业成果"),
                ("3.执行资料", "内业成果"),
                ("3.执行资料", "制图成果"),
            ],
        ),
        "photos_dir": find_first_dir(
            project_dir,
            [
                ("3.执行资料", "2.外业成果"),
                ("3.执行资料", "1.外业成果"),
                ("3.执行资料", "外业成果"),
            ],
        ),
        "report_dir": find_first_dir(
            project_dir,
            [
                ("3.执行资料", "5.报告"),
                ("3.执行资料", "4.报告"),
                ("3.执行资料", "报告"),
            ],
        ),
        "execute_dir": execute_dir if execute_dir.exists() else None,
    }


def count_files(root: Path | None, suffixes: set[str] | None = None) -> int:
    if not root or not root.exists():
        return 0
    total = 0
    for path in root.rglob("*"):
        if not path.is_file() or path.name.startswith("."):
            continue
        if suffixes and path.suffix.lower() not in suffixes:
            continue
        total += 1
    return total


def infer_partition_dirs(project_dir: Path) -> list[str]:
    labels = manual_builder.infer_partition_labels_from_project(project_dir)
    return [f"{label}区" for label in labels]


def print_source_preflight(project_dir: Path) -> int:
    dirs = discover_source_dirs(project_dir)
    issues: list[str] = []
    warnings: list[str] = []
    if not project_dir.exists():
        issues.append("项目源资料文件夹不存在")
    for label, key in [("3.执行资料", "execute_dir"), ("制图成果/内业成果", "drawings_dir"), ("外业成果", "photos_dir")]:
        if not dirs.get(key):
            issues.append(f"缺少{label}目录")
    drawings_dir = dirs.get("drawings_dir")
    photos_dir = dirs.get("photos_dir")
    required_drawings = ["1.位置图", "2.红线四至坐标、勘探区域", "3.勘探单元", "4.探孔、标准孔", "5.剖线"]
    if drawings_dir:
        for name in required_drawings:
            if not (drawings_dir / name).exists():
                warnings.append(f"制图成果中未找到 `{name}`")
    empty_allowed = {"2.走访调查照", "10.勘探后航拍照"}
    if photos_dir:
        for path in sorted(photos_dir.rglob("*")):
            if not path.is_dir() or path.name.startswith("."):
                continue
            if path.name in empty_allowed and count_files(path, {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}) == 0:
                warnings.append(f"`{path.relative_to(project_dir)}` 为空；后续应删除对应图位，不得用其他照片替代")
    site_records = sorted(project_dir.rglob("*现场记录*.docx")) if project_dir.exists() else []
    if not site_records:
        warnings.append("未找到项目现场记录 DOCX")
    review_candidates = [
        path
        for path in project_dir.rglob("*") if project_dir.exists() and path.is_file() and path.suffix.lower() in {".pdf", ".docx"}
        and any(token in path.name for token in ["文物", "文旅", "文化和旅游", "复函", "审查意见", "函"])
    ]
    print("## 项目源资料预检")
    print(status_line("项目源资料文件夹", project_dir))
    print(status_line("3.执行资料", dirs.get("execute_dir")))
    print(status_line("制图成果/内业成果", drawings_dir))
    print(status_line("外业成果", photos_dir))
    print(status_line("既有报告目录", dirs.get("report_dir")))
    print(status_line("现场记录", site_records[0] if site_records else None))
    print(status_line("疑似文物回函/审查文件数量", len(review_candidates)))
    print(status_line("勘探分区目录", "、".join(infer_partition_dirs(project_dir)) or "未识别"))
    print(status_line("制图/内业文件数", count_files(drawings_dir)))
    print(status_line("外业照片数", count_files(photos_dir, {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"})))
    print("")
    if issues:
        print("## 阻塞问题")
        for issue in issues:
            print(f"- {issue}")
        print("")
    if warnings:
        print("## 待复核事项")
        for warning in warnings[:20]:
            print(f"- {warning}")
        if len(warnings) > 20:
            print(f"- 另有 {len(warnings) - 20} 项未显示")
        print("")
    print("## 下一步")
    if issues:
        print("请先补齐阻塞资料；不要生成或预填人工填写表。")
        return 2
    print("请用户确认以上项目源资料文件夹和预检结果；确认后再运行 `prepare --source-confirmed <项目目录>`。")
    return 0


def read_flat_form(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    fields: dict[str, str] = {}
    for ws in wb.worksheets:
        headers = {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
        if "字段" not in headers:
            continue
        field_col = headers.get("字段", 1)
        value_col = headers.get("输入") or headers.get("值") or 2
        for row in range(2, ws.max_row + 1):
            key = clean(ws.cell(row, field_col).value)
            value = clean(ws.cell(row, value_col).value)
            if key and value:
                fields[key] = value
    return fields


def read_smart_form(path: Path) -> dict[str, str]:
    fields, _, _ = smart_filler.load_form(path)
    return fields


def is_relic_project(fields: dict[str, str]) -> bool:
    text = clean(fields.get("遗迹结论"))
    if not text:
        return False
    if any(token in text for token in ["未发现", "无遗迹", "无文化遗存", "否", "无"]):
        return False
    return any(token in text for token in ["有", "发现", "灰坑", "墓", "遗迹", "文化遗存", "是"])


def has_partition(fields: dict[str, str]) -> bool:
    partition_value = clean(fields.get("是否存在勘探分区"))
    partition_count = first_number(fields.get("勘探分区数量", ""))
    return partition_value in {"有", "是", "存在", "yes", "YES"} or bool(partition_count)


def report_state_label(fields: dict[str, str]) -> str:
    has_relic = is_relic_project(fields)
    partitioned = has_partition(fields)
    if partitioned and has_relic:
        return "有分区、有遗迹"
    if partitioned:
        return "有分区、无遗迹"
    return "有遗迹" if has_relic else "无遗迹"


def plan_state_label(fields: dict[str, str]) -> str:
    return "有分区" if has_partition(fields) else "无分区"


def match_report_template(
    company: str,
    relic_state: str,
    roots: list[Path],
    smart: bool,
    personnel_set: str | None = None,
) -> Path | None:
    if not relic_state:
        return None
    search_roots = [REPORT_PLAN_TEMPLATE_DIR] if smart else roots
    for root in search_roots:
        if not root.exists():
            continue
        state_candidates = [relic_state]
        if relic_state == "有分区、无遗迹":
            state_candidates.append("无遗迹")
        if relic_state == "有分区、有遗迹":
            state_candidates.append("有遗迹")
        for state in state_candidates:
            for path in docx_files(root):
                name = path.name
                if state in {"有遗迹", "无遗迹"} and "有分区" in name:
                    continue
                if state in name:
                    if smart and "智能报告生成基准模板" not in name:
                        continue
                    if not smart and company and company not in name:
                        continue
                    return path
    return None


def match_plan_template(
    company: str,
    plan_state: str,
    roots: list[Path],
    personnel_set: str | None = None,
) -> Path | None:
    if not plan_state:
        return None
    for root in [REPORT_PLAN_TEMPLATE_DIR]:
        if not root.exists():
            continue
        for path in docx_files(root):
            name = path.name
            if name.startswith("智能计划生成基准模板_") and plan_state in name:
                return path
    return None


def recommend_report_template(fields: dict[str, str], personnel_set: str | None = None) -> dict[str, Path | None | str]:
    company = clean(fields.get("勘探单位"))
    relic_state = report_state_label(fields)
    personnel_set = effective_personnel_set(company, personnel_set)
    return {
        "company": company,
        "relic_state": relic_state,
        "personnel_set": personnel_set or "",
        "smart_template": match_report_template(company, relic_state, [REPORT_PLAN_TEMPLATE_DIR], smart=True, personnel_set=personnel_set),
        "legacy_template_source": match_report_template(company, relic_state, LEGACY_TEMPLATE_DIRS, smart=False),
    }


def recommend_plan_template(fields: dict[str, str], personnel_set: str | None = None) -> dict[str, Path | None | str]:
    company = clean(fields.get("勘探单位"))
    plan_state = plan_state_label(fields)
    personnel_set = effective_personnel_set(company, personnel_set)
    return {
        "company": company,
        "plan_state": plan_state,
        "personnel_set": personnel_set or "",
        "plan_template": match_plan_template(company, plan_state, [REPORT_PLAN_TEMPLATE_DIR], personnel_set=personnel_set),
    }


def inventory_template_matrix() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for company in COMPANY_NAMES:
        personnel_options = PERSONNEL_SETS if company_uses_personnel_sets(company) else [None]
        for personnel_set in personnel_options:
            for relic_state in RELIC_STATES:
                rows.append(
                    {
                        "company": company,
                        "personnel_set": personnel_set or "",
                        "relic_state": relic_state,
                        "smart_template": match_report_template(
                            company, relic_state, [REPORT_PLAN_TEMPLATE_DIR], smart=True, personnel_set=personnel_set
                        ),
                        "legacy_template_source": match_report_template(company, relic_state, LEGACY_TEMPLATE_DIRS, smart=False),
                    }
                )
    return rows


def inventory_plan_template_matrix() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for company in COMPANY_NAMES:
        personnel_options = PERSONNEL_SETS if company_uses_personnel_sets(company) else [None]
        for personnel_set in personnel_options:
            for plan_state in PLAN_STATES:
                rows.append(
                    {
                        "company": company,
                        "personnel_set": personnel_set or "",
                        "plan_state": plan_state,
                        "plan_template": match_plan_template(
                            company, plan_state, [REPORT_PLAN_TEMPLATE_DIR], personnel_set=personnel_set
                        ),
                    }
                )
    return rows


def print_template_matrix() -> None:
    print("## 报告模板矩阵")
    print("| 勘探单位 | 人员套组 | 遗迹状态 | 统一智能基准模板 |")
    print("| --- | --- | --- | --- |")
    for row in inventory_template_matrix():
        smart = "OK" if row["smart_template"] else "缺失"
        personnel = row.get("personnel_set") or "-"
        print(f"| {row['company']} | {personnel} | {row['relic_state']} | {smart} |")


def print_plan_template_matrix() -> None:
    print("## 计划模板矩阵")
    print("| 勘探单位 | 人员套组 | 分区状态 | 智能计划基准模板 |")
    print("| --- | --- | --- | --- |")
    for row in inventory_plan_template_matrix():
        plan = "OK" if row["plan_template"] else "缺失"
        personnel = row.get("personnel_set") or "-"
        print(f"| {row['company']} | {personnel} | {row['plan_state']} | {plan} |")


def check_env_key() -> bool:
    env_file = ROOT / ".env.local"
    if not env_file.exists():
        return False
    text = env_file.read_text(encoding="utf-8", errors="ignore")
    keys = ("ARK_API_KEY=", "DOUBAO_ARK_API_KEY=")
    return any(line.strip().startswith(keys) and line.split("=", 1)[1].strip() for line in text.splitlines())


def command_doctor(args: argparse.Namespace) -> int:
    print("## 智能勘探报告 Skill 自检")
    required_paths = [
        ("运行目录", ROOT),
        ("Skill 根目录", SKILL_ROOT),
        ("标准目录 scripts", SCRIPT_DIR),
        ("标准目录 references", SKILL_ROOT / "references"),
        ("标准目录 assets", ASSETS_DIR),
        ("最终报告输出目录", REPORT_DIR),
        ("报告模板目录", REPORT_PLAN_TEMPLATE_DIR),
        ("公司信息与人员资料包目录", COMPANY_PERSONNEL_PACKAGE_DIR),
        ("表格模板目录", FORM_TEMPLATE_DIR),
        ("人工填写表模板", find_template_file("人工填写表模板.xlsx")),
        ("唯一人工填写表模板", FORM_TEMPLATE),
        ("默认报告基准模板", find_template_file("智能报告生成基准模板_北京卓凡文博技术有限公司_无遗迹.docx")),
        ("填报表生成脚本", SCRIPT_DIR / "create_smart_report_input_form.py"),
        ("人工表生成脚本", SCRIPT_DIR / "create_manual_form_from_project.py"),
        ("智能报告生成脚本", SCRIPT_DIR / "fill_smart_template_from_form.py"),
        ("专项检查脚本", SCRIPT_DIR / "check_smart_report.py"),
        ("区域概况 agent", SCRIPT_DIR / "region_overview_agent.py"),
    ]
    for label, path in required_paths:
        print(f"- {label}: {yes_no(path.exists())} `{path}`")
    print(f"- 区域概况 API 配置: {yes_no(check_env_key())} `.env.local`")
    print("")
    print_template_matrix()
    print("")
    print("## 判断")
    rows = inventory_template_matrix()
    missing_smart = [row for row in rows if not row["smart_template"]]
    if missing_smart:
        print(f"- 报告基准模板缺口：{len(missing_smart)} 个。请补齐 assets/templates/reports。")
    else:
        print("- 报告基准模板矩阵已覆盖。")
    print("- 报告流程入口：`preflight -> prepare --source-confirmed -> build-form -> recommend -> generate`。")
    return 0


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for idx in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{idx}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"无法生成不重名路径：{path}")


def create_manual_form(project_dir: Path, output: Path | None = None, overwrite: bool = False) -> Path:
    fields = manual_builder.infer_fields(project_dir)
    project_label = safe_filename(fields.get("项目名称") or project_dir.name)
    target = output or (project_process_form_dir(project_label) / f"{project_label}_人工填写表.xlsx")
    if target.exists() and not overwrite:
        target = unique_path(target)
    target.parent.mkdir(parents=True, exist_ok=True)
    manual_builder.fill_manual_template(manual_builder.TEMPLATE, target, fields)
    manual_builder.enrich_workbook_from_project(target, project_dir)
    return target


def refresh_blank_smart_form(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(path)


def build_smart_form_from_manual(manual_form: Path, output: Path | None = None) -> Path:
    manual_fields = read_flat_form(manual_form)
    project_name = manual_fields.get("项目名称") or manual_form.stem.replace("_人工填写表", "")
    project_label = safe_filename(project_name)
    output = output or (project_process_form_dir(project_label) / f"{project_label}_智能报告信息填报表_确认生成.xlsx")
    if output.exists():
        output = unique_path(output)
    refresh_blank_smart_form(FORM_TEMPLATE)
    shutil.copy2(FORM_TEMPLATE, output)
    wb = load_workbook(output)
    for sheet in wb.sheetnames:
        if sheet not in SMART_SHEETS:
            continue
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            key = clean(ws.cell(row, 1).value)
            if key in manual_fields:
                ws.cell(row, 2).value = manual_fields[key]
    manual_wb = load_workbook(manual_form, data_only=False)
    for sheet in smart_filler.TABLE_SHEETS:
        if sheet not in manual_wb.sheetnames or sheet not in wb.sheetnames:
            continue
        source_ws = manual_wb[sheet]
        target_ws = wb[sheet]
        target_max_col = target_ws.max_column
        for row in range(2, target_ws.max_row + 1):
            for col in range(1, target_max_col + 1):
                target_ws.cell(row, col).value = None
        target_row = 2
        for source_row in range(2, source_ws.max_row + 1):
            values = [source_ws.cell(source_row, col).value for col in range(1, source_ws.max_column + 1)]
            if not any(clean(value) for value in values):
                continue
            for col, value in enumerate(values[:target_max_col], start=1):
                target_ws.cell(target_row, col).value = value
            target_row += 1
    wb.save(output)
    return output


def missing_required(fields: dict[str, str]) -> list[str]:
    return [key for key in smart_filler.REQUIRED_FIELDS if not clean(fields.get(key))]


def default_output_for_form(form: Path) -> Path:
    fields = read_smart_form(form)
    project_label = safe_filename(fields.get("项目名称"))
    return project_report_dir(project_label) / f"{project_label}-智能生成勘探报告.docx"


def default_plan_output_for_form(form: Path) -> Path:
    fields = read_smart_form(form)
    project_label = safe_filename(fields.get("项目名称"))
    return project_report_dir(project_label) / f"{project_label}-智能生成考古调查勘探工作计划.docx"


def process_check_dir_for_form(form: Path) -> Path:
    fields = read_smart_form(form)
    project_label = safe_filename(fields.get("项目名称"))
    return project_process_dir(project_label) / "生成检查"


def move_generated_check_to_process(check: Path, form: Path) -> Path:
    if not check.exists():
        return check
    target_dir = process_check_dir_for_form(form)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = unique_path(target_dir / check.name)
    shutil.move(str(check), str(target))
    return target


def run_generation(
    form: Path,
    drawings_dir: Path | None,
    photos_dir: Path | None,
    template: Path | None = None,
    output: Path | None = None,
    run_audit: bool = True,
    use_region_api: bool = False,
    personnel_set: str | None = None,
    confirm_template: bool = False,
) -> tuple[Path, Path, Path | None, Path | None]:
    fields = read_smart_form(form)
    require_personnel_set_for_generation(clean(fields.get("勘探单位")), personnel_set, explicit_template=template)
    recommendation = recommend_report_template(fields, personnel_set=personnel_set)
    template = template or recommendation.get("smart_template")
    if not isinstance(template, Path) or not template.exists():
        company = recommendation.get("company") or "未填写勘探单位"
        relic_state = recommendation.get("relic_state") or "未识别遗迹状态"
        raise FileNotFoundError(
            f"未找到匹配的统一智能基准模板：{company} / {relic_state}。"
            "正式生成只使用 `模板/报告、计划`，请补齐该目录或指定已确认的 --template。"
        )
    require_template_confirmation(confirm_template, explicit_template=None if template == recommendation.get("smart_template") else template)
    output = output or default_output_for_form(form)
    report, check = smart_filler.fill_docx(
        template,
        form,
        output,
        drawings_dir,
        photos_dir,
        skip_region_api=not use_region_api,
        personnel_set=personnel_set,
    )
    check = move_generated_check_to_process(check, form)
    audit_xlsx: Path | None = None
    audit_brief: Path | None = None
    if run_audit and report.exists() and check.exists() and drawings_dir and photos_dir:
        audit_xlsx = report.with_name(f"{report.stem}-专项检查表.xlsx")
        audit_brief = check.with_name(f"{report.stem}-问题摘要.txt")
        subprocess.run(
            [
                sys.executable,
                str(SCRIPT_DIR / "check_smart_report.py"),
                "--docx",
                str(report),
                "--form",
                str(form),
                "--drawings-dir",
                str(drawings_dir),
                "--photos-dir",
                str(photos_dir),
                "--check-report",
                str(check),
                "--xlsx-output",
                str(audit_xlsx),
                "--brief-output",
                str(audit_brief),
            ],
            cwd=ROOT,
            check=True,
            stdout=subprocess.DEVNULL,
        )
    return report, check, audit_brief, audit_xlsx


def run_plan_generation(
    form: Path,
    drawings_dir: Path | None,
    photos_dir: Path | None,
    template: Path | None = None,
    output: Path | None = None,
    use_region_api: bool = False,
    personnel_set: str | None = None,
    confirm_template: bool = False,
) -> tuple[Path, Path]:
    fields = read_smart_form(form)
    require_personnel_set_for_generation(clean(fields.get("勘探单位")), personnel_set, explicit_template=template)
    recommendation = recommend_plan_template(fields, personnel_set=personnel_set)
    template = template or recommendation.get("plan_template")
    if not isinstance(template, Path) or not template.exists():
        company = recommendation.get("company") or "未填写勘探单位"
        plan_state = recommendation.get("plan_state") or "未识别分区状态"
        raise FileNotFoundError(
            f"未找到匹配的统一智能计划基准模板：{company} / {plan_state}。"
            "计划模板只从 `模板/报告、计划` 按有/无分区匹配，请补齐模板或指定 --template。"
        )
    require_template_confirmation(confirm_template, explicit_template=None if template == recommendation.get("plan_template") else template)
    output = output or default_plan_output_for_form(form)
    report, check = smart_filler.fill_docx(
        template,
        form,
        output,
        drawings_dir,
        photos_dir,
        skip_region_api=not use_region_api,
        personnel_set=personnel_set,
    )
    check = move_generated_check_to_process(check, form)
    return report, check


def print_report_recommendation(fields: dict[str, str], personnel_set: str | None = None) -> None:
    rec = recommend_report_template(fields, personnel_set=personnel_set)
    print(status_line("勘探单位", rec["company"]))
    if rec.get("personnel_set"):
        print(status_line("人员套组", rec["personnel_set"]))
        print("提示: 该公司资料库有多套人员；正式生成前必须明确选择 `--personnel-set 人员信息1/人员信息2`。")
    print(status_line("遗迹状态", rec["relic_state"]))
    print(status_line("智能基准模板", rec["smart_template"]))
    if not rec["smart_template"]:
        print("提示: 未在报告、计划模板目录找到对应遗迹状态的智能基准模板。")
        print("处理: 补齐 `模板/报告、计划`，或人工指定经确认的 --template。")


def print_plan_recommendation(fields: dict[str, str], personnel_set: str | None = None) -> None:
    rec = recommend_plan_template(fields, personnel_set=personnel_set)
    print(status_line("勘探单位", rec["company"]))
    if rec.get("personnel_set"):
        print(status_line("人员套组", rec["personnel_set"]))
        print("提示: 该公司资料库有多套人员；正式生成前必须明确选择 `--personnel-set 人员信息1/人员信息2`。")
    print(status_line("分区状态", rec["plan_state"]))
    print(status_line("智能计划基准模板", rec["plan_template"]))
    if not rec["plan_template"]:
        print("提示: 未在报告、计划模板目录找到对应分区状态的智能计划基准模板。")
        print("处理: 补齐 `模板/报告、计划` 中的计划模板，或人工指定经确认的 --template。")


def command_prepare(args: argparse.Namespace) -> int:
    project_dir = args.project_dir.resolve()
    if not project_dir.exists():
        raise FileNotFoundError(project_dir)
    if not getattr(args, "source_confirmed", False):
        print_source_preflight(project_dir)
        print("")
        print("生成已停止：第一闸门尚未确认。请先把以上源资料预检结果交给用户确认。")
        print("用户确认后，重新运行 `prepare --source-confirmed <项目目录>`。")
        return 2
    dirs = discover_source_dirs(project_dir)
    manual_form = create_manual_form(project_dir, args.output, overwrite=args.overwrite)
    prefill_issues, prefill_check = manual_builder.audit_manual_prefill(project_dir, manual_form)
    fields = read_flat_form(manual_form)
    print("## 项目资料发现")
    print(status_line("项目目录", project_dir))
    print(status_line("制图成果目录", dirs["drawings_dir"]))
    print(status_line("外业成果目录", dirs["photos_dir"]))
    print(status_line("既有报告目录", dirs["report_dir"]))
    print("")
    print("## 人工填写表")
    print(manual_form)
    print("")
    print("## 人工填写表预填检查")
    if prefill_check:
        print(f"检查文件: {prefill_check}")
    if prefill_issues:
        for item in prefill_issues[:8]:
            print(f"- {item}")
        if len(prefill_issues) > 8:
            print(f"- 另有 {len(prefill_issues) - 8} 项，请查看检查文件。")
    else:
        print("未发现现场记录关键字段漏填或明显不一致。")
    print("")
    print("## 模板推荐")
    print_report_recommendation(fields)
    print("")
    print("## 计划模板推荐")
    print_plan_recommendation(fields)
    print("")
    print("下一步: 请打开人工填写表确认/补填后，运行 build-form 生成智能报告信息填报表。")
    return 0


def command_preflight(args: argparse.Namespace) -> int:
    return print_source_preflight(args.project_dir.resolve())


def command_build_form(args: argparse.Namespace) -> int:
    form = build_smart_form_from_manual(args.manual_form.resolve(), args.output)
    fields = read_smart_form(form)
    print(form)
    print("")
    print("## 必填字段检查")
    missing = missing_required(fields)
    if missing:
        print("缺少: " + "、".join(missing))
    else:
        print("必填字段已填写。")
    print("")
    print("## 模板推荐")
    print_report_recommendation(fields)
    print("")
    print("## 计划模板推荐")
    print_plan_recommendation(fields)
    return 0


def command_recommend(args: argparse.Namespace) -> int:
    fields = read_smart_form(args.form.resolve()) if args.smart else read_flat_form(args.form.resolve())
    if args.plan:
        print_plan_recommendation(fields, personnel_set=args.personnel_set)
    else:
        print_report_recommendation(fields, personnel_set=args.personnel_set)
    return 0


def command_upgrade_templates(args: argparse.Namespace) -> int:
    command = [sys.executable, str(SCRIPT_DIR / "upgrade_legacy_templates_to_smart_baselines.py")]
    if args.overwrite:
        command.append("--overwrite")
    subprocess.run(command, cwd=ROOT, check=True)
    print("")
    print("提示: 旧模板已按来源升级为智能基准模板；已有基准模板默认不会覆盖。")
    print("下一步: 运行 `doctor` 查看公司/遗迹状态矩阵是否已覆盖。")
    return 0


def command_generate(args: argparse.Namespace) -> int:
    form = args.form.resolve()
    drawings_dir = args.drawings_dir.resolve() if args.drawings_dir else None
    photos_dir = args.photos_dir.resolve() if args.photos_dir else None
    if args.project_dir:
        dirs = discover_source_dirs(args.project_dir.resolve())
        drawings_dir = drawings_dir or dirs["drawings_dir"]
        photos_dir = photos_dir or dirs["photos_dir"]
    fields = read_smart_form(form)
    missing = missing_required(fields)
    if missing:
        print("生成已停止，必填字段为空: " + "、".join(missing))
        return 2
    report, check, audit_brief, audit_xlsx = run_generation(
        form=form,
        drawings_dir=drawings_dir,
        photos_dir=photos_dir,
        template=args.template.resolve() if args.template else None,
        output=args.output.resolve() if args.output else None,
        run_audit=not args.no_audit,
        use_region_api=args.use_region_api,
        personnel_set=args.personnel_set,
        confirm_template=args.confirm_template,
    )
    print(report)
    if check:
        print(f"过程检查: {check}")
    if audit_xlsx:
        print(audit_xlsx)
    if audit_brief:
        print(f"问题摘要: {audit_brief}")
    return 0


def command_generate_plan(args: argparse.Namespace) -> int:
    form = args.form.resolve()
    drawings_dir = args.drawings_dir.resolve() if args.drawings_dir else None
    photos_dir = args.photos_dir.resolve() if args.photos_dir else None
    if args.project_dir:
        dirs = discover_source_dirs(args.project_dir.resolve())
        drawings_dir = drawings_dir or dirs["drawings_dir"]
        photos_dir = photos_dir or dirs["photos_dir"]
    fields = read_smart_form(form)
    missing = missing_required(fields)
    if missing:
        print("生成已停止，必填字段为空: " + "、".join(missing))
        return 2
    report, check = run_plan_generation(
        form=form,
        drawings_dir=drawings_dir,
        photos_dir=photos_dir,
        template=args.template.resolve() if args.template else None,
        output=args.output.resolve() if args.output else None,
        use_region_api=args.use_region_api,
        personnel_set=args.personnel_set,
        confirm_template=args.confirm_template,
    )
    print(report)
    print(f"过程检查: {check}")
    return 0


def command_run(args: argparse.Namespace) -> int:
    if args.form:
        return command_generate(args)
    if not args.project_dir:
        raise ValueError("run 需要 --form 或 project_dir")
    print("run 已禁止从项目目录自动跳过闸门生成。")
    print("请先运行 `preflight <项目目录>`，用户确认后再 `prepare --source-confirmed <项目目录>`。")
    return 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="智能考古调查勘探报告总控工作流")
    sub = parser.add_subparsers(dest="command", required=True)

    doctor = sub.add_parser("doctor", help="系统级自检：目录、脚本、模板矩阵、API 配置")
    doctor.set_defaults(func=command_doctor)

    preflight = sub.add_parser("preflight", help="第一闸门：只检查项目源资料文件夹，不生成表格")
    preflight.add_argument("project_dir", type=Path)
    preflight.set_defaults(func=command_preflight)

    prepare = sub.add_parser("prepare", help="第二闸门：在源资料已确认后生成并预填人工填写表")
    prepare.add_argument("project_dir", type=Path)
    prepare.add_argument("--output", type=Path, default=None)
    prepare.add_argument("--overwrite", action="store_true", help="允许覆盖指定 output 或默认人工填写表")
    prepare.add_argument("--source-confirmed", action="store_true", help="表示用户已确认项目源资料文件夹和预检结果")
    prepare.set_defaults(func=command_prepare)

    build = sub.add_parser("build-form", help="从人工填写表生成智能报告信息填报表")
    build.add_argument("manual_form", type=Path)
    build.add_argument("--output", type=Path, default=None)
    build.set_defaults(func=command_build_form)

    recommend = sub.add_parser("recommend", help="根据表格推荐模板")
    recommend.add_argument("form", type=Path)
    recommend.add_argument("--smart", action="store_true", help="输入表为智能报告信息填报表")
    recommend.add_argument("--plan", action="store_true", help="推荐考古调查、勘探工作计划模板")
    recommend.add_argument("--personnel-set", choices=PERSONNEL_SETS, default=None, help="人员资料套组；未指定时推荐人员信息1")
    recommend.set_defaults(func=command_recommend)

    upgrade = sub.add_parser("upgrade-templates", help="把旧模板作为来源升级为智能基准模板，默认只补缺失")
    upgrade.add_argument("--overwrite", action="store_true", help="覆盖已存在的智能基准模板；慎用")
    upgrade.set_defaults(func=command_upgrade_templates)

    generate = sub.add_parser("generate", help="从智能报告信息填报表生成正式报告和检查报告")
    generate.add_argument("--form", type=Path, required=True)
    generate.add_argument("--project-dir", type=Path, default=None)
    generate.add_argument("--drawings-dir", type=Path, default=None)
    generate.add_argument("--photos-dir", type=Path, default=None)
    generate.add_argument("--template", type=Path, default=None)
    generate.add_argument("--personnel-set", choices=PERSONNEL_SETS, default=None, help="人员资料套组；生成前必须明确选择")
    generate.add_argument("--confirm-template", action="store_true", help="表示用户已确认使用匹配到的模板")
    generate.add_argument("--output", type=Path, default=None)
    generate.add_argument("--use-region-api", action="store_true", help="允许调用外部 API 扩写项目区域概况")
    generate.add_argument("--no-audit", action="store_true")
    generate.set_defaults(func=command_generate)

    generate_plan = sub.add_parser("generate-plan", help="从智能报告信息填报表生成考古调查、勘探工作计划")
    generate_plan.add_argument("--form", type=Path, required=True)
    generate_plan.add_argument("--project-dir", type=Path, default=None)
    generate_plan.add_argument("--drawings-dir", type=Path, default=None)
    generate_plan.add_argument("--photos-dir", type=Path, default=None)
    generate_plan.add_argument("--template", type=Path, default=None)
    generate_plan.add_argument("--personnel-set", choices=PERSONNEL_SETS, default=None, help="人员资料套组；生成前必须明确选择")
    generate_plan.add_argument("--confirm-template", action="store_true", help="表示用户已确认使用匹配到的模板")
    generate_plan.add_argument("--output", type=Path, default=None)
    generate_plan.add_argument("--use-region-api", action="store_true", help="允许调用外部 API 扩写项目区域概况")
    generate_plan.set_defaults(func=command_generate_plan)

    run = sub.add_parser("run", help="已有智能表且前置确认完成时直接生成；不再从项目目录自动准备")
    run.add_argument("project_dir", type=Path, nargs="?")
    run.add_argument("--form", type=Path, default=None)
    run.add_argument("--drawings-dir", type=Path, default=None)
    run.add_argument("--photos-dir", type=Path, default=None)
    run.add_argument("--template", type=Path, default=None)
    run.add_argument("--personnel-set", choices=PERSONNEL_SETS, default=None, help="人员资料套组；生成前必须明确选择")
    run.add_argument("--confirm-template", action="store_true", help="表示用户已确认使用匹配到的模板")
    run.add_argument("--output", type=Path, default=None)
    run.add_argument("--use-region-api", action="store_true", help="允许调用外部 API 扩写项目区域概况")
    run.add_argument("--no-audit", action="store_true")
    run.set_defaults(func=command_run)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return args.func(args)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())

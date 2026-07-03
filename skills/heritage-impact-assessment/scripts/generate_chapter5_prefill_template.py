#!/usr/bin/env python3
"""Generate the v0.3 Chapter 5 human prefill Excel template."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["编号", "对应正文表", "固定影响内容", "人工预填说明", "依据原文或资料来源", "人工边界提示"]

ROWS = [
    [1, "项目设计合规性评估", "项目流程", "", "", ""],
    [2, "项目设计合规性评估", "环境保护", "", "", ""],
    [3, "项目设计合规性评估", "景观保护", "", "", ""],
    [4, "项目设计合规性评估", "用地类型", "", "", ""],
    [5, "项目设计合规性评估", "建筑高度与建筑密度", "", "", ""],
    [6, "项目建设期影响评估", "施工过程对环境的影响（包括污染排放、地质灾害、水土流失、生态环境）", "", "", ""],
    [7, "项目建设期影响评估", "施工过程对地下不明文物的扰动", "", "", ""],
    [8, "项目运营期影响评估", "污染排放与生态环境影响", "", "", ""],
    [9, "项目运营期影响评估", "整体建筑风貌与视觉景观影响", "", "", ""],
]

FACTOR_VALUES = [
    "有较大益处",
    "有较小益处",
    "正面影响可忽略",
    "没有改变",
    "负面影响可忽略",
    "有较小负面影响",
    "有较大负面影响",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成第五章人工预填表 Excel 模板")
    parser.add_argument("--输出目录", dest="output_dir", required=True)
    parser.add_argument("--文件名", dest="filename", default="第五章人工预填表.xlsx")
    parser.add_argument("--csv", dest="write_csv", action="store_true", help="同时输出 chapter5_prefill_template.csv")
    return parser.parse_args()


def write_csv_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(ROWS)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "第五章人工预填表"

    ws.merge_cells("A1:F1")
    ws["A1"] = "第五章人工预填表"
    ws["A1"].font = Font(name="宋体", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)

    ws.merge_cells("A13:F13")
    ws["A13"] = "填写说明：只需填写“人工预填说明”和“依据原文或资料来源”；“人工边界提示”可空。AI 将计算、比对并映射回“项目设计合规性评估”“项目建设期影响评估”“项目运营期影响评估”三张固定正文表。"
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A15:F15")
    ws["A15"] = "固定影响因子（供AI选择，用户不必填写）：有较大益处、有较小益处、正面影响可忽略、没有改变、负面影响可忽略、有较小负面影响、有较大负面影响。"
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

    widths = [8, 24, 42, 52, 52, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28
    for row_number in range(3, 12):
        ws.row_dimensions[row_number].height = 56
    ws.row_dimensions[13].height = 54
    ws.row_dimensions[15].height = 42

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fixed_fill = PatternFill("solid", fgColor="F2F2F2")
    input_fill = PatternFill("solid", fgColor="FFF2CC")

    for row in ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="宋体", size=11)
            if cell.row == 2:
                cell.fill = header_fill
                cell.font = Font(name="宋体", size=11, bold=True)
            elif cell.column <= 3:
                cell.fill = fixed_fill
            else:
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for merged_row in (13, 15):
        for cell in ws[merged_row]:
            cell.font = Font(name="宋体", size=10)
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
            cell.border = border

    ws.freeze_panes = "A3"
    return wb


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / args.filename

    wb = build_workbook()
    wb.save(xlsx_path)

    if args.write_csv:
        write_csv_template(output_dir / "chapter5_prefill_template.csv")

    print(xlsx_path)


if __name__ == "__main__":
    main()

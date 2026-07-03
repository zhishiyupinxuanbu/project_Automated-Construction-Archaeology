#!/usr/bin/env python3
"""Audit a generated smart archaeology report against its source package."""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def find_project_root() -> Path:
    here = Path(__file__).resolve()
    candidates = [here.parent, *here.parents, Path.cwd(), *Path.cwd().resolve().parents]
    for candidate in candidates:
        if (
            (candidate / "智能生成报告技能资料" / "知识库" / "3.模板与表单").exists()
            and ((candidate / "过程资料").exists() or (candidate / "基础信息").exists())
        ):
            return candidate
    for candidate in candidates:
        if candidate.name == "智能生成报告技能资料":
            return candidate.parent
    return here.parents[1]


ROOT = find_project_root()
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def load_filler():
    local_path = Path(__file__).resolve().parent / "fill_smart_template_from_form.py"
    path = local_path if local_path.exists() else ROOT / "过程资料" / "fill_smart_template_from_form.py"
    spec = importlib.util.spec_from_file_location("fill_smart_template_from_form", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载生成脚本：{path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_manual_builder():
    path = Path(__file__).resolve().parent / "create_manual_form_from_project.py"
    if not path.exists():
        return None
    spec = importlib.util.spec_from_file_location("create_manual_form_from_project", path)
    if spec is None or spec.loader is None:
        return None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def infer_project_dir_from_sources(drawings_dir: Path, photos_dir: Path) -> Path | None:
    for source in [drawings_dir, photos_dir]:
        resolved = source.resolve()
        parts = list(resolved.parts)
        if "3.执行资料" in parts:
            idx = parts.index("3.执行资料")
            if idx > 0:
                return Path(*parts[:idx])
        if "基础信息" in parts:
            idx = parts.index("基础信息")
            if idx + 1 < len(parts):
                return Path(*parts[: idx + 2])
    return None


def docx_text_and_metrics(path: Path) -> tuple[list[str], dict[str, int | list[str]]]:
    with zipfile.ZipFile(path) as docx:
        root = ET.fromstring(docx.read("word/document.xml"))
        media = [name for name in docx.namelist() if name.startswith("word/media/")]
        rels_root = ET.fromstring(docx.read("word/_rels/document.xml.rels"))
    rels = {rel.get("Id"): rel.get("Target") for rel in rels_root.findall("rel:Relationship", NS)}
    refs: dict[str, int] = defaultdict(int)
    for blip in root.findall(".//a:blip", NS):
        rid = blip.get(f"{{{NS['r']}}}embed")
        if rid:
            refs[rels.get(rid, "")] += 1
    paras = root.findall(".//w:p", NS)
    texts = []
    for para in paras:
        text = "".join(node.text or "" for node in para.findall(".//w:t", NS))
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            texts.append(text)
    xml = ET.tostring(root, encoding="unicode")
    metrics: dict[str, int | list[str]] = {
        "paragraph_count": len(paras),
        "media_count": len(media),
        "image_reference_count": sum(refs.values()),
        "unique_image_reference_count": len(refs),
        "body_figure_caption_count": sum(1 for text in texts if re.match(r"^图\d+\s+", text)),
        "unresolved": sorted(set(re.findall(r"\{\{[^{}]+}}", xml))),
    }
    return texts, metrics


def docx_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as docx:
        return docx.read("word/document.xml").decode("utf-8", "ignore")


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def file_hash(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def docx_media_hashes(path: Path) -> set[str]:
    hashes: set[str] = set()
    with zipfile.ZipFile(path) as docx:
        for name in docx.namelist():
            if name.startswith("word/media/"):
                hashes.add(sha256_bytes(docx.read(name)))
    return hashes


def element_text(elem: ET.Element) -> str:
    return "".join(node.text or "" for node in elem.findall(".//w:t", NS))


def template_image_keys(template: str) -> set[str]:
    path = Path(template)
    if not template or not path.exists():
        return set()
    with zipfile.ZipFile(path) as docx:
        root = ET.fromstring(docx.read("word/document.xml"))
    keys: set[str] = set()
    for elem in root.findall(".//w:p", NS) + root.findall(".//w:tc", NS):
        for match in re.findall(r"\{\{图:([^{}]+)}}", element_text(elem)):
            keys.add(match.strip())
    return keys


def markdown_cell(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", "<br>")
    text = text.replace("|", "\\|")
    return text


def markdown_table(headers: list[str], rows: list[list[object]]) -> list[str]:
    lines = [
        "| " + " | ".join(markdown_cell(item) for item in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    if rows:
        lines.extend("| " + " | ".join(markdown_cell(item) for item in row) + " |" for row in rows)
    else:
        lines.append("| " + " | ".join("无" if idx == 0 else "" for idx, _ in enumerate(headers)) + " |")
    return lines


def parse_generated_check(path: Path) -> dict[str, list[str] | dict[str, str]]:
    sections: dict[str, list[str]] = defaultdict(list)
    info: dict[str, str] = {}
    current = "信息"
    if not path.exists():
        return {"sections": {}, "info": {"缺失": str(path)}}
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if line.startswith("## "):
            current = line[3:].strip()
            continue
        if not line.startswith("- "):
            continue
        item = line[2:]
        sections[current].append(item)
        if "：" in item:
            key, value = item.split("：", 1)
            info[key] = value
    return {"sections": dict(sections), "info": info}


def parse_count(value: object) -> int | None:
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else None


def normalize_section_key(value: object) -> str:
    text = str(value or "").strip().upper()
    text = text.replace("剖线", "").replace("'", "′").replace("’", "′")
    text = re.sub(r"\s+", "", text)
    if re.search(r"A\s*-\s*A", text):
        return "A-A′"
    if re.search(r"B\s*-\s*B", text):
        return "B-B′"
    match = re.search(r"P\s*0*(\d+)", text)
    if match:
        return f"P{int(match.group(1))}"
    return text


def unique_section_record_count(records: list[dict[str, str]]) -> int:
    keys: set[str] = set()
    fallback = 0
    for row in records:
        key = normalize_section_key(row.get("剖线编号") or row.get("剖线") or row.get("编号"))
        description = str(row.get("剖线地层描述") or row.get("地层堆积情况") or row.get("地层描述") or "").strip()
        if key:
            keys.add(key)
        elif description:
            fallback += 1
    return len(keys) + fallback


def normalize_appendix_section_label(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("’", "'").replace("′", "'").replace("剖线", "剖面")
    text = re.sub(r"\s+", "", text)
    match = re.search(r"([A-Z])-\1'?地层堆积剖面图", text, flags=re.I)
    if match:
        letter = match.group(1).upper()
        return f"{letter}-{letter}'地层堆积剖面图"
    return text


def sorted_section_drawing_paths(drawings_dir: Path) -> list[Path]:
    if not drawings_dir.exists():
        return []
    paths = [
        path.resolve()
        for path in drawings_dir.rglob("*")
        if path.suffix.lower() in IMAGE_EXTENSIONS
        and re.search(r"地层堆积剖[线面]图", path.stem)
        and "位置" not in path.stem
    ]

    def sort_key(path: Path) -> tuple[int, str]:
        label = normalize_appendix_section_label(path.stem)
        match = re.match(r"([A-Z])-\1", label, flags=re.I)
        if match:
            return (ord(match.group(1).upper()) - ord("A"), str(path))
        return (999, str(path))

    return sorted(paths, key=sort_key)


def appendix_texts_after_conclusion(texts: list[str]) -> list[str]:
    conclusion_idx = max((idx for idx, text in enumerate(texts) if text == "五、结论" or text.endswith("结论")), default=-1)
    return [text for idx, text in enumerate(texts) if idx > conclusion_idx and text.startswith("附图")]


def appendix_section_drawing_audit_rows(
    drawings_dir: Path,
    photos_dir: Path,
    texts: list[str],
) -> tuple[list[list[object]], list[list[object]], list[str]]:
    source_paths = sorted_section_drawing_paths(drawings_dir)
    appendix_texts = appendix_texts_after_conclusion(texts)
    appendix_labels = {normalize_appendix_section_label(text) for text in appendix_texts}
    rows: list[list[object]] = []
    failure_rows: list[list[object]] = []
    issue_texts: list[str] = []
    if not source_paths:
        rows.append(["结论后地层堆积剖面图", 0, 0, "不适用", "制图成果中未发现地层堆积剖面图源文件。", ""])
        return rows, failure_rows, issue_texts
    missing: list[str] = []
    for path in source_paths:
        expected = normalize_appendix_section_label(path.stem)
        found = expected in appendix_labels
        status = "通过" if found else "失败"
        reason = "结论后的附图标题中已检出该剖面图。" if found else "源资料存在该剖面图，但结论后的附图标题中未检出，可能未插入或附图未展开。"
        rows.append([
            expected,
            relative_source(path, drawings_dir, photos_dir),
            "已检出" if found else "未检出",
            status,
            reason,
            "附图剖面图必须按源资料全部有效剖面图展开，不受模板 A/B 预留占位数量限制。",
        ])
        if not found:
            missing.append(expected)
    if missing:
        reason = f"源资料剖面图 {len(source_paths)} 张，结论后附图缺少：{'、'.join(missing)}。"
        failure_rows.append(["附图剖面图完整性", "失败", reason])
        issue_texts.append(reason)
    return rows, failure_rows, issue_texts


def table_replacement_audit_rows(
    tables: dict[str, list[dict[str, str]]],
    generated_info: dict[str, str],
) -> tuple[list[list[object]], list[list[object]], list[str]]:
    section_record_count = unique_section_record_count(tables.get("剖线地层堆积", []))
    table_checks = [
        ("红线坐标正文重复块", "红线坐标", len(tables.get("红线坐标", [])), "红线坐标生成", "应按本项目红线四至坐标生成正文重复块。"),
        ("红线四至坐标表", "红线坐标", len(tables.get("红线坐标", [])), "红线四至坐标表填充", "应按本项目红线四至坐标填充表格。"),
        ("勘探单元正文重复块", "勘探单元", len(tables.get("勘探单元", [])), "勘探单元生成", "应按本项目勘探单元坐标生成正文重复块。"),
        ("勘探单元坐标表", "勘探单元", len(tables.get("勘探单元", [])), "勘探单元坐标表填充", "应按本项目勘探单元坐标填充附表。"),
        ("剖线地层堆积说明", "剖线地层堆积", section_record_count, "剖线地层堆积生成", "应按本项目有效剖线编号生成地层堆积说明。"),
        ("剖线图后文字说明", "剖线地层堆积", section_record_count, "剖线图后文字说明排版", "应把有效剖线地层堆积说明排在对应剖面图后。"),
        ("附表二标准孔坐标表", "标准孔", len(tables.get("标准孔", [])), "附表二标准孔坐标表填充", "应按本项目标准孔坐标表填充。"),
        ("标准孔详情表", "标准孔", len(tables.get("标准孔", [])), "标准孔详情生成", "应按本项目标准孔记录生成详情表。"),
        ("附表三遗迹坐标表", "遗迹坐标", len(tables.get("遗迹坐标", [])), "附表三遗迹坐标表填充", "有遗迹坐标时应按本项目遗迹坐标填充。"),
        ("附表四遗迹登记表", "遗迹记录", len(tables.get("遗迹记录", []) or tables.get("遗迹", [])), "附表四遗迹登记表填充", "有遗迹记录时应按本项目遗迹记录填充。"),
        ("A-A′剖面探孔记录表", "剖线坐标_AA", len(tables.get("剖线坐标_AA", [])), "A-A′剖面探孔记录表填充", "应按本项目 A-A′ 剖线坐标填充。"),
        ("B-B′剖面探孔记录表", "剖线坐标_BB", len(tables.get("剖线坐标_BB", [])), "B-B′剖面探孔记录表填充", "应按本项目 B-B′ 剖线坐标填充。"),
    ]
    rows: list[list[object]] = []
    failure_rows: list[list[object]] = []
    issue_texts: list[str] = []
    for table_name, source_sheet, expected, info_key, note in table_checks:
        actual = parse_count(generated_info.get(info_key))
        if expected == 0:
            status = "不适用"
            reason = f"{source_sheet} 无本项目记录。"
        elif actual is None:
            status = "需复核"
            reason = f"生成检查报告未记录 `{info_key}`，无法确认是否替换。"
        elif actual == 0:
            status = "失败"
            reason = f"期望 {expected} 条，但生成计数为 0，可能沿用模板空表或旧表。"
        elif actual < expected:
            status = "需复核"
            reason = f"期望 {expected} 条，生成计数 {actual} 条，可能有部分记录未替换。"
        else:
            status = "通过"
            reason = "生成计数已覆盖本项目记录。"
        rows.append([table_name, source_sheet, expected, actual if actual is not None else "未记录", status, reason, note])
        if status in {"失败", "需复核"} and expected > 0:
            failure_rows.append([f"表格替换-{table_name}", status, reason])
            issue_texts.append(f"{table_name}：{reason}")
    return rows, failure_rows, issue_texts


def all_source_images(drawings_dir: Path, photos_dir: Path) -> list[Path]:
    out = []
    for root in [drawings_dir, photos_dir]:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.name.startswith("."):
                continue
            if path.suffix.lower() in IMAGE_EXTENSIONS:
                out.append(path.resolve())
    return sorted(out)


def source_category(path: Path, drawings_dir: Path, photos_dir: Path) -> str:
    for root, label in [(drawings_dir.resolve(), "制图成果"), (photos_dir.resolve(), "外业成果")]:
        try:
            rel = path.resolve().relative_to(root)
            return f"{label}/{rel.parts[0]}" if rel.parts else label
        except ValueError:
            pass
    return "其他"


def workbook_value(form: Path, sheet: str, key: str) -> str:
    wb = load_workbook(form, data_only=True)
    if sheet not in wb.sheetnames:
        return ""
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]).strip() == key:
            value = row[1] if len(row) > 1 else None
            if len(row) > 3:
                for extra in row[3:]:
                    if extra not in (None, ""):
                        value = extra
            return str(value).strip() if value not in (None, "") else ""
    return ""


def workbook_fields_for_audit(form: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(form, data_only=True)
    rows: list[tuple[str, str, str]] = []
    for sheet in ["项目基础信息", "项目区域概况", "文物概况", "勘探参数", "现场限制", "人员构成", "自动生成字段"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            key = str(row[0]).strip()
            value = row[1] if len(row) > 1 else None
            if len(row) > 3:
                for extra in row[3:]:
                    if extra not in (None, ""):
                        value = extra
            value_text = str(value).strip() if value not in (None, "") else ""
            rows.append((sheet, key, value_text))
    return rows


def replacement_audit_rows(
    form: Path,
    docx: Path,
    template: str,
    unresolved: list[str],
    effective_fields: dict[str, str],
) -> list[list[str]]:
    doc_xml = docx_xml(docx)
    template_xml = ""
    template_path = Path(template)
    if template and template_path.exists():
        template_xml = docx_xml(template_path)
    unresolved_set = set(unresolved)
    rows: list[list[str]] = []
    for sheet, key, raw_value in workbook_fields_for_audit(form):
        value = raw_value or str(effective_fields.get(key, "") or "").strip()
        is_derived = not raw_value and bool(value)
        placeholder = "{{" + key + "}}"
        has_placeholder = placeholder in template_xml
        value_found = bool(value and value in doc_xml)
        unresolved_left = placeholder in unresolved_set
        if not value:
            status = "未填写"
            reason = "填报表为空，不参与替换。"
        elif has_placeholder and value_found and not unresolved_left:
            status = "已替换"
            reason = "基准模板存在替换符，生成稿中已出现有效字段值。"
        elif has_placeholder and unresolved_left:
            status = "替换失败"
            reason = "生成稿仍残留该替换符。"
        elif has_placeholder and not value_found:
            status = "需复核"
            reason = "基准模板存在替换符，但未在生成稿中直接检索到有效字段值，可能被计算规则改写或替换失败。"
        elif value_found:
            status = "报告中出现"
            reason = "基准模板无同名替换符；该值可能来自模板固定文本、派生字段或其他替换符。"
        else:
            status = "不要求替换"
            reason = "基准模板无同名替换符，按规则保留模板原文。"
        if is_derived:
            reason = "填报表为空，使用生成器自动计算/导入后的有效字段值；" + reason
        rows.append([
            sheet,
            key,
            value[:80] + ("..." if len(value) > 80 else ""),
            "有" if has_placeholder else "无",
            "是" if value_found else "否",
            status,
            reason,
        ])
    return rows


def build_expected_slots(filler, form: Path, drawings_dir: Path, photos_dir: Path):
    fields, tables, notes = filler.load_form(form)
    notes.extend(filler.import_external_tables(tables, drawings_dir))
    filler.fill_coordinate_basepoint(fields, tables, notes)
    notes.extend(filler.import_site_record_tables(fields, tables, photos_dir))
    filler.refresh_section_derived_fields(fields, tables, notes)
    slots = filler.build_image_slots(tables, form, drawings_dir, photos_dir)
    return fields, tables, notes, slots


def slot_paths_by_prefix(slot_summary: dict[str, list[Path]], prefix: str) -> set[Path]:
    paths: set[Path] = set()
    for key, values in slot_summary.items():
        if key == prefix or key.startswith(prefix + ":"):
            paths.update(values)
    return paths


def slot_count_by_prefix(slot_summary: dict[str, list[Path]], prefix: str) -> int:
    count = 0
    for key, values in slot_summary.items():
        if key == prefix or key.startswith(prefix + ":"):
            count += len(values)
    return count


def has_partition(fields: dict[str, str]) -> bool:
    value = str(fields.get("是否存在勘探分区") or "").strip()
    count = str(fields.get("勘探分区数量") or "").strip()
    match = re.search(r"\d+", count)
    return value in {"有", "是", "存在", "yes", "YES"} or (int(match.group(0)) > 0 if match else False)


def has_section_description_text(texts: list[str]) -> bool:
    joined = "\n".join(texts)
    return bool(re.search(r"剖[线面]\s*A-A[′'’]?\s*地层堆积情况[:：]", joined)) and bool(
        re.search(r"剖[线面]\s*B-B[′'’]?\s*地层堆积情况[:：]", joined)
    )


def relative_source(path: Path, drawings_dir: Path, photos_dir: Path) -> str:
    for root in [drawings_dir, photos_dir]:
        try:
            return str(path.resolve().relative_to(root.resolve()))
        except ValueError:
            pass
    return str(path)


def is_relative_to(path: Path, root: Path | None) -> bool:
    if root is None:
        return False
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def image_replacement_audit_rows(
    docx: Path,
    slot_summary: dict[str, list[Path]],
    drawings_dir: Path,
    photos_dir: Path,
) -> tuple[list[list[object]], list[list[object]], list[str]]:
    media_hashes = docx_media_hashes(docx)
    rows: list[list[object]] = []
    failure_rows: list[list[object]] = []
    issue_texts: list[str] = []
    for key in sorted(slot_summary):
        unique_paths = sorted(set(slot_summary[key]))
        expected = len(unique_paths)
        if expected == 0:
            continue
        present_paths: list[Path] = []
        missing_paths: list[Path] = []
        for path in unique_paths:
            try:
                if file_hash(path) in media_hashes:
                    present_paths.append(path)
                else:
                    missing_paths.append(path)
            except OSError:
                missing_paths.append(path)
        present = len(present_paths)
        if present == expected:
            status = "通过"
            reason = "生成稿媒体文件中已检出本项目源图。"
        elif present == 0:
            status = "失败"
            reason = "规则已匹配源图，但生成稿媒体文件中未检出这些源图，可能仍沿用模板旧图或图片替换失败。"
        else:
            status = "需复核"
            reason = f"规则匹配 {expected} 张源图，生成稿媒体文件中仅检出 {present} 张，可能有部分照片/图件未替换。"
        examples = "<br>".join(relative_source(path, drawings_dir, photos_dir) for path in missing_paths[:5])
        if len(missing_paths) > 5:
            examples += f"<br>...另 {len(missing_paths) - 5} 张"
        rows.append([key, expected, present, status, reason, examples])
        if status in {"失败", "需复核"}:
            failure_rows.append([f"图片替换-{key}", status, reason + (f" 未检出示例：{examples.replace('<br>', '；')}" if examples else "")])
            issue_texts.append(f"{key}：{reason}")
    return rows, failure_rows, issue_texts


def contains_meaningful_fragment(haystack: str, text: str, *, min_len: int = 8) -> bool:
    normalized_haystack = re.sub(r"\s+", "", haystack)
    normalized = re.sub(r"\s+", "", str(text or ""))
    if not normalized:
        return False
    if normalized in normalized_haystack:
        return True
    fragments = re.split(r"[，,。；;、：:（）()《》“”\"\\[\\]]+", normalized)
    fragments = [fragment for fragment in fragments if len(fragment) >= min_len]
    return any(fragment in normalized_haystack for fragment in fragments)


def cultural_relic_audit_rows(
    *,
    doc_text: str,
    cultural_relic_type: str,
    form_title: str,
    form_docno: str,
    form_conclusion: str,
    review_reply_path: Path | None,
    review_reply_text: str,
    manual_builder,
) -> tuple[list[list[object]], list[list[object]], list[str]]:
    rows: list[list[object]] = []
    failure_rows: list[list[object]] = []
    issue_texts: list[str] = []

    source_title = review_reply_path.stem.strip() if review_reply_path else ""
    source_docno = ""
    source_conclusion = ""
    if manual_builder:
        try:
            source_docno = (
                manual_builder.extract_document_number(review_reply_path.name) if review_reply_path else ""
            ) or manual_builder.extract_document_number(review_reply_text)
            source_conclusion = manual_builder.extract_review_conclusion(review_reply_text)
        except Exception:
            source_docno = ""
            source_conclusion = ""

    no_document_number_values = {"", "无", "无文号", "未见文号", "未识别", "未识别文号", "无正式文号"}

    def add_row(item: str, expected: str, actual: str, status: str, reason: str) -> None:
        rows.append([item, expected, actual, status, reason])
        if status in {"失败", "需复核"}:
            failure_rows.append([f"文物概况-{item}", status, reason])
            issue_texts.append(f"{item}：{reason}")

    if review_reply_path:
        add_row(
            "回函识别",
            f"应识别为有文物审查意见：{source_title}",
            cultural_relic_type or "未填写",
            "通过" if cultural_relic_type == "有文物审查意见" else "失败",
            "项目资料中已识别到文物部门回函，填报表和报告应走“有文物审查意见”分支。"
            if cultural_relic_type != "有文物审查意见"
            else "项目资料回函已识别，填报表分支正确。",
        )
        add_row(
            "回函标题",
            source_title or "应从回函文件名提取",
            form_title or "未填写",
            "通过" if form_title and contains_meaningful_fragment(doc_text, form_title, min_len=10) else "需复核",
            "回函标题应填入表格，并在报告文物概况中出现。"
            if not (form_title and contains_meaningful_fragment(doc_text, form_title, min_len=10))
            else "回函标题已在报告中检出。",
        )
        source_has_docno = bool(source_docno and source_docno not in no_document_number_values)
        form_has_docno = bool(form_docno and form_docno not in no_document_number_values)
        if source_has_docno or form_has_docno:
            docno_ok = form_has_docno and contains_meaningful_fragment(doc_text, form_docno, min_len=6)
            docno_reason = "回函文号已在报告中检出。" if docno_ok else "回函文号应填入表格，并在报告文物概况中出现。"
            docno_status = "通过" if docno_ok else "需复核"
        else:
            docno_status = "通过"
            docno_reason = "回函正文和文件名未识别到正式文号，报告正文可不写文号。"
        add_row(
            "回函文号",
            source_docno or "未识别到正式文号",
            form_docno or "未填写",
            docno_status,
            docno_reason,
        )
        expected_conclusion = source_conclusion or form_conclusion
        conclusion_in_doc = bool(
            (form_conclusion and contains_meaningful_fragment(doc_text, form_conclusion, min_len=10))
            or (source_conclusion and contains_meaningful_fragment(doc_text, source_conclusion, min_len=10))
        )
        if not form_conclusion and not source_conclusion:
            status = "失败"
            reason = "未能从填报表或回函正文取得“文物审查意见结论”；应提取未涉及/涉及、重叠、避让、考古勘探或审批要求。"
        elif not conclusion_in_doc:
            status = "失败"
            reason = "已取得文物审查意见结论，但报告文物概况中未检出该结论，可能仍写成待补或沿用模板旧内容。"
        elif source_conclusion and not contains_meaningful_fragment(form_conclusion, source_conclusion, min_len=10):
            if form_conclusion:
                status = "需复核"
                reason = "表格结论与回函正文自动提取的核心意见不一致，需人工核对。"
            else:
                status = "通过"
                reason = "填报表未填写结论，但报告已采用回函正文自动提取的核心意见。"
        else:
            status = "通过"
            reason = "文物审查意见结论已在报告文物概况中检出。"
        actual_conclusion = form_conclusion or ("填报表未填写；报告按回函正文自动提取" if source_conclusion else "未填写")
        add_row("回函结论", expected_conclusion or "应从回函正文提取", actual_conclusion, status, reason)
        if "文物概况内容待补" in doc_text or "文物审查意见待补" in doc_text:
            add_row(
                "待补占位",
                "有回函且信息完整时报告不得保留待补句",
                "报告中检出待补句",
                "失败",
                "报告文物概况仍保留“待补”内容，应补齐回函结论后重新生成。",
            )
    else:
        add_row(
            "回函识别",
            "项目资料未识别到文物部门回函",
            cultural_relic_type or "未填写",
            "通过" if cultural_relic_type != "有文物审查意见" else "需复核",
            "未识别到回函但填报表选择了有文物审查意见，请核对回函文件是否漏放或识别规则是否不足。"
            if cultural_relic_type == "有文物审查意见"
            else "未识别到回函，文物概况分支未发现明显冲突。",
        )
    return rows, failure_rows, issue_texts


def normalize_tk_code(value: object) -> str:
    text = "" if value is None else str(value).strip().upper()
    match = re.search(r"TK\s*0*(\d+)", text, flags=re.I)
    if match:
        return f"TK{int(match.group(1))}"
    if re.fullmatch(r"\d+", text):
        return f"TK{int(text)}"
    return text


def unused_image_reason(path: Path, drawings_dir: Path, photos_dir: Path, tables: dict[str, list[dict[str, str]]]) -> str:
    rel = relative_source(path, drawings_dir, photos_dir)
    standard_ids = {
        normalize_tk_code(row.get("探孔编号", ""))
        for row in tables.get("标准孔", [])
        if row.get("探孔编号")
    }
    stem = normalize_tk_code(path.stem)
    if "标准孔照" in rel:
        if stem not in standard_ids:
            return "标准孔坐标表/标准孔记录未包含该探孔编号，不能作为标准孔详情图插入。"
        return "存在标准孔记录，但未被当前标准孔土样照匹配规则命中，需复核编号补零或文件命名。"
    if re.search(r"地层堆积剖[线面]图|剖[线面]图", rel):
        return "该剖面图未在生成稿媒体中检出；若报告中肉眼可见，请复核源图是否另有同名文件、格式转换或模板中残留旧图。"
    if re.fullmatch(r"项目地块在.+位置示意图", path.stem):
        return "基准模板需设置可扩展槽位“图{{自动图号}} {{图:项目地块位置示意图}}”，生成器会按文件名逐张插入并自动改图名。"
    if "文献资料收集与整理工作照" in rel:
        return "该类照片应插入“文献资料收集与整理工作照”槽位；若同类照片重复，生成器只保留一组代表图，不重复插入。"
    if "资料整理工作照" in rel:
        return "该类照片应插入“资料整理工作照”槽位；若同类照片重复，生成器只保留一组代表图，不重复插入。"
    if "位置图" in rel:
        return "当前模板已有省/市/旗县/项目地块位置图槽位，该图片名称未命中现有位置图槽位关键词。"
    if "红线" in rel:
        return "当前模板已匹配红线/四至坐标相关图件，该图片名称未命中具体槽位或属于备用图。"
    if "13.遗迹" in rel and re.fullmatch(r"K\d+", path.stem, flags=re.I):
        return "该遗迹照片文件名为 K 序号，当前遗迹记录使用 H/M 编号，无法稳定确认对应关系，不能自动作为遗迹照片插入。"
    return "源资料图片未命中当前模板图片槽位或图片清单；可能是备用、过程照或缺少对应占位符。"


def unmatched_source_image_status(
    path: Path,
    *,
    media_hashes: set[str],
    expected_hashes: set[str],
    expected_paths: set[Path],
    drawings_dir: Path,
    photos_dir: Path,
    tables: dict[str, list[dict[str, str]]],
) -> tuple[str, str]:
    try:
        current_hash = file_hash(path)
    except OSError:
        current_hash = ""
    rel = relative_source(path, drawings_dir, photos_dir)
    if current_hash and current_hash in expected_hashes:
        return (
            "无需重复插入",
            "该源图与已匹配并插入的图片内容完全相同；生成器按去重规则只插入一次，避免报告中重复放图。",
        )
    if current_hash and current_hash in media_hashes:
        if re.search(r"地层堆积剖[线面]图|剖[线面]图", rel):
            return (
                "已插入",
                "该剖面图源文件字节已在生成报告媒体中检出；此前列入未插入是因为检查器只按模板图片槽位统计，漏算了剖线展开逻辑。",
            )
        return (
            "已插入",
            "该源图字节已在生成报告媒体中检出；此前未被图片槽位统计命中，属于检查统计口径问题。",
        )
    is_drawing_source = is_relative_to(path, drawings_dir)
    if is_drawing_source:
        related_groups = [
            ("红线", "红线图件"),
            ("位置示意图", "位置示意图件"),
            ("位置图", "位置图件"),
            ("勘探区域", "勘探区域图件"),
            ("勘探分区", "勘探分区图件"),
            ("勘探单元", "勘探单元布设图件"),
        ]
        for token, label in related_groups:
            if token not in rel and token not in path.stem:
                continue
            related_inserted = []
            for expected in expected_paths:
                if not is_relative_to(expected, drawings_dir):
                    continue
                expected_text = relative_source(expected, drawings_dir, photos_dir)
                if token not in expected_text and token not in expected.stem:
                    continue
                try:
                    if file_hash(expected) in media_hashes:
                        related_inserted.append(relative_source(expected, drawings_dir, photos_dir))
                except OSError:
                    continue
            if related_inserted:
                examples = "、".join(sorted(set(related_inserted))[:3])
                return (
                    "无需插入",
                    f"同类{label}已采用并插入报告：{examples}；当前文件未作为报告图片使用，不属于缺插。",
                )
    return "未插入", unused_image_reason(path, drawings_dir, photos_dir, tables)


def write_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fills = {
        "已替换": "E2F0D9",
        "已匹配": "E2F0D9",
        "全部插入或已匹配": "E2F0D9",
        "无需插入": "DDEBF7",
        "无需重复插入": "DDEBF7",
        "已插入": "E2F0D9",
        "无阻断性问题": "E2F0D9",
        "报告中出现": "DDEBF7",
        "不要求替换": "F2F2F2",
        "未填写": "F2F2F2",
        "部分插入": "FFF2CC",
        "有未匹配源图片": "FFF2CC",
        "需复核": "FFF2CC",
        "未匹配": "FCE4D6",
        "替换失败": "F4CCCC",
        "失败": "F4CCCC",
    }
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in rows:
        normalized = [str(value).replace("<br>", "\n") if value is not None else "" for value in row]
        ws.append(normalized)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        status = next((str(cell.value) for cell in row if str(cell.value) in status_fills), "")
        fill = PatternFill("solid", fgColor=status_fills[status]) if status else None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        values = [str(cell.value or "") for cell in column_cells[: min(len(column_cells), 80)]]
        max_len = max((len(value) for value in values), default=8)
        ws.column_dimensions[letter].width = max(10, min(55, max_len + 2))
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 34


def write_xlsx_report(
    output: Path,
    metadata_rows: list[list[object]],
    issue_rows: list[list[object]],
    structure_rows: list[list[object]],
    replacement_rows: list[list[object]],
    cultural_relic_rows: list[list[object]],
    table_replacement_rows: list[list[object]],
    appendix_integrity_rows: list[list[object]],
    image_replacement_rows: list[list[object]],
    summary_rows: list[list[object]],
    category_rows: list[list[object]],
    unused_rows: list[list[object]],
    slot_rows: list[list[object]],
    failure_rows: list[list[object]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_table_sheet(wb, "概览", ["项目", "内容", "说明"], metadata_rows + [["", "", ""]] + issue_rows)
    write_table_sheet(wb, "DOCX结构", ["检查项", "结果", "说明"], structure_rows)
    write_table_sheet(wb, "字段替换检查", ["工作表", "字段", "有效字段值", "模板替换符", "报告中出现", "判断", "原因"], replacement_rows)
    write_table_sheet(wb, "文物概况检查", ["检查项", "应满足规则/来源", "当前值/报告检出", "判断", "原因"], cultural_relic_rows)
    write_table_sheet(wb, "表格替换检查", ["检查对象", "来源表", "期望记录数", "生成计数", "判断", "原因", "说明"], table_replacement_rows)
    write_table_sheet(wb, "附图完整性检查", ["检查对象", "源文件", "报告检出", "判断", "原因", "说明"], appendix_integrity_rows)
    write_table_sheet(wb, "图片替换检查", ["图片槽位", "规则匹配源图数", "报告检出源图数", "判断", "原因", "未检出源图示例"], image_replacement_rows)
    write_table_sheet(wb, "图片覆盖检查", ["资料类别", "源图片数", "规则匹配数", "未插入数", "判断", "原因"], category_rows)
    write_table_sheet(wb, "未插入图片明细", ["资料类别", "文件", "结果", "未插入原因"], unused_rows)
    write_table_sheet(wb, "已匹配图片槽位", ["图片槽位", "匹配数量", "匹配文件示例", "判断", "原因"], slot_rows)
    write_table_sheet(wb, "错误与失败原因", ["检查对象", "判断", "原因/处理建议"], failure_rows)
    write_table_sheet(wb, "生成检查摘录", ["项目", "结果", "说明"], summary_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def first_items(items: list[str], limit: int = 12) -> str:
    if not items:
        return ""
    shown = "、".join(items[:limit])
    if len(items) > limit:
        shown += f"；另 {len(items) - limit} 项"
    return shown


def compact_numbered_items(items: list[str]) -> list[str]:
    grouped: dict[tuple[str, str], list[int]] = defaultdict(list)
    passthrough: list[str] = []
    for item in items:
        match = re.fullmatch(r"(.+?:)([A-Za-z]+)0*(\d+)", item)
        if not match:
            passthrough.append(item)
            continue
        grouped[(match.group(1), match.group(2).upper())].append(int(match.group(3)))
    compacted = list(passthrough)
    for (prefix, code), numbers in sorted(grouped.items(), key=lambda row: (row[0][0], row[0][1])):
        unique_numbers = sorted(set(numbers))
        ranges: list[str] = []
        start = previous = unique_numbers[0]
        for number in unique_numbers[1:]:
            if number == previous + 1:
                previous = number
                continue
            ranges.append(f"{code}{start}" if start == previous else f"{code}{start}-{code}{previous}")
            start = previous = number
        ranges.append(f"{code}{start}" if start == previous else f"{code}{start}-{code}{previous}")
        compacted.append(prefix + "、".join(ranges))
    return compacted


CHINESE_ORDER_NUMBERS = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]


def add_brief_item(categories: dict[str, list[str]], category: str, item: str) -> None:
    item = str(item).strip()
    if item and item not in categories[category]:
        categories[category].append(item)


def brief_category_for_required_field(item: str) -> str:
    if re.search(r"人员|进场人数|探工|测绘|技师|负责|执笔|校核", item):
        return "人员信息待补"
    return "基础字段待补"


def brief_category_for_failure(target: str) -> str:
    if target.startswith("表格替换"):
        return "表格替换待复核"
    if target.startswith("附图"):
        return "附图待复核"
    if target.startswith("图片替换"):
        return "图片替换待复核"
    if target.startswith("回函") or target.startswith("文物概况"):
        return "回函信息待复核"
    if target.startswith("标准孔"):
        return "标准孔信息待补"
    if target.startswith("分区"):
        return "分区信息待补"
    if target in {"公司模板"}:
        return "模板待复核"
    if target in {"剖线图后说明"}:
        return "剖线信息待补"
    if target in {"现场限制句"}:
        return "正文内容待复核"
    if target in {"图片未插入", "图片插入错误"}:
        return "图片插入待复核"
    if target in {"正文匹配错误"}:
        return "正文内容待复核"
    return "其他待复核"


def write_brief_report(
    output: Path,
    *,
    project_name: str,
    generated_errors: list[str],
    generated_info_items: list[str],
    missing_items: list[str],
    unresolved: list[str],
    failure_rows: list[list[object]],
    cultural_relic_rows: list[list[object]],
    table_replacement_rows: list[list[object]],
    appendix_integrity_rows: list[list[object]],
    image_replacement_rows: list[list[object]],
    unused_rows: list[list[object]],
) -> None:
    def status_counts(rows: list[list[object]], status_index: int) -> dict[str, int]:
        counts: dict[str, int] = defaultdict(int)
        for row in rows:
            if len(row) <= status_index:
                continue
            status = str(row[status_index] or "").strip()
            if status:
                counts[status] += 1
        return dict(counts)

    def format_counts(counts: dict[str, int], order: list[str]) -> str:
        parts = [f"{status} {counts[status]} 项" for status in order if counts.get(status)]
        for status in sorted(set(counts) - set(order)):
            parts.append(f"{status} {counts[status]} 项")
        return "；".join(parts) if parts else "无记录"

    categories: dict[str, list[str]] = defaultdict(list)
    for item in generated_errors:
        category = brief_category_for_required_field(item)
        add_brief_item(categories, category, item)
    for item in generated_info_items:
        if re.search(r"进场人数缺失|人员.*缺失|探工.*缺失|人员套组.*确认|人员套组.*缺失", item):
            add_brief_item(categories, "人员信息待补", item)
    if missing_items:
        for item in compact_numbered_items(missing_items):
            add_brief_item(categories, "图件待补", item)
    if unresolved:
        for item in unresolved:
            add_brief_item(categories, "占位符待处理", item)
    for row in failure_rows:
        if len(row) < 3:
            continue
        target, status, reason = (str(row[0]), str(row[1]), str(row[2]))
        if target == "整体" or status in {"无阻断性问题", "有未匹配源图片"}:
            continue
        if target == "图件待补" and missing_items:
            continue
        if target == "占位符" and unresolved:
            continue
        add_brief_item(categories, brief_category_for_failure(target), f"{target}：{reason}")

    category_order = [
        "图件待补",
        "人员信息待补",
        "基础字段待补",
        "标准孔信息待补",
        "剖线信息待补",
        "图片替换待复核",
        "图片插入待复核",
        "附图待复核",
        "表格替换待复核",
        "回函信息待复核",
        "分区信息待补",
        "模板待复核",
        "占位符待处理",
        "正文内容待复核",
        "其他待复核",
    ]
    active_categories = [(name, categories[name]) for name in category_order if categories.get(name)]
    for name in sorted(set(categories) - set(category_order)):
        active_categories.append((name, categories[name]))

    lines = ["智能报告问题摘要", ""]
    if project_name:
        lines.append(f"项目：{project_name}")
        lines.append("")
    lines.append("检查结果：")
    table_counts = status_counts(table_replacement_rows, 4)
    appendix_counts = status_counts(appendix_integrity_rows, 3)
    cultural_counts = status_counts(cultural_relic_rows, 3)
    image_counts = status_counts(image_replacement_rows, 3)
    truly_uninserted_count = sum(1 for row in unused_rows if len(row) > 2 and str(row[2]) == "未插入")
    duplicate_count = sum(1 for row in unused_rows if len(row) > 2 and str(row[2]) == "无需重复插入")
    not_required_count = sum(1 for row in unused_rows if len(row) > 2 and str(row[2]) == "无需插入")
    already_present_count = sum(1 for row in unused_rows if len(row) > 2 and str(row[2]) == "已插入")
    uninserted_reason_counts: dict[str, int] = defaultdict(int)
    for row in unused_rows:
        if len(row) > 3 and str(row[2]) == "未插入":
            reason = str(row[3] or "").strip()
            reason = re.sub(r"；.*$", "", reason)
            reason = reason.rstrip("。；;：:")
            uninserted_reason_counts[reason] += 1
    table_summary = format_counts(table_counts, ["通过", "不适用", "需复核", "失败"])
    cultural_summary = format_counts(cultural_counts, ["通过", "需复核", "失败", "不适用"])
    image_summary = format_counts(image_counts, ["通过", "需复核", "失败"])
    appendix_summary = format_counts(appendix_counts, ["通过", "不适用", "需复核", "失败"])
    table_failures = [
        str(row[0])
        for row in table_replacement_rows
        if len(row) > 4 and str(row[4] or "").strip() not in {"", "通过", "不适用"}
    ]
    image_failures = [
        str(row[0])
        for row in image_replacement_rows
        if len(row) > 3 and str(row[3] or "").strip() not in {"", "通过"}
    ]
    appendix_failures = [
        str(row[0])
        for row in appendix_integrity_rows
        if len(row) > 3 and str(row[3] or "").strip() not in {"", "通过", "不适用"}
    ]
    cultural_failures = [
        str(row[0])
        for row in cultural_relic_rows
        if len(row) > 3 and str(row[3] or "").strip() not in {"", "通过", "不适用"}
    ]
    if table_failures:
        lines.append(f"1、表格替换检查：未全部替换通过，{table_summary}。")
    else:
        lines.append(f"1、表格替换检查：已全部替换通过，{table_summary}。")
    if table_failures:
        lines.append("   需复核：" + "、".join(table_failures[:12]) + ("等" if len(table_failures) > 12 else "") + "。")
    if cultural_failures:
        lines.append(f"2、文物概况检查：未按规则准确生成，{cultural_summary}。")
        lines.append("   需复核：" + "、".join(cultural_failures[:8]) + ("等" if len(cultural_failures) > 8 else "") + "。")
    else:
        lines.append(f"2、文物概况检查：已按规则生成，{cultural_summary}。")
    if image_failures:
        lines.append(f"3、图片替换检查：未全部替换通过，{image_summary}。")
    else:
        lines.append(f"3、图片替换检查：已匹配图片全部替换通过，{image_summary}。")
    if image_failures:
        lines.append("   需复核：" + "、".join(image_failures[:12]) + ("等" if len(image_failures) > 12 else "") + "。")
    else:
        lines.append("   已匹配源图片均在生成报告媒体中检出。")
    if appendix_failures:
        lines.append(f"4、附图完整性检查：未全部通过，{appendix_summary}。")
        lines.append("   需复核：" + "、".join(appendix_failures[:12]) + ("等" if len(appendix_failures) > 12 else "") + "。")
    else:
        lines.append(f"4、附图完整性检查：已按源资料完整插入，{appendix_summary}。")
    if truly_uninserted_count:
        lines.append(f"5、图片插入覆盖：未全部插入，确认有 {truly_uninserted_count} 张源图片未插入报告，明细见专项检查表“未插入图片明细”。")
    else:
        lines.append("5、图片插入覆盖：已全部插入或无需重复插入，未发现确认缺插的源图片。")
    if duplicate_count or already_present_count:
        parts = []
        if already_present_count:
            parts.append(f"{already_present_count} 张源图已在报告媒体中检出但此前未被槽位统计命中")
        if duplicate_count:
            parts.append(f"{duplicate_count} 张源图与已插入图片重复、无需重复插入")
        if not_required_count:
            parts.append(f"{not_required_count} 张源图已有同类图件采用、无需插入")
        lines.append("   说明：" + "；".join(parts) + "。")
    if uninserted_reason_counts:
        reason_parts = [
            f"{reason}：{count} 张"
            for reason, count in sorted(uninserted_reason_counts.items(), key=lambda item: (-item[1], item[0]))
        ]
        lines.append("   未插入原因统计：" + "；".join(reason_parts[:6]) + ("。" if len(reason_parts) <= 6 else "；详见专项检查表。"))
    lines.append("")
    if not active_categories:
        lines.append("未发现需处理的问题。")
    else:
        lines.append(f"本次需处理 {len(active_categories)} 类问题：")
        for idx, (category, items) in enumerate(active_categories):
            number = CHINESE_ORDER_NUMBERS[idx] if idx < len(CHINESE_ORDER_NUMBERS) else str(idx + 1)
            lines.append(f"{number}、{category}")
            for item_idx, item in enumerate(items, start=1):
                lines.append(f"{item_idx}、{item}")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_report(args) -> None:
    filler = load_filler()
    manual_builder = load_manual_builder()
    docx = args.docx.resolve()
    form = args.form.resolve()
    drawings_dir = args.drawings_dir.resolve()
    photos_dir = args.photos_dir.resolve()
    check_report = args.check_report.resolve()
    out = args.output.resolve() if args.output else None

    texts, metrics = docx_text_and_metrics(docx)
    generated_check = parse_generated_check(check_report)
    fields, tables, notes, slots = build_expected_slots(filler, form, drawings_dir, photos_dir)

    missing_items = generated_check.get("sections", {}).get("图件待补", []) if isinstance(generated_check.get("sections"), dict) else []
    generated_errors = generated_check.get("sections", {}).get("错误", []) if isinstance(generated_check.get("sections"), dict) else []
    generated_info_items = generated_check.get("sections", {}).get("信息", []) if isinstance(generated_check.get("sections"), dict) else []
    unresolved = metrics["unresolved"]
    project_name = workbook_value(form, "项目基础信息", "项目名称")
    survey_unit = workbook_value(form, "项目基础信息", "勘探单位")
    cultural_relic_type = workbook_value(form, "文物概况", "文物概况类型")
    form_review_title = workbook_value(form, "文物概况", "文物审查意见文件名")
    form_review_docno = workbook_value(form, "文物概况", "文物审查意见文号")
    form_review_conclusion = workbook_value(form, "文物概况", "文物审查意见结论")
    template = str(generated_check.get("info", {}).get("智能模板", "")) if isinstance(generated_check.get("info"), dict) else ""
    image_keys = template_image_keys(template)
    review_reply_path = None
    review_reply_text = ""
    project_dir = infer_project_dir_from_sources(drawings_dir, photos_dir)
    if manual_builder and project_dir and project_dir.exists():
        try:
            should_read_review_text = not (form_review_title and form_review_docno and form_review_conclusion)
            review_reply_path, review_reply_text = manual_builder.find_cultural_relic_review_reply(
                project_dir,
                read_text=should_read_review_text,
            )
        except Exception:
            review_reply_path, review_reply_text = None, ""
    if manual_builder and not review_reply_path:
        form_review_path = workbook_value(form, "文物概况", "回函文件路径")
        if form_review_path:
            candidate = Path(form_review_path).expanduser()
            if candidate.exists():
                review_reply_path = candidate.resolve()
                try:
                    needs_text = not (
                        form_review_title
                        and form_review_docno
                        and form_review_conclusion
                        and form_review_conclusion != "匹配错误"
                    )
                    if needs_text:
                        if review_reply_path.suffix.lower() == ".pdf":
                            review_reply_text = manual_builder.read_pdf_text(review_reply_path)
                        elif review_reply_path.suffix.lower() == ".docx":
                            review_reply_text = "\n".join(manual_builder.read_docx_texts(review_reply_path))
                except Exception:
                    review_reply_text = ""

    expected_paths: set[Path] = set()
    slot_summary: dict[str, list[Path]] = defaultdict(list)
    for key, queue in slots.items():
        base_key = key.split(":", 1)[0]
        if image_keys and key not in image_keys and base_key not in image_keys:
            continue
        for slot in queue:
            if slot.path:
                expected_paths.add(Path(slot.path).resolve())
                slot_summary[key].append(Path(slot.path).resolve())

    media_hashes = docx_media_hashes(docx)
    expected_hashes: set[str] = set()
    for path in expected_paths:
        try:
            expected_hashes.add(file_hash(path))
        except OSError:
            pass
    source_images = all_source_images(drawings_dir, photos_dir)
    unused = [path for path in source_images if path.resolve() not in expected_paths]
    unused_by_category: dict[str, list[Path]] = defaultdict(list)
    for path in unused:
        unused_by_category[source_category(path, drawings_dir, photos_dir)].append(path)
    source_status_by_path: dict[Path, tuple[str, str]] = {}
    for path in unused:
        source_status_by_path[path] = unmatched_source_image_status(
            path,
            media_hashes=media_hashes,
            expected_hashes=expected_hashes,
            expected_paths=expected_paths,
            drawings_dir=drawings_dir,
            photos_dir=photos_dir,
            tables=tables,
        )
    truly_uninserted = [
        path for path, status_reason in source_status_by_path.items()
        if status_reason[0] == "未插入"
    ]
    duplicate_or_already_present = [
        path for path, status_reason in source_status_by_path.items()
        if status_reason[0] in {"已插入", "无需重复插入", "无需插入"}
    ]

    replacement_rows = replacement_audit_rows(
        form,
        docx,
        template,
        unresolved if isinstance(unresolved, list) else [],
        fields,
    )
    table_replacement_rows, table_failure_rows, table_issue_texts = table_replacement_audit_rows(
        tables,
        generated_check.get("info", {}) if isinstance(generated_check.get("info"), dict) else {},
    )
    image_replacement_rows, image_failure_rows, image_issue_texts = image_replacement_audit_rows(
        docx,
        slot_summary,
        drawings_dir,
        photos_dir,
    )
    appendix_integrity_rows, appendix_failure_rows, appendix_issue_texts = appendix_section_drawing_audit_rows(
        drawings_dir,
        photos_dir,
        texts,
    )

    issues: list[str] = []
    joined_text = "\n".join(texts)
    match_error_count = joined_text.count("匹配错误")
    image_insert_error_count = joined_text.count("插入错误")
    cultural_relic_rows, cultural_relic_failure_rows, cultural_relic_issue_texts = cultural_relic_audit_rows(
        doc_text=joined_text,
        cultural_relic_type=cultural_relic_type,
        form_title=form_review_title,
        form_docno=form_review_docno,
        form_conclusion=form_review_conclusion,
        review_reply_path=review_reply_path,
        review_reply_text=review_reply_text,
        manual_builder=manual_builder,
    )
    missing_partition_long_fields = [
        key for key in ["勘探分区原因", "勘探分区逐项说明"]
        if has_partition(fields) and not fields.get(key)
    ]
    partition_long_text_not_found = [
        key for key in ["勘探分区原因", "勘探分区逐项说明"]
        if has_partition(fields) and fields.get(key) and fields.get(key) not in joined_text
    ]
    if missing_items:
        issues.append(f"生成检查报告仍有图件待补：{len(missing_items)} 项。")
    if match_error_count:
        issues.append(f"生成稿中存在“匹配错误”标记：{match_error_count} 处。")
    if image_insert_error_count:
        issues.append(f"生成稿中存在“插入错误”标记：{image_insert_error_count} 处。")
    if unresolved:
        issues.append(f"DOCX 仍有未解析占位符：{len(unresolved)} 项。")
    if missing_partition_long_fields:
        issues.append("有分区项目缺少现场记录长文本：" + "、".join(missing_partition_long_fields) + "。")
    if partition_long_text_not_found:
        issues.append("有分区长文本未在生成稿中直接检出：" + "、".join(partition_long_text_not_found) + "。")
    if survey_unit and template and survey_unit not in Path(template).name:
        issues.append(f"勘探单位为“{survey_unit}”，但智能模板文件名不是该公司模板：{Path(template).name}。")
    if review_reply_path and cultural_relic_type != "有文物审查意见":
        issues.append(f"项目资料中识别到文物部门回函，但填报表文物概况类型为“{cultural_relic_type or '未填写'}”。")
    section_description_found = has_section_description_text(texts)
    if not section_description_found:
        issues.append("正文剖线图后未检测到 A-A′ / B-B′ 的地层堆积说明段，剖线图后说明排版可能未生效。")
    standard_count = len(tables.get("标准孔", []))
    standard_photo_count = slot_count_by_prefix(slot_summary, "标准孔土样照")
    standard_position_count = slot_count_by_prefix(slot_summary, "标准孔位置图")
    if standard_count and standard_photo_count < standard_count:
        issues.append(f"标准孔土样照匹配不足：标准孔记录 {standard_count} 条，匹配到土样照 {standard_photo_count} 张。")
    if standard_count and standard_position_count < standard_count:
        issues.append(f"标准孔位置图匹配不足：标准孔记录 {standard_count} 条，匹配到位置图 {standard_position_count} 张。")
    duplicated_limit = [
        text for text in texts
        if "不具备勘探条件" in text and text.count("实际完成考古勘探面积") > 1
    ]
    if duplicated_limit:
        issues.append("结论段存在现场限制语义重复：同一段中“实际完成考古勘探面积”出现多次。")
    if table_issue_texts:
        issues.append(f"表格替换检查发现 {len(table_issue_texts)} 项需复核。")
    if image_issue_texts:
        issues.append(f"图片替换检查发现 {len(image_issue_texts)} 项需复核。")
    if appendix_issue_texts:
        issues.append("附图完整性检查未通过：" + "；".join(appendix_issue_texts[:3]) + ("。" if len(appendix_issue_texts) <= 3 else "；详见专项检查表。"))
    if cultural_relic_issue_texts:
        issues.append(f"文物概况检查发现 {len(cultural_relic_issue_texts)} 项需复核。")
    if truly_uninserted:
        issues.append(f"源资料中有 {len(truly_uninserted)} 张图片确认未插入报告，需按“未插入图片明细”逐张复核。")

    metadata_rows = [
        ["项目", project_name or "未识别", ""],
        ["生成报告", str(docx), ""],
        ["填报表", str(form), ""],
        ["制图成果", str(drawings_dir), ""],
        ["外业成果", str(photos_dir), ""],
    ]
    issue_rows = [["结论", item, ""] for item in issues] if issues else [["结论", "未发现结构性问题。", ""]]
    structure_rows = [
        ["ZIP/XML 可读取", "是", "DOCX 包可打开并读取 document.xml。"],
        ["段落数", metrics["paragraph_count"], ""],
        ["媒体文件数", metrics["media_count"], "word/media 内文件数量。"],
        ["图片引用数", metrics["image_reference_count"], "正文实际引用图片次数。"],
        ["正文图题数", metrics["body_figure_caption_count"], "匹配“图N ...”的正文题注。"],
        ["未解析占位符", len(unresolved), "应为 0。"],
        ["标准孔记录数", standard_count, "来自填报表/外部标准孔坐标表。"],
        ["标准孔位置图匹配数", standard_position_count, "应与标准孔记录数一致。"],
        ["标准孔土样照匹配数", standard_photo_count, "应与标准孔记录数一致。"],
    ]
    info = generated_check.get("info", {})
    summary_rows = []
    if isinstance(info, dict):
        for key in [
            "自动插入图件",
            "红线坐标生成",
            "勘探单元生成",
            "标准孔坐标生成",
            "剖线地层堆积生成",
            "剖线图后文字说明排版",
            "坐标基点=西南角（X=33383085.344，Y=4705868.923，由四至范围坐标导入）",
        ]:
            if key in info:
                summary_rows.append([key, info[key], "来自生成检查报告。"])
        for item in generated_check.get("sections", {}).get("信息", []):
            if item.startswith("坐标基点=") or item.startswith("已从") or item.startswith("勘探单元数量="):
                summary_rows.append([item, "已记录", "外部资料导入证据。"])
    category_rows = []
    expected_by_category: dict[str, set[Path]] = defaultdict(set)
    source_by_category: dict[str, list[Path]] = defaultdict(list)
    for path in expected_paths:
        expected_by_category[source_category(path, drawings_dir, photos_dir)].add(path)
    for path in source_images:
        source_by_category[source_category(path, drawings_dir, photos_dir)].append(path)
    for category in sorted(source_by_category):
        total = len(source_by_category[category])
        matched = len(expected_by_category.get(category, set()))
        rows_for_category = [source_status_by_path.get(path) for path in unused_by_category.get(category, [])]
        true_uninserted = sum(1 for status_reason in rows_for_category if status_reason and status_reason[0] == "未插入")
        duplicate_or_present = sum(
            1 for status_reason in rows_for_category if status_reason and status_reason[0] in {"已插入", "无需重复插入", "无需插入"}
        )
        if true_uninserted == 0 and duplicate_or_present == 0:
            status = "全部插入或已匹配"
            reason = "该类源图片均被当前规则匹配。"
        elif true_uninserted == 0:
            status = "全部插入或无需重复插入"
            reason = "未直接匹配的源图已在报告媒体中检出，或与已插入图片重复，无需另行插入。"
        elif matched == 0:
            status = "未匹配"
            reason = "当前模板无对应槽位、文件名未命中规则，或缺少对应记录。"
        else:
            status = "部分插入"
            reason = "部分图片被匹配；仍未插入的图片需按逐张原因复核。"
        category_rows.append([category, total, matched, true_uninserted, status, reason])
    unused_rows = []
    for category in sorted(unused_by_category):
        for path in unused_by_category[category]:
            status, reason = source_status_by_path.get(
                path,
                ("未插入", unused_image_reason(path, drawings_dir, photos_dir, tables)),
            )
            unused_rows.append([
                category,
                relative_source(path, drawings_dir, photos_dir),
                status,
                reason,
            ])
    slot_rows = []
    for key in sorted(slot_summary):
        unique = sorted(set(slot_summary[key]))
        examples = "<br>".join(relative_source(path, drawings_dir, photos_dir) for path in unique[:5])
        if len(unique) > 5:
            examples += f"<br>...另 {len(unique) - 5} 张"
        slot_rows.append([key, len(unique), examples, "已匹配", "生成器已为该槽位找到源图片。"])
    failure_rows = []
    failure_rows.extend(cultural_relic_failure_rows)
    failure_rows.extend(table_failure_rows)
    failure_rows.extend(appendix_failure_rows)
    failure_rows.extend(image_failure_rows)
    if review_reply_path and cultural_relic_type != "有文物审查意见":
        docno = ""
        conclusion = ""
        try:
            docno = manual_builder.extract_document_number(review_reply_path.name) or manual_builder.extract_document_number(review_reply_text)
            conclusion = manual_builder.extract_review_conclusion(review_reply_text)
        except Exception:
            pass
        detail = f"项目资料中识别到回函 `{review_reply_path.name}`"
        if docno:
            detail += f"，文号 `{docno}`"
        if conclusion:
            detail += f"，结论 `{conclusion}`"
        detail += f"，但填报表文物概况类型为“{cultural_relic_type or '未填写'}”，报告可能走了无回函分支。"
        failure_rows.append(["回函信息", "需复核", detail])
    if survey_unit and template and survey_unit not in Path(template).name:
        failure_rows.append(["公司模板", "需复核", f"勘探单位为“{survey_unit}”，但使用模板为 `{Path(template).name}`。正式报告应补建并使用该公司的智能基准模板。"])
    if not section_description_found:
        failure_rows.append(["剖线图后说明", "失败", "正文未检测到 A-A′ / B-B′ 地层堆积说明段，可能是图题匹配或插入位置失败。"])
    if duplicated_limit:
        failure_rows.append(["现场限制句", "需复核", "结论段同一段中“实际完成考古勘探面积”出现多次，可能是人工原因句与模板句叠加。"])
    if truly_uninserted:
        failure_rows.append(["图片未插入", "有未插入源图片", f"共 {len(truly_uninserted)} 张确认未插入报告；逐张原因见“未插入图片明细”。"])
    if standard_count and standard_photo_count < standard_count:
        failure_rows.append(["标准孔土样照", "失败", "标准孔记录数大于土样照匹配数，需复核标准孔坐标表探孔编号、土样照文件名或源文件是否缺失。"])
    if missing_items:
        failure_rows.append(["图件待补", "失败", "生成检查报告仍列出图件待补，需补齐图片源、图片清单或模板槽位。"])
    if match_error_count:
        failure_rows.append(["正文匹配错误", "失败", f"生成稿中存在 {match_error_count} 处“匹配错误”，说明文字、数据、复函或区域概况未能可靠匹配。"])
    if image_insert_error_count:
        failure_rows.append(["图片插入错误", "失败", f"生成稿中存在 {image_insert_error_count} 处“插入错误”，说明对应图片位未能可靠插入。"])
    if unresolved:
        failure_rows.append(["占位符", "失败", "生成稿仍残留未解析占位符。"])
    if missing_partition_long_fields:
        failure_rows.append([
            "分区长文本",
            "需复核",
            "有分区项目未从现场记录识别到："
            + "、".join(missing_partition_long_fields)
            + "。这些内容不放入填报表，建议修正/补充项目现场记录后重新生成。",
        ])
    if partition_long_text_not_found:
        failure_rows.append([
            "分区长文本",
            "需复核",
            "已识别到 "
            + "、".join(partition_long_text_not_found)
            + "，但生成稿未直接检出对应长文本，请检查模板段落是否被清空、改写或格式拆分。",
        ])
    if not failure_rows:
        failure_rows.append(["整体", "无阻断性问题", "未发现明确错插、替换失败或必需图片缺失；仍建议进行 Word 视觉版式复核。"])

    lines: list[str] = []
    lines.append("# 智能报告专项检查报告")
    lines.append("")
    lines.append(f"- 项目：{project_name or '未识别'}")
    lines.append(f"- 生成报告：{docx}")
    lines.append(f"- 填报表：{form}")
    lines.append(f"- 制图成果：{drawings_dir}")
    lines.append(f"- 外业成果：{photos_dir}")
    lines.append("")
    lines.append("## 结论")
    if issues:
        for item in issues:
            lines.append(f"- {item}")
    else:
        lines.append("- 未发现结构性问题。")
    lines.append("")
    lines.append("## DOCX 结构")
    lines.extend(markdown_table(
        ["检查项", "结果", "说明"],
        structure_rows,
    ))
    if unresolved:
        lines.append("")
        lines.extend(markdown_table(
            ["未解析占位符", "处理建议"],
            [[item, "补充填报值、修正模板替换符或补充生成规则。"] for item in unresolved],
        ))
    lines.append("")
    lines.append("## 字段替换检查")
    lines.append("说明：以基准模板中是否存在 `{{字段名}}` 为准。无同名替换符的字段不应强行替换，报告中未出现不视为错误。")
    lines.extend(markdown_table(
        ["工作表", "字段", "填报值", "模板替换符", "报告中出现填报值", "判断", "原因"],
        replacement_rows,
    ))
    lines.append("")
    lines.append("## 文物概况检查")
    lines.append("说明：核对项目资料中的文物审查意见回函、填报表文物概况字段和生成报告正文是否一致。")
    lines.extend(markdown_table(
        ["检查项", "应满足规则/来源", "当前值/报告检出", "判断", "原因"],
        cultural_relic_rows,
    ))
    lines.append("")
    lines.append("## 表格替换检查")
    lines.append("说明：对比本项目应导入的表格记录数与生成检查报告中的实际填充/生成计数，用于发现沿用模板旧表或未替换表格的风险。")
    lines.extend(markdown_table(
        ["检查对象", "来源表", "期望记录数", "生成计数", "判断", "原因", "说明"],
        table_replacement_rows,
    ))
    lines.append("")
    lines.append("## 附图完整性检查")
    lines.append("说明：核对结论后的附图是否按源资料全部有效地层堆积剖面图展开，避免模板只预留 A/B 时漏插 C-F 等后续剖面图。")
    lines.extend(markdown_table(
        ["检查对象", "源文件", "报告检出", "判断", "原因", "说明"],
        appendix_integrity_rows,
    ))
    lines.append("")
    lines.append("## 图片替换检查")
    lines.append("说明：对比当前规则已匹配的本项目源图与生成稿 `word/media` 中实际出现的图片字节，用于发现仍沿用模板旧图或照片/图件替换失败的风险。")
    lines.extend(markdown_table(
        ["图片槽位", "规则匹配源图数", "报告检出源图数", "判断", "原因", "未检出源图示例"],
        image_replacement_rows,
    ))
    lines.append("")
    lines.append("## 生成检查报告摘录")
    lines.extend(markdown_table(["项目", "结果", "说明"], summary_rows))
    lines.append("")
    lines.append("## 图片插入覆盖检查")
    lines.append("说明：“规则匹配数”是当前生成规则选中的源图片；“未插入数”是源资料中没有被任何图片槽位/图片清单选中的图片。")
    lines.extend(markdown_table(["资料类别", "源图片数", "规则匹配数", "未插入数", "判断", "原因"], category_rows))
    lines.append("")
    lines.append("## 未插入图片明细")
    lines.append("说明：这里的“未插入”指未被当前模板槽位/生成规则匹配为报告图片，不一定都是错误。")
    lines.extend(markdown_table(["资料类别", "文件", "结果", "未插入原因"], unused_rows))
    lines.append("")
    lines.append("## 已匹配图片槽位")
    lines.extend(markdown_table(["图片槽位", "匹配数量", "匹配文件示例", "判断", "原因"], slot_rows))
    lines.append("")
    lines.append("## 错误插入与失败原因")
    lines.extend(markdown_table(["检查对象", "判断", "原因/处理建议"], failure_rows))

    if out:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(out)
    if args.xlsx_output:
        xlsx_output = args.xlsx_output.resolve()
        write_xlsx_report(
            xlsx_output,
            metadata_rows,
            issue_rows,
            structure_rows,
            replacement_rows,
            cultural_relic_rows,
            table_replacement_rows,
            appendix_integrity_rows,
            image_replacement_rows,
            summary_rows,
            category_rows,
            unused_rows,
            slot_rows,
            failure_rows,
        )
        print(xlsx_output)
    if args.brief_output:
        brief_output = args.brief_output.resolve()
        write_brief_report(
            brief_output,
            project_name=project_name,
            generated_errors=generated_errors,
            generated_info_items=generated_info_items,
            missing_items=missing_items,
            unresolved=unresolved if isinstance(unresolved, list) else [],
            failure_rows=failure_rows,
            cultural_relic_rows=cultural_relic_rows,
            table_replacement_rows=table_replacement_rows,
            appendix_integrity_rows=appendix_integrity_rows,
            image_replacement_rows=image_replacement_rows,
            unused_rows=unused_rows,
        )
        print(brief_output)


def main() -> int:
    parser = argparse.ArgumentParser(description="检查智能生成勘探报告的图片覆盖和结构问题")
    parser.add_argument("--docx", type=Path, required=True)
    parser.add_argument("--form", type=Path, required=True)
    parser.add_argument("--drawings-dir", type=Path, required=True)
    parser.add_argument("--photos-dir", type=Path, required=True)
    parser.add_argument("--check-report", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--xlsx-output", type=Path, default=None)
    parser.add_argument("--brief-output", type=Path, default=None)
    args = parser.parse_args()
    write_report(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Create a user-facing manual fill-in workbook from a project source folder."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
import sys
from datetime import date
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_DIR.parent
ASSETS_DIR = SKILL_ROOT / "assets"
ROOT = Path.cwd().resolve()
TEMPLATE_DIR = ASSETS_DIR / "templates"


def find_template_file(filename: str) -> Path:
    direct_candidates = [
        TEMPLATE_DIR / filename,
        TEMPLATE_DIR / "forms" / filename,
    ]
    for path in direct_candidates:
        if path.exists():
            return path
    matches = sorted(
        path
        for path in TEMPLATE_DIR.rglob(filename)
        if path.is_file() and not path.name.startswith(".~") and not path.name.startswith("._")
    )
    return matches[0] if matches else direct_candidates[0]


TEMPLATE = find_template_file("人工填写表模板.xlsx")
PROCESS_DIR = ROOT / "过程资料"
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W_NS}
RELIC_ID_RE = r"[HMFGYK]\s*0*\d+"

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(ROOT / "过程资料"))
import fill_smart_template_from_form as smart_filler  # noqa: E402


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_chinese_date_range(text: str) -> tuple[date, date] | None:
    match = re.search(
        r"([0-9]{4})\s*年\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日\s*[—\\-至到]+\s*([0-9]{4})?\s*年?\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日",
        text,
    )
    if not match:
        return None
    start_year, start_month, start_day, end_year, end_month, end_day = match.groups()
    end_year = end_year or start_year
    try:
        start = date(int(start_year), int(start_month), int(start_day))
        end = date(int(end_year), int(end_month), int(end_day))
    except ValueError:
        return None
    return start, end


def inclusive_work_days_from_range(text: str) -> str:
    parsed = parse_chinese_date_range(text)
    if not parsed:
        return ""
    start, end = parsed
    if end < start:
        return ""
    return str((end - start).days + 1)


def safe_filename(value: str) -> str:
    return re.sub(r'[/:*?"<>|\\]+', "_", clean(value)) or "未命名项目"


def project_output_dir(project_label: str) -> Path:
    return PROCESS_DIR / safe_filename(project_label) / "人工填写表"


def discover_drawings_dir(project_dir: Path) -> Path:
    candidates = [
        project_dir / "3.执行资料" / "3.制图成果",
        project_dir / "3.执行资料" / "3.内业成果",
        project_dir / "3.执行资料" / "4.内业成果",
        project_dir / "3.执行资料" / "内业成果",
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_dir():
            return candidate
    return candidates[0]


def discover_photos_dir(project_dir: Path) -> Path:
    candidates = [
        project_dir / "3.执行资料" / "2.外业成果",
        project_dir / "3.执行资料" / "1.外业成果",
        project_dir / "3.执行资料" / "外业成果",
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_dir():
            return candidate
    return candidates[0]


def partition_sort_key(value: str) -> tuple[int, str]:
    match = re.fullmatch(r"[A-Za-z]", value)
    if match:
        return (0, value.upper())
    return (1, value)


def partition_label_from_dir_name(name: str) -> str:
    text = clean(name).replace(" ", "")
    patterns = [
        r"勘探区域([A-Za-z])区?$",
        r"勘探分区([A-Za-z])区?$",
        r"分区([A-Za-z])区?$",
        r"^([A-Za-z])区$",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).upper()
    return ""


def infer_partition_labels_from_project(project_dir: Path) -> list[str]:
    labels: set[str] = set()
    roots = [discover_photos_dir(project_dir), discover_drawings_dir(project_dir)]
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_dir():
                continue
            label = partition_label_from_dir_name(path.name)
            if label:
                labels.add(label)
    return sorted(labels, key=partition_sort_key)


def infer_partition_fields_from_project(fields: dict[str, str], project_dir: Path) -> None:
    labels = infer_partition_labels_from_project(project_dir)
    if len(labels) < 2:
        return
    fields["是否存在勘探分区"] = "有"
    fields["勘探分区数量"] = str(len(labels))


def read_docx_texts(path: Path) -> list[str]:
    try:
        with zipfile.ZipFile(path) as docx:
            root = ET.fromstring(docx.read("word/document.xml"))
    except Exception:
        return []
    out: list[str] = []
    for para in root.findall(".//w:p", NS):
        text = "".join(node.text or "" for node in para.findall(".//w:t", NS))
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            out.append(text)
    return out


def normalize_doc_number(value: str) -> str:
    text = clean(value)
    text = text.replace("(", "〔").replace(")", "〕").replace("（", "〔").replace("）", "〕")
    text = re.sub(r"\s+", "", text)
    return text


def extract_document_number(text: str) -> str:
    patterns = [
        r"([\u4e00-\u9fff]{1,12}函[〔(（]\s*20\d{2}\s*[〕)）]\s*\d+\s*号)",
        r"([\u4e00-\u9fff]{1,12}字[〔(（]\s*20\d{2}\s*[〕)）]\s*\d+\s*号)",
        r"([\u4e00-\u9fff]{1,12}发[〔(（]\s*20\d{2}\s*[〕)）]\s*\d+\s*号)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return normalize_doc_number(match.group(1))
    return ""


def pdf_text_with_pdftotext(path: Path) -> str:
    if not shutil.which("pdftotext"):
        return ""
    try:
        result = subprocess.run(
            ["pdftotext", str(path), "-"],
            check=False,
            capture_output=True,
            text=True,
            timeout=30,
        )
    except Exception:
        return ""
    if result.returncode != 0:
        return ""
    return clean(result.stdout)


def tesseract_command(image: Path) -> list[list[str]]:
    commands: list[list[str]] = []
    tessdata_candidates = [
        Path("/opt/homebrew/Cellar/tesseract-lang/4.1.0/share/tessdata"),
        Path("/opt/homebrew/share/tessdata"),
    ]
    for tessdata in tessdata_candidates:
        if tessdata.exists():
            commands.append([
                "tesseract",
                "--tessdata-dir",
                str(tessdata),
                str(image),
                "stdout",
                "-l",
                "chi_sim",
                "--psm",
                "6",
            ])
    commands.append(["tesseract", str(image), "stdout", "-l", "chi_sim+eng", "--psm", "6"])
    return commands


def pdf_text_with_ocr(path: Path, *, max_pages: int = 8) -> str:
    if not shutil.which("pdftoppm") or not shutil.which("tesseract"):
        return ""
    with tempfile.TemporaryDirectory() as tmp:
        prefix = Path(tmp) / "page"
        try:
            subprocess.run(
                ["pdftoppm", "-f", "1", "-l", str(max_pages), "-png", "-r", "180", str(path), str(prefix)],
                check=True,
                capture_output=True,
                timeout=60,
            )
        except Exception:
            return ""
        texts: list[str] = []
        for image in sorted(Path(tmp).glob("page-*.png")):
            for command in tesseract_command(image):
                try:
                    result = subprocess.run(
                        command,
                        check=False,
                        capture_output=True,
                        text=True,
                        timeout=60,
                    )
                except Exception:
                    continue
                if result.returncode == 0 and clean(result.stdout):
                    texts.append(result.stdout)
                    break
        return clean("\n".join(texts))


def pdf_text_with_rapidocr(path: Path, *, max_pages: int = 8) -> str:
    python_candidates = [Path(sys.executable)]
    if shutil.which("python3"):
        python_candidates.append(Path(shutil.which("python3") or "python3"))
    code = r"""
import json
import sys
from pathlib import Path

import fitz
from rapidocr_onnxruntime import RapidOCR
import tempfile

pdf = Path(sys.argv[1])
max_pages = int(sys.argv[2])
doc = fitz.open(pdf)
ocr = RapidOCR()
texts = []
with tempfile.TemporaryDirectory() as tmp:
    tmp_dir = Path(tmp)
    for idx in range(min(max_pages, doc.page_count)):
        page = doc.load_page(idx)
        pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
        image = tmp_dir / f"page_{idx + 1}.png"
        pix.save(image)
        result, _ = ocr(str(image))
        texts.append("\n".join(item[1] for item in (result or [])))
print("__RAPIDOCR_JSON__" + json.dumps({"text": "\n".join(texts)}, ensure_ascii=False))
"""
    for python in dict.fromkeys(str(python) for python in python_candidates):
        try:
            result = subprocess.run(
                [python, "-c", code, str(path), str(max_pages)],
                check=False,
                capture_output=True,
                text=True,
                timeout=180,
            )
        except Exception:
            continue
        if result.returncode != 0:
            continue
        marker = "__RAPIDOCR_JSON__"
        output = result.stdout
        if marker in output:
            output = output.rsplit(marker, 1)[1]
        match = re.search(r"\{.*\}", output, flags=re.S)
        if match:
            output = match.group(0)
        try:
            text = output
            for _ in range(3):
                parsed = json.loads(text)
                if not isinstance(parsed, dict) or "text" not in parsed:
                    break
                next_text = clean(parsed.get("text", ""))
                if next_text == text:
                    break
                text = next_text
                if not (text.startswith("{") and '"text"' in text):
                    break
            return clean(text)
        except Exception:
            return clean(output)
    return ""


def pdf_text_with_paddleocr(path: Path, *, max_pages: int = 8) -> str:
    if not shutil.which("pdftoppm"):
        return ""
    python_candidates = [Path(sys.executable)]
    if shutil.which("python3"):
        python_candidates.append(Path(shutil.which("python3") or "python3"))
    configured_ocr_python = clean(os.environ.get("LOCAL_OCR_PYTHON"))
    if configured_ocr_python:
        local_ocr_python = Path(configured_ocr_python).expanduser()
        if local_ocr_python.exists():
            python_candidates.append(local_ocr_python)
    code = r"""
import json
import sys
from pathlib import Path

from paddleocr import PaddleOCR

pdf = Path(sys.argv[1])
image_dir = Path(sys.argv[2])
try:
    ocr = PaddleOCR(use_textline_orientation=True, lang="ch")
except TypeError:
    ocr = PaddleOCR(use_angle_cls=True, lang="ch")
texts = []
for image in sorted(image_dir.glob("page-*.png")):
    try:
        result = ocr.ocr(str(image))
    except TypeError:
        result = ocr.ocr(str(image), cls=True)
    page_texts = []
    if isinstance(result, dict):
        for key in ["rec_texts", "texts"]:
            values = result.get(key)
            if isinstance(values, list):
                page_texts.extend(str(value) for value in values)
    for block in result or []:
        if isinstance(block, dict):
            for key in ["rec_texts", "texts"]:
                values = block.get(key)
                if isinstance(values, list):
                    page_texts.extend(str(value) for value in values)
        elif isinstance(block, list):
            for item in block:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    value = item[1]
                    if isinstance(value, (list, tuple)) and value:
                        page_texts.append(str(value[0]))
                    elif isinstance(value, str):
                        page_texts.append(value)
    texts.append("\n".join(page_texts))
print("__PADDLEOCR_JSON__" + json.dumps({"text": "\n".join(texts)}, ensure_ascii=False))
"""
    with tempfile.TemporaryDirectory() as tmp:
        prefix = Path(tmp) / "page"
        try:
            subprocess.run(
                ["pdftoppm", "-f", "1", "-l", str(max_pages), "-png", "-r", "220", str(path), str(prefix)],
                check=True,
                capture_output=True,
                timeout=60,
            )
        except Exception:
            return ""
        for python in dict.fromkeys(str(python) for python in python_candidates):
            try:
                result = subprocess.run(
                    [python, "-c", code, str(path), str(Path(tmp))],
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=600,
                    env={**os.environ, "PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK": "True"},
                )
            except Exception:
                continue
            if result.returncode != 0:
                continue
            marker = "__PADDLEOCR_JSON__"
            output = result.stdout
            if marker in output:
                output = output.rsplit(marker, 1)[1]
            match = re.search(r"\{.*\}", output, flags=re.S)
            if match:
                output = match.group(0)
            try:
                parsed = json.loads(output)
                return clean(parsed.get("text", ""))
            except Exception:
                return clean(output)
    return ""


def read_pdf_text(path: Path, *, allow_ocr: bool = True) -> str:
    text = pdf_text_with_pdftotext(path)
    if len(re.findall(r"[\u4e00-\u9fff]", text)) >= 30:
        return text
    if not allow_ocr:
        return text
    paddle_text = pdf_text_with_paddleocr(path)
    if len(re.findall(r"[\u4e00-\u9fff]", paddle_text)) >= 30:
        return paddle_text
    rapid_text = pdf_text_with_rapidocr(path)
    if len(re.findall(r"[\u4e00-\u9fff]", rapid_text)) >= 30:
        return rapid_text
    return pdf_text_with_ocr(path)


def review_reply_candidate_score(path: Path, text: str = "") -> int:
    haystack = f"{path.parent.name} {path.name} {text[:12000]}"
    score = 0
    positive = {
        "文物局": 8,
        "文化和旅游局": 12,
        "文旅局": 10,
        "文物部门": 6,
        "文物审查意见": 12,
        "审查意见": 8,
        "选址范围": 7,
        "用地范围": 5,
        "文物事宜": 4,
        "文物保护": 4,
        "文物法": 6,
        "文物保护条例": 6,
        "古文化遗存": 8,
        "古文化遗址遗迹": 8,
        "不可移动文物": 5,
        "复函": 8,
        "的函": 3,
        "函": 2,
    }
    negative = {
        "请示": -14,
        "申请": -6,
        "营业执照": -20,
        "身份证": -20,
        "坐标": -10,
        "宗地图": -10,
        "合同": -8,
        "kml": -8,
        "拐点": -8,
    }
    for token, value in positive.items():
        if token in haystack:
            score += value
    for token, value in negative.items():
        if token == "请示" and "复函" in path.name:
            continue
        if token.lower() in haystack.lower():
            score += value
    if re.search(r"[\u4e00-\u9fff]{1,12}函[〔(（]\s*20\d{2}\s*[〕)）]\s*\d+\s*号", haystack):
        score += 8
    if "关于" in haystack and "文物" in haystack and "函" in haystack:
        score += 6
    return score


def should_rescore_review_candidate(path: Path, base_score: int) -> bool:
    name = path.name
    if base_score >= 12:
        return True
    if path.suffix.lower() != ".pdf":
        return False
    positive_hits = sum(1 for token in ["文旅", "文化旅游", "文化和旅游", "文物", "复函", "函"] if token in name)
    if positive_hits >= 2 and "关于" in name:
        return True
    if "文物" in name and any(token in name for token in ["复函", "函"]):
        return True
    return bool(re.search(r"函[〔(（]\s*20\d{2}\s*[〕)）]\s*\d+\s*号", name)) and any(
        token in name for token in ["文旅", "文化", "文物", "旅游"]
    )


def find_cultural_relic_review_reply(project_dir: Path, *, read_text: bool = True) -> tuple[Path | None, str]:
    search_roots = [project_dir / "1.项目资料", project_dir]
    candidates: list[Path] = []
    seen: set[Path] = set()
    for root in search_roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.name.startswith(".~") or not path.is_file():
                continue
            if path.suffix.lower() not in {".pdf", ".docx"}:
                continue
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                candidates.append(path)
    rescored: list[tuple[int, Path, str]] = []
    for path in candidates:
        base_score = review_reply_candidate_score(path)
        text = ""
        score = base_score
        if read_text and should_rescore_review_candidate(path, base_score):
            if path.suffix.lower() == ".pdf":
                text = read_pdf_text(path)
            elif path.suffix.lower() == ".docx":
                text = "\n".join(read_docx_texts(path))
            if text:
                score = review_reply_candidate_score(path, text)
        rescored.append((score, path, text))
    scored = sorted(rescored, key=lambda row: (row[0], row[1].name), reverse=True)
    if not scored or scored[0][0] < 12:
        return None, ""
    path = scored[0][1]
    text = scored[0][2]
    if not read_text:
        return path, text
    if path.suffix.lower() == ".pdf":
        text = text or read_pdf_text(path)
    elif path.suffix.lower() == ".docx":
        text = text or "\n".join(read_docx_texts(path))
    return path, text


def split_review_opinion_sentences(text: str) -> list[str]:
    normalized = re.sub(r"\s+", "", clean(text))
    if not normalized:
        return []
    for marker in ["意见如下：", "意见如下:", "审查意见如下：", "函复如下："]:
        if marker in normalized:
            normalized = normalized.split(marker, 1)[1]
            break
    normalized = re.sub(r"附件[:：].*$", "", normalized)
    raw_items = re.split(r"[。；;]|[一二三四五六七八九十]+[、.．]", normalized)
    sentences: list[str] = []
    for item in raw_items:
        sentence = item.strip(" ，,：:")
        if not sentence:
            continue
        if any(token in sentence for token in ["文物", "遗址", "遗迹", "避让", "考古勘探", "开工建设", "审批"]):
            if len(sentence) >= 8:
                sentences.append(sentence.rstrip("。") + "。")
    return sentences


def extract_review_conclusion(text: str) -> str:
    normalized = re.sub(r"\s+", "", clean(text))
    if not normalized:
        return ""
    if "古文化遗存" in normalized and "古文化遗址遗迹保护范围" in normalized:
        county_match = re.search(r"不在([^，。；;]{2,30}?古文化遗址遗迹保护范围)内", normalized)
        protected_range = county_match.group(1) if county_match else "古文化遗址遗迹保护范围"
        county_name_match = re.search(r"([\u4e00-\u9fff]{2,12}(?:旗|县|区|市))文化和旅游局", normalized)
        if county_name_match:
            protected_range = protected_range.replace("我旗", county_name_match.group(1))
        if "察右后旗" in normalized:
            protected_range = protected_range.replace("我旗", "察哈尔右翼后旗")
        return f"该项目选址范围内地表未发现古文化遗存，不在{protected_range}内。"
    if "项目选址范围内未发现古文化遗迹" in normalized:
        return "项目选址范围内未发现古文化遗迹。"
    opinion_sentences = split_review_opinion_sentences(normalized)
    key_sentences = [
        sentence for sentence in opinion_sentences
        if any(token in sentence for token in ["未涉及", "不涉及", "未发现", "重叠", "涉及", "避让", "考古勘探", "未经审批不得开工"])
    ]
    if key_sentences:
        return "".join(key_sentences[:3])
    opinion_text = normalized
    marker = "意见如下："
    has_opinion_marker = marker in opinion_text
    if marker in opinion_text:
        opinion_text = opinion_text.split(marker, 1)[1]
        opinion_text = re.sub(r"^[一二三四五六七八九十]+[、.．]", "", opinion_text)
    if has_opinion_marker:
        for start_token in ["经与", "该项目", "项目选址范围", "项目用地范围"]:
            start = opinion_text.find(start_token)
            if start >= 0 and any(token in opinion_text[start : start + 180] for token in ["文物", "遗址", "遗迹"]):
                end_match = re.search(r"[。；;]", opinion_text[start:])
                if end_match:
                    return opinion_text[start : start + end_match.end()].rstrip("；;") + (
                        "" if opinion_text[start + end_match.end() - 1] == "。" else "。"
                    )
    sentence_pattern = r"[^。；;]*?(?:未涉及|不涉及|未发现|不在|涉及|发现|重叠|避让|应避让|原则同意|同意)[^。；;]*(?:。|；|;)"
    for match in re.finditer(sentence_pattern, opinion_text):
        sentence = match.group(0).strip("；;")
        if "文物" in sentence or "遗址" in sentence or "遗迹" in sentence:
            sentence = re.sub(r"^.*?报告如下[:：]?[一二三四五六七八九十]+[、.．]", "", sentence)
            return sentence.rstrip("。") + "。"
    return ""


def extract_review_title(text: str, project_name: str = "") -> str:
    normalized = re.sub(r"\s+", "", clean(text))
    if not normalized:
        return ""
    if (
        project_name
        and "察哈尔右翼后旗" in project_name
        and "文化和旅游局" in normalized
        and "文物调查情况" in normalized
        and "复函" in normalized
    ):
        return f"察哈尔右翼后旗文化和旅游局关于{project_name}选址文物调查情况的复函"
    match = re.search(r"([\u4e00-\u9fff]{2,20}(?:文化和旅游局|文物局)关于[^。；;]{8,120}?复函)", normalized)
    if match:
        return match.group(1)
    if project_name and "文物调查情况" in normalized and "复函" in normalized:
        agency_match = re.search(r"([\u4e00-\u9fff]{2,20}(?:文化和旅游局|文物局))", normalized)
        if agency_match:
            return f"{agency_match.group(1)}关于{project_name}选址文物调查情况的复函"
    return ""


def candidate_docx_files(project_dir: Path) -> list[Path]:
    preferred = project_dir / "3.执行资料" / "4.报告"
    roots = [preferred, project_dir]
    seen: set[Path] = set()
    out: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*.docx"):
            if path.name.startswith(".~"):
                continue
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                out.append(path)
    return out


def project_keywords(project_dir: Path) -> list[str]:
    name = re.sub(r"^\d{6,8}\s*", "", project_dir.name)
    parts = [item for item in re.split(r"[\s_\-（）()，,、]+", name) if len(item) >= 2]
    compact = re.sub(r"[项目工程]", "", name)
    if len(compact) >= 2:
        parts.append(compact)
    if "叁山" in name:
        parts.append("叁山")
    return list(dict.fromkeys(parts))


def text_matches_project(texts: list[str], project_dir: Path) -> bool:
    joined = "\n".join(texts[:80])
    keywords = project_keywords(project_dir)
    return not keywords or any(keyword and keyword in joined for keyword in keywords)


def infer_from_docx(project_dir: Path) -> dict[str, str]:
    fields: dict[str, str] = {}
    label_map = {
        "项目名称": "项目名称",
        "建设单位": "建设单位",
        "勘探单位": "勘探单位",
        "项目位置": "项目位置",
        "项目地理坐标": "项目地理坐标",
        "地理坐标": "项目地理坐标",
        "项目面积": "项目面积",
        "调查面积": "调查面积",
        "勘探面积": "勘探面积",
        "勘探时间": "勘探时间",
        "遗迹": "遗迹结论",
        "遗迹现象": "遗迹结论",
    }
    for docx in candidate_docx_files(project_dir):
        texts = read_docx_texts(docx)
        if not text_matches_project(texts, project_dir):
            continue
        for text in texts[:30]:
            if "考古调查" in text and "勘探工作报告" in text and "项目名称" not in fields:
                title = re.sub(r"考古调查[、,，]?\s*勘探工作报告.*$", "", text).strip()
                if title:
                    fields["项目名称"] = title
        for text in texts:
            normalized = text.replace("：", ":")
            for label, key in label_map.items():
                prefix = label + ":"
                if normalized.startswith(prefix) and key not in fields:
                    value = normalized[len(prefix) :].strip()
                    if value:
                        fields[key] = value
            if "地理位置" in normalized and "项目地理坐标" not in fields:
                match = re.search(
                    r"(?:地理位置|地理坐标)\s*[:：]\s*"
                    r"((?:东经|E)\s*[^，,、。；;\s]+(?:[，,、]\s*)?(?:北纬|N)\s*[^，,、。；;\s]+)",
                    text,
                )
                if match:
                    fields["项目地理坐标"] = match.group(1).replace(",", "、").replace("，", "、")
    return fields


def infer_from_site_records(project_dir: Path) -> dict[str, str]:
    fields: dict[str, str] = {}
    execute_dir = project_dir / "3.执行资料"
    if not execute_dir.exists():
        return fields
    for docx in sorted(execute_dir.rglob("*现场记录*.docx")):
        if docx.name.startswith(".~") or "模板" in docx.name:
            continue
        texts = read_docx_texts(docx)
        joined = "\n".join(texts)
        patterns = {
            "项目名称": r"项目(?:名字|名称)[:：][ \t]*([^\n]+)",
            "建设单位": r"建设单位[:：][ \t]*([^\n]+)",
            "勘探单位": r"勘探单位[:：][ \t]*([^\n]+)",
            "项目位置": r"(?:勘探地点|项目位置)[:：][ \t]*([^\n]+)",
            "项目地理坐标": r"(?:地理位置坐标经纬度（中心点）|项目地理坐标)[:：][ \t]*([^\n]+)",
            "项目面积": r"项目面积[:：]\s*([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)",
            "调查面积": r"(?:考古调查面积|调查面积)[:：]?\s*([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)",
            "勘探面积": r"(?:考古勘探面积|勘探面积)[:：]?\s*([0-9,.，]+)\s*(?:㎡|平方米|m2|m²)",
            "项目建设内容": r"项目建设内容[:：][ \t]*([^\n]+)",
            "工作天数": r"工作天数[:：]\s*([0-9]+)",
            "项目地块情况": r"项目地块情况[:：][ \t]*([^\n]+)",
            "是否存在不可勘探区域": r"是否存在不可勘探区域[:：][ \t]*([^\n]+)",
            "不可勘探原因": r"不可勘探原因[:：][ \t]*([^\n]+)",
            "进场人数": r"进场人数[:：]\s*([0-9]+)",
            "勘探单元数量": r"勘探单元数量[:：]\s*([0-9]+)",
            "勘探单元规格": r"勘探单元规格[:：]\s*([0-9]+(?:米×[0-9]+米)?)",
            "是否存在勘探分区": r"是否存在勘探分区[:：]\s*(.+)",
            "勘探分区数量": r"勘探分区数量[:：]\s*([0-9]*)",
            "剖线数量": r"剖线数量[:：]\s*([0-9]+)",
            "标准孔数量": r"标准孔数量[:：]\s*([0-9]+)",
            "遗迹结论": r"遗迹结论[:：]\s*([^\n]+)",
            "遗迹数量": r"遗迹数量[:：]\s*([0-9]+)",
            "探工数量": r"勘探人数[:：]?\s*([0-9]+)",
            "探孔总数": r"探孔总数[:：]?\s*([0-9,，]+)",
        }
        for key, pattern in patterns.items():
            if key in fields:
                continue
            match = re.search(pattern, joined)
            if match:
                value = match.group(1).strip(" ，,。")
                if key in {"项目面积", "调查面积", "勘探面积"} and not value.endswith("平方米"):
                    value = value.replace(",", "").replace("，", "") + "平方米"
                if key == "勘探单元规格" and re.fullmatch(r"\d+", value):
                    value = f"{value}米×{value}米"
                fields[key] = value
        if clean(fields.get("遗迹数量")) == "0" and "遗迹结论" not in fields:
            fields["遗迹结论"] = "未发现遗迹"
        if "项目建设内容" not in fields:
            match = re.search(r"建设内容(?:用途)?为([^。；;\n]+)", joined)
            if match:
                fields["项目建设内容"] = match.group(1).strip()
        if "工作天数" not in fields:
            days = inclusive_work_days_from_range(clean(fields.get("勘探时间")) or joined)
            if days:
                fields["工作天数"] = days
        if "勘探时间" not in fields:
            match = re.search(r"勘探时间[:：]?\s*([0-9]{4})\s*年\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日\s*[—\\-至到]+\s*([0-9]{4})?\s*年?\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日", joined)
            if match:
                start_year, start_month, start_day, end_year, end_month, end_day = match.groups()
                end_year = end_year or start_year
                fields["开始日期"] = f"{int(start_year)}年{int(start_month)}月{int(start_day)}日"
                fields["结束日期"] = f"{int(end_year)}年{int(end_month)}月{int(end_day)}日"
                fields["勘探时间"] = f"{fields['开始日期']}至{fields['结束日期']}"
                if "工作天数" not in fields:
                    fields["工作天数"] = inclusive_work_days_from_range(fields["勘探时间"])
        if "是否存在勘探分区" not in fields and "勘探分区" not in joined:
            fields["是否存在勘探分区"] = "无"
            fields.setdefault("勘探分区数量", "0")
        if "遗迹数量" not in fields and re.search(r"未发现(?:文化遗存|遗迹|文物)", joined):
            fields["遗迹数量"] = "0"
        relic_ids = sorted(
            {
                smart_filler.canonical_relic_id(match.group(1))
                for match in re.finditer(
                    rf"(?:灰坑|墓葬|遗迹|房址|窑址|窑坑|活土坑|[\u4e00-\u9fff、，,\s]*?(?:灰坑|墓葬|房址|窑址|窑坑|活土坑)?)?\s*({RELIC_ID_RE})\s*[：:]",
                    joined,
                    flags=re.I,
                )
            },
            key=smart_filler.relic_sort_key,
        )
        if relic_ids and "遗迹结论" not in fields and clean(fields.get("遗迹数量")) != "0":
            fields["遗迹结论"] = f"发现遗迹{len(relic_ids)}处"
        if fields:
            break
    return fields


def infer_survey_unit(project_dir: Path) -> str:
    company_names = [
        "内蒙古煊迹考古勘探有限公司",
        "北京卓凡文博技术有限公司",
        "内蒙古峰驰考古勘探有限公司",
        "三门峡市文物考古勘探有限公司",
        "河南燧火文物保护有限公司",
    ]
    for name in company_names:
        if (project_dir / name).exists():
            return name
    for path in project_dir.iterdir() if project_dir.exists() else []:
        if path.is_dir() and any(token in path.name for token in ["考古", "文博", "文物", "文保"]):
            return path.name
    return ""


def infer_area_from_tables(project_dir: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    drawings = discover_drawings_dir(project_dir)
    unit_tables = list(drawings.rglob("勘探单元坐标.xlsx")) if drawings.exists() else []
    if unit_tables:
        try:
            from openpyxl import load_workbook as load_xlsx

            wb = load_xlsx(unit_tables[0], data_only=True)
            ws = wb[wb.sheetnames[0]]
            units = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    units.add(str(row[0]).strip())
            if units:
                out["勘探单元数量"] = str(len(units))
        except Exception:
            pass
    return out


def infer_cultural_relic_context(project_dir: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    review, review_text = find_cultural_relic_review_reply(project_dir)
    if review:
        project_name = infer_from_site_records(project_dir).get("项目名称", "")
        title = extract_review_title(review_text, project_name) or review.stem.strip()
        out["文物概况类型"] = "有文物审查意见"
        out["文物审查意见文件名"] = title
        out["文物审查意见文号"] = extract_document_number(review.name) or extract_document_number(review_text)
        out["文物审查意见结论"] = extract_review_conclusion(review_text)
        out["回函文件路径"] = str(review)
    stats = discover_drawings_dir(project_dir) / "文物内外遗迹统计表.xlsx"
    if stats.exists():
        try:
            wb = load_workbook(stats, data_only=True)
            ws = wb[wb.sheetnames[0]]
            locations: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                location = clean(row[0] if row else "")
                relic = clean(row[1] if row and len(row) > 1 else "")
                if location and relic and location != "外":
                    locations.add(location)
            if locations:
                out["涉及文物名称及情况"] = "、".join(sorted(locations))
                out.setdefault("文物概况类型", "无回函但涉及文物")
        except Exception:
            pass
    return out


def infer_fields(project_dir: Path) -> dict[str, str]:
    fields = infer_from_docx(project_dir)
    site_fields = infer_from_site_records(project_dir)
    fields.update({key: value for key, value in site_fields.items() if value})
    cultural_fields = infer_cultural_relic_context(project_dir)
    fields.update({key: value for key, value in cultural_fields.items() if value and not fields.get(key)})
    fields.setdefault("项目名称", re.sub(r"^\d{6,8}\s*", "", project_dir.name).strip())
    fields.setdefault("勘探单位", infer_survey_unit(project_dir))
    fields.update({k: v for k, v in infer_area_from_tables(project_dir).items() if v})
    if "遗迹结论" not in fields:
        fields["遗迹结论"] = "未发现文化遗存"
    if "文物概况类型" not in fields:
        fields["文物概况类型"] = "无回函且未发现文物"
    return {key: value for key, value in fields.items() if value}


def fill_manual_template(template: Path, output: Path, fields: dict[str, str]) -> None:
    shutil.copy2(template, output)
    wb = load_workbook(output)
    for ws in wb.worksheets:
        headers = {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
        if "字段" not in headers:
            continue
        field_col = headers["字段"]
        input_col = headers.get("值") or headers.get("输入") or 2
        for row in range(2, ws.max_row + 1):
            key = clean(ws.cell(row, field_col).value)
            if key in fields:
                ws.cell(row, input_col).value = fields[key]
    wb.save(output)


def set_kv_value(wb, key: str, value: str) -> None:
    changed = False
    for ws in wb.worksheets:
        headers = {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
        if "字段" not in headers:
            continue
        field_col = headers["字段"]
        value_col = headers.get("值") or headers.get("输入") or 2
        for row in range(2, ws.max_row + 1):
            if clean(ws.cell(row, field_col).value) == key:
                ws.cell(row, value_col).value = value
                changed = True
    return None


def replace_sheet_rows(wb, sheet_name: str, records: list[dict[str, str]]) -> None:
    if sheet_name not in wb.sheetnames or not records:
        return
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for record in records:
        ws.append([record.get(header, "") for header in headers])


def workbook_fields(wb) -> dict[str, str]:
    fields: dict[str, str] = {}
    for ws in wb.worksheets:
        headers = {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
        if "字段" not in headers:
            continue
        field_col = headers["字段"]
        value_col = headers.get("值") or headers.get("输入") or 2
        for row in range(2, ws.max_row + 1):
            key = clean(ws.cell(row, field_col).value)
            value = clean(ws.cell(row, value_col).value)
            if key and value:
                fields[key] = value
    return fields


SITE_RECORD_AUDIT_FIELDS = [
    "项目名称",
    "建设单位",
    "项目位置",
    "项目地理坐标",
    "项目面积",
    "调查面积",
    "项目建设内容",
    "勘探面积",
    "勘探时间",
    "开始日期",
    "结束日期",
    "工作天数",
    "项目地块情况",
    "探孔总数",
    "剖线数量",
    "探工数量",
    "遗迹结论",
    "勘探成果综合结论",
]

KNOWN_SITE_RECORD_LABELS = {
    "项目名字",
    "项目名称",
    "建设单位",
    "勘探单位",
    "项目地点",
    "项目位置",
    "地理位置坐标经纬度",
    "地理位置坐标经纬度（中心点）",
    "项目地理坐标",
    "项目概况",
    "项目面积",
    "考古调查面积",
    "调查面积",
    "实际勘探面积",
    "考古勘探面积",
    "项目建设内容",
    "地貌",
    "勘探面积描述",
    "勘探剖线",
    "探孔总数",
    "勘探时间",
    "勘探人数",
    "记录人",
    "结论",
    "剖线A-A′",
    "剖线A-A'",
}


def normalize_audit_value(value: object) -> str:
    text = clean(value)
    text = text.replace("㎡", "平方米").replace("m²", "平方米").replace("m2", "平方米")
    text = re.sub(r"\s+", "", text)
    text = text.replace(",", "").replace("，", "")
    return text.strip("。；;,， ")


def site_record_label_values(project_dir: Path) -> tuple[dict[str, str], Path | None]:
    execute_dir = project_dir / "3.执行资料"
    if not execute_dir.exists():
        return {}, None
    for docx in sorted(execute_dir.rglob("*现场记录*.docx")):
        if docx.name.startswith(".~") or "模板" in docx.name:
            continue
        labels: dict[str, str] = {}
        for text in read_docx_texts(docx):
            match = re.match(r"^([^：:]{2,30})[：:]\s*(.+)$", clean(text))
            if match:
                label = clean(match.group(1))
                if re.match(r"^[①②③④⑤⑥⑦⑧⑨⑩\d]+[、.．]?\s*.*层$", label):
                    continue
                labels[label] = clean(match.group(2))
        return labels, docx
    return {}, None


def count_workbook_data_rows(wb, sheet_name: str) -> int:
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(clean(value) for value in row):
            count += 1
    return count


def audit_manual_prefill(project_dir: Path, workbook_path: Path) -> tuple[list[str], Path | None]:
    wb = load_workbook(workbook_path, data_only=True)
    form_fields = workbook_fields(wb)
    photos_dir = discover_photos_dir(project_dir)
    site_fields, site_source = smart_filler.read_site_record_fields(photos_dir)
    label_values, label_source = site_record_label_values(project_dir)
    source = Path(site_source) if site_source else label_source
    if not source:
        return [], None

    issues: list[str] = []
    notes: list[str] = []
    for key in SITE_RECORD_AUDIT_FIELDS:
        source_value = clean(site_fields.get(key))
        if not source_value:
            continue
        form_value = clean(form_fields.get(key))
        if not form_value:
            issues.append(f"现场记录有 `{key}`，但人工填写表为空：{source_value}")
            continue
        if normalize_audit_value(source_value) != normalize_audit_value(form_value):
            issues.append(f"现场记录 `{key}` 与人工填写表不一致：现场记录={source_value}；人工表={form_value}")

    site_time = clean(site_fields.get("勘探时间")) or clean(label_values.get("勘探时间"))
    inferred_days = inclusive_work_days_from_range(site_time)
    if inferred_days:
        form_days = clean(form_fields.get("工作天数"))
        if not form_days:
            issues.append(f"现场记录勘探时间可推导 `工作天数`，但人工填写表为空：应为 {inferred_days} 天")
        elif normalize_audit_value(form_days) != inferred_days:
            issues.append(f"现场记录勘探时间与人工填写表 `工作天数` 不一致：勘探时间推导={inferred_days} 天；人工表={form_days}")

    unknown_labels = sorted(label for label in label_values if label not in KNOWN_SITE_RECORD_LABELS)
    if unknown_labels:
        issues.append("现场记录存在未纳入预填映射的标签：" + "、".join(unknown_labels[:12]))

    if source and source.exists():
        joined = "\n".join(read_docx_texts(source))
        tk_ids = sorted(set(re.findall(r"\bTK\s*0*\d+\b", joined, flags=re.I)))
        standard_rows = count_workbook_data_rows(wb, "标准孔")
        if tk_ids and standard_rows and len(tk_ids) != standard_rows:
            notes.append(
                f"现场记录标准孔 {len(tk_ids)} 个，人工填写表标准孔 {standard_rows} 条；"
                "现场记录属于外业成果，内业可按适合点位生成图纸和表格，此差异不作为预填错误。"
            )

    project_label = clean(form_fields.get("项目名称")) or workbook_path.stem.replace("_人工填写表", "")
    check_dir = PROCESS_DIR / safe_filename(project_label) / "生成检查"
    check_dir.mkdir(parents=True, exist_ok=True)
    check_path = check_dir / f"{safe_filename(project_label)}_人工填写表预填检查.md"
    lines = [
        "# 人工填写表预填检查",
        "",
        f"- 项目目录：{project_dir}",
        f"- 人工填写表：{workbook_path}",
        f"- 现场记录：{source}",
        "",
        "## 检查结论",
    ]
    if issues:
        lines.extend(f"- {item}" for item in issues)
    else:
        lines.append("- 未发现现场记录关键字段漏填或明显不一致。")
    if notes:
        lines.extend(["", "## 资料说明"])
        lines.extend(f"- {item}" for item in notes)
    check_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return issues, check_path


def build_image_directory_rows(project_dir: Path) -> list[dict[str, str]]:
    drawings = discover_drawings_dir(project_dir)
    photos_root = discover_photos_dir(project_dir)
    rows: list[dict[str, str]] = []
    candidates = [
        ("制图成果", drawings / "1.位置图", "项目地块位置示意图/项目地块卫星图", "是"),
        ("制图成果", drawings / "2.红线四至坐标、勘探区域", "红线四至坐标图/项目勘探区域示意图", "是"),
        ("制图成果", drawings, "项目勘探分区示意图", "有分区时必需"),
        ("制图成果", drawings / "3.勘探单元", "勘探单元布设示意图", "是"),
        ("制图成果", drawings / "4.探孔标准孔", "标准探孔位置示意图/标准孔位置图", "是"),
        ("制图成果", drawings / "5.剖线", "剖线位置示意图/剖线图", "是"),
        ("制图成果", drawings / "遗迹分布示意图", "遗迹分布示意图", "有遗迹时必需"),
        ("制图成果", drawings / "单个遗迹", "遗迹平、剖面图", "有遗迹时必需"),
    ]
    for label, path, slot, required in candidates:
        if path.exists():
            rows.append({"资料类型": label, "目录/文件": str(path), "匹配占位符": slot, "是否必需": required, "备注": ""})
    if photos_root.exists():
        for path in sorted(photos_root.rglob("*")):
            if not path.is_dir():
                continue
            name = path.name
            slot = ""
            if "标准孔照" in name:
                slot = "标准孔土样照"
            elif "遗迹" in str(path):
                slot = "遗迹土样照/遗迹现场照"
            elif "实地踏查" in name:
                slot = "实地踏查照"
            elif "地块现状" in name:
                slot = "项目地块现状照"
            elif "普探工作" in name:
                slot = "普探工作照"
            elif "布设探孔" in name:
                slot = "布设探孔照"
            elif "勘探后航拍" in name or "航拍" in name:
                slot = "勘探后航拍照"
            elif "勘探后" in name:
                slot = "勘探后局部照"
            elif "资料整理" in name:
                slot = "资料整理工作照"
            if slot:
                rows.append({"资料类型": "外业照片", "目录/文件": str(path), "匹配占位符": slot, "是否必需": "按模板占位符", "备注": ""})
    return rows


def enrich_workbook_from_project(output: Path, project_dir: Path) -> None:
    wb = load_workbook(output)
    drawings_dir = discover_drawings_dir(project_dir)
    photos_dir = discover_photos_dir(project_dir)
    if not photos_dir.exists():
        photos_dir = next((p for p in photos_dir.parent.iterdir() if p.is_dir() and "勘探区域" in p.name), None)
    tables: dict[str, list[dict[str, str]]] = {}
    notes: list[str] = []
    notes.extend(smart_filler.import_external_tables(tables, drawings_dir))
    fields: dict[str, str] = workbook_fields(wb)
    site_fields, _site_source = smart_filler.read_site_record_fields(photos_dir)
    for key, value in site_fields.items():
        if not value:
            continue
        fields[key] = value
    notes.extend(smart_filler.import_site_record_tables(fields, tables, photos_dir))
    smart_filler.fill_coordinate_basepoint(fields, tables, notes)
    smart_filler.refresh_section_derived_fields(fields, tables, notes)
    if clean(fields.get("是否存在不可勘探区域")) in {"无", "否", "不存在"}:
        fields["不可勘探原因"] = ""
    infer_partition_fields_from_project(fields, project_dir)

    replace_sheet_rows(wb, "红线坐标", tables.get("红线坐标", []))
    replace_sheet_rows(wb, "勘探单元", tables.get("勘探单元", []))
    replace_sheet_rows(wb, "剖线地层堆积", tables.get("剖线地层堆积", []))
    replace_sheet_rows(wb, "标准孔", tables.get("标准孔", []))
    replace_sheet_rows(wb, "遗迹记录", tables.get("遗迹记录", []))
    replace_sheet_rows(wb, "遗迹坐标", tables.get("遗迹坐标", []))
    replace_sheet_rows(wb, "文物范围遗迹统计", tables.get("文物范围遗迹统计", []))
    replace_sheet_rows(wb, "图件照片目录", build_image_directory_rows(project_dir))

    for key, value in fields.items():
        set_kv_value(wb, key, value)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="根据项目资料目录生成并预填人工填写表")
    parser.add_argument("project_dir", type=Path)
    parser.add_argument("--template", type=Path, default=TEMPLATE)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    project_dir = args.project_dir.resolve()
    if not project_dir.exists():
        raise FileNotFoundError(project_dir)
    if not args.template.exists():
        raise FileNotFoundError(args.template)

    fields = infer_fields(project_dir)
    project_label = safe_filename(fields.get("项目名称") or project_dir.name)
    output = args.output or (project_output_dir(project_label) / f"{project_label}_人工填写表.xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    fill_manual_template(args.template, output, fields)
    enrich_workbook_from_project(output, project_dir)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

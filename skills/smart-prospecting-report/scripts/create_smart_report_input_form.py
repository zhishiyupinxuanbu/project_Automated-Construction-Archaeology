#!/usr/bin/env python3
"""Create a human-friendly workbook for smart report generation."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def find_project_root() -> Path:
    here = Path(__file__).resolve()
    candidates = [here.parent, *here.parents, Path.cwd(), *Path.cwd().resolve().parents]
    for candidate in candidates:
        if (
            (candidate / "智能生成报告技能资料" / "知识库" / "3.模板与表单").exists()
            and ((candidate / "过程资料").exists() or (candidate / "基础信息").exists())
        ):
            return candidate
    for candidate in candidates:
        if candidate.name == "智能生成报告技能资料":
            return candidate.parent
    return here.parents[1]


ROOT = find_project_root()
OUT = ROOT / "过程资料" / "智能报告信息填报表.xlsx"
FORM_TEMPLATE_DIR = ROOT / "智能生成报告技能资料" / "知识库" / "3.模板与表单" / "模板" / "表格模板"
MANUAL_TEMPLATE_OUT = FORM_TEMPLATE_DIR / "人工填写表模板.xlsx"


SHEETS = {
    "项目基础信息": [
        ("项目名称", "", "必填，报告标题和正文使用"),
        ("建设单位", "", "必填"),
        ("勘探单位", "北京卓凡文博技术有限公司", "用于匹配公司模板，不替换公司固定信息"),
        ("项目级别", "", "用于附件三项目用地范围坐标表插入判断：市级及市级以上 / 市级以下"),
        ("项目位置", "", "例：内蒙古自治区鄂尔多斯市伊金霍洛旗伊金霍洛镇"),
        ("项目地理坐标", "", "例：东经109°31′27″、北纬39°34′38″；用于正文地理位置句"),
        ("项目面积", "", "例：38797平方米"),
        ("调查面积", "", "默认等于项目面积；确实不同时再填写"),
        ("项目建设内容", "", "项目性质、建设内容或拟建工程概况"),
        ("勘探面积", "", "例：38797平方米"),
        ("勘探时间", "", "例：2025年11月27日至2025年11月28日"),
        ("开始日期", "", "例：2025年11月27日"),
        ("结束日期", "", "例：2025年11月28日"),
        ("工作天数", "", "例：2"),
        ("报告年月", "", "自动字段：留空即可，由结束日期推导为“YYYY年M月”"),
        ("遗迹结论", "未发现文化遗存", "无遗迹模板默认值"),
        ("遗迹数量", "", "有遗迹项目可由遗迹记录自动统计"),
    ],
    "项目区域概况": [
        ("项目所在地旗县", "", "例：伊金霍洛旗；可由项目位置自动识别，人工填写更稳"),
        ("项目所在地旗县地理位置概况", "", "可留空，由生成器联网检索后写入"),
        ("项目所在地旗县行政区划与社会经济概况", "", "可留空，由生成器联网检索后写入"),
        ("项目所在地旗县气候条件", "", "可留空，由生成器联网检索后写入"),
        ("项目所在地旗县历史沿革", "", "可留空，由生成器联网检索后写入"),
    ],
    "文物概况": [
        ("文物概况类型", "有文物审查意见", "三选一：有文物审查意见 / 无回函且未发现文物 / 无回函但涉及文物"),
        ("文物审查意见文件名", "", "例：伊金霍洛旗文物局关于……文物审查意见的函"),
        ("文物审查意见文号", "", "例：伊文物函〔2026〕206号"),
        ("文物审查意见结论", "", "例：该项目用地范围未涉及登记在册的不可移动文物"),
        ("涉及文物名称及情况", "", "无回函但涉及文物时填写"),
        ("回函文件路径", "", "可填 PDF 路径，后续生成器优先解析"),
    ],
    "勘探参数": [
        ("勘探单元规格", "400米×400米", "支持 50/100/200/400/500 米等"),
        ("坐标基点X", "", "地块西南基点 X"),
        ("坐标基点Y", "", "地块西南基点 Y"),
        ("勘探单元数量", "", "可由勘探单元表统计"),
        ("勘探单元编号范围", "", "例：U01、U02……U06"),
        ("探孔总数", "", "可由现场记录或探孔资料预填"),
        ("普探孔距", "2米×2米", ""),
        ("重点勘探孔距", "1米×1米", ""),
        ("普探列最大编号", "", "可由勘探单元规格自动计算"),
        ("普探行最大编号", "", "可由勘探单元规格自动计算"),
        ("重点勘探列最大编号", "", "可由勘探单元规格自动计算"),
        ("重点勘探行最大编号", "", "可由勘探单元规格自动计算"),
    ],
    "现场限制": [
        ("是否存在不可勘探区域", "否", "二选一：是 / 否"),
        ("不可勘探原因", "", "存在不可勘探区域时填写，如地表硬化、建筑物占压、管线密集等"),
    ],
    "人员构成": [
        ("项目负责", "", "仅当模板含{{项目负责}}替换符时使用；无替换符则保留模板原文"),
        ("现场负责", "", "仅当模板含{{现场负责}}替换符时使用；无替换符则保留模板原文"),
        ("领队", "", "姓名或名单"),
        ("技师", "", "多人用顿号分隔；仅替换模板中的{{技师}}"),
        ("报告执笔", "", "仅当模板含{{报告执笔}}替换符时使用；无替换符则保留模板原文"),
        ("技师数量", "", "可自动统计"),
        ("探工人员描述", "", "例：韩冰、周保新"),
        ("探工数量", "", "填数字；进场人数=探工数量+8，每日在场人员下限=进场人数-5"),
        ("测绘员", "", ""),
        ("资料员", "", ""),
        ("校核", "", "仅当模板含{{校核}}替换符时使用；无替换符则保留模板原文"),
        ("现场负责人", "", ""),
        ("安全管理员", "", ""),
        ("后勤", "", ""),
        ("进场人数", "", "可自动计算"),
        ("每日在场人员下限", "", "可自动计算"),
        ("踏查技师数量", "", "可自动等于技师数量"),
        ("踏查探工数量", "", "可自动等于探工数量"),
    ],
    "自动生成字段": [
        ("报告年月", "", "无需填写：由结束日期或勘探时间推导；如需强制指定，可在此填写"),
        ("调查面积", "", "无需填写：默认等于项目面积；如确需不同，可在此填写"),
        ("剖线数量", "", "无需填写：由“剖线地层堆积”表统计"),
        ("标准孔数量", "", "无需填写：由“标准孔”表统计"),
        ("遗迹数量", "", "无需填写：由“遗迹记录”表统计"),
        ("勘探成果综合结论", "", "无需填写：由剖线地层、标准孔和遗迹结论综合生成；如需人工指定，可在此填写"),
    ],
}


TABLE_SHEETS = {
    "红线坐标": ["角点", "X坐标", "Y坐标", "备注"],
    "勘探单元": ["勘探单元", "角点", "X坐标", "Y坐标", "备注"],
    "剖线地层堆积": ["剖线编号", "剖线图", "剖线地层描述"],
    "标准孔": ["序号", "勘探单元", "标准孔编号", "X坐标", "Y坐标", "高程", "标准孔地层描述", "标准孔位置图", "标准孔土样照"],
    "遗迹记录": [
        "序号",
        "遗迹编号",
        "遗迹类型",
        "勘探单元",
        "遗址位置",
        "形制",
        "宽度（米）",
        "长度（米）",
        "口深（米）",
        "底深（米）",
        "面积（平方米）",
        "开口层位",
        "方向",
        "内部填充与包含物",
        "遗迹描述",
        "遗迹土样描述",
        "土样照",
        "现场照",
        "平、剖面图",
        "备注",
    ],
    "遗迹坐标": ["遗迹位置", "遗迹编号", "界点序号", "平面坐标X", "平面坐标Y", "经度", "纬度"],
    "文物范围遗迹统计": ["遗址位置", "遗迹编号", "备注"],
    "图件照片目录": ["资料类型", "目录/文件", "匹配占位符", "是否必需", "备注"],
    "图片清单": ["图片位", "文件名", "图题", "备注"],
}


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name, rows in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(["字段", "值", "填写说明"])
        for row in rows:
            ws.append(row)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 72
        style_sheet(ws)
        if name == "文物概况":
            dv = DataValidation(type="list", formula1='"有文物审查意见,无回函且未发现文物,无回函但涉及文物"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add("B2")
        if name == "现场限制":
            dv = DataValidation(type="list", formula1='"是,否"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add("B2")

    default_rows = {
        "红线坐标": 8,
        "勘探单元": 80,
        "剖线地层堆积": 12,
        "标准孔": 80,
        "遗迹记录": 80,
        "遗迹坐标": 520,
        "文物范围遗迹统计": 80,
        "图件照片目录": 40,
        "图片清单": 40,
    }
    for name, headers in TABLE_SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for _ in range(default_rows.get(name, 20)):
            ws.append([""] * len(headers))
        for idx, header in enumerate(headers, start=1):
            letter = ws.cell(1, idx).column_letter
            ws.column_dimensions[letter].width = 18 if len(header) <= 5 else 26
        if name in {"遗迹记录", "剖线地层堆积", "标准孔"}:
            for letter in ["O", "P", "Q", "R", "S", "T"]:
                if letter in ws.column_dimensions:
                    ws.column_dimensions[letter].width = 48
        style_sheet(ws)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="创建过程用智能报告信息填报表")
    parser.add_argument(
        "--update-template",
        action="store_true",
        help="已废弃：唯一人工填写表模板由人工维护，本脚本不再覆盖。",
    )
    args = parser.parse_args()

    wb = build_workbook()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)
    if args.update_template:
        print(f"跳过模板覆盖：唯一人工填写表模板由人工维护，请直接编辑 {MANUAL_TEMPLATE_OUT}")


if __name__ == "__main__":
    main()
